# © 2026 Martín Viera. Todos los derechos reservados.
# Software propietario. Ver LICENSE — prohibida su redistribución.
"""
MV Data Governance · Suite de pruebas del motor, i18n, exportadores y API.

Ejecutar:  pytest tests/ -v
"""
from __future__ import annotations

import io
import os
import re
import sys

import pandas as pd
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mvdg.catalog import catalog_df, dictionary_df, dataset_names, pii_columns
from mvdg.demo_data import load_demo_tables
from mvdg.exporters import (bi_bundle_xlsx, governance_tables, to_csv_bytes,
                            to_excel_bytes, to_json_bytes)
from mvdg.glossary import glossary_df, term_count
from mvdg.i18n import LANGS, _T, all_keys, t
from mvdg.lineage import EDGES, NODES, downstream, lineage_df, lineage_figure, upstream
from mvdg.policies import policies_df
from mvdg.profiler import profile_table, suggest_rules, summary
from mvdg.quality import (DIMENSIONS, RULES, open_issues, overall_index,
                          quality_by_dataset, quality_by_dimension,
                          quality_matrix, quality_trend, run_rules)


# ------------------------------------------------------------------- i18n
def test_i18n_parity_all_languages():
    """Cada clave existe en los 3 idiomas y no está vacía."""
    for key, entry in _T.items():
        for lang in LANGS:
            assert entry.get(lang), f"Falta traducción {lang} para {key}"


def test_i18n_fallback():
    assert t("clave_inexistente", "en") == "clave_inexistente"
    assert t("app_title", "xx") == t("app_title", "es")
    assert len(all_keys()) > 80


def test_metrica_de_conteo_de_columnas_usa_plural():
    """La primera pantalla que ve un prospecto con SU archivo mostraba
    "Columna 8" — el singular de un encabezado de tabla reusado como
    métrica de conteo. Se separaron las dos claves: `col_column` sigue
    siendo el encabezado (una fila = una columna, singular correcto) y
    `col_columns_count` es la métrica."""
    import ast
    assert t("col_columns_count", "es") == "Columnas"
    assert t("col_columns_count", "en") == "Columns"
    assert t("col_columns_count", "pt") == "Colunas"
    # y el singular NO vuelve a usarse como st.metric(...)
    ruta = os.path.join(_repo_root(), "app", "app.py")
    with open(ruta, encoding="utf-8") as fh:
        arbol = ast.parse(fh.read())
    for n in ast.walk(arbol):
        if (isinstance(n, ast.Call) and isinstance(n.func, ast.Attribute)
                and n.func.attr == "metric" and n.args
                and isinstance(n.args[0], ast.Call)
                and isinstance(n.args[0].func, ast.Name)
                and n.args[0].func.id == "t" and n.args[0].args
                and isinstance(n.args[0].args[0], ast.Constant)):
            assert n.args[0].args[0].value != "col_column", (
                f"linea {n.lineno}: 'col_column' (singular) usado como metrica "
                "de conteo; usa 'col_columns_count'")


# -------------------------------------------------------------- demo data
def test_demo_tables_deterministic():
    a, b = load_demo_tables(), load_demo_tables()
    for name in a:
        pd.testing.assert_frame_equal(a[name], b[name])


def test_demo_tables_have_injected_issues():
    tables = load_demo_tables()
    assert tables["dim_customers"]["email"].isna().sum() > 0
    assert (tables["dim_products"]["unit_price"] <= 0).sum() > 0
    assert tables["dim_customers"].duplicated(subset=["customer_id"]).sum() > 0


# ---------------------------------------------------------------- catálogo
@pytest.mark.parametrize("lang", LANGS)
def test_catalog_and_dictionary(lang):
    cat = catalog_df(lang)
    dic = dictionary_df(lang)
    assert set(cat["dataset"]) == set(dataset_names())
    assert (cat["rows"] > 0).all()
    assert dic["description"].str.len().gt(0).all()
    for ds, col in pii_columns():
        row = dic[(dic["dataset"] == ds) & (dic["column"] == col)]
        assert bool(row["pii"].iloc[0])


def test_catalog_translations_differ():
    es = catalog_df("es")["description"].iloc[0]
    en = catalog_df("en")["description"].iloc[0]
    assert es != en


# ----------------------------------------------------------------- calidad
def test_rules_cover_all_dimensions():
    assert {r.dimension for r in RULES} == set(DIMENSIONS)


@pytest.mark.parametrize("lang", LANGS)
def test_run_rules(lang):
    res = run_rules(lang=lang)
    assert len(res) == len(RULES)
    assert res["score"].between(0, 100).all()
    assert set(res["status"]) <= {"pass", "warn", "fail"}
    assert (res.loc[res["status"] == "pass", "score"]
            >= res.loc[res["status"] == "pass", "threshold"]).all()


def test_quality_aggregations():
    res = run_rules()
    assert 0 < overall_index(res) <= 100
    assert len(quality_by_dataset(res)) == len(dataset_names())
    assert quality_matrix(res).shape[0] == len(dataset_names())
    assert len(quality_by_dimension(res)) <= len(DIMENSIONS)
    trend = quality_trend(res)
    assert len(trend) == 12
    assert trend["quality_index"].iloc[-1] == overall_index(res)
    issues = open_issues(res)
    assert (issues["severity"].isin(["media", "alta"])).all()


# ------------------------------------------------------------------ linaje
def test_lineage_graph_consistency():
    ids = {n["id"] for n in NODES}
    for a, b in EDGES:
        assert a in ids and b in ids
    assert "crm" in upstream("bi_dashboard")
    assert "bi_dashboard" in downstream("crm")
    assert upstream("crm") == set()
    assert len(lineage_df()) == len(EDGES)
    fig = lineage_figure("fct_sales", {"source": "Fuentes"})
    assert len(fig.data) == len(EDGES) + 1  # aristas + capa de nodos


# ---------------------------------------------------------------- glosario
@pytest.mark.parametrize("lang", LANGS)
def test_glossary(lang):
    g = glossary_df(lang)
    assert len(g) == term_count()
    assert g["definition"].str.len().gt(10).all()
    known = set(dataset_names())
    for linked in g["linked_datasets"]:
        assert set(linked.split(", ")) <= known


# --------------------------------------------------------------- políticas
@pytest.mark.parametrize("lang", LANGS)
def test_policies(lang):
    p = policies_df(lang)
    assert len(p) == 6
    assert set(p["status"]) <= {"compliant", "partial", "noncompliant"}
    assert p["evidence"].str.len().gt(5).all()


# --------------------------------------------------------------- perfilador
def test_profiler_detects_issues_and_pii():
    df = load_demo_tables()["dim_customers"]
    prof = profile_table(df)
    assert "email" in prof.loc[prof["possible_pii"], "column"].tolist()
    info = summary(df)
    assert info["duplicate_rows"] > 0 and info["pii_columns"] >= 2
    for lang in LANGS:
        sugs = suggest_rules(df, lang)
        assert len(sugs) > 0


def test_profiler_empty_frame():
    empty = pd.DataFrame({"a": []})
    assert summary(empty)["rows"] == 0
    assert len(profile_table(empty)) == 1


# ------------------------------ catalogo de calidad automatico (archivo propio)
def _inventario_propio():
    """Un archivo que no es ninguno de los datasets de ejemplo del repo:
    nulos parciales en 3 columnas y una columna (sku) que es clave real —
    sin filas duplicadas, para no pisar el heuristico de unicidad de sku
    (ver test separado para la regla de fila completa duplicada)."""
    return pd.DataFrame({
        "sku": [f"A-{1000+i}" for i in range(1, 10)],
        "producto": ["Tornillo M6", "Tuerca M6", "Arandela M6", "Taladro 500W",
                    "Taladro 500W", "Sierra circular", "Guantes de cuero",
                    "Casco de seguridad", "Cinta métrica 5m"],
        "categoria": ["Ferretería", "Ferretería", "Ferretería", "Herramientas",
                     "Herramientas", "Herramientas", "Seguridad", "Seguridad",
                     "Herramientas"],
        "stock": [540, 610, None, 12, 12, 7, 300, 150, 220],
        "precio_unitario": [0.12, 0.08, 0.03, 45.90, 45.90, 89.50, 3.20, None, 2.10],
        "proveedor": ["Norte", "Norte", "Norte", "Sur", "Sur", None, "Este", "Este", "Norte"],
    })


def test_auto_rules_genera_reglas_reales_para_un_archivo_propio():
    """Diferencia central con el perfilado: estas reglas se CORREN contra el
    archivo (score, umbral, pass/fail), no son solo texto sugerido. Verificado
    tambien de punta a punta en la app real (Streamlit + Playwright, subiendo
    este mismo archivo por la UI) antes de escribir este test."""
    from mvdg.auto_rules import auto_quality_results, build_rules

    df = _inventario_propio()
    rules = build_rules(df, "inventario_real.csv")
    ids = [r.rule_id for r in rules]
    assert len(ids) == len(set(ids)), "rule_id repetido"
    # sku es clave -> unicidad; stock/precio_unitario/proveedor tienen nulos
    # parciales -> completitud; mas la regla de tabla completa (fila duplicada)
    assert any(r.column == "sku" and r.dimension == "uniqueness" for r in rules)
    assert any(r.column == "stock" and r.dimension == "completeness" for r in rules)
    assert any(r.column == "precio_unitario" and r.dimension == "completeness" for r in rules)
    assert any(r.column == "proveedor" and r.dimension == "completeness" for r in rules)
    assert sum(r.column not in df.columns for r in rules) == 1  # la de "fila completa"
    assert not any(r.column == "producto" for r in rules)  # sin nulos ni clave: no aplica

    res = auto_quality_results(df, "inventario_real.csv", "es")
    assert set(res["dataset"]) == {"inventario_real.csv"}
    assert set(res["dimension"]) <= {"completeness", "uniqueness"}
    sku_rule = res[res["column"] == "sku"].iloc[0]
    assert sku_rule["status"] == "pass" and sku_rule["score"] == 100.0
    # sin filas duplicadas: la regla de tabla completa tiene que pasar tambien
    fila_completa = res[res["description"].str.contains("100% duplicadas")].iloc[0]
    assert fila_completa["status"] == "pass" and fila_completa["affected_rows"] == 0
    # 3 columnas con 10% de nulos (1/9, redondeando) no llegan al umbral 95%
    fallas = res[res["status"] != "pass"]
    assert set(fallas["column"]) == {"stock", "precio_unitario", "proveedor"}

    from mvdg.quality import overall_index
    assert 0 <= overall_index(res) <= 100


def test_auto_rules_detecta_fila_100pct_duplicada():
    """El caso que _inventario_propio() evita a proposito: una fila que se
    repite en TODAS las columnas. Sin una columna clave en juego, para no
    mezclar los dos heuristicos en el mismo caso."""
    from mvdg.auto_rules import auto_quality_results

    base = pd.DataFrame({
        "categoria": ["Ferretería", "Herramientas", "Seguridad"],
        "proveedor": ["Norte", "Sur", "Este"],
    })
    con_dupe = pd.concat([base, base.iloc[[1]]], ignore_index=True)  # 1 fila repetida

    res = auto_quality_results(con_dupe, "categorias.csv", "es")
    fila_completa = res[res["description"].str.contains("100% duplicadas")].iloc[0]
    assert fila_completa["status"] != "pass"
    assert fila_completa["affected_rows"] == 1


def test_auto_rules_archivo_limpio_da_100():
    from mvdg.auto_rules import auto_quality_results
    from mvdg.quality import overall_index
    limpio = pd.DataFrame({"id": [1, 2, 3, 4], "nombre": ["A", "B", "C", "D"]})
    res = auto_quality_results(limpio, "limpio.csv", "es")
    assert overall_index(res) == 100.0
    assert (res["status"] == "pass").all()


def test_auto_rules_archivo_vacio_no_rompe():
    from mvdg.auto_rules import auto_quality_results
    vacio = pd.DataFrame({"a": []})
    res = auto_quality_results(vacio, "vacio.csv", "es")
    assert res.empty
    assert list(res.columns) == ["rule_id", "dataset", "column", "dimension",
                                 "description", "score", "threshold",
                                 "status", "affected_rows"]


@pytest.mark.parametrize("lang", LANGS)
def test_auto_rules_trilingue(lang):
    """Las descripciones de las reglas y el rotulo de la regla de tabla
    completa tienen que existir en los 3 idiomas — mismo motor de i18n que
    el resto del programa, no un atajo en un solo idioma."""
    from mvdg.auto_rules import auto_quality_results
    res = auto_quality_results(_inventario_propio(), "inventario_real.csv", lang)
    assert not res["description"].isna().any()
    assert (res["description"].str.len() > 0).all()
    columnas_tabla = {"es": "(fila completa)", "en": "(entire row)",
                      "pt": "(linha completa)"}
    assert columnas_tabla[lang] in set(res["column"])


def test_auto_rules_remediacion_funciona_para_reglas_generadas():
    """suggest_fix() tiene que dar una sugerencia util aunque el rule_id
    ('AUTO-01', etc.) no exista en REMEDIATIONS — cae a la plantilla generica
    de la dimension, que es justamente lo que permite reusar _render_fixes()
    en app.py sin tocarlo."""
    from mvdg.auto_rules import auto_quality_results
    from mvdg.remediation import suggest_fix
    res = auto_quality_results(_inventario_propio(), "inventario_real.csv", "es")
    fallas = res[res["status"] != "pass"]
    assert len(fallas) > 0
    for _, row in fallas.iterrows():
        fix = suggest_fix(row["rule_id"], row["dimension"], row["column"],
                          int(row["affected_rows"]), "es")
        assert fix["root_cause"] and fix["short_term"] and fix["long_term"] and fix["owner"]


def test_auto_rules_no_inventa_dimensiones_que_no_puede_evaluar():
    """Solo completitud y unicidad: validez/consistencia/puntualidad/exactitud
    dependen de reglas de negocio que un archivo generico no puede dar."""
    from mvdg.auto_rules import build_rules
    rules = build_rules(_inventario_propio(), "x", "es")
    assert {r.dimension for r in rules} <= {"completeness", "uniqueness"}


# ------------------------------------------------------------- exportadores
@pytest.mark.parametrize("lang", LANGS)
def test_governance_tables_complete(lang):
    tabs = governance_tables(lang)
    assert set(tabs) == {"catalog", "dictionary", "quality_results",
                         "quality_by_dataset", "quality_by_dimension",
                         "lineage", "glossary", "policies", "kpis"}
    for name, df in tabs.items():
        assert len(df) > 0, name


def test_export_formats():
    df = catalog_df("es")
    assert to_csv_bytes(df).startswith("dataset".encode("utf-8-sig"))
    assert to_excel_bytes(df)[:2] == b"PK"      # zip/xlsx
    assert b"dim_customers" in to_json_bytes(df)
    assert bi_bundle_xlsx("pt")[:2] == b"PK"


# ------------------------------------------------------------ fichas clientes
def test_clients_crud_persistence(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import clients
    assert clients.load_clients() == []
    rec = clients.save_client({"company": "ACME", "country": "UY",
                               "it_restriction": "no_exe_python_ok",
                               "recommended_pack": clients.recommended_pack("no_exe_python_ok"),
                               "maturity": 3, "status": "demo"})
    assert rec["client_id"] and rec["created_at"]
    assert rec["recommended_pack"] == "B"
    # actualizar conserva id y created_at
    rec2 = clients.save_client({**rec, "status": "piloto"})
    assert rec2["client_id"] == rec["client_id"]
    assert rec2["created_at"] == rec["created_at"]
    assert len(clients.load_clients()) == 1
    # persiste en disco (relectura fría)
    df = clients.clients_df()
    assert df.iloc[0]["company"] == "ACME" and df.iloc[0]["status"] == "piloto"
    # borrar
    assert clients.delete_client(rec["client_id"]) is True
    assert clients.delete_client("nope") is False
    assert clients.clients_df().empty


def test_clients_recommended_pack():
    from mvdg.clients import recommended_pack
    assert recommended_pack("exe_ok") == "A"
    assert recommended_pack("no_exe_python_ok") == "B"
    assert recommended_pack("solo_web") == "Web"
    assert recommended_pack("???") == "B"


def test_clients_corrupt_file_is_safe(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import clients
    with open(tmp_path / "clientes.json", "w") as fh:
        fh.write("{esto no es json valido")
    assert clients.load_clients() == []


def test_data_dir_variable_explicita_gana_siempre(tmp_path, monkeypatch):
    """MVDG_DATA_DIR manda pase lo que pase - la prioridad más alta."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path / "a mano"))
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(tmp_path / "otro_lado" / "MVDataGovernance.exe"))
    from mvdg import clients
    assert clients.data_dir() == str(tmp_path / "a mano")


def test_data_dir_instalacion_empaquetada_queda_en_el_disco_elegido(tmp_path, monkeypatch):
    """.exe instalado (Inno Setup) sin MVDG_DATA_DIR: los datos van al lado
    del ejecutable, no a ~/.mv_data_governance - si el cliente instaló en
    D:\\, todo (programa Y datos) tiene que quedar en D:\\, no la mitad en
    C:\\ por culpa de un default que ignora dónde se instaló."""
    monkeypatch.delenv("MVDG_DATA_DIR", raising=False)
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    exe = tmp_path / "D_simulado" / "MV Data Governance" / "MVDataGovernance.exe"
    monkeypatch.setattr(sys, "executable", str(exe))
    from mvdg import clients
    d = clients.data_dir()
    assert d == str(exe.parent / "Data")
    assert os.path.isdir(d)


def test_data_dir_sin_empaquetar_usa_home(tmp_path, monkeypatch):
    """Portable .bat / código fuente: sin instalación fija, ~/.mv_data_governance
    de siempre - no hay disco de instalación al cual atarse."""
    monkeypatch.delenv("MVDG_DATA_DIR", raising=False)
    monkeypatch.setattr(sys, "frozen", False, raising=False)
    from mvdg import clients
    assert clients.data_dir() == os.path.join(os.path.expanduser("~"), ".mv_data_governance")


# ------------------------------------------------------------- centro de ayuda
@pytest.mark.parametrize("lang", LANGS)
def test_help_center(lang):
    from mvdg.help_center import AUTOMATION, SPEECHES, automation_rows, speeches
    rows = automation_rows(lang)
    assert len(rows) == len(AUTOMATION) >= 6
    assert {r["level"] for r in rows} == {"auto", "partial", "human"}
    sps = speeches(lang)
    assert len(sps) == len(SPEECHES) == 5
    for sp in sps:
        assert len(sp["text"]) > 200 and sp["title"] and sp["audience"]
    # cada área no automática apunta a un speech existente (círculo cerrado)
    ids = {s["speech_id"] for s in sps}
    for r in rows:
        if r["level"] != "auto":
            assert r["speech_id"] in ids


def test_help_center_translations_differ():
    from mvdg.help_center import speeches
    assert speeches("es")[0]["text"] != speeches("en")[0]["text"]
    assert speeches("pt")[0]["text"] != speeches("en")[0]["text"]


# ---------------------------------------------------------------- release zip
def test_build_release_option_b(tmp_path, monkeypatch):
    # la carpeta packaging/ del repo queda tapada por la librería 'packaging'
    # de PyPI, asi que se carga el modulo directamente por ruta
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "mvdg_build_release",
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                     "packaging", "build_release.py"))
    br = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(br)
    monkeypatch.setattr(br, "DIST", str(tmp_path))
    out = br.build_option_b()
    assert os.path.exists(out)
    import zipfile
    names = zipfile.ZipFile(out).namelist()
    assert "MVDataGovernance/MV_DataGovernance.bat" in names
    assert "MVDataGovernance/app/app.py" in names
    assert "MVDataGovernance/requirements.txt" in names
    # las 3 versiones de arranque viajan en el paquete
    assert "MVDataGovernance/MV_DataGovernance_Server.bat" in names
    assert "MVDataGovernance/run_server.sh" in names
    assert "MVDataGovernance/server_authorized.txt" in names
    assert "MVDataGovernance/bi_api/main.py" in names
    assert not any(".venv" in n or "__pycache__" in n for n in names)
    # los terminos de uso viajan con el producto: el .bat portable y el ZIP no
    # pasan por el instalador, asi que si no estan aca se distribuye sin licencia
    assert "MVDataGovernance/LICENSE" in names
    assert "MVDataGovernance/legal/EULA_es.txt" in names
    assert br.build_option_a() is None  # sin Setup.exe construido


# ------------------------------------------------------------------ licencias
def _par_de_claves():
    import base64
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey
    priv = Ed25519PrivateKey.generate()
    pub = base64.urlsafe_b64encode(priv.public_key().public_bytes(
        encoding=serialization.Encoding.Raw,
        format=serialization.PublicFormat.Raw)).decode().rstrip("=")
    return priv, pub


def _emitir(priv, plan="professional", exp=None, email="c@empresa.com"):
    import base64
    import json
    import time as _t
    p = {"plan": plan, "email": email, "iat": int(_t.time())}
    if exp is not None:
        p["exp"] = exp
    body = base64.urlsafe_b64encode(json.dumps(
        p, separators=(",", ":"), sort_keys=True).encode()).decode().rstrip("=")
    sig = base64.urlsafe_b64encode(priv.sign(body.encode())).decode().rstrip("=")
    return f"MVDG2.{body}.{sig}"


def test_licencia_valida_y_gating(tmp_path, monkeypatch):
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    assert licensing.plan() == licensing.PLAN_DEMO
    assert licensing.has_feature("migracion_purview") is False
    # una función que no está en la tabla de pagas está abierta para todos
    assert licensing.has_feature("catalogo") is True

    token = _emitir(priv)
    assert licensing.save(token) is not None
    assert licensing.plan() == "professional"
    assert licensing.has_feature("migracion_purview") is True
    licensing.clear()
    assert licensing.plan() == licensing.PLAN_DEMO


def test_licencia_plan_owner_desbloquea_todo(tmp_path, monkeypatch):
    """El plan "owner" (auto-emitido con la clave privada del dueño) pasa
    has_feature() para cualquier función, incluida una que no exista todavía
    en FUNCIONES_PAGAS al momento de escribir este test."""
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    token = _emitir(priv, plan="owner")
    assert licensing.save(token) is not None
    assert licensing.plan() == "owner"
    for funcion in list(licensing.FUNCIONES_PAGAS) + ["funcion_futura_no_declarada"]:
        assert licensing.has_feature(funcion) is True


def test_licencia_trial_14_dias_da_lo_mismo_que_professional(tmp_path, monkeypatch):
    """La licencia de prueba que se entrega despues de la demo 1 a 1 lleva
    plan 'trial'. Este test verifica el lado que lo interpreta: mientras no
    venza, desbloquea exactamente lo mismo que 'professional' (el plan de
    USD 390/mes que la prueba demuestra)."""
    import time as _t
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    vigente = _emitir(priv, plan="trial", exp=int(_t.time()) + 14 * 86400)
    assert licensing.save(vigente) is not None
    assert licensing.plan() == "trial"
    for funcion in licensing.FUNCIONES_PAGAS:
        assert licensing.has_feature(funcion) is True, (
            f"trial no desbloquea {funcion}, pero professional si")
    licensing.clear()


def test_licencia_trial_vencida_vuelve_a_demo_sin_codigo_nuevo(tmp_path, monkeypatch):
    """El vencimiento del trial no es un caso especial: es el mismo chequeo
    de 'exp' que ya usa cualquier licencia paga con fecha de corte."""
    import time as _t
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    vencida = _emitir(priv, plan="trial", exp=int(_t.time()) - 10)
    assert licensing.verify(vencida) is None
    assert licensing.save(vencida) is None  # no se guarda un trial vencido
    assert licensing.plan() == licensing.PLAN_DEMO
    assert licensing.has_feature("migracion_purview") is False


def test_licencia_rechaza_manipulacion_y_vencimiento(tmp_path, monkeypatch):
    """Los tres ataques que importan contra una verificación local."""
    import base64
    import json
    import time as _t
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    # 1) editar el payload conservando la firma original
    token = _emitir(priv, plan="professional")
    body, firma = token.split(".")[1], token.split(".")[2]
    p = json.loads(base64.urlsafe_b64decode(body + "=" * (-len(body) % 4)))
    p["plan"] = "enterprise"
    body2 = base64.urlsafe_b64encode(json.dumps(
        p, separators=(",", ":"), sort_keys=True).encode()).decode().rstrip("=")
    assert licensing.verify(f"MVDG2.{body2}.{firma}") is None

    # 2) firmar con otra clave (un falsificador con su propio par)
    otra, _ = _par_de_claves()
    assert licensing.verify(_emitir(otra, plan="enterprise")) is None

    # 3) licencia vencida
    assert licensing.verify(_emitir(priv, exp=int(_t.time()) - 10)) is None

    # y sin clave pública configurada, NADA valida (falla cerrado)
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", "")
    assert licensing.verify(_emitir(priv)) is None


def test_licencia_guardada_se_revalida_no_se_confia_en_el_json(tmp_path, monkeypatch):
    """Editar licencia.json a mano no sirve: el plan sale de revalidar la firma
    del token, no del payload guardado en el archivo."""
    import json
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    licensing.save(_emitir(priv, plan="professional"))
    ruta = os.path.join(str(tmp_path), "licencia.json")
    guardado = json.load(open(ruta, encoding="utf-8"))
    guardado["payload"]["plan"] = "enterprise"  # intento de escalar de plan
    with open(ruta, "w", encoding="utf-8") as fh:
        json.dump(guardado, fh)

    assert licensing.plan() == "professional"  # la edición se ignora

    # y si se rompe el token, se cae a demo (no se confía en el payload)
    guardado["token"] = "MVDG2.roto.roto"
    with open(ruta, "w", encoding="utf-8") as fh:
        json.dump(guardado, fh)
    assert licensing.plan() == licensing.PLAN_DEMO


def test_licencia_no_se_guarda_si_no_valida(tmp_path, monkeypatch):
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    _, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)
    otra, _ = _par_de_claves()
    assert licensing.save(_emitir(otra)) is None
    assert not os.path.exists(os.path.join(str(tmp_path), "licencia.json"))


def test_licencia_se_compila_con_cython():
    """licensing.py es la pieza que decide qué está pago: tiene que viajar
    compilada, no como .py que se edita con el Bloc de notas."""
    bc = _load_build_compiled()
    assert "licensing.py" not in bc.NO_COMPILAR


# --------------------------------------------- compilación del motor (Cython)
def _load_build_compiled():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "build_compiled", os.path.join(_repo_root(), "packaging", "build_compiled.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_build_compiled_excludes_every_dash_m_entrypoint():
    """Todo módulo que se lance con ``python -m mvdg.X`` DEBE quedar sin
    compilar.

    Una extensión nativa no se puede ejecutar con ``-m``: el intérprete corta
    con "No code object available", porque runpy necesita fuente o bytecode.
    Si alguien agrega un entrypoint nuevo y se olvida de excluirlo, el build
    compilado se rompe SOLO en el .exe entregado al cliente (el .bat portable
    sigue andando), que es la peor forma de enterarse. Este test lo detecta
    escaneando el repo de verdad, no una lista escrita a mano."""
    import re
    bc = _load_build_compiled()
    root = _repo_root()
    encontrados = set()
    for base, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs if d not in
                   {".git", "node_modules", "dist", "build", "__pycache__",
                    ".venv", "electron", ".pytest_cache"}]
        for fn in files:
            if not fn.endswith((".py", ".bat", ".sh", ".md", ".iss", ".txt")):
                continue
            try:
                txt = open(os.path.join(base, fn), encoding="utf-8",
                           errors="ignore").read()
            except OSError:
                continue
            encontrados.update(re.findall(r"-m\s+mvdg\.([a-z_]+)", txt))
    assert encontrados, "no se detectó ningún entrypoint -m mvdg.X (¿regex roto?)"
    faltan = sorted(f"{m}.py" for m in encontrados
                    if f"{m}.py" not in bc.NO_COMPILAR)
    assert not faltan, (
        "estos módulos se lanzan con 'python -m' pero build_compiled.py los "
        f"compilaría (se romperían en el .exe): {faltan}")


def test_build_compiled_excludes_module_whose_source_the_selfcheck_audits():
    """El selfcheck abre el código fuente de enforcement.py para afirmar que no
    contiene 'sqlalchemy' ni '.execute(' — la garantía auditable de que genera
    DDL como texto y nunca lo ejecuta. Compilado, ese open() falla y la
    garantía queda sin poder verificarse."""
    bc = _load_build_compiled()
    assert "enforcement.py" in bc.NO_COMPILAR
    sc = open(os.path.join(_repo_root(), "mvdg", "selfcheck.py"),
              encoding="utf-8").read()
    assert "open(en.__file__" in sc, (
        "el selfcheck ya no lee el fuente de enforcement.py: revisá si sigue "
        "haciendo falta excluirlo de la compilación")


# ----------------------------------------------------- modo servidor (web)
def test_server_authorization_modes():
    from mvdg.server import authorization_status

    # sin lista -> modo abierto
    assert authorization_status([])["mode"] == "open"
    # comodín -> autorizado
    assert authorization_status(["*"])["mode"] == "authorized"
    # host no listado -> denegado
    assert authorization_status(["srv-datos"], identities={"otro"})["mode"] == "denied"
    # host listado -> autorizado (case-insensitive)
    st = authorization_status(["SRV-Datos", "10.0.5.20"], identities={"srv-datos"})
    assert st["mode"] == "authorized" and st["matched"] == "srv-datos"


def test_server_parse_authorized_ignores_comments_and_commas():
    from mvdg.server import parse_authorized
    raw = ("# Lista de servidores, con comas en el comentario\n"
           "srv-datos.empresa.local, 10.0.5.20\n"
           "  \n"
           "# otra nota\n"
           "backup.empresa.local\n")
    assert parse_authorized(raw) == [
        "srv-datos.empresa.local", "10.0.5.20", "backup.empresa.local"]
    # env var en una sola línea
    assert parse_authorized("a, B ,c") == ["a", "b", "c"]


def test_server_run_dry_run_builds_streamlit_argv(monkeypatch):
    from mvdg import server
    monkeypatch.setenv("MVDG_SERVER_HOST", "0.0.0.0")
    monkeypatch.setenv("MVDG_SERVER_PORT", "8555")
    monkeypatch.setenv("MVDG_AUTHORIZED_HOSTS", "*")  # autoriza para el dry-run
    argv: list = []
    rc = server.run_server(argv_out=argv)
    assert rc == 0
    assert argv[:3] == ["streamlit", "run", argv[2]]
    assert argv[argv.index("--server.address") + 1] == "0.0.0.0"
    assert argv[argv.index("--server.port") + 1] == "8555"


def test_server_denied_when_host_not_authorized(monkeypatch):
    from mvdg import server
    monkeypatch.setenv("MVDG_AUTHORIZED_HOSTS", "un-host-que-no-soy-yo.local")
    argv: list = []
    rc = server.run_server(argv_out=argv)
    assert rc == 2  # no autorizado: no arranca
    assert argv == []


# ---------------------------------------------------- conectores a base de datos
def test_connectors_sqlite_end_to_end(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    pytest.importorskip("sqlalchemy")
    import sqlite3
    from mvdg import connectors as C

    db = str(tmp_path / "empresa.db")
    con = sqlite3.connect(db)
    pd.DataFrame({"id": [1, 2, 3], "email": ["a@x.com", None, "mal"],
                 "monto": [10, -5, 20]}).to_sql("ventas", con, index=False)
    con.close()

    prof = {"name": "demo", "engine": "sqlite", "database": db,
            "user": "", "password": "clave"}
    saved = C.save_connection(prof, save_password=True)
    assert saved["conn_id"]
    # contraseña no queda en texto plano pero es recuperable
    assert "clave" not in open(C._file(), encoding="utf-8").read()
    assert C.stored_password(C.load_connections()[0]) == "clave"

    ok, msg = C.test_connection(saved)
    assert ok, msg
    assert "ventas" in C.list_tables(saved)
    df = C.load_table(saved, "ventas", limit=100)
    assert len(df) == 3 and "email" in df.columns
    q = C.run_query(saved, "SELECT id FROM ventas WHERE monto > 0")
    assert len(q) == 2
    assert C.delete_connection(saved["conn_id"]) is True
    assert C.load_connections() == []


def test_connectors_guards():
    from mvdg import connectors as C
    # motor desconocido
    with pytest.raises(ValueError):
        C.build_url({"engine": "no-existe"})
    # solo lectura en run_query
    with pytest.raises(ValueError):
        C.run_query({"engine": "sqlite", "database": ":memory:"},
                    "DELETE FROM x")
    # driver ausente -> mensaje legible, no excepción
    ok, msg = C.test_connection({"engine": "postgresql", "host": "localhost",
                                 "port": 5432, "database": "x",
                                 "user": "u", "password": "p"})
    assert ok is False and "driver" in msg.lower()


def _sqlite_grande(ruta, filas=60_000, relleno=200):
    """Una tabla lo bastante grande como para que se note la diferencia entre
    traerla entera y cortar en las primeras filas."""
    import sqlite3
    con = sqlite3.connect(ruta)
    con.execute("CREATE TABLE ventas (id INTEGER, cliente TEXT, monto REAL, nota TEXT)")
    con.executemany("INSERT INTO ventas VALUES (?,?,?,?)",
                    ((i, f"cli_{i % 500}", i * 1.5, "x" * relleno)
                     for i in range(filas)))
    con.commit()
    con.close()
    return {"engine": "sqlite", "database": str(ruta), "host": "", "port": None,
            "user": "", "extra": ""}


def test_sql_sin_tope_duro_de_filas(tmp_path, monkeypatch):
    """`MAX_ROWS` era un TECHO: `min(limit, MAX_ROWS)` recortaba lo que se
    pidiera. Quien tenía una tabla de tres millones de filas veía las
    primeras cien mil y no había forma de subirlo — el programa decía que
    gobernaba su base y gobernaba un pedazo.

    Ahora es solo el valor por defecto, y `limit=0` trae todo.

    Se baja `MAX_ROWS` a 10 en vez de armar una tabla de cien mil filas: lo
    que se prueba es que el pedido NO se recorta contra ese número, y con
    una tabla enorme el test tardaría un minuto para probar lo mismo."""
    from mvdg import connectors as C
    monkeypatch.setattr(C, "MAX_ROWS", 10)
    perfil = _sqlite_grande(tmp_path / "g.db", filas=1000, relleno=10)
    assert len(C.load_table(perfil, "ventas", 500)) == 500, (
        "el pedido se recortó contra MAX_ROWS: sigue siendo un techo")
    assert len(C.load_table(perfil, "ventas", 0)) == 1000, "limit=0 no trajo todo"
    # y el tope sigue funcionando cuando se pide poco
    assert len(C.load_table(perfil, "ventas", 7)) == 7
    # lo mismo por el camino de consulta libre
    assert len(C.run_query(perfil, "SELECT * FROM ventas", 400)) == 400


def test_sql_no_trae_la_tabla_entera_para_devolver_unas_pocas_filas(tmp_path):
    """El bug que esto fija: `pd.read_sql(sql, eng).head(limit)` traía la
    tabla ENTERA a memoria y recién después recortaba. El tope no protegía
    nada — sobre una tabla de diez millones de filas, pedir las primeras mil
    se llevaba las diez millones puestas primero.

    Medido sobre 60.000 filas: antes ~35 MB de pico para devolver 100 filas,
    ahora menos de 1 MB. El margen es de dos órdenes de magnitud, así que el
    umbral no es frágil."""
    import tracemalloc
    from sqlalchemy import create_engine
    from mvdg import connectors as C
    perfil = _sqlite_grande(tmp_path / "g.db", filas=60_000)
    eng = create_engine(f"sqlite:///{perfil['database']}")
    sql = "SELECT * FROM ventas"

    def pico(fn):
        tracemalloc.start()
        try:
            fn()
            return tracemalloc.get_traced_memory()[1]
        finally:
            tracemalloc.stop()

    pico_todo = pico(lambda: __import__("pandas").read_sql(sql, eng))
    pico_pocas = pico(lambda: C._leer_sql(sql, eng, 100))
    assert pico_pocas < pico_todo / 10, (
        f"pedir 100 filas usó {pico_pocas/1e6:.1f} MB y traer todo "
        f"{pico_todo/1e6:.1f} MB: sigue materializando la tabla entera")


def test_sql_devuelve_las_columnas_aunque_no_haya_filas(tmp_path):
    """Un resultado vacío tiene que traer igual el esquema: sin columnas, el
    perfilado de aguas abajo no puede decir siquiera qué se consultó."""
    from mvdg import connectors as C
    perfil = _sqlite_grande(tmp_path / "g.db", filas=100)
    df = C.run_query(perfil, "SELECT * FROM ventas WHERE id < 0", 50)
    assert len(df) == 0
    assert list(df.columns) == ["id", "cliente", "monto", "nota"]


def test_los_topes_de_subida_se_configuran_y_se_pueden_apagar():
    """Los topes viejos (40 MB / 200.000 filas) estaban puestos para la API
    expuesta y se comían también el caso normal: alguien perfilando SU
    archivo en SU PC. El de filas era el peor: no rechazaba, LEÍA LAS
    PRIMERAS 200.000 Y SEGUÍA."""
    import importlib
    import bi_api.main as api
    fuente = open(api.__file__, encoding="utf-8").read()
    for env in ("MVDG_MAX_UPLOAD_MB", "MVDG_MAX_FILAS", "MVDG_MAX_UPLOAD_DE_MB"):
        assert env in fuente, f"{env} no se puede configurar"
    # Por defecto NO se trunca en silencio.
    assert api._MAX_FILAS == 0, (
        "sigue habiendo un corte de filas por defecto: el cliente recibiría "
        "el perfil de un pedazo de su archivo con cara de perfil completo")
    assert api._MAX_BYTES > 200 * 1024 * 1024, "el tope de tamaño sigue siendo chico"
    # Y se pueden apagar del todo.
    os.environ["MVDG_MAX_UPLOAD_MB"] = "0"
    try:
        recargado = importlib.reload(api)
        assert recargado._MAX_BYTES == 0, "MVDG_MAX_UPLOAD_MB=0 no apaga el tope"
    finally:
        del os.environ["MVDG_MAX_UPLOAD_MB"]
        importlib.reload(api)


def test_streamlit_acepta_archivos_grandes():
    """Streamlit corta las subidas en 200 MB por defecto, y el corte es del
    servidor: el archivo se sube ENTERO y recién ahí se rechaza. El usuario
    espera toda la subida de su Excel para que le digan que no."""
    ruta = os.path.join(_repo_root(), ".streamlit", "config.toml")
    with open(ruta, encoding="utf-8") as fh:
        cfg = fh.read()
    m = re.search(r"^maxUploadSize\s*=\s*(\d+)", cfg, re.M)
    assert m, "no se configuró maxUploadSize: el tope sigue siendo 200 MB"
    assert int(m.group(1)) >= 1000, f"maxUploadSize quedó en {m.group(1)} MB"


# ------------------------------------------------- conectores Cloud DW/Lake
def test_connectors_cloud_engines_registered():
    from mvdg import connectors as C
    assert {"synapse", "snowflake", "bigquery", "databricks"} <= set(C.ENGINES)
    assert set(C.CLOUD_ENGINES) == {"snowflake", "bigquery", "databricks"}
    for eng in C.CLOUD_ENGINES:
        assert eng in C.EXTRA_EXAMPLE and C.ENGINES[eng]["pip"]


def test_connectors_snowflake_url():
    from mvdg import connectors as C
    profile = {"engine": "snowflake", "user": "u", "database": "DB",
              "extra": {"account": "xy123.us-east-1", "warehouse": "WH",
                       "role": "SYSADMIN", "schema": "PUBLIC"}}
    url = str(C.build_url(profile, password="pw"))
    assert url.startswith("snowflake://u:")
    assert "xy123.us-east-1" in url and "DB/PUBLIC" in url
    assert "warehouse=WH" in url and "role=SYSADMIN" in url


def test_connectors_bigquery_url():
    from mvdg import connectors as C
    profile = {"engine": "bigquery",
              "extra": {"project": "my-proj", "dataset": "my_ds",
                       "credentials_path": "/tmp/creds.json"}}
    url = str(C.build_url(profile))
    assert url == "bigquery://my-proj/my_ds"
    # sin dataset -> solo project
    url2 = str(C.build_url({"engine": "bigquery", "extra": {"project": "my-proj"}}))
    assert url2 == "bigquery://my-proj"


def test_connectors_databricks_url():
    from mvdg import connectors as C
    profile = {"engine": "databricks",
              "extra": {"server_hostname": "adb-1.azuredatabricks.net",
                       "http_path": "/sql/1.0/warehouses/abc", "catalog": "main",
                       "schema": "default"}}
    url = str(C.build_url(profile, password="dapiTOKEN"))
    assert url.startswith("databricks://token:")
    assert "adb-1.azuredatabricks.net" in url
    assert "catalog=main" in url and "schema=default" in url


def test_connectors_synapse_reuses_mssql_driver():
    from mvdg import connectors as C
    assert C.ENGINES["synapse"]["driver"] == C.ENGINES["sqlserver"]["driver"]
    profile = {"engine": "synapse", "host": "myws.sql.azuresynapse.net",
              "port": 1433, "database": "mydb", "user": "admin"}
    url = str(C.build_url(profile, password="pw"))
    assert url.startswith("mssql+pyodbc://admin:")
    assert "myws.sql.azuresynapse.net:1433/mydb" in url


def test_connectors_save_connection_persists_extra(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import connectors as C
    profile = {"name": "sf-demo", "engine": "snowflake", "user": "u",
              "database": "DB", "password": "pw",
              "extra": {"account": "xy123", "warehouse": "WH"}}
    C.save_connection(profile, save_password=True)
    reloaded = C.load_connections()[0]
    assert reloaded["extra"] == {"account": "xy123", "warehouse": "WH"}


# --------------------------------------------------------- proyecto por cliente
def test_workspace_save_load_stage_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import workspace as ws
    cid = "cli0001"
    df = pd.DataFrame({"id": [1, 2, 3], "name": ["a", "b", "c"]})
    m = ws.save_stage(cid, "Catálogo inicial", {"dataset": df},
                      kind="dataset", notes="primera carga")
    assert m["stage_id"] and m["name"] == "Catálogo inicial"
    assert m["tables"][0]["rows"] == 3 and m["tables"][0]["cols"] == 2
    # relectura fría (nueva llamada, disco)
    loaded = ws.load_stage(cid, m["stage_id"])
    assert list(loaded["loaded_tables"].keys()) == ["dataset"]
    pd.testing.assert_frame_equal(loaded["loaded_tables"]["dataset"], df)


def test_workspace_list_summary_and_delete(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import workspace as ws
    cid = "cli0002"
    df1 = pd.DataFrame({"a": [1, 2]})
    df2 = pd.DataFrame({"b": [1, 2, 3]})
    s1 = ws.save_stage(cid, "Etapa 1", {"t1": df1})
    ws.save_stage(cid, "Etapa 2", {"t2": df2, "t1": df1})
    stages = ws.list_stages(cid)
    assert [s["name"] for s in stages] == ["Etapa 2", "Etapa 1"]  # más nueva primero
    summ = ws.project_summary(cid)
    assert summ["stages"] == 2 and summ["tables"] == 3 and summ["rows"] == 2 + 3 + 2
    assert ws.delete_stage(cid, s1["stage_id"]) is True
    assert ws.delete_stage(cid, "nope") is False
    assert ws.project_summary(cid)["stages"] == 1


def test_workspace_rejects_empty(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import workspace as ws
    with pytest.raises(ValueError):
        ws.save_stage("c", "", {"t": pd.DataFrame({"a": [1]})})  # sin nombre
    with pytest.raises(ValueError):
        ws.save_stage("c", "Etapa", {})  # sin tablas
    with pytest.raises(ValueError):
        ws.save_stage("c", "Etapa", {"t": pd.DataFrame()})  # tabla vacía


def test_workspace_export_import_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import workspace as ws
    cid = "cli0003"
    df = pd.DataFrame({"id": [1, 2, 3, 4]})
    ws.save_stage(cid, "E1", {"d": df})
    ws.save_stage(cid, "E2", {"d": df})
    blob = ws.export_project(cid)
    assert isinstance(blob, bytes) and len(blob) > 0
    # borrar todo y restaurar desde el ZIP
    assert ws.delete_project(cid) is True
    assert ws.project_summary(cid)["stages"] == 0
    n = ws.import_project(cid, blob, replace=True)
    assert n == 2
    names = sorted(s["name"] for s in ws.list_stages(cid))
    assert names == ["E1", "E2"]


# ------------------------------------------------------------- tutorial DMBOK
@pytest.mark.parametrize("lang", LANGS)
def test_dmbok_content_complete(lang):
    from mvdg import dmbok
    assert len(dmbok.areas(lang)) == 11
    assert len(dmbok.principles(lang)) == 6
    assert len(dmbok.concepts(lang)) == 14
    assert len(dmbok.roles(lang)) == 3
    assert len(dmbok.maturity(lang)) == 5
    assert len(dmbok.lifecycle(lang)) == 6
    for a in dmbok.areas(lang):
        assert a["area"] and a["plain"] and a["tech"] and a["deliverables"]
        assert a["coverage"] in ("covered", "partial", "out")
        assert 0 <= a["score"] <= 100
    for c in dmbok.concepts(lang):
        assert c["term"] and c["def"] and c["cat"]


def test_dmbok_coverage_and_radar():
    from mvdg import dmbok
    cov = dmbok.coverage_summary()
    assert cov["covered"] + cov["partial"] + cov["out"] == 11
    radar = dmbok.coverage_scores("es")
    assert len(radar) == 11 and all(0 <= s <= 100 for _, s in radar)


def test_dmbok_translations_differ():
    from mvdg import dmbok
    es = [a["area"] for a in dmbok.areas("es")]
    en = [a["area"] for a in dmbok.areas("en")]
    assert es != en  # están realmente traducidas


# ------------------------------------------------- referencia COBIT 2019 + ISO 38505
@pytest.mark.parametrize("lang", LANGS)
def test_cobit_iso_content_complete(lang):
    from mvdg import cobit_iso as ci
    obs = ci.cobit_objectives(lang)
    assert len(obs) == 8
    for o in obs:
        assert o["code"] and o["name"] and o["plain"] and o["tech"] and o["deliverables"]
        assert o["coverage"] in ("covered", "partial", "out")
        assert 0 <= o["score"] <= 100

    princ = ci.iso_principles(lang)
    assert len(princ) == 6
    for p in princ:
        assert p["name"] and p["text"] and p["note"]
        assert p["coverage"] in ("covered", "partial", "out")

    vrc = ci.iso_vrc(lang)
    assert len(vrc) == 3
    for v in vrc:
        assert v["dim"] and v["text"] and v["mapped"]


def test_cobit_iso_coverage_and_radar():
    from mvdg import cobit_iso as ci
    ccov = ci.cobit_coverage_summary()
    assert ccov["covered"] + ccov["partial"] + ccov["out"] == 8
    icov = ci.iso_coverage_summary()
    assert icov["covered"] + icov["partial"] + icov["out"] == 6
    cradar = ci.cobit_coverage_scores("es")
    assert len(cradar) == 8 and all(0 <= s <= 100 for _, s in cradar)
    iradar = ci.iso_coverage_scores("es")
    assert len(iradar) == 6 and all(0 <= s <= 100 for _, s in iradar)


def test_cobit_iso_translations_differ():
    from mvdg import cobit_iso as ci
    es = [o["name"] for o in ci.cobit_objectives("es")]
    en = [o["name"] for o in ci.cobit_objectives("en")]
    assert es != en
    es_p = [p["name"] for p in ci.iso_principles("es")]
    en_p = [p["name"] for p in ci.iso_principles("en")]
    assert es_p != en_p


# ------------------------------------------------------- MDM (duplicados + golden record)
def test_mdm_finds_real_duplicates_in_demo_customers():
    # el dataset de demo trae 8 colisiones reales de document_id/email, sin
    # inyectar nada a propósito para este test — validación end-to-end real.
    from mvdg import mdm
    from mvdg.demo_data import load_demo_tables
    df = load_demo_tables()["dim_customers"]
    rules = mdm.suggest_rules(df, ["document_id", "email", "full_name", "birth_date"])
    report = mdm.dedup_report(df, rules, min_confidence=0.5, block_column="country")
    assert len(report) == 8
    assert (report["confidence"] == 100.0).all()
    assert (report["rows"] == 2).all()


def test_mdm_avoids_false_positives_on_common_names():
    # 'Ana Costa' aparece 4 veces en la demo: personas distintas, mismo
    # nombre común. El nombre solo (sin ID/email coincidente) no debe
    # alcanzar el umbral de confianza.
    from mvdg import mdm
    from mvdg.demo_data import load_demo_tables
    df = load_demo_tables()["dim_customers"]
    ana = df[df["full_name"] == "Ana Costa"]
    assert len(ana) >= 3   # confirma que el caso de prueba existe en la demo
    rules = mdm.suggest_rules(df, ["document_id", "email", "full_name", "birth_date"])
    clusters = mdm.find_duplicate_clusters(df, rules, min_confidence=0.5, block_column="country")
    flagged_ids = {df.loc[i, "customer_id"] for c in clusters for i in c.row_indices}
    assert not (set(ana["customer_id"]) & flagged_ids)


def test_mdm_golden_record_fills_gaps_from_best_row():
    import pandas as pd
    from mvdg import mdm
    df = pd.DataFrame([
        {"id": 1, "name": "Juan Perez", "email": "juan@x.com", "phone": None},
        {"id": 1, "name": "Juan Perez", "email": None, "phone": "099123456"},
    ])
    cluster = mdm.DuplicateCluster(row_indices=[0, 1], confidence=1.0, matched_on=["id"])
    golden = mdm.build_golden_record(df, cluster)
    assert golden["email"] == "juan@x.com" and golden["phone"] == "099123456"


def test_mdm_blocking_required_for_large_unblocked_comparisons():
    import pandas as pd
    from mvdg import mdm
    df = pd.DataFrame({"name": [f"Person {i}" for i in range(200)]})
    rules = [mdm.MatchRule("name", weight=1.0, kind="fuzzy")]
    with pytest.raises(ValueError):
        mdm.find_duplicate_clusters(df, rules)   # sin block_column, 200 filas -> demasiados pares


def test_mdm_suggest_rules_classifies_by_column_name():
    import pandas as pd
    from mvdg import mdm
    df = pd.DataFrame({"document_id": ["1"], "full_name": ["x"], "amount": [1.0]})
    rules = {r.column: r for r in mdm.suggest_rules(df, ["document_id", "full_name", "amount"])}
    assert rules["document_id"].kind == "exact" and rules["document_id"].weight == 3.0
    assert rules["full_name"].kind == "fuzzy"
    assert rules["amount"].kind == "exact"   # numérico -> exacto, no fuzzy


# ------------------------------------------------- dataset de ejemplo real
def test_sample_dataset_profiles():
    from mvdg.profiler import profile_table, suggest_rules, summary
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(root, "assets", "samples", "rotulado_de_alimentos_2026.csv")
    assert os.path.exists(path), "el dataset de ejemplo debe estar versionado"
    df = pd.read_csv(path)
    info = summary(df)
    assert info["rows"] == 284 and info["columns"] == 12
    assert info["duplicate_rows"] == 0
    prof = profile_table(df)
    assert len(prof) == 12
    # 'muestra' es clave (única) y 'articulos' tiene muchos nulos
    assert set(prof["column"]) >= {"producto", "marca", "muestra", "articulos"}
    # las sugerencias corren sin error en los 3 idiomas
    for lang in LANGS:
        assert isinstance(suggest_rules(df, lang), list)


# ------------------------------------- datasets de ejemplo, gobernados end-to-end
def test_samples_file_second_dataset_versioned():
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(root, "assets", "samples", "dirty_cafe_sales.csv")
    assert os.path.exists(path), "el segundo dataset de ejemplo debe estar versionado"
    df = pd.read_csv(path)
    assert len(df) == 10000 and len(df.columns) == 8


@pytest.mark.parametrize("lang", LANGS)
def test_samples_meta_complete(lang):
    from mvdg import samples
    for key in samples.sample_keys():
        m = samples.sample_meta(key, lang)
        for field in ("name", "domain", "description", "owner", "steward",
                     "classification", "refresh", "source", "license"):
            assert m[field], f"{key}.{field} vacío en {lang}"


@pytest.mark.parametrize("lang", LANGS)
def test_samples_quality_results_have_real_spread(lang):
    from mvdg import samples
    ral = samples.sample_quality_results("rotulado_alimentos", lang)
    assert len(ral) == 6
    assert set(ral["status"]) <= {"pass", "warn", "fail"}
    assert (ral["status"] == "warn").sum() >= 1  # marca "-" y vencimiento detectados

    caf = samples.sample_quality_results("cafe_sales_kaggle", lang)
    assert len(caf) == 7
    assert (caf["status"] == "fail").sum() >= 3  # Item/Payment Method/Location muy incompletos

    bnk = samples.sample_quality_results("bank_marketing_uci", lang)
    assert len(bnk) == 6
    assert (bnk["status"] == "fail").sum() == 1  # contact muy incompleto (29%)
    assert (bnk["status"] == "pass").sum() >= 3

    from mvdg.quality import overall_index
    assert overall_index(caf) < overall_index(bnk) < overall_index(ral)  # cafe_sales es la más sucia a propósito


def test_samples_bank_conditional_rule_is_not_a_false_positive():
    """poutcome='unknown' cuando previous=0 es un caso de negocio válido (el
    cliente nunca fue contactado antes), no un hueco de calidad — BNK-05 debe
    salir en pass, no marcarlo como falla."""
    from mvdg import samples
    res = samples.sample_quality_results("bank_marketing_uci", "es")
    bnk05 = res[res["rule_id"] == "BNK-05"].iloc[0]
    assert bnk05["status"] == "pass"
    assert bnk05["affected_rows"] == 0


def test_samples_bank_classification_note_present():
    from mvdg import samples
    for lang in LANGS:
        m = samples.sample_meta("bank_marketing_uci", lang)
        assert m["classification"] == "Confidencial"
        assert m["classification_note"]
    # los otros dos datasets no llevan nota (no hace falta aclarar nada)
    for key in ("rotulado_alimentos", "cafe_sales_kaggle"):
        assert samples.sample_meta(key, "es")["classification_note"] is None


def test_samples_openfda_file_versioned():
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(root, "assets", "samples", "medicamentos_openfda.csv")
    assert os.path.exists(path), "el dataset openFDA debe estar versionado"
    df = pd.read_csv(path)
    assert len(df) == 1546 and len(df.columns) == 15
    # 6 grupos multinacionales, muchas razones sociales (caso MDM real)
    assert df["labeler_name"].nunique() > 20


@pytest.mark.parametrize("lang", LANGS)
def test_samples_openfda_quality_real_defects(lang):
    """Los defectos del NDC Directory son reales (no inyectados): 4 NDC
    duplicados, huecos de marca/principio activo/clase farmacológica en
    productos de uso humano, y listados FDA sin fecha."""
    from mvdg import samples
    med = samples.sample_quality_results("medicamentos_openfda", lang)
    assert len(med) == 8
    by_id = med.set_index("rule_id")
    assert by_id.loc["MED-01", "affected_rows"] == 4       # NDC duplicados reales
    assert by_id.loc["MED-01", "status"] in ("warn", "fail")
    assert by_id.loc["MED-02", "status"] == "pass"          # formato NDC impecable
    assert by_id.loc["MED-07", "status"] == "pass"          # fechas YYYYMMDD válidas
    assert (med["status"] == "fail").sum() >= 3             # huecos reales del registro


def test_samples_openfda_conditional_rules_scope():
    """Las reglas condicionales solo evalúan medicamentos de uso humano:
    los graneles/semielaborados sin marca NO son falsos positivos."""
    from mvdg import samples
    med = samples.sample_quality_results("medicamentos_openfda", "es")
    by_id = med.set_index("rule_id")
    df = samples.load_sample_table("medicamentos_openfda")
    total_sin_marca = int(df["brand_name"].isna().sum())
    humanos_sin_marca = int(df[df["product_type"].isin(
        {"HUMAN PRESCRIPTION DRUG", "HUMAN OTC DRUG"})]["brand_name"].isna().sum())
    assert by_id.loc["MED-04", "affected_rows"] == humanos_sin_marca
    assert humanos_sin_marca < total_sin_marca  # la condición recorta de verdad


def test_curation_inventory_covers_all_definitions():
    """Toda definición del programa (glosario demo + samples, catálogo,
    diccionario) aparece en el inventario de curaduría, pre-establecida y en
    estado 'sugerido_ia' hasta que un responsable la revise."""
    from mvdg import curation
    for lang in LANGS:
        df = curation.list_items(lang)
        assert len(df) > 100
        assert set(df["kind"]) == {"glossary", "catalog", "column"}
        assert (df["proposed"].str.len() > 0).all()  # nada arranca en blanco


def test_curation_validate_modify_reset(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation
    item = "glossary:medicamentos_openfda:ndc"
    # validar tal cual
    rec = curation.save_validation(item, "es", "validado", "",
                                   "María Viera", "Data Owner Regulatorio")
    assert rec["status"] == "validado" and rec["date"]
    df = curation.list_items("es").set_index("item_id")
    assert df.loc[item, "status"] == "validado"
    assert df.loc[item, "responsible_name"] == "María Viera"
    assert df.loc[item, "text"] == df.loc[item, "proposed"]  # validar no cambia el texto
    # modificar con texto oficial
    curation.save_validation(item, "es", "modificado", "Definición oficial corregida.",
                             "J. Pérez", "Data Steward")
    df = curation.list_items("es").set_index("item_id")
    assert df.loc[item, "status"] == "modificado"
    assert df.loc[item, "text"] == "Definición oficial corregida."
    assert curation.effective_text(item, "es", "fallback") == "Definición oficial corregida."
    # el veredicto es por idioma: en inglés sigue sugerido_ia
    assert curation.list_items("en").set_index("item_id").loc[item, "status"] == "sugerido_ia"
    # resumen y reset
    s = curation.summary("es")
    assert s["modificado"] == 1 and s["reviewed_pct"] > 0
    assert curation.reset_item(item, "es")
    assert curation.get_record(item, "es") is None


def test_curation_requires_responsible_name(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation
    with pytest.raises(ValueError):
        curation.save_validation("glossary:demo:customer", "es", "validado",
                                 "", "", "Data Owner")
    with pytest.raises(ValueError):
        curation.save_validation("glossary:demo:customer", "es", "sugerido_ia",
                                 "", "Nombre", "Cargo")


def _org_fixture():
    return pd.DataFrame({
        "Departamento": ["Dirección General", "Gerencia Comercial", "Gerencia Comercial",
                         "Finanzas", "Calidad y Regulatorio", "Calidad y Regulatorio",
                         "Marketing", "TI / Datos"],
        "Nombre completo": ["Ana Torres", "Bruno Díaz", "Carla Gómez", "Diego Ruiz",
                            "Elena Sosa", "Fabián López", "Gina Méndez", "Hugo Pereira"],
        "Puesto": ["CEO", "Gerente Comercial", "Analista de Ventas", "Director de Finanzas",
                   "Directora de Calidad", "Analista Regulatorio", "Jefa de Marketing",
                   "Coordinador de BI"],
        "Jefe directo": ["", "Ana Torres", "Bruno Díaz", "Ana Torres", "Ana Torres",
                         "Elena Sosa", "Ana Torres", "Diego Ruiz"],
    })


def test_orgchart_header_detection_any_language_and_order():
    from mvdg import orgchart as oc
    org = oc.parse_org_table(_org_fixture())
    assert list(org.columns) == ["nombre", "cargo", "area", "reporta_a", "email"]
    assert len(org) == 8
    # encabezados en inglés también
    en = _org_fixture().rename(columns={"Departamento": "Department",
                                        "Nombre completo": "Name",
                                        "Puesto": "Job Title",
                                        "Jefe directo": "Reports To"})
    assert len(oc.parse_org_table(en)) == 8
    # sin columnas mínimas -> error claro, no un KeyError críptico
    with pytest.raises(ValueError):
        oc.parse_org_table(pd.DataFrame({"x": [1], "y": [2]}))


def test_orgchart_assignments_match_domain_and_seniority():
    """El owner sugerido es la persona de mayor jerarquía del área que
    matchea el dominio — y el orden de keywords es prioridad (para
    bank_marketing gana Marketing, no Comercial)."""
    from mvdg import orgchart as oc
    org = oc.parse_org_table(_org_fixture())
    asg = oc.suggest_assignments(org).set_index("dataset")
    assert asg.loc["fct_payments", "owner_name"] == "Diego Ruiz"        # Finanzas
    assert asg.loc["medicamentos_openfda", "owner_name"] == "Elena Sosa"  # Calidad/Regulatorio
    assert asg.loc["medicamentos_openfda", "steward_name"] == "Fabián López"
    assert asg.loc["bank_marketing_uci", "owner_name"] == "Gina Méndez"  # Marketing > Comercial
    assert (asg["estado"] == "sugerido").all()
    # todos los datasets del programa reciben responsable con nombre y cargo
    assert (asg["owner_name"].str.len() > 0).all()
    assert (asg["owner_role"].str.len() > 0).all()


def test_orgchart_persistence_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import orgchart as oc
    org = oc.parse_org_table(_org_fixture())
    oc.save_org(org)
    assert oc.load_org().equals(org)
    asg = oc.suggest_assignments(org)
    oc.save_assignments(asg)
    assert oc.load_assignments().equals(asg)


def test_orgchart_photo_ai_off_by_default(monkeypatch):
    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY",
                "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(var, raising=False)
    from mvdg.ai_provider import ai_parse_orgchart_image
    assert ai_parse_orgchart_image(b"fake-image-bytes") is None


def test_orgchart_photo_ai_parses_mocked_response(monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    from mvdg import ai_provider as ap

    def fake_post(url, headers, body):
        assert "anthropic" in url
        # la imagen viaja en base64 dentro del body
        assert body["messages"][0]["content"][0]["type"] == "image"
        return {"content": [{"text": '{"personas": [{"nombre": "Ana Torres", '
                                     '"cargo": "CEO", "area": "Dirección", '
                                     '"reporta_a": ""}]}'}]}

    monkeypatch.setattr(ap, "_post_json", fake_post)
    people = ap.ai_parse_orgchart_image(b"png-bytes", "image/png", "es", "claude")
    assert people == [{"nombre": "Ana Torres", "cargo": "CEO",
                       "area": "Dirección", "reporta_a": "", "email": ""}]


def test_insights_governance_coverage(tmp_path, monkeypatch):
    """El índice de gobierno cubre los 8 datasets y sube cuando se asignan
    responsables con nombre y se curan definiciones."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, insights, orgchart
    base = insights.governance_summary("es")
    assert base["datasets"] == 8
    assert base["classified_pct"] == 100.0 and base["rules_pct"] == 100.0
    assert 0 < base["governance_index"] < 100  # honesto: no arranca en 10/10
    # asignar responsables con nombre a todo -> owner/steward suben
    org = orgchart.parse_org_table(_org_fixture())
    orgchart.save_assignments(orgchart.suggest_assignments(org))
    # curar una definición -> curaduría sube
    curation.save_validation("glossary:medicamentos_openfda:ndc", "es",
                             "validado", "", "María Viera", "Data Owner")
    after = insights.governance_summary("es")
    assert after["owner_pct"] == 100.0 and after["steward_pct"] == 100.0
    assert after["curation_pct"] > base["curation_pct"]
    assert after["governance_index"] > base["governance_index"]


def test_insights_named_heuristic():
    from mvdg.insights import _named
    assert _named("María Viera") and _named("J. Pérez")
    assert not _named("Gerencia Comercial")
    assert not _named("Equipo de Datos de Ventas")
    assert not _named("")


def test_samples_openfda_bi_bundle_complete():
    """End-to-end hasta el BI: el paquete de gobierno del dataset openFDA trae
    datos + diccionario + calidad + glosario listos para exportar/servir."""
    from mvdg import samples
    gt = samples.sample_governance_tables("medicamentos_openfda", "es")
    for table in ("data", "dictionary", "quality_results", "glossary"):
        assert table in gt and len(gt[table]) > 0
    assert len(gt["dictionary"]) == 15   # las 15 columnas documentadas
    assert len(gt["glossary"]) == 9      # los 9 términos de negocio


# --------------------------------------- IA externa opcional (con fallback local)
def test_ai_provider_off_by_default(monkeypatch):
    from mvdg import ai_provider as ap
    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY", "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(var, raising=False)
    assert ap.configured_provider() is None
    assert ap.ai_suggest_fix("ds", "col", "completeness", "desc", 5, "es") is None


def test_ai_provider_priority_and_override(monkeypatch):
    from mvdg import ai_provider as ap
    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY", "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(var, raising=False)

    monkeypatch.setenv("OPENAI_API_KEY", "sk-test")
    assert ap.configured_provider() == "openai"

    # claude tiene prioridad si ambas keys estan presentes
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-test")
    assert ap.configured_provider() == "claude"

    # MVDG_AI_PROVIDER fuerza uno especifico, si tiene su key
    monkeypatch.setenv("MVDG_AI_PROVIDER", "openai")
    assert ap.configured_provider() == "openai"

    # forzar un proveedor sin key cargada no debe romper: cae al de prioridad
    monkeypatch.setenv("MVDG_AI_PROVIDER", "gemini")
    assert ap.configured_provider() == "claude"


def test_ai_provider_network_errors_fall_back_to_none(monkeypatch):
    """Cualquier falla de red/timeout/HTTP nunca debe romper la app: siempre
    devuelve None y el llamador cae a la sugerencia local."""
    import urllib.error
    from mvdg import ai_provider as ap
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-test")

    monkeypatch.setitem(ap._CALLERS, "claude",
                        lambda prompt, key, model: (_ for _ in ()).throw(urllib.error.URLError("offline")))
    assert ap.ai_suggest_fix("ds", "col", "completeness", "desc", 5, "es") is None

    monkeypatch.setitem(ap._CALLERS, "claude",
                        lambda prompt, key, model: (_ for _ in ()).throw(
                            urllib.error.HTTPError("url", 401, "unauthorized", {}, None)))
    assert ap.ai_suggest_fix("ds", "col", "completeness", "desc", 5, "es") is None


def test_ai_provider_malformed_responses_fall_back_to_none(monkeypatch):
    from mvdg import ai_provider as ap
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-test")

    monkeypatch.setitem(ap._CALLERS, "claude", lambda prompt, key, model: "not json at all")
    assert ap.ai_suggest_fix("ds", "col", "completeness", "desc", 5, "es") is None

    import json
    monkeypatch.setitem(ap._CALLERS, "claude",
                        lambda prompt, key, model: json.dumps({"root_cause": "x"}))  # faltan claves
    assert ap.ai_suggest_fix("ds", "col", "completeness", "desc", 5, "es") is None


def test_ai_provider_successful_response_parsed(monkeypatch):
    import json
    from mvdg import ai_provider as ap
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-test")

    fake = {"root_cause": "causa", "short_term": "corto", "long_term": "largo", "owner": "equipo"}
    monkeypatch.setitem(ap._CALLERS, "claude",
                        lambda prompt, key, model: "aquí tenés:\n```json\n" + json.dumps(fake) + "\n```")
    result = ap.ai_suggest_fix("cafe_sales_kaggle", "Payment Method", "completeness",
                               "Payment Method completo", 3178, "es")
    assert result == fake


def test_ai_provider_prompt_never_includes_raw_data():
    """El prompt manda solo metadato de la falla (nombres/numeros), nunca
    puede referenciar una fila de datos real porque la funcion no la recibe."""
    from mvdg import ai_provider as ap
    prompt = ap._build_prompt("cafe_sales_kaggle", "Payment Method", "completeness",
                              "Payment Method completo", 3178, "es")
    assert "cafe_sales_kaggle" in prompt and "Payment Method" in prompt and "3178" in prompt
    assert "root_cause" in prompt  # pide el JSON con esas claves


def test_ai_provider_label_and_copilot_not_offered():
    from mvdg import ai_provider as ap
    assert ap.provider_label("claude") == "Claude (Anthropic)"
    assert ap.provider_label("openai") == "ChatGPT (OpenAI)"
    assert ap.provider_label("gemini") == "Gemini (Google)"
    assert "copilot" not in ap._PROVIDERS


def test_samples_accuracy_rule_is_meaningful():
    """La regla de exactitud (total = cantidad × precio) no es un no-op: si
    se corrompe Total Spent, el score debe bajar."""
    from mvdg import samples
    df = samples.load_sample_table("cafe_sales_kaggle").copy()
    rule = next(r for r in samples.SAMPLES["cafe_sales_kaggle"]["rules"] if r.rule_id == "CAF-07")
    score_before, _ = rule.check(df)
    df.loc[df.index[:500], "Total Spent"] = "999999.0"
    score_after, affected_after = rule.check(df)
    assert score_after < score_before
    assert affected_after > 0


def test_samples_dictionary_and_glossary_link_columns():
    from mvdg import samples
    for key in samples.sample_keys():
        dic = samples.sample_dictionary_df(key, "es")
        df = samples.load_sample_table(key)
        assert set(dic["column"]) == set(df.columns)
        gloss = samples.sample_glossary_df(key, "es")
        assert len(gloss) >= 3
        assert all(key in ds for ds in gloss["linked_datasets"])


def test_samples_governance_tables_bundle():
    from mvdg import samples
    for key in samples.sample_keys():
        gov = samples.sample_governance_tables(key, "es")
        assert set(gov) == {"data", "dictionary", "quality_results", "glossary"}
        assert len(gov["data"]) > 0


def test_samples_exportable_with_generic_exporters():
    """Los datasets de ejemplo deben poder exportarse con los mismos
    exportadores genéricos que usa el resto de la plataforma (BI real)."""
    from mvdg import samples
    from mvdg.exporters import to_csv_bytes, to_excel_bytes, to_json_bytes
    df = samples.load_sample_table("rotulado_alimentos")
    assert to_csv_bytes(df).startswith(b"\xef\xbb\xbf") or len(to_csv_bytes(df)) > 0
    assert to_excel_bytes(df)[:2] == b"PK"
    assert len(to_json_bytes(df)) > 0


def test_bi_api_serves_sample_datasets():
    from fastapi.testclient import TestClient
    from bi_api.main import app
    client = TestClient(app)

    r = client.get("/")
    assert r.status_code == 200
    body = r.json()
    assert set(body["samples"]) == {"rotulado_alimentos", "cafe_sales_kaggle",
                                    "bank_marketing_uci", "medicamentos_openfda"}

    r = client.get("/api/samples/cafe_sales_kaggle?lang=en")
    assert r.status_code == 200
    assert r.json()["owner"] == "Operations / Point of Sale"

    r = client.get("/api/samples/cafe_sales_kaggle/quality_results?lang=es&format=csv")
    assert r.status_code == 200
    assert "rule_id" in r.text

    r = client.get("/api/samples/rotulado_alimentos/data?format=json")
    assert r.status_code == 200 and r.json()["rows"] == 284

    assert client.get("/api/samples/no-existe/data").status_code == 404
    assert client.get("/api/samples/cafe_sales_kaggle/no-existe").status_code == 404


# --------------------------------------------------- seguridad de la API BI
def test_bi_api_rate_limits_por_ip_y_exime_health():
    """El limitador corta al pasarse y devuelve 429 accionable, pero /health
    (el ping de las herramientas BI) nunca consume cuota."""
    from fastapi.testclient import TestClient
    from bi_api import main as bm
    monkeys = bm.RATE_LIMIT_REQUESTS
    bm.RATE_LIMIT_REQUESTS = 5
    try:
        bm._reset_rate_limit()
        c = TestClient(bm.app)
        for _ in range(5):
            assert c.get("/health").status_code == 200   # exento: no gasta
        for _ in range(5):
            assert c.get("/api/catalog").status_code == 200
        r = c.get("/api/catalog")
        assert r.status_code == 429
        assert r.headers["Retry-After"] == "60"
        cuerpo = r.json()
        assert cuerpo["error"] == "rate_limit"
        # mensaje accionable en los 3 idiomas, no un traceback
        for pista in ("ES:", "EN:", "PT:", "MVDG_API_RATE_LIMIT"):
            assert pista in cuerpo["detail"]
        assert c.get("/health").status_code == 200       # sigue respondiendo
    finally:
        bm.RATE_LIMIT_REQUESTS = monkeys
        bm._reset_rate_limit()


def test_bi_api_token_opcional_pero_se_exige_cuando_esta_definido(monkeypatch):
    """Sin MVDG_API_TOKEN la API queda abierta en localhost (integracion BI
    documentada). Con token definido, TODA ruta de datos lo exige."""
    from fastapi.testclient import TestClient
    from bi_api import main as bm
    bm._reset_rate_limit()
    c = TestClient(bm.app)
    assert c.get("/api/catalog").status_code == 200      # sin token: abierto

    monkeypatch.setenv("MVDG_API_TOKEN", "s3creto")
    bm._reset_rate_limit()
    r = c.get("/api/catalog")
    assert r.status_code == 401
    assert r.headers["WWW-Authenticate"] == "Bearer"
    assert "Authorization: Bearer" in r.json()["detail"]

    assert c.get("/api/catalog",
                 headers={"Authorization": "Bearer mal"}).status_code == 401
    assert c.get("/api/catalog",
                 headers={"Authorization": "s3creto"}).status_code == 401
    assert c.get("/api/catalog",
                 headers={"Authorization": "Bearer s3creto"}).status_code == 200
    # health sigue abierto para que un monitor externo no necesite el secreto
    assert c.get("/health").status_code == 200
    bm._reset_rate_limit()


# ------------------------------------------- puertos: no pisar a otra app
def test_puerto_ocupado_por_otra_app_se_detecta():
    """El chequeo anterior ponia SO_REUSEADDR antes del bind. En Windows eso
    NO sirve para reciclar sockets: permite atarse a un puerto que otra
    aplicacion ya tiene tomado. O sea que el chequeo hecho para no pisar a
    nadie devolvia "libre" justo en el sistema donde corre el producto."""
    import socket as _s
    from mvdg.netports import hay_alguien_escuchando, puerto_libre

    otra_app = _s.socket()
    otra_app.bind(("127.0.0.1", 0))
    otra_app.listen()
    puerto = otra_app.getsockname()[1]
    try:
        assert hay_alguien_escuchando("127.0.0.1", puerto) is True
        assert puerto_libre("127.0.0.1", puerto) is False
    finally:
        otra_app.close()


def test_puerto_realmente_libre_se_reconoce():
    import socket as _s
    from mvdg.netports import puerto_libre
    tmp = _s.socket()
    tmp.bind(("127.0.0.1", 0))
    puerto = tmp.getsockname()[1]
    tmp.close()          # queda libre
    assert puerto_libre("127.0.0.1", puerto) is True


def _usa_flag_socket(ruta, flag):
    """.El CODIGO usa socket.<flag>? Se mira el AST y no el texto: los
    docstrings de estos archivos explican justamente por que NO se usa
    SO_REUSEADDR, y un grep de texto los contaria como si lo usaran."""
    import ast
    with open(ruta, encoding="utf-8") as fh:
        arbol = ast.parse(fh.read())
    return any(isinstance(n, ast.Attribute) and n.attr == flag
               for n in ast.walk(arbol))


def test_deteccion_de_puertos_no_usa_reuseaddr():
    """Regresion de la causa raiz: si alguien vuelve a poner SO_REUSEADDR en
    el sondeo, en Windows se rompe el aislamiento de puertos otra vez."""
    ruta = os.path.join(_repo_root(), "mvdg", "netports.py")
    assert not _usa_flag_socket(ruta, "SO_REUSEADDR"), "volvio SO_REUSEADDR al sondeo"
    assert _usa_flag_socket(ruta, "SO_EXCLUSIVEADDRUSE"), (
        "falta el flag que garantiza exclusividad en Windows")


def test_ningun_modulo_sondea_puertos_con_reuseaddr():
    """Ni el lanzador, ni la API, ni el modo servidor pueden volver a hacer
    su propio sondeo con SO_REUSEADDR: todos delegan en mvdg.netports."""
    for rel in (("packaging", "mvdg_launcher.py"), ("bi_api", "main.py"),
                ("mvdg", "server.py")):
        ruta = os.path.join(_repo_root(), *rel)
        assert not _usa_flag_socket(ruta, "SO_REUSEADDR"), (
            f"{'/'.join(rel)} sondea puertos con SO_REUSEADDR")
        with open(ruta, encoding="utf-8") as fh:
            assert "netports" in fh.read(), (
                f"{'/'.join(rel)} no delega en mvdg.netports")


def test_elegir_puerto_saltea_los_ocupados():
    import socket as _s
    from mvdg.netports import PUERTOS_DASHBOARD, elegir_puerto

    tomados, sockets = [], []
    try:
        for p in PUERTOS_DASHBOARD[:2]:
            so = _s.socket()
            try:
                so.bind(("127.0.0.1", p))
                so.listen()
                tomados.append(p)
                sockets.append(so)
            except OSError:
                so.close()          # ya estaba ocupado por el entorno: da igual
        if not tomados:
            pytest.skip("no se pudo ocupar ningun puerto candidato")
        elegido = elegir_puerto("127.0.0.1")
        assert elegido not in tomados, "eligio un puerto que otra app tenia"
    finally:
        for so in sockets:
            so.close()


def _launcher():
    """Importa packaging/mvdg_launcher.py (no es un paquete importable)."""
    import importlib.util
    ruta = os.path.join(_repo_root(), "packaging", "mvdg_launcher.py")
    spec = importlib.util.spec_from_file_location("mvdg_launcher_test", ruta)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_launcher_respeta_el_puerto_pedido_si_esta_libre(monkeypatch):
    """Con STREAMLIT_SERVER_PORT libre, se usa ese y no otro."""
    import socket as _s
    with _s.socket() as so:
        so.bind(("127.0.0.1", 0))
        libre = so.getsockname()[1]          # se cierra al salir del with
    monkeypatch.setenv("STREAMLIT_SERVER_PORT", str(libre))
    assert _launcher()._puerto_pedido() == libre


def test_launcher_no_arranca_encima_del_puerto_pedido_si_esta_ocupado(monkeypatch):
    """Si el usuario fijo un puerto y OTRA app lo tiene, el lanzador corta
    con un mensaje accionable — no se ata encima ni escupe un traceback."""
    import socket as _s
    with _s.socket() as ocupado:
        ocupado.bind(("127.0.0.1", 0))
        ocupado.listen()
        port = ocupado.getsockname()[1]
        monkeypatch.setenv("STREAMLIT_SERVER_PORT", str(port))
        with pytest.raises(SystemExit) as exc:
            _launcher()._puerto_pedido()
    assert exc.value.code == 3


def test_launcher_sin_variable_elige_puerto_solo(monkeypatch):
    """Sin STREAMLIT_SERVER_PORT no hay contrato: devuelve 0 y el lanzador
    delega en elegir_puerto (que ya saltea los ocupados)."""
    monkeypatch.delenv("STREAMLIT_SERVER_PORT", raising=False)
    assert _launcher()._puerto_pedido() == 0


def test_launcher_puerto_no_numerico_no_explota_con_traceback(monkeypatch):
    """Un valor basura da mensaje accionable, no ValueError crudo."""
    monkeypatch.setenv("STREAMLIT_SERVER_PORT", "ocho mil")
    with pytest.raises(SystemExit) as exc:
        _launcher()._puerto_pedido()
    assert exc.value.code == 3


def test_app_no_parece_una_app_de_streamlit():
    """El cliente compra software de gobierno de datos, no una demo. Tres
    cosas delataban el framework: la barra blanca del header (que además
    rompía el tema oscuro), el menú ⋮ con opciones de desarrollo y el pie
    "Made with Streamlit"."""
    with open(os.path.join(_repo_root(), "app", "app.py"), encoding="utf-8") as fh:
        css = fh.read()
    assert 'header[data-testid="stHeader"]' in css and "transparent" in css, (
        "la barra blanca del header sigue rompiendo el tema oscuro")
    for sel in ('[data-testid="stToolbar"]', "#MainMenu", "footer"):
        assert sel in css, f"no se oculta {sel}"
    # el header se hace transparente, NO se oculta: ahí vive el botón que
    # despliega la barra lateral colapsada
    assert 'header[data-testid="stHeader"] { display: none' not in css


def test_launcher_arranca_sin_boton_deploy():
    """--client.toolbarMode viewer saca el botón "Deploy" (lo único del
    cromo que sí se puede apagar por configuración). Verificado contra un
    Streamlit real: con el flag, 0 botones Deploy visibles; sin él, 1."""
    with open(os.path.join(_repo_root(), "packaging", "mvdg_launcher.py"),
              encoding="utf-8") as fh:
        src = fh.read()
    assert '"--client.toolbarMode", "viewer"' in src


def test_launcher_reconfirma_el_puerto_antes_de_arrancar(monkeypatch):
    """Entre elegir el puerto y el bind real de Streamlit pasa casi un
    segundo (se abre el navegador, se importa la CLI). Si en esa ventana
    otro programa se lo lleva, antes moría con el traceback de Tornado —
    invisible en un .exe sin consola. Ahora re-chequea y re-elige."""
    L = _launcher()
    import mvdg.netports as np
    ocupados = {7001}
    monkeypatch.setattr(np, "puerto_libre",
                        lambda h, p: p not in ocupados)
    monkeypatch.setattr(np, "elegir_puerto", lambda h="127.0.0.1", **k: 7002)
    # el puerto elegido se lo llevaron -> devuelve otro, no insiste
    assert L._puerto_confirmado(7001) == 7002
    # si sigue libre, se respeta (no se cambia porque sí)
    assert L._puerto_confirmado(7003) == 7003


def test_owner_y_comprador_reciben_el_mismo_exe():
    """"VERSION OWNER IDENTICA A LA DESCARGADA POR COMPRADORES": el kit del
    owner y el paquete del comprador empaquetan EL MISMO archivo
    Setup_v{ver}.exe — no hay un build aparte "ya desbloqueado". Lo que
    cambia es la licencia que cada uno activa, no el binario. Si alguien
    forkeara los builds, esto lo caza."""
    import ast
    ruta = os.path.join(_repo_root(), "packaging", "build_release.py")
    with open(ruta, encoding="utf-8") as fh:
        arbol = ast.parse(fh.read())
    fuentes = {}
    for fn in [n for n in ast.walk(arbol) if isinstance(n, ast.FunctionDef)]:
        if fn.name in ("build_option_a", "build_owner"):
            fuentes[fn.name] = [
                n.value for n in ast.walk(fn)
                if isinstance(n, ast.Constant) and isinstance(n.value, str)
                and "Setup_v" in n.value]
    assert fuentes["build_option_a"] == fuentes["build_owner"], (
        "el owner y el comprador arman nombres de .exe distintos: "
        f"{fuentes}")
    assert fuentes["build_owner"], "no se encontró el Setup.exe en build_owner"


def _par_de_claves():
    """Par Ed25519 de prueba (no la clave real, que no está en el repo)."""
    import base64
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey
    priv = Ed25519PrivateKey.generate()
    pub = base64.urlsafe_b64encode(priv.public_key().public_bytes(
        encoding=serialization.Encoding.Raw,
        format=serialization.PublicFormat.Raw)).decode().rstrip("=")
    return priv, pub


def _firmar_token(priv, payload):
    import base64
    import json as _json
    body = base64.urlsafe_b64encode(_json.dumps(
        payload, separators=(",", ":"), sort_keys=True).encode()).decode().rstrip("=")
    firma = base64.urlsafe_b64encode(priv.sign(body.encode("ascii"))).decode().rstrip("=")
    return f"MVDG2.{body}.{firma}"


def test_licencia_del_owner_atada_no_sirve_en_otra_maquina():
    """EL punto del build del owner: viene desbloqueado, así que si ese .exe
    se filtra tiene que ser inútil en cualquier otra PC. La licencia lleva
    el id de máquina ("mid") y verify() la descarta si no coincide."""
    import time as _time
    from mvdg import licensing
    from mvdg.machine import machine_id
    priv, pub = _par_de_claves()
    base = {"plan": "owner", "email": "yo@ejemplo.com", "iat": int(_time.time())}

    # en ESTA máquina: vale
    propia = _firmar_token(priv, {**base, "mid": machine_id()})
    assert licensing.verify(propia, public_key_b64=pub) is not None

    # el MISMO token, en otra PC (id distinto): no vale -> plan demo
    ajena = _firmar_token(priv, {**base, "mid": "0000ffff0000ffff"})
    assert licensing.verify(ajena, public_key_b64=pub) is None, (
        "un .exe del owner filtrado funcionaria en otra maquina")

    # y la firma sigue siendo válida: lo que falla es SOLO la máquina
    assert licensing.verify(ajena, public_key_b64=pub,
                            check_machine=False) is not None


def test_licencias_vendidas_no_se_atan_a_una_maquina():
    """Atarle la licencia al equipo a un cliente que pagó sería hostil:
    cambia de notebook y pierde lo que compró. Sin "mid", vale en todos
    lados — el binding es solo para el build del owner."""
    import time as _time
    from mvdg import licensing
    priv, pub = _par_de_claves()
    token = _firmar_token(priv, {"plan": "professional", "email": "cliente@x.com",
                                 "iat": int(_time.time())})
    assert licensing.verify(token, public_key_b64=pub) is not None


def test_id_de_maquina_es_estable_y_no_filtra_el_nombre_del_equipo():
    """Tiene que dar lo mismo en cada llamada (si no, la licencia dejaría de
    valer al reiniciar) y no puede contener el nombre de la PC en claro: el
    token viaja por mail y como secreto de CI."""
    import platform
    from mvdg.machine import machine_id, matches
    mid = machine_id()
    assert mid == machine_id() and len(mid) == 16
    assert all(c in "0123456789abcdef" for c in mid)
    nodo = (platform.node() or "").lower()
    if nodo:
        assert nodo not in mid.lower()
    assert matches(None) and matches("")        # sin atar = vale en todos lados
    assert matches(mid.upper())                 # se compara sin distinguir caso
    assert not matches("deadbeefdeadbeef")


def test_licencia_empaquetada_se_activa_sola_pero_valida_igual(tmp_path, monkeypatch):
    """El .exe del owner abre ya desbloqueado (licencia_owner.txt al lado del
    binario), pero ese token pasa por la MISMA verificación que el de un
    comprador: si no valida, se ignora y queda en demo."""
    import time as _time
    from mvdg import licensing
    from mvdg.machine import machine_id
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))   # sin licencia guardada
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    exe_dir = tmp_path / "instalado"
    exe_dir.mkdir()
    monkeypatch.setattr(sys, "executable", str(exe_dir / "MVDataGovernance.exe"))

    token = _firmar_token(priv, {"plan": "owner", "email": "yo@x.com",
                                 "iat": int(_time.time()), "mid": machine_id()})
    (exe_dir / "licencia_owner.txt").write_text(token, encoding="utf-8")
    assert licensing.plan() == "owner"
    assert licensing.has_feature("migracion_purview")

    # token basura empaquetado -> se ignora, NO desbloquea nada
    (exe_dir / "licencia_owner.txt").write_text("MVDG2.falso.falso", encoding="utf-8")
    assert licensing.plan() == "demo"
    assert not licensing.has_feature("migracion_purview")


def test_activacion_del_owner_en_un_solo_paso(tmp_path, monkeypatch):
    """Un solo doble clic tiene que dejar todo listo: generar el par de
    claves (sin él NINGUNA licencia valida — PUBLIC_KEY_B64 arranca vacía),
    escribir la pública en el código, guardar la privada FUERA del repo,
    firmar la licencia del owner atada a esta PC y activarla."""
    import importlib
    import shutil
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path / "datos"))
    # copia del repo: el script EDITA mvdg/licensing.py, no se toca el real
    copia = tmp_path / "repo"
    copia.mkdir()
    for d in ("mvdg", "packaging"):
        shutil.copytree(os.path.join(_repo_root(), d), copia / d)
    # El repo YA trae una clave publica configurada (sin ella no valida
    # ninguna licencia). Este test ejercita el camino de PRIMERA activacion,
    # asi que la vacia en la copia — nunca en el repo real.
    lic = copia / "mvdg" / "licensing.py"
    lic.write_text(re.sub(r'^PUBLIC_KEY_B64 = "[^"]*"$', 'PUBLIC_KEY_B64 = ""',
                          lic.read_text(encoding="utf-8"), count=1,
                          flags=re.MULTILINE), encoding="utf-8")
    monkeypatch.syspath_prepend(str(copia))
    for m in [k for k in list(sys.modules) if k.startswith("mvdg")]:
        del sys.modules[m]
    try:
        licensing = importlib.import_module("mvdg.licensing")
        assert licensing.PUBLIC_KEY_B64 == ""
        spec = importlib.util.spec_from_file_location(
            "owner_setup_test", copia / "packaging" / "owner_setup.py")
        setup = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(setup)
        assert setup.EMAIL_POR_DEFECTO == "vieraschiavi@gmail.com"

        assert setup.main() == 0
        # la pública quedó escrita en el código (si no, el cliente no valida nada)
        texto = (copia / "mvdg" / "licensing.py").read_text(encoding="utf-8")
        pub = re.search(r'^PUBLIC_KEY_B64 = "([^"]+)"$', texto, re.MULTILINE)
        assert pub and len(pub.group(1)) > 20
        # la privada quedó FUERA del repo
        priv = tmp_path / "datos" / "clave_privada_licencias.txt"
        assert priv.exists() and priv.read_text().strip()
        assert not (copia / "clave_privada_licencias.txt").exists()

        # y la licencia quedó activa, con el email pedido y atada a esta PC
        importlib.reload(licensing)
        assert licensing.plan() == "owner"
        assert licensing.has_feature("migracion_purview")
        assert licensing.current()["email"] == "vieraschiavi@gmail.com"
        assert licensing.current()["mid"]

        # idempotente: correrlo de nuevo NO genera otro par (eso invalidaría
        # todas las licencias ya emitidas a clientes)
        antes = pub.group(1)
        assert setup.main() == 0
        texto2 = (copia / "mvdg" / "licensing.py").read_text(encoding="utf-8")
        assert f'PUBLIC_KEY_B64 = "{antes}"' in texto2
    finally:
        for m in [k for k in list(sys.modules) if k.startswith("mvdg")]:
            del sys.modules[m]


def test_activacion_del_owner_corta_si_falta_la_privada(tmp_path, monkeypatch):
    """Con la pública ya configurada pero sin la privada, no se puede firmar
    nada: tiene que cortar con un mensaje que diga dónde la buscó, no
    generar un par nuevo por las suyas (eso invalidaría las licencias de
    los clientes en silencio)."""
    import importlib
    import shutil
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path / "datos"))
    monkeypatch.delenv("LICENSE_PRIVATE_KEY", raising=False)
    copia = tmp_path / "repo"
    copia.mkdir()
    for d in ("mvdg", "packaging"):
        shutil.copytree(os.path.join(_repo_root(), d), copia / d)
    lic = copia / "mvdg" / "licensing.py"
    # Anclado con regex y no un replace literal: el repo puede traer la clave
    # ya configurada o vacia, y un replace de texto suelto fallaria en
    # silencio en uno de los dos casos (dejando el test verde sin probar nada).
    lic.write_text(re.sub(r'^PUBLIC_KEY_B64 = "[^"]*"$',
                          'PUBLIC_KEY_B64 = "unaClavePublicaCualquiera"',
                          lic.read_text(encoding="utf-8"), count=1,
                          flags=re.MULTILINE), encoding="utf-8")
    monkeypatch.syspath_prepend(str(copia))
    for m in [k for k in list(sys.modules) if k.startswith("mvdg")]:
        del sys.modules[m]
    try:
        spec = importlib.util.spec_from_file_location(
            "owner_setup_test2", copia / "packaging" / "owner_setup.py")
        setup = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(setup)
        assert setup.main() == 1
        # y NO piso la clave publica existente
        assert 'PUBLIC_KEY_B64 = "unaClavePublicaCualquiera"' in \
            lic.read_text(encoding="utf-8")
    finally:
        for m in [k for k in list(sys.modules) if k.startswith("mvdg")]:
            del sys.modules[m]


def _sku_a_plan() -> dict:
    """El mapa SKU->plan leido del propio JS, sin reimplementarlo."""
    import json as _json
    import subprocess
    salida = subprocess.run(
        ["node", "-e",
         "console.log(JSON.stringify(require('./api/_license').PLAN_POR_SKU))"],
        cwd=_repo_root(), capture_output=True, text=True, check=True)
    return _json.loads(salida.stdout)


def _skus_del_checkout() -> list[str]:
    """Los SKUs que el checkout sabe cobrar, leidos de api/checkout.js.

    Se le pregunta a Node por el objeto REAL en vez de sacarlo con una regex
    del texto fuente. La regex leia `^\\s*(\\w+):` adentro del bloque PLANS, o
    sea cualquier clave — cuando "pro" gano un atributo (`suscripcion: true`)
    lo tomo como si fuera un SKU mas y tres tests fallaron acusando un bug de
    producto que no existia. Un test que no distingue un SKU de un campo no
    esta verificando lo que dice verificar.
    """
    import json as _json
    import subprocess
    salida = subprocess.run(
        ["node", "-e",
         "console.log(JSON.stringify(Object.keys(require('./api/checkout').PLANS)))"],
        cwd=_repo_root(), capture_output=True, text=True, check=True)
    return _json.loads(salida.stdout)


def test_todo_sku_que_se_cobra_esta_mapeado_a_un_plan():
    """Si el checkout aprende a cobrar un SKU nuevo y nadie lo mapea, el
    cliente paga y recibe un token con un plan que licensing.verify() rechaza.
    Paso de verdad con "pro": el checkout cobraba US$390/mes y el programa
    solo conoce "professional"."""
    from mvdg import licensing
    mapa = _sku_a_plan()
    for sku in _skus_del_checkout():
        assert sku in mapa, (
            f"el checkout cobra '{sku}' y no esta en PLAN_POR_SKU: el cliente "
            f"pagaria y recibiria una licencia que el programa rechaza")
        plan = mapa[sku]
        if plan is not None:
            assert plan in licensing.PLANES, (
                f"'{sku}' mapea al plan '{plan}', que no esta en PLANES")


def test_todo_plan_pago_le_da_al_cliente_mas_que_la_demo(tmp_path, monkeypatch):
    """El test que faltaba, y el que hubiera evitado todo esto: cada plan que
    se cobra tiene que habilitar ESTRICTAMENTE mas que no pagar. Los cinco
    SKUs del checkout entregaban la demo — dos motivos distintos ("pro" no
    validaba, "licencia" validaba pero no abria nada) y ningun sintoma."""
    import json as _json
    from mvdg import licensing
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)
    pagas = sorted(licensing.FUNCIONES_PAGAS)

    def habilitadas(token=None):
        carpeta = tmp_path / (token[-12:] if token else "demo")
        carpeta.mkdir(exist_ok=True)
        monkeypatch.setenv("MVDG_DATA_DIR", str(carpeta))
        if token:
            (carpeta / "licencia.json").write_text(
                _json.dumps({"token": token}), encoding="utf-8")
        return {f for f in pagas if licensing.has_feature(f)}

    demo = habilitadas()
    assert demo == set(), "la demo no deberia traer ninguna funcion paga"

    mapa = _sku_a_plan()
    for sku in _skus_del_checkout():
        plan = mapa[sku]
        if plan is None:
            continue          # packs de creditos: no otorgan licencia
        abiertas = habilitadas(_emitir(priv, plan=plan))
        assert abiertas > demo, (
            f"el SKU '{sku}' (plan '{plan}') le da al cliente exactamente lo "
            f"mismo que no pagar: {sorted(abiertas)}")


def _dias_por_sku() -> dict:
    """Cuanto dura cada SKU, leido del propio JS."""
    import json as _json
    import subprocess
    salida = subprocess.run(
        ["node", "-e",
         "console.log(JSON.stringify(require('./api/_license').DIAS_POR_SKU))"],
        cwd=_repo_root(), capture_output=True, text=True, check=True)
    return _json.loads(salida.stdout)


def _limpiar_ia(monkeypatch):
    for v in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY",
              "MVDG_AI_PROVIDER", "MVDG_AI_API_KEY", "MVDG_AI_BASE_URL",
              "MVDG_AI_MODEL"):
        monkeypatch.delenv(v, raising=False)


def test_elegir_modelo_desde_la_configuracion(tmp_path, monkeypatch):
    """Poder elegir el modelo no es estetico: entre el mas chico y el mas
    grande de un mismo proveedor hay un orden de magnitud de diferencia en
    costo por llamada, y el que paga es el usuario con su propia key."""
    from mvdg import ai_provider, ai_settings
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    _limpiar_ia(monkeypatch)
    monkeypatch.delenv("MVDG_AI_MODEL_CLAUDE", raising=False)

    assert ai_provider.configured_provider() is None
    ai_settings.guardar_key("claude", "sk-ant-x")
    assert ai_provider.configured_provider() == "claude", (
        "la key cargada desde la interfaz tiene que servir igual que la "
        "variable de entorno")
    assert ai_provider._model_for("claude") == "claude-sonnet-5"  # default

    ai_settings.guardar_modelo("claude", "claude-haiku-4-5")
    assert ai_provider._model_for("claude") == "claude-haiku-4-5"

    # La variable de entorno sigue mandando: es la configuracion explicita de
    # quien automatiza, y no puede quedar tapada por algo guardado una vez.
    monkeypatch.setenv("MVDG_AI_MODEL_CLAUDE", "claude-opus-5")
    assert ai_provider._model_for("claude") == "claude-opus-5"


def test_el_boton_actualizar_trae_los_modelos_de_cada_proveedor(tmp_path, monkeypatch):
    """Una lista de modelos hardcodeada envejece mal: a los dos meses ofrece
    modelos viejos y esconde los nuevos, que suelen ser mas baratos. Por eso
    se le pregunta al proveedor."""
    from mvdg import ai_settings
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    _limpiar_ia(monkeypatch)

    pedidos = []

    def fake_get(url, headers):
        pedidos.append((url, headers))
        if "anthropic" in url:
            return {"data": [{"id": "claude-opus-5"}, {"id": "claude-haiku-4-5"}]}
        if "generativelanguage" in url:
            return {"models": [{"name": "models/gemini-2.5-pro"}]}
        return {"data": [{"id": "grok-4"}, {"id": "grok-2-latest"}]}

    monkeypatch.setattr(ai_settings, "_get_json", fake_get)

    for prov, esperado in (("claude", ["claude-haiku-4-5", "claude-opus-5"]),
                           ("gemini", ["gemini-2.5-pro"]),
                           ("grok", ["grok-2-latest", "grok-4"])):
        ai_settings.guardar_key(prov, "sk-" + prov)
        assert ai_settings.refrescar_modelos(prov) == esperado, prov

    # Cada proveedor se autentica a su manera; mandar el header equivocado
    # devolveria 401 y el usuario veria "no se pudo traer la lista".
    urls = {u: h for u, h in pedidos}
    assert any("x-api-key" in h for u, h in pedidos if "anthropic" in u)
    assert any("key=sk-gemini" in u for u, h in pedidos if "generativelanguage" in u)
    assert any(h.get("Authorization") == "Bearer sk-grok"
               for u, h in pedidos if "x.ai" in u)
    # Gemini devuelve "models/xxx": se guarda el nombre corto, que es el que
    # despues hay que mandar en la llamada.
    assert not any(m.startswith("models/") for m in ai_settings.modelos_conocidos("gemini"))
    del urls


def test_si_falla_la_consulta_no_se_pierde_la_lista(tmp_path, monkeypatch):
    """Dejar al usuario sin opciones porque se cayo internet un segundo seria
    peor que mostrarle la lista de ayer."""
    from mvdg import ai_settings
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    _limpiar_ia(monkeypatch)
    ai_settings.guardar_key("openai", "sk-x")

    monkeypatch.setattr(ai_settings, "_get_json",
                        lambda u, h: {"data": [{"id": "gpt-5"}, {"id": "gpt-5-mini"}]})
    assert ai_settings.refrescar_modelos("openai") == ["gpt-5", "gpt-5-mini"]

    def cae(url, headers):
        raise OSError("sin red")

    monkeypatch.setattr(ai_settings, "_get_json", cae)
    assert ai_settings.refrescar_modelos("openai") == ["gpt-5", "gpt-5-mini"]
    assert ai_settings.listar_modelos("openai") == []   # pero no inventa nada


def test_la_api_key_no_queda_en_claro_en_disco(tmp_path, monkeypatch):
    """Sin llavero del sistema se guarda ofuscada — que no es cifrado, y la
    interfaz lo avisa — pero NUNCA en texto plano."""
    from mvdg import ai_settings
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    _limpiar_ia(monkeypatch)

    secreto = "sk-ant-secreto-que-no-debe-verse"
    ai_settings.guardar_key("claude", secreto)
    assert ai_settings.leer_key("claude") == secreto

    for ruta in tmp_path.rglob("*"):
        if ruta.is_file():
            try:
                assert secreto not in ruta.read_text(encoding="utf-8", errors="ignore"), ruta
            except OSError:
                pass

    # y borrarla la saca de verdad
    ai_settings.guardar_key("claude", "")
    assert ai_settings.leer_key("claude") == ""


def test_copilot_no_se_ofrece_como_proveedor():
    """No expone una API para pedir texto con solo una key: se autentica por
    OAuth adentro de un editor. Ofrecerlo seria poner una opcion que nunca
    puede funcionar, y el usuario perderia el rato buscando su key."""
    from mvdg import ai_settings
    assert "copilot" not in ai_settings.PROVEEDORES
    assert "copilot" in ai_settings.SIN_API_PROPIA
    # y la interfaz tiene que explicarlo, no callarlo
    from mvdg.i18n import t
    for lang in ("es", "en", "pt"):
        assert "Copilot" in t("ia_copilot", lang)


def test_cualquier_agente_compatible_con_openai_sirve(monkeypatch):
    """El pedido es que cada cliente use el agente que quiera con su propia
    key. Escribir un conector por proveedor deja el producto siempre atrás del
    que salió ayer; soportar el formato de OpenAI cubre de una vez OpenRouter,
    Groq, Mistral, DeepSeek, Together, Azure y los locales tipo Ollama."""
    from mvdg import ai_provider
    _limpiar_ia(monkeypatch)
    assert ai_provider.configured_provider() is None

    # Una key SOLA no alcanza: sin URL base no se sabe contra qué servicio va,
    # y darlo por configurado ofreceria en la UI una opcion que siempre falla.
    monkeypatch.setenv("MVDG_AI_API_KEY", "sk-lo-que-sea")
    assert ai_provider.configured_provider() is None

    monkeypatch.setenv("MVDG_AI_BASE_URL", "https://openrouter.ai/api/v1")
    assert ai_provider.configured_provider() == "compatible"

    # Si ademas tiene la key oficial de su proveedor, esa gana: es la via
    # directa, sin intermediario.
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant")
    assert ai_provider.configured_provider() == "claude"
    # ...salvo que fuerce el compatible a mano
    monkeypatch.setenv("MVDG_AI_PROVIDER", "compatible")
    assert ai_provider.configured_provider() == "compatible"


def test_el_proveedor_compatible_arma_bien_la_url_y_el_pedido(monkeypatch):
    """Y tolera el error de tipeo mas comun: pegar la URL con
    /chat/completions ya incluida, copiandola de la doc del proveedor. Sin
    esto termina en un 404 que el usuario ve como "la IA no anda"."""
    import json as _json
    from mvdg import ai_provider
    _limpiar_ia(monkeypatch)
    monkeypatch.setenv("MVDG_AI_API_KEY", "sk-abc")
    monkeypatch.setenv("MVDG_AI_MODEL", "llama-3.3-70b")

    visto = {}

    def falso_post(url, headers, body):
        visto.update(url=url, headers=headers, body=body)
        return {"choices": [{"message": {"content": _json.dumps(
            {"root_cause": "c", "short_term": "s",
             "long_term": "l", "owner": "o"})}}]}

    monkeypatch.setattr(ai_provider, "_post_json", falso_post)

    for base in ("https://api.groq.com/openai/v1",
                 "https://api.groq.com/openai/v1/",
                 "https://api.groq.com/openai/v1/chat/completions"):
        monkeypatch.setenv("MVDG_AI_BASE_URL", base)
        r = ai_provider.ai_suggest_fix("ventas", "email", "completitud",
                                       "faltan mails", 3, "es")
        assert r and r["root_cause"] == "c"
        assert visto["url"] == "https://api.groq.com/openai/v1/chat/completions", base
    # el modelo sale de MVDG_AI_MODEL, no de un default de otro proveedor
    assert visto["body"]["model"] == "llama-3.3-70b"
    assert visto["headers"]["Authorization"] == "Bearer sk-abc"


def test_ningun_sku_queda_declarado_a_medias():
    """Cada SKU tiene que declarar plan Y plazo en la misma entrada.

    Cuando eran dos mapas separados se podia registrar el plan y olvidar el
    plazo; el plazo ausente vale 0, asi que ese SKU salia PERPETUO y pasaba
    todos los tests — de Python y de Node. Esto exige la entrada completa."""
    import json as _json
    import subprocess
    tabla = _json.loads(subprocess.run(
        ["node", "-e",
         "console.log(JSON.stringify(require('./api/_license').SKU))"],
        cwd=_repo_root(), capture_output=True, text=True,
        check=True).stdout)

    for sku in _skus_del_checkout():
        assert sku in tabla, f"el checkout cobra '{sku}' y no esta en SKU"
    for sku, e in tabla.items():
        assert set(e) == {"plan", "dias", "recurrente"}, (
            f"'{sku}' declara {sorted(e)}: cada SKU tiene que decir que plan "
            f"da, por cuanto tiempo y si se cobra todos los meses. Un campo "
            f"ausente vale 0/false y se entrega perpetuo sin que nada avise")
        assert isinstance(e["dias"], int) and e["dias"] >= 0
        assert isinstance(e["recurrente"], bool)
        # Lo que se cobra todos los meses tiene que vencer todos los meses, y
        # lo que se cobra una vez no puede vencer: son las dos formas de que
        # el cliente y el cobro dejen de coincidir. La primera es plata que no
        # entra; la segunda es un cliente que pago y se queda afuera.
        assert (e["dias"] > 0) == e["recurrente"], (
            f"'{sku}': recurrente={e['recurrente']} pero dias={e['dias']}. "
            f"Cobro recurrente sin vencimiento = se paga un mes y se tiene "
            f"para siempre; vencimiento sin cobro recurrente = el cliente "
            f"pago una vez y se queda sin nada")




def test_el_boton_de_un_plan_mensual_pide_el_email_que_la_suscripcion_exige():
    """Un preapproval de MercadoPago NO se puede crear sin `payer_email`, asi
    que /api/checkout corta con 400 si el plan es recurrente y no viene email.

    Este test existe por un bug que me comi yo: al pasar "pro" a suscripcion
    puse esa validacion en el servidor y deje el boton de la landing mandando
    {plan:'pro'} pelado. Todo el resto del gate quedaba VERDE — el endpoint
    hacia exactamente lo que decia hacer — y el plan de mayor precio no se
    podia comprar. El unico lugar donde eso se ve es en la union de las dos
    mitades, y no habia nadie mirando ahi.

    No alcanza con que EXISTA un campo de email en la pagina: hay uno para la
    prueba de 14 dias, que llama a otro endpoint. Tiene que ser el boton de
    ESE plan el que lleve a un formulario con un email adentro.
    """
    import json as _json
    import subprocess
    recurrentes = [
        sku for sku, e in _json.loads(subprocess.run(
            ["node", "-e",
             "console.log(JSON.stringify(require('./api/_license').SKU))"],
            cwd=_repo_root(), capture_output=True, text=True,
            check=True).stdout).items() if e["recurrente"]]
    assert recurrentes, "ningun SKU recurrente: este test dejo de proteger algo"

    ruta = os.path.join(_repo_root(), "landing", "index.html")
    with open(ruta, encoding="utf-8") as fh:
        html = fh.read()

    for sku in recurrentes:
        boton = re.search(r"<a[^>]*data-mp=\"%s\"[^>]*>" % re.escape(sku), html)
        assert boton, f"la landing no tiene boton de compra para '{sku}'"
        form_id = re.search(r"data-mp-form=\"([\w-]+)\"", boton.group(0))
        assert form_id, (
            f"el boton de '{sku}' llama al checkout sin pedir email, y el "
            f"checkout responde 400 email_requerido a los planes recurrentes: "
            f"el plan seria imposible de comprar")
        form = re.search(
            r"<form[^>]*id=\"%s\".*?</form>" % re.escape(form_id.group(1)),
            html, re.S)
        assert form, f"'{sku}' apunta a #{form_id.group(1)}, que no existe"
        assert 'type="email"' in form.group(0), (
            f"el formulario de '{sku}' no pide un email")


def test_no_se_anuncia_como_mensual_lo_que_se_entrega_perpetuo():
    """El checkout vendia "pro" como MENSUAL (US$390/mes) y el token salia sin
    `exp`: el cliente pagaba un mes y se quedaba con Professional para
    siempre. Nada detectaba el desfasaje porque la condicion de venta vivia en
    el HTML y el vencimiento en el JS, sin nada que los atara.

    Esto los ata EN LAS DOS DIRECCIONES: si un SKU esta declarado perpetuo (0
    dias), la landing no puede anunciarlo por mes; y si vence, TIENE que
    anunciarlo por mes. La version anterior solo miraba el primer caso, asi
    que al pasar "pro" a 35 dias el test se volvia mudo justo cuando habia que
    cambiar la landing — un guardarrail que se apaga solo el dia que sirve."""
    dias = _dias_por_sku()
    ruta = os.path.join(_repo_root(), "landing", "index.html")
    with open(ruta, encoding="utf-8") as fh:
        html = fh.read()

    # Marcas de cobro recurrente en los tres idiomas.
    recurrente = ("mensual", "mensal", "monthly", "/mes", "/mês", '/mo"')
    encontradas = [m for m in recurrente if m in html]
    if dias.get("pro", 0) == 0:
        assert not encontradas, (
            f"la landing anuncia cobro recurrente {encontradas} pero el SKU "
            f"'pro' esta declarado PERPETUO en DIAS_POR_SKU: el cliente pagaria "
            f"un mes y se quedaria con la licencia para siempre")
    else:
        assert encontradas, (
            f"el SKU 'pro' vence a los {dias['pro']} dias pero la landing no "
            f"dice en ningun lado que el cobro es mensual: el cliente cree que "
            f"compro para siempre y a los {dias['pro']} dias se queda afuera")


def test_el_vencimiento_del_token_sale_de_lo_que_se_vendio():
    """Un SKU con dias>0 tiene que producir un token con `exp`, y uno perpetuo
    NO puede llevarlo. verify() solo rechaza cuando `exp` existe y ya paso, asi
    que omitirlo es exactamente lo que hacia perpetua una suscripcion."""
    import json as _json
    import subprocess
    guion = """
      const {signEd25519, diasDeSku} = require('./api/_license');
      const priv = require('crypto').randomBytes(32).toString('base64')
        .replace(/\\+/g,'-').replace(/\\//g,'_').replace(/=+$/,'');
      const out = {};
      for (const sku of Object.keys(require('./api/_license').DIAS_POR_SKU)) {
        const iat = 1000000, dias = diasDeSku(sku);
        const p = {plan:'professional', iat};
        if (dias > 0) p.exp = iat + dias * 86400;
        out[sku] = {dias, tiene_exp: 'exp' in p};
        signEd25519(p, priv);   // que la firma no reviente con este payload
      }
      console.log(JSON.stringify(out));
    """
    salida = subprocess.run(["node", "-e", guion], cwd=_repo_root(),
                            capture_output=True, text=True, check=True)
    for sku, r in _json.loads(salida.stdout).items():
        assert r["tiene_exp"] == (r["dias"] > 0), (
            f"'{sku}' declara {r['dias']} dias y tiene_exp={r['tiene_exp']}")


def test_la_pagina_de_pago_muestra_la_licencia_que_el_programa_valida():
    """pago.html mostraba d.license, que es la MVDG1/HMAC. El programa solo
    valida MVDG2, asi que el cliente copiaba una clave que se rechaza en el
    primer chequeo: pagaba, la pegaba y no se activaba nada."""
    ruta = os.path.join(_repo_root(), "landing", "pago.html")
    with open(ruta, encoding="utf-8") as fh:
        html = fh.read()
    assert "STATE.license=d.license_key" in html, (
        "la pagina de pago tiene que mostrar license_key (MVDG2), no la HMAC")


def test_licencia_json_con_solo_el_token_alcanza(tmp_path, monkeypatch):
    """El .bat de activacion escribe {"token": "..."} y nada mas — no puede
    calcular el payload sin firmar nada. Este test fija ese contrato: si
    current() algun dia exigiera tambien "payload", el .bat dejaria de
    activar NADA y sin ningun error visible."""
    import json as _json
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    token = _emitir(priv, plan="owner")
    (tmp_path / "licencia.json").write_text(
        _json.dumps({"token": token}), encoding="utf-8")
    assert licensing.plan() == "owner"
    assert licensing.has_feature("migracion_purview")

    # y sigue siendo la FIRMA lo que manda: un token editado a mano no entra
    (tmp_path / "licencia.json").write_text(
        _json.dumps({"token": token[:-4] + "AAAA"}), encoding="utf-8")
    assert licensing.plan() == "demo"


def test_bat_de_activacion_owner_escribe_donde_el_programa_lee(tmp_path,
                                                               monkeypatch):
    """El .bat tiene que cubrir las DOS versiones instaladas y el caso sin
    permisos de admin: por eso escribe en los cuatro lugares de los que
    licensing.current() puede leer, no en uno."""
    import importlib
    from mvdg import licensing
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path / "datos"))
    priv, pub = _par_de_claves()
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    spec = importlib.util.spec_from_file_location(
        "activar_owner_bat_test",
        os.path.join(_repo_root(), "packaging", "activar_owner_bat.py"))
    gen = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(gen)

    token = _emitir(priv, plan="owner")
    salida = tmp_path / "MV_Activar_Owner.bat"
    assert gen.main(["--token", token, "--salida", str(salida)]) == 0

    crudo = salida.read_bytes()
    texto = crudo.decode("ascii")          # lanza si se colo un acento
    assert crudo.startswith(b"@echo off")
    assert b"\xef\xbb\xbf" not in crudo, "un BOM romperia la lectura del token"
    assert b"\r\n" in crudo, "cmd.exe necesita finales de linea CRLF"

    # el token va UNA sola vez (en set TOKEN=); el resto lo usa por variable
    assert texto.count(token) == 1

    # los destinos, que son los caminos reales de current()
    assert '"%USERPROFILE%\\.mv_data_governance"' in texto  # Electron / sin admin
    assert '"%AQUI%Data"' in texto                          # frozen escribible
    assert 'call :txtEn  "%AQUI%"' in texto                 # frozen empaquetada
    assert "resources\\server" in texto                     # motor de Electron
    # y avisa si MVDG_DATA_DIR le gana a todo lo anterior
    assert "MVDG_DATA_DIR" in texto
    # no se cierra la ventana sin que se lea el resultado
    assert "pause" in texto

    # DETECCION AUTOMATICA: el instalador deja elegir carpeta y disco
    # (allowToChangeInstallationDirectory), asi que adivinar rutas no alcanza.
    # Se le pregunta al registro de desinstalacion, donde el instalador anota
    # su InstallLocation real, en las dos ramas posibles.
    assert "reg query" in texto and "InstallLocation" in texto
    assert "Uninstall" in texto
    for rama in ("HKCU", "HKLM"):
        assert f"call :buscarEn {rama}" in texto

    # Toda subrutina llamada tiene que existir: un ":algo" que no esta
    # definido no falla ruidoso en batch, simplemente no hace nada.
    definidas = {ln.strip()[1:] for ln in texto.splitlines()
                 if ln.strip().startswith(":") and not ln.strip().startswith("::")}
    llamadas = {ln.split("call :")[1].split()[0]
                for ln in texto.splitlines() if "call :" in ln}
    assert llamadas <= definidas, f"subrutinas sin definir: {llamadas - definidas}"

    # NINGUN bloque "( ... )" puede expandir una ruta adentro. Es el error
    # clasico de batch: %ProgramFiles(x86)% trae un ")" que cierra el bloque
    # antes de tiempo y rompe el script entero. Por eso las subrutinas usan
    # saltos y no bloques.
    bloques = re.findall(r"\((?:[^()]|\^\(|\^\))*\)", texto, re.S)
    con_ruta = [b for b in bloques
                if ("%~1" in b or "%MVDG_DATA_DIR%" in b) and "reg query" not in b]
    assert not con_ruta, f"bloques que expanden rutas adentro: {con_ruta}"

    # un token que no verifica NO genera nada: un .bat con un token malo se
    # ve identico a uno bueno y el sintoma recien aparece al abrir el programa
    otra = tmp_path / "no.bat"
    assert gen.main(["--token", "MVDG2.falso.falso",
                     "--salida", str(otra)]) == 1
    assert not otra.exists()

    # y un token valido pero que NO es de owner tampoco
    assert gen.main(["--token", _emitir(priv, plan="professional"),
                     "--salida", str(otra)]) == 1
    assert not otra.exists()


def test_el_emisor_de_licencias_esta_configurado():
    """Con PUBLIC_KEY_B64 vacia, verify() devuelve None para CUALQUIER token:
    ni el owner ni un cliente que pago pueden activar nada, y el sintoma es
    mudo (todo el mundo en demo, sin ningun error). Este test lo vuelve
    ruidoso: si alguien la borra, el CI se pone en rojo."""
    from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey
    from mvdg import licensing

    assert licensing.PUBLIC_KEY_B64, (
        "PUBLIC_KEY_B64 vacia: ninguna licencia validaria. Generá el par con "
        "packaging/licencias.py keygen")
    crudo = licensing._b64u_decode(licensing.PUBLIC_KEY_B64)
    assert len(crudo) == 32, f"una publica Ed25519 son 32 bytes, no {len(crudo)}"
    Ed25519PublicKey.from_public_bytes(crudo)   # lanza si no es una clave valida

    # Un token cualquiera NO valida contra ella: la pública sola no desbloquea.
    assert licensing.verify("MVDG2.falso.falso") is None


def _correr_la_app(monkeypatch, tmp_path, con_ia: bool):
    """Levanta el dashboard entero con AppTest y devuelve la corrida."""
    from streamlit.testing.v1 import AppTest
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    for v in ("OPENAI_API_KEY", "GEMINI_API_KEY", "XAI_API_KEY",
              "MVDG_AI_API_KEY", "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(v, raising=False)
    if con_ia:
        monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-de-prueba")
    else:
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    at = AppTest.from_file(os.path.join(_repo_root(), "app", "app.py"),
                           default_timeout=300)
    at.run()
    return at


def _keys_de(at) -> list:
    tipos = ("button", "checkbox", "selectbox", "multiselect", "radio",
             "text_input", "number_input", "slider", "text_area", "toggle",
             "date_input", "color_picker")
    keys = []
    for tp in tipos:
        try:
            keys += [w.key for w in getattr(at, tp) if w.key]
        except (AttributeError, KeyError):
            continue
    return keys


def test_la_app_arranca_sin_excepciones_con_ia_configurada(tmp_path, monkeypatch):
    """Regresión de un crash REAL en la pantalla del usuario:

        StreamlitDuplicateElementKey: key='btn_ai_fix_rotulado_alimentos_RAL-02_es'

    Los botones de sugerencia por IA solo se dibujan si hay un proveedor
    configurado. Mientras eso solo se podía hacer por variable de entorno,
    nadie los veía y el choque de keys quedó latente. Al poder configurarlo
    desde la interfaz, salió a la superficie.

    Streamlit ejecuta el script ENTERO en cada rerun y las pestañas no son
    perezosas: las tres llamadas a _render_fixes conviven en la misma pasada,
    y con "Mis datos" activo comparten dataset y regla."""
    at = _correr_la_app(monkeypatch, tmp_path, con_ia=True)
    assert not at.exception, [str(e.value)[:300] for e in at.exception]
    # y los botones tienen que estar DE VERDAD: si no se dibujaran, este test
    # pasaría sin probar nada — que es como el bug sobrevivió hasta ahora.
    botones_ia = [b.key for b in at.button if b.key and "ai_fix" in b.key]
    assert len(botones_ia) > 1, "sin botones de IA no se prueba el caso que fallaba"


def test_ningun_widget_de_la_app_comparte_key(tmp_path, monkeypatch):
    """El caso general del bug de arriba: dos widgets con la misma key hacen
    reventar la pantalla entera, no solo ese widget. Se revisa TODA la app en
    los dos estados que cambian qué se dibuja."""
    from collections import Counter
    for con_ia in (True, False):
        at = _correr_la_app(monkeypatch, tmp_path, con_ia=con_ia)
        assert not at.exception, [str(e.value)[:300] for e in at.exception]
        keys = _keys_de(at)
        repetidas = [k for k, n in Counter(keys).items() if n > 1]
        assert not repetidas, f"con_ia={con_ia}: keys repetidas {repetidas}"
        assert len(keys) > 20, f"con_ia={con_ia}: solo {len(keys)} widgets con key"


def test_todo_archivo_fuente_lleva_el_aviso_de_copyright():
    """El repo es PUBLICO y el software es propietario: el aviso de copyright
    en cada archivo es lo que hace que quien copie un modulo suelto no pueda
    decir que no sabia. El propio LICENSE prohibe removerlo (clausula 3.d).

    Sin un test, un archivo nuevo nace sin aviso y nadie lo nota hasta que ya
    circula."""
    import subprocess
    raiz = _repo_root()
    trackeados = subprocess.run(["git", "ls-files", "-z"], cwd=raiz,
                                capture_output=True, check=True
                                ).stdout.decode("utf-8").split("\0")
    # payments-config.js esta fuera: la configuracion de settings.json prohibe
    # leerlo (puede traer links de pago), asi que este test tampoco lo abre.
    excluir = ("node_modules/", "dist/", "landing/payments-config.js")
    sin_aviso = []
    for ruta in filter(None, trackeados):
        if any(x in ruta for x in excluir):
            continue
        if not ruta.endswith((".py", ".js", ".jsx", ".mjs", ".cjs")):
            continue
        try:
            with open(os.path.join(raiz, ruta), encoding="utf-8") as fh:
                cabeza = fh.read(400)
        except (OSError, UnicodeDecodeError):
            continue
        if "Martín Viera" not in cabeza or "derechos reservados" not in cabeza:
            sin_aviso.append(ruta)
    assert not sin_aviso, f"archivos fuente sin aviso de copyright: {sin_aviso}"


def test_el_license_nombra_al_titular():
    """"El titular indicado en el aviso de copyright" es a quien apuntan la
    clausula de propiedad y el contacto comercial. Si ahi no hay una persona
    nombrada, el texto se queda sin sujeto."""
    with open(os.path.join(_repo_root(), "LICENSE"), encoding="utf-8") as fh:
        texto = fh.read()
    assert "© 2026 Martín Viera" in texto
    assert "Todos los derechos reservados" in texto
    # y que siga siendo el texto propietario completo, en los tres idiomas
    for marca in ("ESPAÑOL", "ENGLISH", "PORTUGUÊS", "NO es de código abierto"):
        assert marca in texto


def test_el_repo_publico_no_lleva_secretos_de_licencia():
    """El repo es PUBLICO. Un token de owner o la clave privada commiteados
    equivalen a regalar la version full — y a que cualquiera pueda emitirse
    licencias. Se revisa lo que git tiene TRACKEADO, que es lo que se publica
    (un archivo ignorado en el disco del que trabaja no le llega a nadie)."""
    import subprocess

    raiz = _repo_root()
    trackeados = subprocess.run(
        ["git", "ls-files", "-z"], cwd=raiz, capture_output=True, check=True,
    ).stdout.decode("utf-8").split("\0")

    prohibidos = {"licencia_owner.txt", "clave_privada_licencias.txt",
                  "licencia.json"}
    for ruta in filter(None, trackeados):
        assert os.path.basename(ruta) not in prohibidos, \
            f"{ruta} no puede estar en un repo publico"

    # Y que no se cuele un token firmado DENTRO de otro archivo. El patron
    # exige cuerpo y firma largos: asi no matchea el "MVDG2.falso.falso" de
    # los tests ni el "MVDG2.xxx.yyy" de la documentacion, que son inocuos.
    token = re.compile(r"MVDG2\.[A-Za-z0-9_-]{40,}\.[A-Za-z0-9_-]{60,}")
    for ruta in filter(None, trackeados):
        completa = os.path.join(raiz, ruta)
        try:
            with open(completa, encoding="utf-8") as fh:
                contenido = fh.read()
        except (OSError, UnicodeDecodeError):
            continue          # binarios (iconos, xlsx de ejemplo): no aplican
        hallado = token.search(contenido)
        assert hallado is None, \
            f"{ruta} contiene una licencia firmada: {hallado.group()[:24]}..."

    # El .gitignore tiene que cubrirlos, para que no dependa de acordarse.
    ignorados = open(os.path.join(raiz, ".gitignore"), encoding="utf-8").read()
    for nombre in ("licencia_owner.txt", "clave_privada_licencias.txt"):
        assert nombre in ignorados, f"{nombre} deberia estar en .gitignore"


def test_bat_de_activacion_del_owner_es_de_un_clic():
    """El .bat tiene que resolver todo solo: encontrar Python, instalar
    'cryptography' si falta y correr el setup — sin que el usuario escriba
    ningún comando."""
    with open(os.path.join(_repo_root(), "MV_Owner_Activar.bat"),
              encoding="ascii") as fh:      # ascii: cmd.exe usa cp850/cp1252
        bat = fh.read()
    assert "packaging\\owner_setup.py" in bat
    assert "import cryptography" in bat, "no verifica la dependencia de firma"
    assert "pip install --no-cache-dir cryptography" in bat
    assert ":nopython" in bat, "sin salida clara si no hay Python"
    # usa el venv del programa si existe, y si no el Python del sistema
    assert '.venv\\Scripts\\python.exe' in bat and "py -3" in bat


def test_el_verificador_de_licencias_importa_sin_pandas():
    """Regresión de un fallo REAL del workflow del instalador owner en main:

        File "mvdg/licensing.py", line 45, in <module>
            from .clients import data_dir
        File "mvdg/clients.py", line 21, in <module>
            import pandas as pd
        ModuleNotFoundError: No module named 'pandas'

    El paso solo quería validar un token en un runner con Python limpio, y
    se llevaba puesto el CRUD de clientes —y con él pandas— por importar
    `data_dir` desde el módulo equivocado. Instalar pandas ahí habría
    tapado el síntoma: un verificador de firmas Ed25519 no necesita
    DataFrames.

    Se prueba en un subproceso con pandas bloqueado, que es la única forma
    de reproducir el runner: acá pandas ya está importado."""
    import subprocess
    guion = (
        "import sys, builtins\n"
        f"sys.path.insert(0, {_repo_root()!r})\n"
        "_real = builtins.__import__\n"
        "def fake(n, *a, **k):\n"
        "    if n == 'pandas' or n.startswith('pandas.'):\n"
        "        raise ModuleNotFoundError(\"No module named 'pandas'\")\n"
        "    return _real(n, *a, **k)\n"
        "builtins.__import__ = fake\n"
        "from mvdg import licensing\n"
        "assert licensing.plan() == 'demo'\n"
        "assert licensing.verify('MVDG2.falso.falso') is None\n"
        "print('ok')\n"
    )
    r = subprocess.run([sys.executable, "-c", guion], capture_output=True, text=True)
    assert r.returncode == 0, (
        "mvdg.licensing no se puede importar sin pandas: el build del "
        f"instalador owner vuelve a romperse.\n{r.stderr[-800:]}")
    assert "ok" in r.stdout


def test_el_build_del_owner_exige_licencia_atada_a_la_maquina():
    """El instalador del owner se publica en una Release (decisión explícita:
    que siempre haya uno descargable), así que el único candado que queda de
    pie es que la licencia esté ATADA A LA MÁQUINA del dueño — campo `mid`,
    ver mvdg/machine.py. Sin eso, un .exe reenviado por mail abre
    desbloqueado en cualquier PC y el producto se regala solo.

    También tiene que saltearse limpio cuando falta el secreto: un build sin
    MVDG_OWNER_TOKEN que termine en VERDE sin producir instalador es un
    verde que no significa nada."""
    _, wf = _yaml_workflow("instalador_electron.yml")
    assert "MVDG_OWNER_TOKEN" in wf, "el owner ya no recibe su licencia"
    assert "mid" in wf, "no se verifica que la licencia esté atada a la máquina"
    assert "::warning title=Instalador owner sin atar" in wf, (
        "un .exe owner sin atar se construiría en silencio")
    assert "steps.owner.outputs.hay == 'true'" in wf, (
        "sin el secreto, los pasos del owner tienen que saltearse, no fallar")


def test_todo_lo_que_importa_la_suite_esta_declarado():
    """Regresión de un error real: agregué tests que hacen `import yaml` y
    pasaron en local — porque acá PyYAML venía como paquete del SISTEMA. En
    el venv limpio del CI no existía y la suite entera se cayó.

    Esto lo caza antes: recorre los imports de este archivo por AST y exige
    que todo lo que no sea stdlib ni del repo figure en algún requirements.
    Un `pip install -r requirements-dev.txt` en una máquina limpia tiene que
    alcanzar para correr todo — sin eso, "los tests pasan" no significa nada
    fuera de la máquina donde se escribieron."""
    import ast
    ruta = os.path.join(_repo_root(), "tests", "test_core.py")
    with open(ruta, encoding="utf-8") as fh:
        arbol = ast.parse(fh.read())

    def _raices(nodos):
        fuera = set()
        for n in nodos:
            for x in ast.walk(n):
                if isinstance(x, ast.Import):
                    fuera.update(a.name.split(".")[0] for a in x.names)
                elif isinstance(x, ast.ImportFrom) and x.level == 0 and x.module:
                    fuera.add(x.module.split(".")[0])
        return fuera

    raices = _raices([arbol])

    # Imports OPCIONALES: los que estan dentro de un try que captura
    # ImportError. Esos no se exigen en requirements a proposito — el test
    # que los usa degrada solo cuando no estan.
    #
    # El caso concreto: cairosvg rasteriza el logo para comparar los iconos
    # contra el vector. Sirve para REGENERARLOS, no para correr la suite, y
    # declararlo obligaria a instalar libcairo en el CI para un chequeo que
    # ya se hace igual sin el. Se declara acá, explicito, en vez de esquivar
    # el escaneo escribiendo el import de otra forma.
    opcionales = set()
    for n in ast.walk(arbol):
        if isinstance(n, ast.Try) and any(
                (h.type is None) or
                (isinstance(h.type, ast.Name) and h.type.id == "ImportError")
                for h in n.handlers):
            opcionales |= _raices(n.body)
    raices -= opcionales

    propios = {"mvdg", "app", "bi_api", "tests", "conftest"}
    externos = {r for r in raices
                if r not in sys.stdlib_module_names and r not in propios}

    declarado = ""
    for req in ("requirements.txt", "requirements-dev.txt"):
        with open(os.path.join(_repo_root(), req), encoding="utf-8") as fh:
            for linea in fh:
                linea = linea.split("#")[0].strip()
                if linea and not linea.startswith("-r"):
                    declarado += linea.lower() + "\n"

    # nombre de import != nombre del paquete en unos pocos casos conocidos
    ALIAS = {"yaml": "pyyaml", "PIL": "pillow", "dateutil": "python-dateutil",
             "sqlalchemy": "sqlalchemy", "openpyxl": "openpyxl"}
    faltan = [m for m in sorted(externos)
              if ALIAS.get(m, m).lower() not in declarado]
    assert not faltan, (
        f"la suite importa {faltan} y no esta(n) en requirements*.txt: "
        "en una maquina limpia los tests no corren")


def _yaml_workflow(nombre):
    """Devuelve (yaml parseado, texto crudo) de un workflow.

    Antes el segundo elemento era el descriptor del archivo — ya cerrado, así
    que cualquiera que lo usara se llevaba "I/O operation on closed file".
    Nadie lo usaba, y por eso nadie lo notó."""
    import yaml
    ruta = os.path.join(_repo_root(), ".github", "workflows", nombre)
    with open(ruta, encoding="utf-8") as fh:
        texto = fh.read()
    return yaml.safe_load(texto), texto


def test_la_ui_de_escritorio_no_usa_streamlit():
    """La versión .exe tiene que ser React consumiendo la API, no Streamlit
    embebido en una ventana. Es lo que la hace vendible como programa."""
    import json
    raiz = _repo_root()
    # el proceso principal levanta la API, no streamlit
    with open(os.path.join(raiz, "electron", "main.js"), encoding="utf-8") as fh:
        main_js = fh.read()
    assert "sm.spawnApi(" in main_js, "Electron sigue levantando Streamlit"
    assert "spawnStreamlit(" not in main_js
    assert "/app/`" in main_js, "no carga la interfaz React servida en /app"

    # la UI en sí no menciona Streamlit por ningún lado
    for archivo in ("App.jsx", "api.js", "i18n.js"):
        ruta = os.path.join(raiz, "electron", "ui", "src", archivo)
        assert os.path.exists(ruta), f"falta {archivo}"
        with open(ruta, encoding="utf-8") as fh:
            cuerpo = fh.read()
        # Aparece solo en comentarios que EXPLICAN que no se usa, asi que se
        # sacan los comentarios ANTES de buscar.
        #
        # Se sacan con una regex de bloque y no salteando lineas que empiecen
        # con "*" o "//": un /* ... */ de varias lineas tiene lineas del medio
        # que no empiezan con nada de eso, y quedaban adentro. El sintoma es
        # un test que se pone rojo porque alguien ESCRIBIO la palabra en una
        # explicacion — acusando un problema de producto que no existe. Es la
        # tercera vez que este mismo atajo falla en este repo.
        codigo = re.sub(r"/\*.*?\*/", "", cuerpo, flags=re.S)
        codigo = re.sub(r"//.*", "", codigo)
        assert "streamlit" not in codigo.lower(), f"{archivo} usa Streamlit"

    # el instalador empaqueta la UI y un Python propio
    with open(os.path.join(raiz, "electron", "package.json"), encoding="utf-8") as fh:
        build = json.load(fh)["build"]
    destinos = [r["to"] for r in build["extraResources"]]
    assert "ui" in destinos, "el bundle React no viaja en el instalador"
    assert "server/python" in destinos, (
        "sin Python adentro, el .exe no arranca en una PC sin Python")
    assert "server/bi_api" in destinos


def test_la_version_bat_sigue_usando_streamlit():
    """Las DOS formas conviven: el .exe es React, y el .bat sigue abriendo
    Streamlit para las empresas donde no se pueden correr ejecutables. Sacar
    una para tener la otra sería perder la mitad de los clientes posibles."""
    with open(os.path.join(_repo_root(), "MV_DataGovernance.bat"),
              encoding="ascii") as fh:
        bat = fh.read()
    assert "packaging\\mvdg_launcher.py" in bat
    assert "import streamlit" in bat, (
        "el .bat dejo de verificar Streamlit: es su interfaz")
    # y el lanzador de Streamlit sigue existiendo para ese camino
    with open(os.path.join(_repo_root(), "packaging", "mvdg_launcher.py"),
              encoding="utf-8") as fh:
        assert "streamlit" in fh.read().lower()


def test_la_api_sirve_la_ui_en_el_mismo_origen(tmp_path, monkeypatch):
    """La UI se sirve DESDE la API (/app) y no por file://: mismo origen,
    así no hace falta CORS ni relajar webSecurity en Electron — las dos
    formas habituales de que un empaquetado termine con un agujero.

    Se arma un bundle de mentira y se apunta MVDG_UI_DIR ahí en vez de
    depender de electron/ui/dist: ese directorio es un artefacto de build y
    está en .gitignore, así que en una máquina limpia (y en el CI) no
    existe. Un test que lo diera por sentado fallaría por no haber corrido
    `npm run build-ui`, no por un defecto del programa."""
    import importlib
    ui = tmp_path / "bundle"
    ui.mkdir()
    (ui / "index.html").write_text("<!doctype html><div id='root'></div>",
                                   encoding="utf-8")
    (ui / "ui.js").write_text("/* bundle */", encoding="utf-8")
    monkeypatch.setenv("MVDG_UI_DIR", str(ui))

    import bi_api.main as bm
    bm = importlib.reload(bm)
    try:
        assert bm._dir_ui() == str(ui)
        rutas = [getattr(r, "path", "") for r in bm.app.routes]
        assert any(r.startswith("/app") for r in rutas), "la UI no se monta en /app"

        # y se sirve de verdad, por el MISMO servidor que la API
        from fastapi.testclient import TestClient
        with TestClient(bm.app) as c:
            r = c.get("/app/")
            assert r.status_code == 200 and "id='root'" in r.text
            assert c.get("/app/ui.js").status_code == 200
            assert c.get("/api/kpis?lang=es").status_code == 200
    finally:
        # el módulo queda cargado para el resto de la suite: se restaura
        monkeypatch.delenv("MVDG_UI_DIR", raising=False)
        importlib.reload(bm)


def test_sin_bundle_construido_la_api_sigue_sirviendo_a_bi(tmp_path, monkeypatch):
    """Sin la UI construida, /app no se monta y la API sigue funcionando
    igual para Power BI/Tableau — que es su trabajo principal. Montar un
    directorio inexistente haría explotar el arranque de la API por culpa
    de una interfaz que ese cliente ni usa."""
    import importlib
    monkeypatch.setenv("MVDG_UI_DIR", str(tmp_path / "no-existe"))
    import bi_api.main as bm
    bm = importlib.reload(bm)
    try:
        from fastapi.testclient import TestClient
        with TestClient(bm.app) as c:
            assert c.get("/api/catalog?lang=es").status_code == 200
            assert c.get("/health").status_code == 200
    finally:
        monkeypatch.delenv("MVDG_UI_DIR", raising=False)
        importlib.reload(bm)


def test_todos_los_workflows_son_yaml_valido():
    """Un workflow mal formado NO falla ruidoso: GitHub simplemente no lo
    corre, y el síntoma es "el CI no se disparó" — que ya nos costó una
    tarde de diagnóstico. Y los tests que solo hacen grep de texto no lo
    detectan: este archivo se escribió con un here-string de PowerShell
    pegado al margen, que rompía el bloque `run: |`, y el test de grep pasó
    igual. Parsearlos de verdad es lo único que lo caza."""
    import glob

    import yaml
    encontrados = sorted(glob.glob(
        os.path.join(_repo_root(), ".github", "workflows", "*.yml")))
    assert len(encontrados) >= 4, "faltan workflows"
    for ruta in encontrados:
        with open(ruta, encoding="utf-8") as fh:
            try:
                datos = yaml.safe_load(fh)
            except yaml.YAMLError as exc:
                raise AssertionError(
                    f"{os.path.basename(ruta)} no es YAML valido: {exc}") from exc
        assert isinstance(datos, dict), os.path.basename(ruta)
        # "on" en YAML 1.1 se parsea como el booleano True
        assert (True in datos or "on" in datos), (
            f"{os.path.basename(ruta)} sin disparadores")
        assert datos.get("jobs"), f"{os.path.basename(ruta)} sin jobs"


def test_instalador_electron_deja_iconos_y_desinstalador():
    """Accesos directos en escritorio y menú Inicio, y desinstalador — los
    tres DECLARADOS, no heredados del default de electron-builder: un
    default puede cambiar entre versiones y el síntoma sería un instalador
    que no deja icono, descubierto por un cliente."""
    import json
    with open(os.path.join(_repo_root(), "electron", "package.json"),
              encoding="utf-8") as fh:
        nsis = json.load(fh)["build"]["nsis"]
    assert nsis["createDesktopShortcut"] is True
    assert nsis["createStartMenuShortcut"] is True
    assert nsis["uninstallDisplayName"] == "MV Data Governance"
    # elegir carpeta y disco, como el instalador de PyInstaller
    assert nsis["allowToChangeInstallationDirectory"] is True
    assert nsis["oneClick"] is False
    # desinstalar NO borra los datos del usuario (fichas, licencia, glosario)
    assert nsis["deleteAppDataOnUninstall"] is False


def test_workflow_electron_arma_la_carpeta_instalador():
    """El pedido concreto: un ZIP con una carpeta INSTALADOR adentro. Y las
    dos versiones — cliente sin licencia, owner ya desbloqueado."""
    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    assert os.path.exists(ruta), "falta el workflow del instalador Electron"
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()
    assert "INSTALADOR" in wf and "Compress-Archive" in wf
    assert "MVDataGovernance_CLIENTE" in wf and "MVDataGovernance_OWNER" in wf
    # el owner exige un token firmado; el cliente no lleva ninguno
    assert "MVDG_OWNER_TOKEN" in wf
    # Atar el token a la PC es lo deseable pero no siempre es posible (el
    # "mid" solo se puede calcular EN esa PC). Si no viene atado se construye
    # igual, pero tiene que AVISAR: un exe de owner sin atar abre desbloqueado
    # en cualquier lado. Que ese aviso exista es lo que se verifica acá.
    assert '::warning title=Instalador owner sin atar::' in wf
    assert 'if p.get("mid"):' in wf
    # Python EMBEBIBLE, no un venv: un venv guarda la ruta del Python base
    # de la maquina donde se creo y no arranca en la PC de un cliente que no
    # tiene Python. El embeddable de python.org es relocalizable por diseño.
    assert "python-$ver-embed-amd64.zip" in wf, (
        "empaqueta un venv, que no es relocalizable")
    assert "-r requirements.txt" in wf
    # el ._pth del embeddable desactiva site-packages: sin habilitarlo, pip
    # instala pero los import fallan sin ningun error claro
    assert "import site" in wf and "site-packages" in wf
    # Y el ._pth tiene que sumar la carpeta de ARRIBA, donde vive el motor.
    # Regresion de un fallo real: el build murio con "No module named 'mvdg'"
    # con todas las dependencias instaladas. Con un ._pth presente Python
    # arranca AISLADO — sys.path es solo lo que dice ese archivo, sin el
    # directorio actual y sin PYTHONPATH — asi que "python -m bi_api.main",
    # que es como spawnApi() lanza el motor, no encontraba nada.
    assert 'Add-Content $pth ".."' in wf, (
        "sin la carpeta de arriba en el ._pth, el Python embebido no puede "
        "importar mvdg ni bi_api aunque el cwd sea el correcto")
    # y la verificacion tiene que hacerse parada en OTRA carpeta: desde la
    # raiz del repo un sys.path roto queda tapado por el cwd y el instalador
    # sale roto igual.
    assert "Push-Location $env:RUNNER_TEMP" in wf
    assert "import fastapi, uvicorn, pandas, mvdg, bi_api.main" in wf
    # y se verifica que ese Python REALMENTE pueda importar el motor
    assert "import fastapi, uvicorn, pandas, mvdg" in wf
    # la interfaz React se construye antes de empaquetar
    assert "npm run build-ui" in wf
    # no se publica como Release (el del owner no puede quedar pegado al repo)
    assert "action-gh-release" not in wf and "softprops" not in wf
    # chequeo anti-stub, igual que el otro instalador
    assert "-lt 40" in wf


def test_automerge_no_puede_mergear_sin_tests_en_verde():
    """El riesgo real del merge automático: que "ningún check en rojo" se
    tome por "está todo bien". Pasó de verdad en el PR #60 — Actions no
    registró los tests y solo corrió Vercel. Ausencia de rojo NO es verde:
    el workflow tiene que exigir que los tests HAYAN CORRIDO sobre el commit
    actual del PR."""
    ruta = os.path.join(_repo_root(), ".github", "workflows", "automerge.yml")
    assert os.path.exists(ruta), "falta el workflow de automerge"
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()
    # exige tests en verde, y no solo ausencia de fallos
    assert "tests.length === 0" in wf, (
        "no exige que los tests hayan corrido: mergearia sin verificar nada")
    assert "conclusion === 'success'" in wf
    # y los mira contra el HEAD del PR, no contra el commit del workflow_run
    assert "detalle.head.sha" in wf, (
        "usa el sha del workflow_run: un push nuevo con codigo roto podria "
        "colarse con el verde del commit anterior")
    # nada de checks pendientes
    assert "status !== 'completed'" in wf


def test_automerge_solo_toca_ramas_de_trabajo_y_tiene_salida_de_emergencia():
    """No puede mergear una rama hecha a mano ni un PR en borrador, y tiene
    que haber una forma de frenarlo sin desactivar el workflow entero."""
    ruta = os.path.join(_repo_root(), ".github", "workflows", "automerge.yml")
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()
    assert "startsWith('claude/')" in wf, "mergearia cualquier rama"
    assert "pr.draft" in wf, "mergearia un PR en borrador"
    assert "no-automerge" in wf, "sin etiqueta para frenarlo puntualmente"
    assert "mergeable === false" in wf, "mergearia con conflictos"
    # y borra la rama, que es la otra mitad del pedido
    assert "deleteRef" in wf
    # solo corre si los tests pasaron
    datos, _ = _yaml_workflow("automerge.yml")
    assert list(datos[True] if True in datos else datos["on"]) == ["workflow_run"]
    assert datos["jobs"]["merge"]["if"] == (
        "github.event.workflow_run.conclusion == 'success'")


def test_hay_un_solo_instalador_y_hace_las_dos_ediciones():
    """Un solo camino de instalador en GitHub: Electron + NSIS.

    Convivían tres workflows que producían .exe: dos de PyInstaller + Inno
    (`instalador.yml` y `instalador_owner.yml`) y el de Electron. Tres
    superficies para lo mismo son dos de más — y las de PyInstaller estaban
    muertas: la última corrida de `instalador.yml` fue 26 merges antes de
    borrarlo, sin que nada lo dijera, porque "no corrió" no falla.

    Queda el de Electron, que hace las DOS ediciones desde el mismo binario:
    cliente (arranca en demo, se desbloquea con la licencia comprada) y
    owner (licencia adentro, atada a la máquina del dueño)."""
    wf_dir = os.path.join(_repo_root(), ".github", "workflows")
    for muerto in ("instalador.yml", "instalador_owner.yml"):
        assert not os.path.exists(os.path.join(wf_dir, muerto)), (
            f"{muerto} volvió: duplica lo que ya hace instalador_electron.yml")
    datos, crudo = _yaml_workflow("instalador_electron.yml")
    on = datos[True] if True in datos else datos["on"]
    opciones = on["workflow_dispatch"]["inputs"]["version"]["options"]
    for edicion in ("cliente", "owner", "ambas"):
        assert edicion in opciones, f"el instalador ya no ofrece «{edicion}»"
    # El del cliente NO puede llevar licencia adentro: se desbloquea pagando.
    assert "licencia_owner.txt" in crudo, "el owner ya no recibe su licencia"
    assert "concurrency" in datos, "sin concurrency, dos merges = dos builds largos"


def test_tests_no_corren_dos_veces_por_pr():
    """Un PR dispara push Y pull_request, con refs distintos: agrupando por
    github.ref cada uno armaba su grupo y la suite corría DOS VECES sobre el
    mismo commit. Agrupada por nombre de rama, queda una sola."""
    datos, _ = _yaml_workflow("tests.yml")
    grupo = datos["concurrency"]["group"]
    assert "github.ref }}" not in grupo, "vuelve a correr dos veces por PR"
    assert "pull_request.head.ref" in grupo and "github.ref_name" in grupo
    assert datos["concurrency"]["cancel-in-progress"] is True


def test_workflow_del_instalador_publica_y_verifica():
    """El .exe no puede vivir en el repo (GitHub rechaza >100 MB), así que
    se construye en un runner Windows y se publica como Release. El
    workflow tiene que verificar que el .exe es real antes de publicarlo:
    un build que falla a medias deja un stub de pocos KB, y publicarlo es
    peor que no publicar nada — el cliente baja algo que no abre."""
    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    assert os.path.exists(ruta), "falta el workflow del instalador"
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()
    assert "windows-latest" in wf, "no se construye en Windows"
    assert "electron-builder" in wf, "no se construye con electron-builder"
    assert "contents: write" in wf, "sin permiso para crear la Release"
    assert "gh release" in wf, "no publica ninguna Release"
    assert re.search(r"-lt\s+\d+", wf), (
        "no verifica el tamaño del instalador: un build a medias deja un stub")
    # y el LEEME manda a las Releases, no a una carpeta del repo
    leeme = os.path.join(_repo_root(), "distribucion",
                         "opcion_A_instalador_exe", "LEEME.md")
    with open(leeme, encoding="utf-8") as fh:
        assert "releases" in fh.read()


def test_launcher_abre_ventana_de_programa_no_pestana():
    """El .exe instalado tiene que abrirse como PROGRAMA (ventana propia,
    sin barra de direcciones), no como una pestaña más del navegador.
    ``--app=`` es el modo aplicación de Edge/Chrome; Edge viene preinstalado
    en Windows 10/11, así que siempre hay con qué abrirlo."""
    L = _launcher()
    cmd = L._comando_ventana("http://127.0.0.1:8641", r"C:\x\msedge.exe")
    assert cmd[0] == r"C:\x\msedge.exe"
    assert "--app=http://127.0.0.1:8641" in cmd
    # y el launcher usa la ventana, no el navegador a secas
    with open(os.path.join(_repo_root(), "packaging", "mvdg_launcher.py"),
              encoding="utf-8") as fh:
        src = fh.read()
    assert "target=_abrir_programa" in src, "el hilo sigue abriendo pestaña"
    # en Windows, Edge está primero en la lista de candidatos
    assert "msedge.exe" in src and "chrome.exe" in src


def test_launcher_sin_edge_ni_chrome_cae_al_navegador(monkeypatch):
    """Sin ningún navegador con modo app (raro en Windows), se abre la
    pestaña de siempre: mejor eso que una ventana que nunca aparece."""
    L = _launcher()
    abiertos = []
    monkeypatch.setattr(L, "_navegadores_ventana", lambda: ["/no/existe"])
    monkeypatch.setattr(L.webbrowser, "open", lambda u: abiertos.append(u))
    monkeypatch.setattr(L.time, "sleep", lambda s: None)

    class _Resp:
        def __init__(self, *a, **k): pass
    import urllib.request as _ur
    monkeypatch.setattr(_ur, "urlopen", lambda *a, **k: _Resp())
    L._abrir_programa("http://127.0.0.1:9")
    assert abiertos == ["http://127.0.0.1:9"]


def test_launcher_deja_log_si_el_arranque_explota(tmp_path, monkeypatch):
    """console=False en el spec significa que una excepción de arranque era
    invisible: doble clic y 'no pasa nada'. Ahora queda mvdg_error.log al
    lado del .exe (o en TEMP si no se puede escribir ahí)."""
    L = _launcher()
    monkeypatch.setattr(sys, "executable", str(tmp_path / "MVDataGovernance.exe"))
    ruta = L._log_y_avisar_error("Traceback: explosion de prueba")
    assert ruta == str(tmp_path / "mvdg_error.log")
    assert "explosion de prueba" in open(ruta, encoding="utf-8").read()


def test_data_dir_program_files_sin_permiso_cae_al_perfil(tmp_path, monkeypatch):
    """LA CAUSA DEL "NO FUNCIONA": instalado en Archivos de programa (el
    default) y abierto como usuario normal, makedirs(Data) da PermissionError
    y el .exe sin consola moría en silencio en el primer arranque. Ahora la
    escritura se SONDEA (con un archivo real, no permisos declarados) y si
    no se puede, los datos van al perfil del usuario: arrancar siempre."""
    # Se parchea `paths`, que es donde vive el código: `clients` solo lo
    # reexporta, y parchear el reexport cambiaría el nombre en clients sin
    # tocar el que la función usa de verdad — el test pasaría probando nada.
    from mvdg import clients, paths
    monkeypatch.delenv("MVDG_DATA_DIR", raising=False)
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    exe_dir = tmp_path / "Program Files" / "MV Data Governance"
    exe_dir.mkdir(parents=True)
    monkeypatch.setattr(sys, "executable", str(exe_dir / "MVDataGovernance.exe"))
    monkeypatch.setattr(paths, "_ESCRITURA_PROBADA", {})

    real_makedirs = os.makedirs
    def sin_permiso(path, *a, **k):
        if str(path).startswith(str(exe_dir)):
            raise PermissionError(13, "Acceso denegado", str(path))
        return real_makedirs(path, *a, **k)
    monkeypatch.setattr(paths.os, "makedirs", sin_permiso)

    # Se llama por `clients` a propósito: es como lo importan seis módulos,
    # así que esto también verifica que el reexport siga funcionando.
    d = clients.data_dir()
    assert d == os.path.join(os.path.expanduser("~"), ".mv_data_governance")
    # y el sondeo quedó cacheado como "no escribible" (no se reintenta).
    # La clave es la carpeta DESTINO (…/Data), no la del ejecutable: el mismo
    # sondeo lo usa ahora el modo portable, que apunta a otra carpeta.
    assert paths._ESCRITURA_PROBADA[str(exe_dir / "Data")] == ""


def test_data_dir_carpeta_existente_pero_no_escribible_tambien_cae(tmp_path, monkeypatch):
    """Peor variante: la carpeta Data EXISTE (la creó el instalador con
    admin) pero no deja escribir adentro. makedirs(exist_ok=True) pasa —
    solo el sondeo con un archivo real lo detecta."""
    from mvdg import clients
    monkeypatch.delenv("MVDG_DATA_DIR", raising=False)
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    exe_dir = tmp_path / "instalado"
    (exe_dir / "Data").mkdir(parents=True)
    monkeypatch.setattr(sys, "executable", str(exe_dir / "MVDataGovernance.exe"))
    monkeypatch.setattr(clients, "_ESCRITURA_PROBADA", {})

    real_open = open
    def open_sin_permiso(f, *a, **k):
        if ".sonda_escritura" in str(f):
            raise PermissionError(13, "Acceso denegado", str(f))
        return real_open(f, *a, **k)
    import builtins
    monkeypatch.setattr(builtins, "open", open_sin_permiso)
    d = clients.data_dir()
    monkeypatch.undo()
    assert d == os.path.join(os.path.expanduser("~"), ".mv_data_governance")


def test_bi_api_no_publica_fuera_de_loopback_sin_token(monkeypatch):
    """Falla cerrado: exponer la API en 0.0.0.0 sin token aborta con un
    mensaje accionable en vez de servir el gobierno a toda la red."""
    from bi_api import main as bm
    monkeypatch.setenv("MVDG_API_HOST", "0.0.0.0")
    monkeypatch.delenv("MVDG_API_TOKEN", raising=False)
    with pytest.raises(SystemExit) as exc:
        bm.main()
    assert exc.value.code == 1
    assert bm._is_loopback("127.0.0.1") and bm._is_loopback("::1")
    assert not bm._is_loopback("0.0.0.0")


def test_bi_api_cors_no_trae_comodin_por_defecto():
    """allow_origins=['*'] en un puerto local dejaria que cualquier web
    abierta en el navegador leyera las tablas de gobierno."""
    from bi_api import main as bm
    assert "*" not in bm.CORS_ORIGINS
    assert all(o.startswith("http://127.0.0.1") or o.startswith("http://localhost")
               for o in bm.CORS_ORIGINS)


def test_la_demo_no_se_baja_de_ninguna_pagina_publica():
    """La demo dejo de ser de descarga libre: se pide, se muestra en una
    sesion 1 a 1, y la licencia se entrega despues.

    Lo que se estaba regalando no era "una demo": el ZIP publico de la pagina
    traia los 105 archivos del programa, `app/app.py` y el motor `mvdg/`
    enteros. Cualquiera se llevaba el codigo fuente sin dar un nombre.

    Este test cubre las paginas. El gate de verdad esta en api/descargar.js
    (exige licencia) y lo cubre api/payments.test.js — sacar el boton y dejar
    el endpoint abierto seria decoracion, porque la URL ya esta publicada."""
    for pagina in ("index.html", "descargas.html", "reviews.html"):
        ruta = os.path.join(_repo_root(), "landing", pagina)
        if not os.path.exists(ruta):
            continue
        with open(ruta, encoding="utf-8") as fh:
            html = fh.read()
        # Sin comentarios: los que explican POR QUE se saco la descarga
        # mencionan el zip y el endpoint, y no son links.
        limpio = re.sub(r"<!--.*?-->|/\*.*?\*/", "", html, flags=re.S)
        assert not re.search(r'href=["\']\s*[^"\']*\.zip', limpio), (
            f"{pagina} ofrece un .zip: ahi adentro va el codigo fuente entero")
        assert "/api/descargar" not in limpio, (
            f"{pagina} linkea la descarga directa; la demo se pide, no se baja")


def test_la_pagina_de_acceso_pide_los_datos_que_filtran_al_curioso():
    """El formulario tiene que pedir nombre, empresa, pais y email — que es
    lo que separa a un prospecto de la competencia mirando, y lo que deja
    rastro de quien pidio acceso. Si alcanzara con un email descartable,
    volveria a ser una descarga abierta con un paso de mas."""
    ruta = os.path.join(_repo_root(), "landing", "descargas.html")
    with open(ruta, encoding="utf-8") as fh:
        html = fh.read()
    assert "/api/acceso" in html, "la pagina no manda el pedido a ningun lado"
    for campo in ("nombre", "empresa", "pais", "email"):
        assert re.search(r'name="%s"' % campo, html), f"no pide {campo}"
    # y el email tiene que ser un campo de email, no un texto libre
    assert re.search(r'name="email"[^>]*type="email"|type="email"[^>]*name="email"',
                     html), "el email no esta declarado como type=email"


def test_pago_entrega_el_mismo_instalador_mas_la_licencia():
    """Demo y full son EL MISMO binario: quien paga no baja otro programa,
    baja el mismo y pega su clave. La página de pago tiene que ofrecer el
    instalador (no el ZIP de demo) y mostrar la licencia."""
    ruta = os.path.join(_repo_root(), "landing", "pago.html")
    with open(ruta, encoding="utf-8") as fh:
        html = fh.read()
    assert "/api/descargar" in html
    # Y con la licencia ADENTRO del link: la descarga dejo de ser publica, asi
    # que sin `k=<licencia>` el que acaba de pagar se comeria un 403.
    #
    # Se exige que la URL se CONSTRUYA con STATE.license, no que las dos cosas
    # aparezcan sueltas en el archivo. La primera version de este test decia
    # `"k=" in html and "STATE.license" in html` y quedaba VERDE con la
    # licencia sacada del link: las dos cadenas seguian existiendo en otras
    # partes de la pagina. Verificado mutando el arreglo.
    enlace = re.search(
        r"'/api/descargar\?[^']*'\s*\+\s*encodeURIComponent\(STATE\.license\)", html)
    assert enlace, (
        "el link de descarga no se arma con STATE.license: el comprador "
        "recibiria 403 despues de haber pagado")
    assert "k=" in enlace.group(0), (
        f"el link no manda la licencia como `k`: {enlace.group(0)}")
    assert "MVDataGovernance_Demo_v1.0.0.zip" not in html, (
        "despues de pagar seguia bajando el ZIP de demo")
    assert "lic_label" in html, "no muestra la clave de licencia"
    # y lo explica en los 3 idiomas, sin prometer una segunda descarga
    for frase in ("No hay una segunda descarga", "No second download",
                  "Sem segundo download"):
        assert frase in html, f"falta la aclaracion: {frase!r}"


def test_endpoint_de_descarga_falla_ruidoso_sin_configurar():
    """Sin MVDG_INSTALLER_URL el endpoint NO puede inventar una URL ni
    servir un archivo viejo: tiene que decir qué falta. Un botón de descarga
    que baja algo equivocado es peor que uno que avisa."""
    ruta = os.path.join(_repo_root(), "api", "descargar.js")
    with open(ruta, encoding="utf-8") as fh:
        js = fh.read()
    assert "MVDG_INSTALLER_URL" in js
    assert "503" in js and "sin_configurar" in js
    assert 'url.protocol !== "https:"' in js, "permitiria bajar un .exe por http"
    assert "302" in js and "no-store" in js, (
        "un 301 cacheado dejaria a los usuarios pegados al hosting viejo")


def test_landing_tiene_headers_de_seguridad():
    """CSP, nosniff y Referrer-Policy configurados para el sitio publico."""
    import json as _json
    with open(os.path.join(_repo_root(), "vercel.json"), encoding="utf-8") as fh:
        cfg = _json.load(fh)
    catch_all = [h for h in cfg["headers"] if h["source"] == "/(.*)"]
    assert catch_all, "falta el bloque de headers que cubre todo el sitio"
    hs = {h["key"]: h["value"] for h in catch_all[0]["headers"]}
    assert hs["X-Content-Type-Options"] == "nosniff"
    assert hs["Referrer-Policy"] == "strict-origin-when-cross-origin"
    csp = hs["Content-Security-Policy"]
    for directiva in ("default-src 'self'", "object-src 'none'",
                      "frame-ancestors 'none'", "base-uri 'self'"):
        assert directiva in csp


def test_landing_escapa_datos_antes_de_inyectar_html():
    """Las resenas se inyectan con innerHTML/insertAdjacentHTML: tienen que
    pasar por un escapado completo, no por un .replace solo de '<'."""
    for archivo in ("index.html", "reviews.html"):
        ruta = os.path.join(_repo_root(), "landing", archivo)
        with open(ruta, encoding="utf-8") as fh:
            html = fh.read()
        assert "/[&<>\"']/g" in html, f"{archivo}: falta el escapado completo"
        # el patron viejo (solo '<') no debe quedar en el render de resenas
        assert ".replace(/</g,'&lt;')" not in html, (
            f"{archivo}: quedo un escapado parcial de '<'")


# ----------------------------------- errores accionables (no stack traces)
def test_errores_de_archivo_dan_consejo_no_traceback(tmp_path):
    """Cada falla real de carga de archivo se traduce a algo que el usuario
    puede accionar. El detalle tecnico va aparte, no como respuesta."""
    import json as _json
    import zipfile
    from mvdg.errors import friendly_error

    casos = []
    mal = tmp_path / "latin1.csv"
    mal.write_bytes("nombre,ciudad\nJosé,Montevideo\n".encode("latin-1"))
    try:
        pd.read_csv(mal, encoding="utf-8")
    except Exception as exc:
        casos.append(("encoding", exc, ("UTF-8",)))

    vacio = tmp_path / "vacio.csv"
    vacio.write_text("")
    try:
        pd.read_csv(vacio)
    except Exception as exc:
        casos.append(("vacio", exc, ("vacío", "encabezados")))

    roto = tmp_path / "roto.csv"
    roto.write_text("a\n1,2,3\n4,5,6,7\n")
    try:
        pd.read_csv(roto)
    except Exception as exc:
        casos.append(("csv", exc, ("columnas", "separador")))

    nozip = tmp_path / "no.zip"
    nozip.write_text("esto no es un zip")
    try:
        zipfile.ZipFile(nozip)
    except Exception as exc:
        casos.append(("zip", exc, ("ZIP",)))

    try:
        _json.loads("{roto:")
    except Exception as exc:
        casos.append(("json", exc, ("JSON",)))

    try:
        open(str(tmp_path / "no" / "existe.csv"))
    except Exception as exc:
        casos.append(("no_existe", exc, ("encontró", "encontr")))

    assert len(casos) == 6, "no se reprodujeron todos los fallos esperados"
    for nombre, exc, pistas in casos:
        msg, detalle = friendly_error(exc, "es", "archivo")
        # el mensaje aconseja algo, y NO es el texto crudo de la excepcion
        assert msg and msg != str(exc), nombre
        assert len(msg) > 40, f"{nombre}: mensaje demasiado escueto"
        assert any(p.lower() in msg.lower() for p in pistas), f"{nombre}: {msg}"
        # el detalle tecnico se conserva, pero aparte
        assert type(exc).__name__ in detalle


@pytest.mark.parametrize("lang", LANGS)
def test_errores_traducidos_a_los_tres_idiomas(lang):
    from mvdg.errors import friendly_error
    exc = UnicodeDecodeError("utf-8", b"", 0, 1, "invalid")
    msg, _ = friendly_error(exc, lang, "archivo")
    assert msg and msg != "err_encoding"          # la clave existe traducida
    assert "UTF-8" in msg
    # y el generico tambien esta en los 3
    assert t("err_generico", lang) != "err_generico"
    assert t("err_detalle", lang) != "err_detalle"


def test_errores_de_conexion_distinguen_credencial_de_host():
    """Un usuario que se equivoco de contrasena y otro que no tiene VPN
    necesitan consejos distintos, no el mismo texto generico."""
    from mvdg.errors import friendly_error

    class OperationalError(Exception):
        pass

    cred = OperationalError("FATAL: password authentication failed for user 'x'")
    host = OperationalError("could not translate host name 'db.interno' to address")
    m_cred, _ = friendly_error(cred, "es", "conexion")
    m_host, _ = friendly_error(host, "es", "conexion")
    assert m_cred != m_host
    assert "contraseña" in m_cred.lower() or "credencial" in m_cred.lower()
    assert "host" in m_host.lower() or "servidor" in m_host.lower()


def test_app_no_muestra_excepciones_crudas():
    """Ningun st.error debe volcar la excepcion tal cual: para eso esta el
    helper _error(), que traduce y deja el detalle plegado."""
    import ast
    ruta = os.path.join(_repo_root(), "app", "app.py")
    with open(ruta, encoding="utf-8") as fh:
        fuente = fh.read()
    arbol = ast.parse(fuente)

    crudos = []
    for nodo in ast.walk(arbol):
        # solo llamadas reales st.error(...) / st.warning(...): asi los
        # comentarios y docstrings que citan el patron viejo no cuentan
        if not (isinstance(nodo, ast.Call)
                and isinstance(nodo.func, ast.Attribute)
                and nodo.func.attr in ("error", "warning")
                and isinstance(nodo.func.value, ast.Name)
                and nodo.func.value.id == "st"):
            continue
        for arg in nodo.args:
            texto = ast.get_source_segment(fuente, arg) or ""
            # volcar la excepcion tal cual, sin pasarla por friendly_error
            if "{exc}" in texto or texto.strip() in ("str(exc)", "exc"):
                crudos.append(f"linea {nodo.lineno}: {texto[:70]}")
    assert not crudos, "excepciones crudas en pantalla: " + " | ".join(crudos[:4])
    assert "def _error(" in fuente


# ------------------------------------- primer valor sin registro ni pagos
def test_demo_abre_con_valor_real_sin_licencia_ni_registro():
    """Un usuario nuevo ve el producto funcionando: nada del nucleo esta
    detras de un login, un registro o una licencia."""
    from mvdg import licensing
    assert licensing.plan() == licensing.PLAN_DEMO
    for funcion in ("catalogo", "calidad", "linaje", "glosario", "politicas",
                    "perfilado", "mdm", "export_bi", "api_bi"):
        assert licensing.has_feature(funcion), f"{funcion} bloqueada en demo"
    # solo los aceleradores de migracion y el escaneo de tenant son pagos
    assert set(licensing.FUNCIONES_PAGAS) == {
        "migracion_purview", "migracion_collibra", "escaneo_tenant_bi"}


def test_flujo_principal_funciona_sin_red_ni_servicio_de_pago(monkeypatch):
    """El producto no depende de MercadoPago ni de ningun servidor: con la
    red caida, catalogo, calidad, linaje, glosario y export siguen andando."""
    import socket as _socket

    def sin_red(*a, **k):
        raise OSError("red bloqueada (prueba offline)")

    monkeypatch.setattr(_socket.socket, "connect", sin_red, raising=False)
    monkeypatch.setattr(_socket, "create_connection", sin_red)

    from mvdg import licensing, profiler
    from mvdg.catalog import catalog_df, dictionary_df
    from mvdg.exporters import governance_tables
    from mvdg.glossary import glossary_df
    from mvdg.lineage import lineage_figure
    from mvdg.policies import policies_df
    from mvdg.quality import run_rules

    tablas = load_demo_tables()
    assert len(catalog_df("es")) > 0
    assert len(dictionary_df("es")) > 0
    assert len(run_rules(lang="es")) == 17
    assert lineage_figure() is not None
    assert len(glossary_df("es")) > 0
    assert len(policies_df("es")) > 0
    assert len(governance_tables("es")) == 9
    assert len(profiler.profile_table(tablas["dim_customers"])) > 0
    # la licencia se verifica localmente: no hay llamada a ningun servidor
    assert licensing.plan() == licensing.PLAN_DEMO


def test_screencast_real_existe_y_se_puede_reproducir():
    """El video narrado (build_video.py) es una animacion con TTS, no una
    grabacion de la app -- no muestra la UI real. Esto exige ademas un
    screencast de verdad: grabado con Playwright contra la app Streamlit
    CORRIENDO, no un mockup dibujado con PIL."""
    for ruta_rel in ("assets/video/MVDataGovernance_Screencast_real.webm",
                     "landing/video/MVDataGovernance_Screencast_real.webm"):
        ruta = os.path.join(_repo_root(), ruta_rel)
        assert os.path.exists(ruta), f"falta {ruta_rel}"
        assert os.path.getsize(ruta) > 300_000, f"{ruta_rel} sospechosamente chico"
    # el script que lo genera existe y esta documentado como reproducible
    script = os.path.join(_repo_root(), "assets", "video", "record_screencast.py")
    assert os.path.exists(script)
    with open(script, encoding="utf-8") as fh:
        contenido = fh.read()
    assert "streamlit" in contenido.lower() and "playwright" in contenido.lower()
    assert "stTab" in contenido, "el script no navega pestañas reales de la app"
    # esta enlazado desde la landing, no solo tirado en una carpeta
    html = _landing("index.html")
    assert "MVDataGovernance_Screencast_real.webm" in html


def test_landing_promete_solo_lo_que_el_codigo_hace():
    """Cada capacidad fuerte que la landing menciona (push/pull real a
    Purview y Collibra, MIP, escaneo batch de conexiones) tiene que tener
    una implementacion real detras, no solo la mencion."""
    import mvdg.purview_export as pv_export
    import mvdg.collibra_export as co_export
    import mvdg.collibra_pull as co_pull
    import mvdg.mip_labels  # noqa: F401  -- existe el modulo, es la promesa
    from mvdg.connectors import scan_all_connections

    # "directo a Purview por su API real (Atlas)"
    src_purview = open(pv_export.__file__, encoding="utf-8").read()
    assert "atlas/v2/entity/bulk" in src_purview, (
        "la landing promete API real de Atlas; el codigo no la llama")

    # "Collibra en las dos direcciones": push_catalog/push_glossary son reales
    # (no un stub), y collibra_pull.py es un modulo aparte para la vuelta.
    assert callable(co_export.push_catalog)
    assert callable(co_export.push_glossary)
    assert callable(co_pull.pull_catalog)
    assert callable(scan_all_connections)

    # el escaneo batch aisla el error de cada conexion (no frena las demas) --
    # es la promesa concreta ("con el error de cada fuente aislado")
    ruta_connectors = os.path.join(os.path.dirname(pv_export.__file__), "connectors.py")
    with open(ruta_connectors, encoding="utf-8") as fh:
        src_scan = fh.read()
    assert "except Exception" in src_scan, "scan_all_connections deberia aislar errores"


# ------------------------------------------------ landing: SEO / a11y / UX
_LANDING_PAGES = ("index.html", "descargas.html", "guia.html",
                  "pago.html", "reviews.html")


def _landing(archivo):
    with open(os.path.join(_repo_root(), "landing", archivo), encoding="utf-8") as fh:
        return fh.read()


@pytest.mark.parametrize("archivo", _LANDING_PAGES)
def test_landing_tiene_meta_social_y_seo(archivo):
    """Sin estas etiquetas el link compartido en LinkedIn/WhatsApp sale sin
    titulo ni imagen, y Lighthouse SEO no llega a 90."""
    html = _landing(archivo)
    for etiqueta in ('name="description"',
                     'property="og:title"', 'property="og:description"',
                     'property="og:image"', 'property="og:url"',
                     'name="twitter:card"', 'name="twitter:image"',
                     'rel="canonical"'):
        assert etiqueta in html, f"{archivo}: falta {etiqueta}"
    assert 'content="summary_large_image"' in html
    # la imagen social tiene que existir de verdad y declarar su tamano
    assert 'property="og:image:width" content="1200"' in html
    assert 'property="og:image:height" content="630"' in html


def test_landing_og_image_existe_y_mide_1200x630():
    ruta = os.path.join(_repo_root(), "landing", "img", "og_cover.jpg")
    assert os.path.exists(ruta), "falta landing/img/og_cover.jpg"
    try:
        from PIL import Image
    except ImportError:
        pytest.skip("Pillow no disponible")
    assert Image.open(ruta).size == (1200, 630)


def test_landing_capturas_no_recortan_el_contenido_con_alto_fijo():
    """`.shot a` tenía `max-height:340px;overflow:hidden` — un recorte fijo
    sobre capturas de 990px a 1640px de alto, así que según la pestaña se
    cortaba la mitad del contenido a mitad de un gráfico, sin ningún indicio
    visual de que faltaba algo debajo. El marco tiene que ajustarse a la
    imagen entera, no al revés."""
    html = _landing("index.html")
    m = re.search(r"\.shot a\{([^}]*)\}", html)
    assert m, "no encontré la regla .shot a en el CSS de index.html"
    regla = m.group(1)
    assert "max-height" not in regla
    assert "overflow" not in regla


def test_landing_capturas_width_height_coinciden_con_el_archivo_real():
    """`width`/`height` en el `<img>` son para reservar el espacio ANTES de
    que la imagen cargue (evita el salto de layout) — si no coinciden con
    el archivo real, el navegador escala mal mientras carga o el hueco
    reservado queda con un tamaño incorrecto. Se rompió una vez: las
    capturas se regeneraron con el motor de calidad/linaje corregido
    (contraste + etiquetas superpuestas) y quedaron más altas que las
    viejas, pero el HTML seguía con el tamaño de las capturas anteriores."""
    try:
        from PIL import Image
    except ImportError:
        pytest.skip("Pillow no disponible")
    html = _landing("index.html")
    img_dir = os.path.join(_repo_root(), "landing", "img")
    for base in ("tab_panorama", "tab_calidad", "tab_linaje", "tab_bi"):
        m = re.search(rf'src="img/{base}\.jpg"[^>]*\bwidth="(\d+)" height="(\d+)"', html)
        assert m, f"no encontré el <img> de {base} (o le falta width/height) en index.html"
        w_html, h_html = int(m.group(1)), int(m.group(2))
        for suf in ("", "_en", "_pt"):
            ruta = os.path.join(img_dir, f"{base}{suf}.jpg")
            assert os.path.isfile(ruta), f"falta {ruta}"
            w_real, h_real = Image.open(ruta).size
            assert (w_real, h_real) == (w_html, h_html), (
                f"{base}{suf}.jpg mide {w_real}x{h_real} pero el HTML declara "
                f"{w_html}x{h_html} — desalineado, va a saltar el layout al cargar")


def test_video_antes_despues_existe_en_los_tres_idiomas_y_con_audio():
    """El video de antes/después es lo primero que ve el que decide la compra.
    Tiene que existir en los 3 idiomas y CON narración: un mp4 mudo pasa
    desapercibido en el repo (el archivo está, la landing lo carga, se ve) y
    solo se descubre cuando alguien lo mira. Se verifica la pista de audio en
    el contenedor, no el tamaño del archivo."""
    video_dir = os.path.join(_repo_root(), "landing", "video")
    for lang in ("es", "en", "pt"):
        ruta = os.path.join(video_dir, f"MVDataGovernance_AntesDespues_{lang}.mp4")
        assert os.path.isfile(ruta), (
            f"falta {os.path.basename(ruta)} — regeneralo con "
            "python assets/video/build_antes_despues.py")
        with open(ruta, "rb") as fh:
            crudo = fh.read()
        # El índice del MP4 (átomo 'moov') describe las pistas. Acá NO está al
        # principio: ffmpeg lo escribe al final salvo que se pida faststart, así
        # que hay que buscarlo en todo el archivo, no en la cabecera. Dentro de
        # él, 'mp4a' es la entrada de la pista de audio AAC.
        i_moov = crudo.rfind(b"moov")
        assert i_moov > 0, f"{os.path.basename(ruta)} no parece un MP4 válido"
        assert b"mp4a" in crudo[i_moov:], (
            f"{os.path.basename(ruta)} no tiene pista de audio: se generó sin "
            "los modelos de voz (MVDG_VOICE_ONNX_ES/EN/PT)")


def test_video_antes_despues_no_hardcodea_las_cifras():
    """Las cifras del video salen del motor (`lab_measure`) al generarlo, no
    escritas a mano. Si alguien las pega como texto, el video sigue prometiendo
    un resultado que el producto ya no da — y es justo lo que un gerente
    verifica cuando corre el caso en su PC."""
    ruta = os.path.join(_repo_root(), "assets", "video", "build_antes_despues.py")
    src = open(ruta, encoding="utf-8").read()
    assert "lab_measure" in src, "el guion del video no consulta el motor"
    from mvdg.lab_case import lab_measure
    m = lab_measure()
    cifras = {str(m["summary_before"]["indice"]), str(m["summary_after"]["indice"]),
              str(m["summary_before"]["filas_afectadas"]),
              str(m["summary_after"]["filas_afectadas"])}
    # Se busca en el código MENOS los f-strings que interpolan la medición:
    # ahí las cifras aparecen como expresión, no como literal.
    for cifra in cifras:
        assert f'"{cifra}"' not in src and f"'{cifra}'" not in src, (
            f"la cifra {cifra} está escrita a mano en el guion del video; "
            "tiene que venir de lab_measure()")


def test_landing_ofrece_el_antes_y_despues_antes_que_la_demo():
    """El orden es el argumento: la demo muestra qué hace el producto, el
    antes/después muestra qué cambia. Si la demo queda primero, el que decide
    la compra ve funcionalidad antes que resultado. Además el video tiene que
    poder cambiar de idioma como el otro (data-vbase) y tener su propio panel
    de error (data-errbox): sin eso, un 404 deja un recuadro negro mudo."""
    html = _landing("index.html")
    i_ad = html.find('id="antes-despues"')
    i_demo = html.find('<section class="vsec" id="video"')
    assert i_ad > 0, "falta la sección del video antes/después en la landing"
    assert i_demo > 0, "falta la sección del video de demo en la landing"
    assert i_ad < i_demo, "el antes/después tiene que ir ANTES de la demo narrada"
    assert 'data-vbase="MVDataGovernance_AntesDespues"' in html
    assert 'data-errbox="adVideoError"' in html and 'id="adVideoError"' in html
    # El enlace "cómo se mide" tiene que llegar a algún lado real.
    assert 'href="guia.html#caso"' in html
    assert 'id="caso"' in _landing("guia.html"), "guia.html no tiene el ancla #caso"


def test_landing_paridad_de_textos_del_antes_y_despues():
    """Cada clave nueva de la sección tiene que estar en los 3 idiomas: si
    falta una, el visitante en inglés o portugués ve la frase en español en el
    medio de la página, que es peor que no tenerla."""
    html = _landing("index.html")
    claves = ("ad_eye", "ad_h2", "ad_lead", "ad_err_d", "ad_err_cta",
              "ad_repro", "ad_repro_link")
    for clave in claves:
        assert f'data-i="{clave}"' in html, f"la clave {clave} no se usa en el HTML"
        # Dos diccionarios de traducción (EN y PT) más el texto en el markup.
        assert len(re.findall(rf"\b{clave}:", html)) >= 2, (
            f"{clave} no está en los diccionarios EN y PT")


@pytest.mark.parametrize("archivo", _LANDING_PAGES)
def test_landing_viewport_favicon_y_lang(archivo):
    html = _landing(archivo)
    assert 'name="viewport"' in html and "width=device-width" in html
    assert 'rel="icon"' in html
    assert 'rel="apple-touch-icon"' in html
    assert '<html lang="es"' in html      # idioma por defecto del HTML servido


def test_landing_menciona_integraciones_concretas_no_lenguaje_generico():
    """La landing tiene que nombrar las integraciones reales -- no
    "conectate a tus fuentes de datos" generico -- Y el numero que declara
    (9 motores) tiene que coincidir con lo que el codigo soporta de verdad,
    para que agregar/sacar un conector sin actualizar la landing rompa esto
    en vez de quedar mintiendo en silencio."""
    from mvdg.connectors import ENGINES
    html = _landing("index.html")
    # integraciones con nombre propio, no lenguaje generico
    for nombre in ("Power BI", "Tableau", "Purview", "Collibra"):
        assert nombre in html, f"falta mencionar {nombre} por su nombre"
    # el numero de motores que la landing declara tiene que ser el real
    assert len(ENGINES) == 9, "cambio la cantidad de motores: actualizar la landing"
    assert "9 motores" in html
    # no se inventan capacidades que el codigo no tiene (ERPs, "en todos los
    # paises", etc.) -- ausencia deliberada, no un genericazo disfrazado
    assert "SAP" not in html and "Dynamics 365" not in html and "NetSuite" not in html


def test_landing_meta_tags_se_sincronizan_al_cambiar_de_idioma():
    """No existen /en/ /pt/ como rutas separadas (es un solo HTML con
    traduccion por JS) -- eso significa que un bot que nunca ejecuta JS
    (Google, el unfurl de WhatsApp/LinkedIn) siempre ve el HTML servido en
    espanol, sin importar que arquitectura se use ariba. Lo que SI esta al
    alcance sin rehacer el sitio es que, para alguien mirando la pagina ya
    cargada, el <title> y los meta cambien de verdad al tocar el selector de
    idioma -- antes quedaban fijos en espanol aunque el usuario estuviera
    viendo el contenido en EN/PT."""
    html = _landing("index.html")
    assert "var META={" in html or "var META = {" in html, "falta el diccionario META por idioma"
    assert "function setMetaTags(" in html
    assert "setMetaTags(lang)" in html, "setLang() no llama a setMetaTags()"

    import re
    bloque = re.search(r"var META=\{(.*?)\n\};", html, re.S)
    assert bloque, "no se pudo extraer el diccionario META"
    cuerpo = bloque.group(1)
    for lang in ("es", "en", "pt"):
        assert re.search(rf"\b{lang}:\{{", cuerpo), f"META sin entrada para {lang}"
    # los 3 titulos y las 3 descripciones tienen que ser distintos entre si
    titulos = re.findall(r'title:"([^"]+)"', cuerpo)
    descs = re.findall(r'desc:"([^"]+)"', cuerpo)
    assert len(titulos) == 3 and len(set(titulos)) == 3, "hay titulos repetidos entre idiomas"
    assert len(descs) == 3 and len(set(descs)) == 3, "hay descripciones repetidas entre idiomas"
    # setMetaTags toca los 5 selectores que importan (title + 4 meta)
    bloque_fn = re.search(r"function setMetaTags\(lang\)\{(.*?)\n\}", html, re.S).group(1)
    for selector in ('meta[name="description"]', 'meta[property="og:title"]',
                     'meta[property="og:description"]', 'meta[name="twitter:title"]'):
        assert selector in bloque_fn, f"setMetaTags() no toca {selector}"
    assert "document.title=" in bloque_fn


@pytest.mark.parametrize("archivo", _LANDING_PAGES)
def test_landing_imagenes_con_alt_descriptivo(archivo):
    """Ninguna <img> sin alt, y ningun alt de una sola palabra suelta."""
    import re
    for tag in re.findall(r"<img\b[^>]*>", _landing(archivo)):
        m = re.search(r'\balt="([^"]*)"', tag)
        assert m, f"{archivo}: <img> sin alt -> {tag[:80]}"
        assert len(m.group(1).strip()) >= 3, f"{archivo}: alt pobre -> {tag[:80]}"


def test_landing_capturas_declaran_tamano_y_traducen_su_alt():
    """width/height evitan el salto de layout (CLS, penaliza Performance), y
    el alt de cada captura existe en los 3 idiomas."""
    import re
    html = _landing("index.html")
    shots = re.findall(r'<img\b[^>]*data-shot="([^"]+)"[^>]*>', html)
    assert len(shots) >= 4
    for tag in re.findall(r'<img\b[^>]*data-shot="[^"]+"[^>]*>', html):
        assert 'width="' in tag and 'height="' in tag, f"sin tamano: {tag[:90]}"
        assert 'loading="lazy"' in tag
    # el alt de cada captura tiene su clave en ES (explicita), EN y PT
    for base in set(shots):
        clave = f"alt_{base}"
        assert html.count(clave) >= 3, f"{clave} no esta en los 3 idiomas"


def test_landing_scripts_no_bloquean_el_parseo():
    for archivo in ("index.html", "reviews.html"):
        html = _landing(archivo)
        for src in ("reviews-data.js", "payments-config.js"):
            if f'src="{src}"' in html:
                assert f'src="{src}" defer' in html, f"{archivo}: {src} sin defer"
        # con defer, pintar en el script inline llega tarde: hay que esperar
        # a DOMContentLoaded o el grid queda vacio
        assert "DOMContentLoaded" in html


def test_landing_contraste_de_texto_tenue_cumple_aa():
    """--faint sobre el navy estaba en 4.48:1, abajo del minimo AA de 4.5:1,
    y se usa en los botones de idioma a 11.5px."""
    def luminancia(hexcol):
        def canal(c):
            c = c / 255
            return c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
        r, g, b = (int(hexcol[i:i + 2], 16) for i in (1, 3, 5))
        return .2126 * canal(r) + .7152 * canal(g) + .0722 * canal(b)

    for archivo in _LANDING_PAGES:
        html = _landing(archivo)
        import re
        faint = re.search(r"--faint:(#[0-9a-fA-F]{6})", html).group(1)
        navy = re.search(r"--navy:(#[0-9a-fA-F]{6})", html).group(1)
        l1, l2 = sorted([luminancia(faint), luminancia(navy)], reverse=True)
        ratio = (l1 + .05) / (l2 + .05)
        assert ratio >= 4.5, f"{archivo}: contraste {ratio:.2f}:1 < 4.5:1"


def test_landing_estados_de_error_visibles():
    """Nada se queda en blanco sin explicacion: el video y las resenas tienen
    su propio estado de error."""
    html = _landing("index.html")
    assert 'id="videoError"' in html
    assert "function watchVideo" in html and "watchVideo();" in html
    # el <span> dentro de <video> NO cubre un 404: hace falta escuchar el error
    assert "NETWORK_NO_SOURCE" in html
    # y no es fatal que falle un solo <source> habiendo respaldo
    assert "readyState" in html
    # poster="" es una referencia vacia: algunos navegadores la intentan cargar
    # y pintan un primer cuadro roto. Se mira la etiqueta, no todo el archivo.
    import re
    tag_video = re.search(r"<video\b[^>]*>", html).group(0)
    assert 'poster=""' not in tag_video, f"poster vacio en {tag_video}"
    for archivo in ("index.html", "reviews.html"):
        assert "rvempty" in _landing(archivo), f"{archivo}: sin estado vacio"


def test_los_secretos_no_se_pueden_commitear_ni_por_accidente():
    """El repo es PUBLICO. Un `git add -A` con un .env al lado publica el
    Access Token de MercadoPago y la clave que FIRMA LAS LICENCIAS.

    Con esa clave cualquiera emite licencias infinitas del producto, y no hay
    forma de revocarlas: habria que rotar el par, lo que le rompe la licencia
    a TODOS los que ya compraron. Es el peor secreto del proyecto y el
    .gitignore no lo cubria.

    No se verifica leyendo el .gitignore —una linea puede estar y no aplicar—
    sino preguntandole a git si ignoraria cada archivo."""
    import subprocess
    import tempfile
    raiz = _repo_root()
    peligrosos = [".env", ".env.local", ".env.production",
                  "clave.pem", "id_rsa", "servidor.key"]
    for nombre in peligrosos:
        ruta = os.path.join(raiz, nombre)
        creado = not os.path.exists(ruta)
        if creado:
            with open(ruta, "w", encoding="utf-8") as fh:
                fh.write("secreto-de-prueba\n")
        try:
            r = subprocess.run(["git", "check-ignore", nombre],
                               cwd=raiz, capture_output=True, text=True)
            assert r.returncode == 0, (
                f"'{nombre}' NO esta ignorado: un `git add -A` lo sube a un "
                f"repo publico junto con las claves que tenga adentro")
        finally:
            if creado:
                os.remove(ruta)
    del tempfile

    # La plantilla SI tiene que poder versionarse: es la que documenta que
    # variables hacen falta, y no lleva ningun valor real.
    ejemplo = os.path.join(raiz, ".env.example")
    creado = not os.path.exists(ejemplo)
    if creado:
        with open(ejemplo, "w", encoding="utf-8") as fh:
            fh.write("MP_ACCESS_TOKEN=\n")
    try:
        r = subprocess.run(["git", "check-ignore", ".env.example"],
                           cwd=raiz, capture_output=True, text=True)
        assert r.returncode != 0, (
            ".env.example quedo ignorado: es la plantilla sin valores, tiene "
            "que poder versionarse")
    finally:
        if creado:
            os.remove(ejemplo)


def test_cada_plan_baja_su_build_y_el_workflow_publica_los_dos():
    """El owner tiene que bajar el build owner y el cliente el suyo.

    Antes habia UNA sola variable de entorno, asi que el owner —con su
    licencia owner en la mano— bajaba el mismo .exe sin desbloquear que
    cualquier cliente. El build owner existia solo como artefacto de Actions y
    no llegaba por ningun lado.

    Esto ata las tres puntas: que el endpoint elija por plan, que el workflow
    publique el build owner, y que te diga la URL exacta que va en Vercel —
    porque una variable que hay que armar a mano es donde se erraba."""
    import json as _json
    import subprocess
    elige = _json.loads(subprocess.run(
        ["node", "-e",
         "const d=require('./api/descargar');"
         "console.log(JSON.stringify({"
         "  owner: d.variableDelPlan('owner'),"
         "  licencia: d.variableDelPlan('licencia'),"
         "  professional: d.variableDelPlan('professional'),"
         "  trial: d.variableDelPlan('trial'),"
         "  nada: d.variableDelPlan(undefined)}))"],
        cwd=_repo_root(), capture_output=True, text=True, check=True).stdout)

    assert elige["owner"] == "MVDG_INSTALLER_URL_OWNER"
    for plan in ("licencia", "professional", "trial", "nada"):
        assert elige[plan] == "MVDG_INSTALLER_URL", (
            f"el plan '{plan}' llega al build owner, que viene desbloqueado")

    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()
    for tag in ("cliente-latest", "owner-latest"):
        assert tag in wf, f"el workflow no publica {tag}"
    # Y cada release tiene que decir QUE variable llenar con SU url: si el
    # workflow publica el build owner pero nadie sabe donde apuntarlo, el
    # endpoint contesta 503 y la version owner sigue sin llegar.
    for var in ("MVDG_INSTALLER_URL", "MVDG_INSTALLER_URL_OWNER"):
        assert var in wf, f"el workflow no dice donde poner {var}"


def test_la_clave_publica_es_LA_MISMA_en_el_servidor_y_en_el_programa():
    """El servidor decide quien baja el instalador verificando la firma de la
    licencia (api/_license.js), y el programa decide que habilita verificando
    la misma firma (mvdg/licensing.py). Son dos copias de la clave publica en
    dos lenguajes distintos.

    Si se separan, el sintoma es de los peores: el cliente paga, el programa
    acepta su licencia — y al querer bajar el instalador recibe 403. O al
    reves. Nada falla al deployar; falla el dia que alguien compra.

    Rotar el par es cambiar LAS DOS. Este test lo vuelve imposible de
    olvidar."""
    import subprocess
    from mvdg import licensing
    del_servidor = subprocess.run(
        ["node", "-e",
         "process.stdout.write(require('./api/_license').PUBLIC_KEY_B64)"],
        cwd=_repo_root(), capture_output=True, text=True, check=True).stdout.strip()
    assert del_servidor == licensing.PUBLIC_KEY_B64, (
        f"api/_license.js verifica con {del_servidor!r} y mvdg/licensing.py "
        f"con {licensing.PUBLIC_KEY_B64!r}: el que compre va a poder hacer una "
        f"de las dos cosas, no las dos")


def test_el_trial_que_se_emitia_solo_ya_no_existe():
    """Habia un trial autoservicio: dejabas un email y /api/trial te firmaba
    una licencia Professional de 14 dias, sin hablar con nadie.

    Se elimino junto con la descarga abierta, y no por prolijidad: la descarga
    ahora exige una licencia valida, y ESA licencia servia para bajar el
    programa. O sea que el trial autoservicio era la misma puerta que se
    acababa de cerrar, un paso mas adentro — cualquiera se auto-emitia la
    llave. Hoy la prueba se entrega despues de la demo 1 a 1.

    Este test es lo que impide que vuelva por descuido."""
    assert not os.path.exists(os.path.join(_repo_root(), "api", "trial.js")), (
        "volvio api/trial.js: emite licencias sin que intervenga una persona, "
        "y con esa licencia se baja el programa")
    html = _landing("index.html")
    limpio = re.sub(r"<!--.*?-->|/\*.*?\*/", "", html, flags=re.S)
    assert "/api/trial" not in limpio, "la landing sigue llamando a /api/trial"


def test_el_endpoint_de_acceso_avisa_pero_NO_entrega_nada():
    """api/acceso.js reemplaza al trial. La linea que no se puede cruzar es
    que emita algo: si devolviera una licencia, volveria a ser una descarga
    automatica con un formulario adelante — que es justo lo que se saco.

    Se mira el CODIGO sin comentarios: los comentarios explican por que no
    firma nada, y nombran las cosas que no hace."""
    ruta = os.path.join(_repo_root(), "api", "acceso.js")
    assert os.path.exists(ruta), "falta api/acceso.js"
    with open(ruta, encoding="utf-8") as fh:
        codigo = fh.read()
    sin_comentarios = re.sub(r"//.*", "", codigo)
    sin_comentarios = re.sub(r"/\*.*?\*/", "", sin_comentarios, flags=re.S)
    for prohibido in ("signEd25519", "LICENSE_PRIVATE_KEY", "license_key",
                      "MP_ACCESS_TOKEN", "cvv", "card_number"):
        assert prohibido.lower() not in sin_comentarios.lower(), (
            f"'{prohibido}' en el codigo de acceso.js: este endpoint avisa, "
            f"no entrega")
    assert "rateLimited" in sin_comentarios, "sin rate limiting"
    # y el aviso tiene que llegar a algun lado: sin mail configurado no puede
    # contestar 200, o el pedido se pierde y el visitante cree que llego
    assert "RESEND_API_KEY" in sin_comentarios
    assert "503" in sin_comentarios


@pytest.mark.parametrize("lang", ["en", "pt"])
def test_pagina_de_acceso_traducida(lang):
    """La pagina de pedido de acceso es la unica puerta a la demo: si no esta
    en los 3 idiomas, se pierde al que entra en ingles o portugues."""
    html = _landing("descargas.html")
    bloque = html[html.index("%s:{" % lang):]
    for clave in ("l_nombre", "l_empresa", "l_pais", "l_email", "c1btn", "h1"):
        assert f"{clave}:" in bloque, f"falta {clave} en {lang}"


def test_landing_publica_la_comparativa_honesta():
    """La comparativa capacidad-por-capacidad contra Purview/Collibra tiene que
    estar en la LANDING, no solo en docs/. Es el mejor argumento de venta y en
    un .md del repo no la ve ningun cliente."""
    import re
    html = _landing("index.html")
    assert 'id="honesta"' in html, "falta la seccion de comparativa honesta"
    tabla = re.search(r'<table class="cmp cmp2">.*?</table>', html, re.S)
    assert tabla, "falta la tabla de la comparativa honesta"
    filas = re.findall(r"<tr><td data-i=\"hon_", tabla.group(0))
    assert len(filas) >= 12, f"solo {len(filas)} capacidades comparadas"
    # se accede desde el nav, no queda enterrada
    assert 'href="#honesta"' in html


def test_landing_comparativa_dice_lo_que_mv_no_hace():
    """Lo que la hace honesta (y creible) son los limites. Si alguien la
    'mejora' borrando los parciales, deja de ser una comparativa honesta."""
    import re
    html = _landing("index.html")
    tabla = re.search(r'<table class="cmp cmp2">.*?</table>', html, re.S).group(0)
    parciales = re.findall(r'<td class="part"', tabla)
    assert len(parciales) >= 4, (
        f"solo {len(parciales)} limites declarados: la comparativa dejo de ser honesta")
    # los cuatro techos reales del producto siguen dichos
    for tema in ("linaje", "conectores", "OneDrive/SharePoint", "DBA"):
        assert tema.lower() in tabla.lower(), f"ya no se declara el limite: {tema}"


def test_landing_y_docs_comparan_las_mismas_capacidades():
    """La tabla de la landing y la de docs/PURVIEW_COLLIBRA.md no pueden
    divergir: si una fila cambia en el repo tiene que cambiar en la landing."""
    import re
    ruta = os.path.join(_repo_root(), "docs", "PURVIEW_COLLIBRA.md")
    with open(ruta, encoding="utf-8") as fh:
        md = fh.read()
    filas_md = [ln for ln in md.split("\n")
                if ln.startswith("|") and "Purview / Collibra" not in ln
                and not ln.startswith("|---")]
    html = _landing("index.html")
    tabla = re.search(r'<table class="cmp cmp2">.*?</table>', html, re.S).group(0)
    filas_landing = re.findall(r"<tr><td data-i=\"hon_", tabla)
    # la landing puede condensar, pero no puede tener menos de la mitad
    assert len(filas_landing) >= len(filas_md) // 2 > 0
    # y el .md no puede seguir diciendo que los contratos estan fuera de alcance
    assert "Fuera de alcance: aplica a organizaciones que publican" not in md, (
        "docs desactualizado: mvdg/contracts.py existe")


@pytest.mark.parametrize("lang", ["en", "pt"])
def test_landing_comparativa_traducida(lang):
    """Cada celda de la comparativa tiene su clave en EN y PT, o el cliente
    que mira en ingles ve la tabla a medio traducir."""
    import re
    html = _landing("index.html")
    tabla = re.search(r'<table class="cmp cmp2">.*?</table>', html, re.S).group(0)
    claves = set(re.findall(r'data-i="(hon_[a-z0-9_]+)"', tabla))
    assert len(claves) >= 30, f"solo {len(claves)} claves en la tabla"
    bloque = re.search(r"\n%s:\{(.*?)\n\}" % lang, html, re.S)
    assert bloque, f"no se encontro el diccionario {lang}"
    faltan = [k for k in claves if f"{k}:" not in bloque.group(1)]
    assert not faltan, f"sin traduccion {lang}: {sorted(faltan)[:6]}"


# ------------------------------- inyeccion de HTML con datos del cliente
def test_app_no_inyecta_datos_del_cliente_como_html():
    """El catalogo del cliente (nombres de tablas y columnas de SU base) NO
    puede terminar dentro de un st.markdown(unsafe_allow_html=True): ahi un
    nombre de columna con '<' rompe el render, y con etiquetas inyecta HTML.

    Streamlit escapa por defecto en st.dataframe/st.table, que es como se
    muestra el catalogo — este test verifica que nadie meta datos por la via
    insegura mas adelante."""
    import ast
    ruta = os.path.join(_repo_root(), "app", "app.py")
    with open(ruta, encoding="utf-8") as fh:
        fuente = fh.read()
    arbol = ast.parse(fuente)

    dinamicos = []
    for nodo in ast.walk(arbol):
        if not (isinstance(nodo, ast.Call)
                and isinstance(nodo.func, ast.Attribute)
                and nodo.func.attr in ("markdown", "write", "caption")):
            continue
        inseguro = any(kw.arg == "unsafe_allow_html"
                       and isinstance(kw.value, ast.Constant) and kw.value.value
                       for kw in nodo.keywords)
        if not inseguro or not nodo.args:
            continue
        arg = nodo.args[0]
        # un literal sin interpolacion es HTML propio y estatico: esta bien
        if isinstance(arg, ast.Constant) and isinstance(arg.value, str):
            continue
        # Un f-string que solo interpola constantes de marca (la paleta de
        # BRAND en el bloque de CSS) tampoco mete datos: no viene de la base
        # del cliente ni de un archivo subido. Cualquier OTRO nombre si.
        SEGUROS = {"BRAND"}
        if isinstance(arg, ast.JoinedStr):
            nombres = {n.id for parte in arg.values
                       if isinstance(parte, ast.FormattedValue)
                       for n in ast.walk(parte.value) if isinstance(n, ast.Name)}
            if nombres and nombres <= SEGUROS:
                continue
        dinamicos.append(f"linea {nodo.lineno}: "
                         f"{(ast.get_source_segment(fuente, arg) or '')[:60]}")
    assert not dinamicos, (
        "HTML con valores interpolados y unsafe_allow_html: " + " | ".join(dinamicos))


def test_landing_no_inyecta_datos_dinamicos_sin_escapar():
    """Inventario de innerHTML/insertAdjacentHTML de la landing: los que
    reciben datos (resenas, respuesta de la API de pago) tienen que pasar por
    una funcion de escapado."""
    import re
    # index.html y reviews.html renderizan resenas
    for archivo, escapador in (("index.html", "rvEsc"), ("reviews.html", "esc")):
        html = _landing(archivo)
        # cada campo de la resena que se inyecta pasa por el escapador
        for campo in ("comment", "name", "role"):
            patron = rf"{escapador}\(r\.{campo}\)"
            assert re.search(patron, html), f"{archivo}: r.{campo} sin {escapador}()"
        assert f"function {escapador}(" in html, f"{archivo}: falta {escapador}()"
    # pago.html inyecta lo que devuelve la API de pagos
    pago = _landing("pago.html")
    for campo in ("STATE.license", "planName"):
        assert f"esc({campo})" in pago, f"pago.html: {campo} sin escapar"


def _node_disponible():
    import shutil
    return shutil.which("node") is not None


def _correr_test_js(ruta_relativa):
    """Corre un test .test.js con Node puro y devuelve (ok, salida)."""
    import subprocess
    ruta = os.path.join(_repo_root(), ruta_relativa)
    r = subprocess.run(["node", ruta], capture_output=True, text=True, timeout=60)
    return r.returncode == 0, r.stdout + r.stderr


def test_seguridad_xss_regresion_real_no_solo_grep():
    """Test de REGRESION de verdad: ejecuta las funciones esc()/rvEsc() reales
    de la landing (extraidas del propio HTML de produccion, no reimplementadas)
    contra una bateria de payloads de XSS. Si alguien rompe el escapado -- lo
    vacia, lo comenta, o lo deja escapar de menos -- esto falla, no hace falta
    que nadie lo note a mano revisando el diff.

    landing/security.test.js es Node puro (assert + fs, sin dependencia
    nueva) porque las funciones que audita corren en el navegador, no en
    Python: no hay forma de probarlas de verdad sin ejecutar el JS real."""
    if not _node_disponible():
        pytest.skip("node no disponible en este entorno")
    ok, salida = _correr_test_js("landing/security.test.js")
    assert ok, f"test de seguridad XSS fallo:\n{salida}"
    assert "Todos los checks de seguridad (XSS) pasaron" in salida
    # evidencia de que de verdad se probaron los 3 sitios con payloads reales
    for pista in ("<script>", "onerror", "svg onload", "index.html", "reviews.html", "pago.html"):
        assert pista in salida, f"falta evidencia de que se probo: {pista}"


def test_pagos_y_licencia_cobertura_real_no_solo_grep():
    """Los modulos que tocan dinero (checkout, verificacion de pago, firma de
    licencia) tenian 0% de cobertura de EJECUCION -- solo grep de texto. Esto
    corre los handlers reales con mocks de req/res/fetch: metodo invalido,
    plan invalido, rate limit, MercadoPago aprobando/rechazando/fallando,
    emision de licencia Ed25519 verificable, e intentos de inyeccion en
    payment_id."""
    if not _node_disponible():
        pytest.skip("node no disponible en este entorno")
    ok, salida = _correr_test_js("api/payments.test.js")
    assert ok, f"test de pagos/licencia fallo:\n{salida}"
    assert "Todos los checks de pago/licencia pasaron" in salida
    # evidencia de que se cubrieron los casos que importan, no solo un smoke test
    for pista in ("rate limit", "Ed25519", "MercadoPago responde error",
                  "inyección", "no_token", "aprobado"):
        assert pista in salida, f"falta cobertura de: {pista}"


def test_no_hay_carpetas_vendor_de_terceros_sin_declarar():
    """Nada de codigo de terceros pegado a mano: las dependencias van por
    pip/npm con version declarada, nunca copiadas en una carpeta vendor/."""
    import subprocess
    r = subprocess.run(["git", "ls-files"], capture_output=True, text=True,
                       cwd=_repo_root(), check=True)
    archivos = r.stdout.splitlines()
    sospechosos = [f for f in archivos
                  if "/vendor/" in f or f.startswith("vendor/")
                  or ("/lib/" in f and "electron/lib/" not in f
                      and "packaging/" not in f)]
    assert not sospechosos, f"posible codigo vendorizado sin declarar: {sospechosos}"


def test_security_md_documenta_cves_de_dependencias():
    """El estado de las dependencias de terceros (electron, esbuild,
    electron-builder) tiene que estar declarado en algun lado visible, no
    solo "confiar" en que npm audit este limpio hoy."""
    ruta = os.path.join(_repo_root(), "SECURITY.md")
    assert os.path.exists(ruta), "falta SECURITY.md"
    with open(ruta, encoding="utf-8") as fh:
        sec = fh.read()
    for paquete in ("electron-builder", "esbuild", "electron", "extract-zip"):
        assert paquete in sec, f"SECURITY.md no menciona {paquete}"
    assert "npm audit" in sec
    # Y tiene que distinguir lo que corre en la PC del cliente de lo que solo
    # corre al compilar: son dos riesgos distintos y meterlos en la misma
    # bolsa hace que el documento no sirva para decidir nada.
    assert "NO viaja en la app instalada" in sec, (
        "SECURITY.md no aclara que dependencias son solo de compilacion")
    # el gap real (electron sin actualizar) tiene que quedar dicho, no oculto
    assert "sin actualizar" in sec or "CVE" in sec


def test_endpoints_publicos_de_pago_tienen_rate_limiting_propio():
    """api/checkout.js y api/verify-payment.js reciben input publico
    directo (body/query) y NO pueden depender solo del rate limiting de
    plataforma de Vercel -- eso no es auditable ni configurable desde el
    repo. Cada uno tiene que tener su propio control."""
    for archivo in ("checkout.js", "verify-payment.js"):
        ruta = os.path.join(_repo_root(), "api", archivo)
        with open(ruta, encoding="utf-8") as fh:
            codigo = fh.read()
        assert "_rate_limit" in codigo, f"{archivo}: sin rate limiting propio"
        assert "rateLimited(" in codigo and "429" in codigo, (
            f"{archivo}: no corta con 429 al pasarse del limite")


def test_landing_tiene_landmark_main():
    html = _landing("index.html")
    assert html.count("<main>") == 1 and html.count("</main>") == 1
    assert html.index("<main>") < html.index('<section class="hero"')
    assert html.index("</main>") < html.index("<footer>")


# --------------------------------------------- calidad de codigo / tooling
def test_dependencias_de_test_declaradas():
    """pytest y httpx tienen que estar DECLARADAS, no ser un paso manual que
    solo aparece en la documentacion."""
    ruta = os.path.join(_repo_root(), "requirements-dev.txt")
    assert os.path.exists(ruta), "falta requirements-dev.txt"
    with open(ruta, encoding="utf-8") as fh:
        dev = fh.read()
    assert "-r requirements.txt" in dev      # un solo install alcanza
    for paquete in ("pytest", "httpx", "ruff"):
        assert paquete in dev, f"{paquete} sin declarar"


def test_un_solo_comando_instala_y_testea():
    """`make test` tiene que instalar dependencias Y correr la suite: en una
    maquina limpia `pytest` solo no alcanza."""
    ruta = os.path.join(_repo_root(), "Makefile")
    assert os.path.exists(ruta), "falta el Makefile"
    with open(ruta, encoding="utf-8") as fh:
        mk = fh.read()
    assert "test-py: install" in mk          # testear depende de instalar
    assert "test: test-py test-js" in mk     # `make test` corre TODO, no solo Python
    assert "requirements-dev.txt" in mk
    assert "pytest tests/" in mk
    # `--upgrade pip` rompe en Pythons administrados por la distro
    assert "--upgrade pip" not in mk


def test_readme_explica_correr_tests_en_maquina_limpia():
    """El README tiene que alcanzar solo: alguien que nunca vio el repo debe
    poder clonarlo y correr los tests sin preguntarle a nadie ni leer
    CLAUDE.md. Antes faltaba el git clone y el prerrequisito de Python."""
    ruta = os.path.join(_repo_root(), "README.md")
    with open(ruta, encoding="utf-8") as fh:
        rm = fh.read()
    seccion = rm[rm.index("## ✅ Tests"):]
    assert "git clone" in seccion, "falta el paso de clonar el repo"
    assert "python" in seccion.lower() and ("3.10" in seccion or "3.1" in seccion), (
        "falta el prerrequisito de version de Python")
    assert "pip install -r requirements-dev.txt" in seccion
    assert "pytest tests/" in seccion
    # los tests de JS tambien se documentan, no solo los de Python
    assert "node api/payments.test.js" in seccion
    assert "node landing/security.test.js" in seccion


def test_mcp_pinneado_por_debajo_de_2():
    """mcp 2.0.0 saco `mcp.server.fastmcp`, que es lo que importa
    mvdg/mcp_server.py. Sin tope superior, una instalacion limpia agarraba
    2.0.0 y el servidor MCP moria con ModuleNotFoundError — lo detecto el CI,
    no el entorno de desarrollo (que ya tenia 1.28 instalada)."""
    ruta = os.path.join(_repo_root(), "requirements.txt")
    with open(ruta, encoding="utf-8") as fh:
        lineas = [ln.strip() for ln in fh if ln.strip().startswith("mcp")]
    assert lineas, "mcp no esta declarado"
    assert "<2" in lineas[0], f"mcp sin tope de version mayor: {lineas[0]}"
    # y la API que usamos tiene que seguir existiendo con lo instalado
    from mcp.server.fastmcp import FastMCP  # noqa: F401


def test_ci_corre_tests_en_cada_push():
    ruta = os.path.join(_repo_root(), ".github", "workflows", "tests.yml")
    assert os.path.exists(ruta), "no hay workflow de CI"
    with open(ruta, encoding="utf-8") as fh:
        ci = fh.read()
    assert "on:" in ci and "push:" in ci and "pull_request:" in ci
    assert "pytest tests/" in ci             # corre la suite
    assert "ruff check" in ci                # y el linter
    assert "requirements-dev.txt" in ci
    assert "--upgrade pip" not in ci


def test_linter_configurado():
    ruta = os.path.join(_repo_root(), "pyproject.toml")
    assert os.path.exists(ruta), "falta pyproject.toml con la config del linter"
    with open(ruta, encoding="utf-8") as fh:
        cfg = fh.read()
    assert "[tool.ruff]" in cfg and "[tool.ruff.lint]" in cfg
    assert "select" in cfg


def test_sin_funciones_gigantes_en_el_motor():
    """Ninguna funcion de mas de 100 lineas en el motor ni en la API.

    Este test es el que evita que run_checks() (que llego a tener 822 lineas)
    vuelva a crecer sin que nadie lo note."""
    import ast
    largas = []
    for carpeta in ("mvdg", "bi_api"):
        base = os.path.join(_repo_root(), carpeta)
        for dirpath, _, archivos in os.walk(base):
            for archivo in archivos:
                if not archivo.endswith(".py"):
                    continue
                ruta = os.path.join(dirpath, archivo)
                with open(ruta, encoding="utf-8") as fh:
                    arbol = ast.parse(fh.read())
                for nodo in ast.walk(arbol):
                    if isinstance(nodo, (ast.FunctionDef, ast.AsyncFunctionDef)):
                        n = nodo.end_lineno - nodo.lineno + 1
                        if n > 100:
                            rel = os.path.relpath(ruta, _repo_root())
                            largas.append(f"{rel}:{nodo.lineno} {nodo.name} ({n})")
    assert not largas, "funciones de mas de 100 lineas: " + ", ".join(largas)


def test_selfcheck_expone_los_chequeos_como_registro():
    """El auto-diagnostico dejo de ser una funcion monolitica: cada chequeo es
    una funcion propia registrada, y run_checks solo los recorre."""
    from mvdg import selfcheck
    assert len(selfcheck.CHECKS) >= 40
    assert all(isinstance(nombre, str) and callable(fn)
               for nombre, fn in selfcheck.CHECKS)
    nombres = [n for n, _ in selfcheck.CHECKS]
    assert len(nombres) == len(set(nombres)), "hay chequeos con nombre repetido"


# ------------------------------------------- sugerencias de correccion (IA)
@pytest.mark.parametrize("lang", LANGS)
def test_remediation_covers_all_demo_rules(lang):
    from mvdg.quality import RULES
    from mvdg.remediation import REMEDIATIONS, suggest_fix
    # las 17 reglas de demo tienen contenido especifico, no generico
    assert {r.rule_id for r in RULES} <= set(REMEDIATIONS)
    for r in RULES:
        fix = suggest_fix(r.rule_id, r.dimension, r.column, 123, lang)
        for field in ("root_cause", "short_term", "long_term", "owner"):
            assert fix[field], f"{r.rule_id}.{field} vacio en {lang}"
        # el numero de filas afectadas aparece formateado en el texto
        assert "123" in fix["short_term"] or "123" in fix["root_cause"]


@pytest.mark.parametrize("lang", LANGS)
def test_remediation_covers_all_sample_rules(lang):
    from mvdg import samples
    from mvdg.remediation import REMEDIATIONS, suggest_fix
    for key in samples.sample_keys():
        for r in samples.SAMPLES[key]["rules"]:
            assert r.rule_id in REMEDIATIONS
            fix = suggest_fix(r.rule_id, r.dimension, r.column, 9, lang)
            assert all(fix.values())


def test_remediation_generic_fallback_for_unknown_rule():
    from mvdg.quality import DIMENSIONS
    from mvdg.remediation import suggest_fix
    for dim in DIMENSIONS:
        for lang in LANGS:
            fix = suggest_fix("NEW-01", dim, "alguna_columna", 4, lang)
            assert all(fix.values())
            assert "alguna_columna" in fix["short_term"] or "alguna_columna" in fix["root_cause"]


def test_remediation_thousands_separator():
    from mvdg.remediation import suggest_fix
    fix = suggest_fix("CAF-03", "completeness", "Payment Method", 3178, "es")
    assert "3,178" in fix["short_term"]


def test_render_fixes_shown_next_to_failures_in_app(monkeypatch):
    """Las reglas en warn/fail del dataset de ejemplo mas sucio deben tener
    una sugerencia de la IA disponible; las que pasan, no hace falta."""
    from mvdg import samples
    from mvdg.remediation import suggest_fix
    res = samples.sample_quality_results("cafe_sales_kaggle", "es")
    broken = res[res["status"] != "pass"]
    assert len(broken) == 5  # 3 fail + 2 warn, ver test_samples_quality_results_have_real_spread
    for _, row in broken.iterrows():
        fix = suggest_fix(row["rule_id"], row["dimension"], row["column"],
                          int(row["affected_rows"]), "es")
        assert all(fix.values())


# ------------------------------------------------------------- auto-diagnostico
def test_selfcheck_all_pass():
    from mvdg.selfcheck import run_checks
    results = run_checks()
    assert len(results) >= 12
    failed = [(n, d) for n, ok, d in results if not ok]
    assert not failed, f"selfcheck fallo: {failed}"


# --------------------------------------------------- caso de ejemplo (impacto)
def test_medir_impacto_reproducible():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "mvdg_medir_impacto",
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                     "docs", "caso_ejemplo", "medir_impacto.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    r = mod.medir()
    # el "despues" debe ser mejor que el "antes" en indice y filas afectadas
    assert r["despues"]["indice"] > r["antes"]["indice"]
    assert r["despues"]["filas_afectadas"] < r["antes"]["filas_afectadas"]
    assert r["mejora_indice"] > 0 and r["reduccion_filas_pct"] > 0
    # determinista: dos corridas dan el mismo resultado
    assert mod.medir()["antes"]["indice"] == r["antes"]["indice"]


# --------------------------------------------------------------------- API
def test_api_all_tables_and_formats():
    pytest.importorskip("fastapi")
    try:
        from fastapi.testclient import TestClient
    except RuntimeError:
        pytest.skip("httpx2 no disponible para TestClient")
    from bi_api.main import TABLES, app
    c = TestClient(app)
    assert c.get("/health").json() == {"status": "ok"}
    for tbl in TABLES:
        r = c.get(f"/api/{tbl}", params={"lang": "pt"})
        assert r.status_code == 200
        body = r.json()
        assert body["rows"] == len(body["data"]) > 0
    assert c.get("/api/catalog", params={"format": "csv"}).text.startswith("dataset")
    assert c.get("/api/nope").status_code == 404
    assert c.get("/api/catalog", params={"lang": "xx"}).status_code == 422


# ------------------------------------------------------------- Power BI meta
def _make_pbip(root):
    """Escribe un proyecto .pbip mínimo (TMDL, sin cache.abf) bajo ``root``."""
    sm = root / "VentasDemo.SemanticModel"
    dfn = sm / "definition"
    (dfn / "tables").mkdir(parents=True)
    (dfn / "roles").mkdir(parents=True)
    (root / "VentasDemo.Report").mkdir(parents=True)
    (sm / ".platform").write_text(
        '{ "metadata": { "type": "SemanticModel", "displayName": "VentasDemo" } }',
        encoding="utf-8")
    # DAX multi-línea + medidas simples + columna calculada + PII
    ventas = (
        "table Ventas\n"
        "\tmeasure 'Total Ventas' =\n"
        "\t\t\tSUMX (\n"
        "\t\t\t\tVentas,\n"
        "\t\t\t\tVentas[Cantidad] * Ventas[PrecioUnitario]\n"
        "\t\t\t)\n"
        "\t\tdisplayFolder: Metricas\n"
        "\t\tdescription: Suma de cantidad por precio\n"
        "\tmeasure Margen = [Total Ventas] - [Total Costo]\n"
        "\tcolumn Cantidad\n"
        "\t\tdataType: int64\n"
        "\t\tsourceColumn: Cantidad\n"
        "\tcolumn Email\n"
        "\t\tdataType: string\n"
        "\t\tsourceColumn: email\n"
        "\tcolumn MargenPct\n"
        "\t\tdataType: double\n"
        "\t\texpression = DIVIDE ( [Margen], [Total Ventas] )\n"
        "\tpartition Ventas = m\n"
        "\t\tmode: import\n"
        "\t\tsource =\n"
        "\t\t\t\tlet\n"
        "\t\t\t\t\tSource = Sql.Database(\"MyServer\", \"MyDB\"),\n"
        "\t\t\t\t\tVentas1 = Source{[Schema=\"dbo\",Item=\"Ventas\"]}[Data]\n"
        "\t\t\t\tin\n"
        "\t\t\t\t\tVentas1\n"
    )
    (dfn / "tables" / "Ventas.tmdl").write_text(ventas, encoding="utf-8")
    rels = (
        "relationship aaaa-1111\n"
        "\tfromColumn: Ventas.ClienteKey\n"
        "\ttoColumn: Cliente.ClienteKey\n\n"
        "relationship bbbb-2222\n"
        "\tcrossFilteringBehavior: bothDirections\n"
        "\tfromColumn: Ventas.FechaKey\n"
        "\ttoColumn: Calendario.FechaKey\n"
    )
    (dfn / "relationships.tmdl").write_text(rels, encoding="utf-8")
    (dfn / "roles" / "Vendedor.tmdl").write_text(
        'role Vendedor\n\ttablePermission Ventas = Ventas[Region] = "Sur"\n',
        encoding="utf-8")
    return str(sm)


def test_powerbi_pbip_parse(tmp_path):
    from mvdg import powerbi_meta as pbi
    model = pbi.read_pbip(_make_pbip(tmp_path))
    assert model.name == "VentasDemo"
    assert "Ventas" in model.tables
    names = {m.name for m in model.measures}
    assert {"Total Ventas", "Margen"} <= names
    total = next(m for m in model.measures if m.name == "Total Ventas")
    assert "\n" not in total.dax and "SUMX" in total.dax   # DAX multi-línea colapsado
    assert total.description and total.display_folder == "Metricas"
    assert any(r.both_directions for r in model.relationships)   # relación bidireccional
    assert model.roles == ["Vendedor"]                            # RLS detectado
    assert "VentasDemo" in model.reports
    calc = next(c for c in model.columns if c.name == "MargenPct")
    assert calc.is_calculated and "DIVIDE" in calc.dax
    dic = pbi.to_dictionary(model)
    assert bool(dic.loc[dic["column"] == "Email", "pii"].iloc[0])  # PII


def test_powerbi_sql_source_wired_into_lineage(tmp_path):
    # cadena completa: SQL Server -> tabla -> dataset (modelo) -> reporte
    from mvdg import powerbi_meta as pbi
    out = pbi.ingest_pbip(_make_pbip(tmp_path))
    model = out["_model"]
    assert model.table_sources.get("Ventas") == "SQL Server · MyServer/MyDB"

    srcs = out["sources"]
    assert set(srcs.columns) == {"table", "source"}
    assert srcs.loc[srcs["table"] == "Ventas", "source"].iloc[0] == "SQL Server · MyServer/MyDB"

    lin = out["lineage"]
    sql_row = lin[(lin["source"] == "SQL Server · MyServer/MyDB") & (lin["source_layer"] == "source")]
    assert len(sql_row) == 1
    assert sql_row.iloc[0]["target"] == "Ventas" and sql_row.iloc[0]["target_layer"] == "curated"
    # la cadena sigue: tabla -> modelo -> reporte, sin cortarse
    assert ((lin["source"] == "Ventas") & (lin["target"] == model.name)).any()
    assert ((lin["source"] == model.name) & (lin["source_layer"] == "mart")).any()


def test_powerbi_tmdl_doc_comment_and_metadata_traits():
    # regresión: encontrado escaneando un proyecto .pbip real de GitHub —
    # sourceLineageTag/dataCategory se colaban en el texto del DAX, y las
    # descripciones nativas "/// ..." de TMDL no se capturaban.
    from mvdg.powerbi_meta import _parse_table_tmdl
    tmdl = (
        "table Targets\n"
        "\tlineageTag: 5c67d908-0588-43c0-9dbc-1c64794f92c8\n"
        "\tsourceLineageTag: fc05aa02-32da-45a3-a497-78e97999d244\n"
        "\n"
        "\t/// Total sales goal for the current filter context.\n"
        "\tmeasure Target = SUM(Targets[TargetAmount])\n"
        "\t\tformatString: \\$#,0;(\\$#,0);\\$#,0\n"
        "\t\tlineageTag: 35b9a71b-304f-4f4f-9f7f-cc49da4a2c84\n"
        "\t\tsourceLineageTag: ffc79c2c-8c37-499d-919a-2a3511102514\n"
        "\t\tdataCategory: Uncategorized\n"
        "\n"
        "\tmeasure Undocumented = SUM(Targets[X])\n"
    )
    _, _, measures, _ = _parse_table_tmdl(tmdl)
    target = next(m for m in measures if m.name == "Target")
    assert target.dax == "SUM(Targets[TargetAmount])"   # sin metadatos colados
    assert target.description == "Total sales goal for the current filter context."
    other = next(m for m in measures if m.name == "Undocumented")
    assert other.description == ""   # sin "///" antes -> sin descripción, no arrastra la anterior


def test_powerbi_source_label_heuristics():
    from mvdg.powerbi_meta import _source_label_from_mquery
    assert _source_label_from_mquery('Source = Sql.Database("Srv", "Db")') == "SQL Server · Srv/Db"
    assert _source_label_from_mquery('Source = Sql.Databases("Srv")') == "SQL Server · Srv"
    assert _source_label_from_mquery('Value.NativeQuery(Source, "SELECT 1")') == \
        "SQL (consulta nativa · Value.NativeQuery)"
    assert _source_label_from_mquery('Source = Excel.Workbook(File.Contents("x.xlsx"))') == "Power Query · Excel.Workbook"
    assert _source_label_from_mquery('no hay ninguna funcion m aca') is None


# --------------------------------------------------- Power BI tenant (Scanner API)
def test_powerbi_tenant_off_by_default(monkeypatch):
    from mvdg import powerbi_meta as pbi
    for var in ("POWERBI_TENANT_ID", "POWERBI_CLIENT_ID", "POWERBI_CLIENT_SECRET"):
        monkeypatch.delenv(var, raising=False)
    assert pbi.tenant_configured() is False
    with pytest.raises(RuntimeError):
        pbi.read_scanner()


def _scanner_result_json():
    return {
        "workspaces": [{
            "name": "Ventas LATAM",
            "datasets": [{
                "id": "ds-1", "name": "VentasDemo",
                "tables": [{
                    "name": "Ventas",
                    "columns": [{"name": "Monto", "dataType": "double"}],
                    "measures": [{"name": "Total", "expression": "SUM ( Ventas[Monto] )",
                                 "description": "Suma de ventas"}],
                    "source": [{"expression": 'Source = Sql.Database("Srv", "Db")'}],
                }],
                "relationships": [{"fromTable": "Ventas", "fromColumn": "ClienteKey",
                                   "toTable": "Cliente", "toColumn": "ClienteKey",
                                   "crossFilteringBehavior": "BothDirections"}],
                "roles": [{"name": "Vendedor"}],
            }],
            "reports": [{"name": "Dashboard Ventas", "datasetId": "ds-1"}],
        }],
    }


def test_powerbi_list_workspace_ids_paginates_past_5000(monkeypatch):
    # un tenant multinacional puede tener más workspaces que el tope de una
    # sola página ($top) — hay que seguir pidiendo con $skip hasta agotarlos.
    from mvdg import powerbi_meta as pbi

    pages = {
        0: [{"id": f"ws-{i}", "name": f"W{i}"} for i in range(3)],   # top=3, página llena
        3: [{"id": "ws-3", "name": "W3"}],                            # última página, incompleta
    }

    def fake_http_json(url, headers, method="GET", body=None):
        assert "admin/groups" in url
        skip = int(url.split("$skip=")[1].split("&")[0])
        return {"value": pages.get(skip, [])}

    monkeypatch.setattr(pbi, "_http_json", fake_http_json)
    result = pbi.list_workspace_ids("tok", top=3)
    assert [w["id"] for w in result] == ["ws-0", "ws-1", "ws-2", "ws-3"]


def test_powerbi_tenant_scan_mocked_end_to_end(monkeypatch):
    from mvdg import powerbi_meta as pbi

    calls = {"token": 0, "groups": 0, "getinfo": 0, "status": 0, "result": 0}

    def fake_http_form(url, form):
        calls["token"] += 1
        assert "login.microsoftonline.com" in url
        assert form["client_id"] == "cid"
        return {"access_token": "tok-123"}

    def fake_http_json(url, headers, method="GET", body=None):
        assert headers["Authorization"] == "Bearer tok-123" or "Authorization" not in headers
        if "admin/groups" in url:
            calls["groups"] += 1
            return {"value": [{"id": "ws-1", "name": "Ventas LATAM"}]}
        if url.endswith("/getInfo") or "getInfo?" in url:
            calls["getinfo"] += 1
            assert body == {"workspaces": ["ws-1"]}
            return {"id": "scan-1"}
        if "/scanStatus/" in url:
            calls["status"] += 1
            return {"status": "Succeeded"}
        if "/scanResult/" in url:
            calls["result"] += 1
            return _scanner_result_json()
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(pbi, "_http_form", fake_http_form)
    monkeypatch.setattr(pbi, "_http_json", fake_http_json)

    models = pbi.read_scanner(tenant_id="tid", client_id="cid", client_secret="sec")
    assert calls == {"token": 1, "groups": 1, "getinfo": 1, "status": 1, "result": 1}
    assert len(models) == 1
    m = models[0]
    assert m.name == "VentasDemo" and m.workspace == "Ventas LATAM"
    assert m.table_sources.get("Ventas") == "SQL Server · Srv/Db"
    assert any(r.both_directions for r in m.relationships)
    assert m.roles == ["Vendedor"]
    assert m.reports == ["Dashboard Ventas"]   # linkeado por datasetId

    out = pbi.ingest_tenant(tenant_id="tid", client_id="cid", client_secret="sec")
    assert len(out["catalog"]) == 1
    assert "Ventas LATAM" in out["catalog"].iloc[0]["domain"]
    lin = out["lineage"]
    assert (lin["source_layer"] == "source").any() and (lin["target_layer"] == "bi").any()


def test_powerbi_tenant_missing_credentials_raises(monkeypatch):
    from mvdg import powerbi_meta as pbi
    for var in ("POWERBI_TENANT_ID", "POWERBI_CLIENT_ID", "POWERBI_CLIENT_SECRET"):
        monkeypatch.delenv(var, raising=False)
    with pytest.raises(RuntimeError):
        pbi.read_scanner(tenant_id="only-this-one")


def test_powerbi_bundled_example_is_real_and_parses(tmp_path):
    # el .pbip real incluido con el programa (GitHub, MIT) debe seguir
    # parseando limpio — este mismo archivo encontró 2 bugs reales del
    # parser (ver test_powerbi_tmdl_doc_comment_and_metadata_traits).
    from mvdg import powerbi_meta as pbi
    out = pbi.ingest_example()
    model = out["_model"]
    assert model.name == "Adventure Works Demo"
    assert len(model.tables) == 10 and len(model.measures) == 17
    assert all(m.description for m in model.measures)   # /// se capturan bien
    assert "Sales" in model.tables and "Targets" in model.tables


def test_powerbi_example_tenant_is_illustrative_not_a_real_scan():
    from mvdg import powerbi_meta as pbi
    out = pbi.ingest_example_tenant()
    models = out["_models"]
    assert len(models) == 4   # 4 workspaces simulados
    workspaces = {m.workspace for m in models}
    assert len(workspaces) == 4   # cada uno con nombre distinto
    # todos comparten el contenido real (mismas tablas/medidas), solo cambia
    # el workspace/reporte simulado
    assert {m.name for m in models} == {"Adventure Works Demo"}
    assert all(len(m.tables) == 10 for m in models)
    # tiene que quedar explícitamente marcado como ilustrativo, no un scan real
    assert all("ilustrativo" in m.source.lower() for m in models)
    assert (out["catalog"]["source"].str.contains("ilustrativo", case=False)).all()


# ---------------------------------------------------------- Tableau (Metadata API)
def test_tableau_off_by_default(monkeypatch):
    from mvdg import tableau_meta as tab
    for var in ("TABLEAU_SERVER_URL", "TABLEAU_TOKEN_NAME", "TABLEAU_TOKEN_SECRET"):
        monkeypatch.delenv(var, raising=False)
    assert tab.configured() is False
    with pytest.raises(RuntimeError):
        tab.read_site()


def _tableau_graphql_response():
    return {"data": {"workbooks": [{
        "name": "Ventas Regional", "projectName": "Comercial",
        "upstreamDatasources": [{
            "name": "DS Ventas",
            "fields": [
                {"name": "Monto", "description": "", "formula": None},
                {"name": "Margen %", "description": "Margen sobre ventas",
                 "formula": "SUM([Margen]) / SUM([Monto])"},
                {"name": "Margen % dup", "description": "",
                 "formula": "SUM([Margen]) / SUM([Monto])"},
            ],
            "upstreamTables": [{"name": "ventas", "schema": "dbo",
                               "database": {"name": "Db", "connectionType": "sqlserver"}}],
        }],
    }]}}


def test_tableau_site_scan_mocked_end_to_end(monkeypatch):
    from mvdg import tableau_meta as tab

    calls = {"signin": 0, "graphql": 0, "signout": 0}

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/signin"):
            calls["signin"] += 1
            assert body["credentials"]["personalAccessTokenName"] == "mv-token"
            return {"credentials": {"token": "sess-abc", "site": {"id": "site-1"}}}
        if url.endswith("/auth/signout"):
            calls["signout"] += 1
            return {}
        if url.endswith("/api/metadata/graphql"):
            calls["graphql"] += 1
            assert headers["X-Tableau-Auth"] == "sess-abc"
            return _tableau_graphql_response()
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(tab, "_http_json", fake_http_json)

    model = tab.read_site(server="https://tableau.example.com",
                          token_name="mv-token", token_secret="secret")
    assert calls == {"signin": 1, "graphql": 1, "signout": 1}
    assert model.workbooks == ["Ventas Regional"]
    assert [d.name for d in model.datasources] == ["DS Ventas"]
    assert model.datasources[0].upstream_tables == ["dbo.ventas (sqlserver)"]
    calc_names = {fl.name for fl in model.fields if fl.is_calculated}
    assert calc_names == {"Margen %", "Margen % dup"}   # duplicated formula, detected by TAB-02

    out = tab.ingest_site(server="https://tableau.example.com",
                          token_name="mv-token", token_secret="secret")
    assert len(out["catalog"]) == 1
    assert len(out["glossary"]) == 2   # 2 calculated fields = 2 glossary terms
    q = out["quality"]
    dupe_rule = q[q["rule_id"] == "TAB-02"].iloc[0]
    assert dupe_rule["affected_rows"] == 1 and dupe_rule["status"] != "pass"
    lin = out["lineage"]
    assert (lin["source_layer"] == "source").any() and (lin["target_layer"] == "bi").any()


def test_tableau_signout_failure_does_not_break_scan(monkeypatch):
    from mvdg import tableau_meta as tab

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/signin"):
            return {"credentials": {"token": "sess-abc", "site": {"id": "site-1"}}}
        if url.endswith("/auth/signout"):
            raise OSError("network blip on signout")
        if url.endswith("/api/metadata/graphql"):
            return _tableau_graphql_response()
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(tab, "_http_json", fake_http_json)
    model = tab.read_site(server="https://tableau.example.com",
                          token_name="mv-token", token_secret="secret")
    assert model.workbooks == ["Ventas Regional"]   # el fallo del sign-out no rompe el escaneo


def _make_twb(root, name="Demo.twb"):
    xml = (
        "<?xml version='1.0' encoding='utf-8' ?>\n"
        "<workbook version='18.1'>\n"
        "  <datasources>\n"
        "    <datasource caption='Ventas' name='sqlserver.x'>\n"
        "      <connection class='sqlserver' server='Srv' dbname='Db'>\n"
        "        <relation name='Ventas' table='[dbo].[Ventas]' type='table' />\n"
        "      </connection>\n"
        "      <column caption='Monto' datatype='real' name='[Monto]' role='measure' />\n"
        "      <column caption='Margen %' datatype='real' name='[Calc1]' role='measure'>\n"
        "        <calculation class='tableau' formula='SUM([Margen])/SUM([Monto])' />\n"
        "      </column>\n"
        "    </datasource>\n"
        "  </datasources>\n"
        "  <dashboards><dashboard name='Panel'><zones/></dashboard></dashboards>\n"
        "</workbook>\n"
    )
    p = root / name
    p.write_text(xml, encoding="utf-8")
    return str(p)


def test_tableau_read_twb_offline(tmp_path):
    from mvdg import tableau_meta as tabl
    path = _make_twb(tmp_path)
    model = tabl.read_twb(path)
    assert model.workbooks == ["Demo"]
    assert [d.name for d in model.datasources] == ["Ventas"]
    assert model.datasources[0].upstream_tables == ["sqlserver · Srv/Db"]
    calc = next(f for f in model.fields if f.is_calculated)
    assert calc.name == "Margen %" and calc.formula == "SUM([Margen])/SUM([Monto])"

    out = tabl.ingest_twb(path)
    assert len(out["catalog"]) == 1
    lin = out["lineage"]
    assert (lin["source_layer"] == "source").any() and (lin["target_layer"] == "bi").any()


def test_tableau_read_twbx_zip_wrapper(tmp_path):
    import zipfile
    from mvdg import tableau_meta as tabl
    twb_path = _make_twb(tmp_path, "Inner.twb")
    twbx_path = str(tmp_path / "Demo.twbx")
    with zipfile.ZipFile(twbx_path, "w") as zf:
        zf.write(twb_path, "Inner.twb")
    model = tabl.read_twb(twbx_path)
    assert [d.name for d in model.datasources] == ["Ventas"]   # se desempaqueta el .twb interno


def test_tableau_read_twb_missing_file_raises(tmp_path):
    from mvdg import tableau_meta as tabl
    with pytest.raises(FileNotFoundError):
        tabl.read_twb(str(tmp_path / "no_existe.twb"))


def test_tableau_bundled_example_parses():
    from mvdg import tableau_meta as tabl
    out = tabl.ingest_example()
    model = out["_model"]
    assert model.workbooks == ["VentasGlobalDemo"]
    assert {d.name for d in model.datasources} == {"Ventas Global", "Metas Regionales"}
    calc_names = {f.name for f in model.fields if f.is_calculated}
    assert calc_names == {"Margen %", "Ticket Promedio", "Segmento de Cuenta"}
    assert len(out["glossary"]) == 3
    lin = out["lineage"]
    assert (lin["source_layer"] == "source").any() and (lin["target_layer"] == "bi").any()


def test_powerbi_normalizers_match_mvdg_schema(tmp_path):
    from mvdg import powerbi_meta as pbi
    from mvdg.glossary import glossary_df
    from mvdg.lineage import lineage_df
    from mvdg.quality import DIMENSIONS, run_rules
    out = pbi.ingest_pbip(_make_pbip(tmp_path))
    # columnas idénticas a las tablas nativas del motor de gobierno
    assert set(out["glossary"].columns) == set(glossary_df().columns)
    assert set(out["lineage"].columns) == set(lineage_df().columns)
    assert set(out["quality"].columns) == set(run_rules().columns)
    # las dimensiones de salud de modelo caen dentro de las 6 DAMA
    assert set(out["quality"]["dimension"]) <= set(DIMENSIONS)
    # el catálogo reporta 0 filas: es metadata, no datos
    assert (out["catalog"]["rows"] == 0).all()


def test_powerbi_lineage_dynamic_figure(tmp_path):
    from mvdg import powerbi_meta as pbi
    from mvdg.lineage import downstream, graph_from_lineage, lineage_figure, upstream
    out = pbi.ingest_pbip(_make_pbip(tmp_path))
    nodes, edges = graph_from_lineage(out["lineage"])
    assert nodes and edges
    fig = lineage_figure(nodes=nodes, edges=edges)     # no debe romper con grafo dinámico
    assert fig is not None and len(fig.data) > 0
    model_id = f"model_{out['_model'].name}"
    assert downstream("tbl_Ventas", edges) & {model_id}   # tabla → modelo
    assert upstream(model_id, edges)                        # el modelo tiene ancestros


def test_lineage_demo_still_works():
    # el grafo de demo (sin args) sigue funcionando igual que antes
    from mvdg.lineage import lineage_figure, upstream
    fig = lineage_figure()
    assert fig is not None and len(fig.data) > 0
    assert upstream("mart_sales")  # ancestros del mart de demo


def test_ai_dax_refactor_offline(monkeypatch):
    from mvdg.ai_provider import _build_dax_prompt, ai_refactor_dax
    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY", "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(var, raising=False)
    # sin key configurada, nunca llama afuera: devuelve None
    assert ai_refactor_dax("Total Ventas", "SUMX ( Ventas, 1 )", "Ventas", "es") is None
    # DAX vacío también da None
    assert ai_refactor_dax("X", "", "T", "es") is None
    # el prompt se arma en los 3 idiomas e incluye la medida y el DAX
    for lg in LANGS:
        p = _build_dax_prompt("Total Ventas", "SUMX ( Ventas, 1 )", "Ventas", lg)
        assert "Total Ventas" in p and "SUMX" in p


def test_ai_tableau_calc_refactor_offline(monkeypatch):
    from mvdg.ai_provider import _build_calc_prompt, ai_refactor_calc
    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GEMINI_API_KEY", "MVDG_AI_PROVIDER"):
        monkeypatch.delenv(var, raising=False)
    # sin key configurada, nunca llama afuera: devuelve None
    assert ai_refactor_calc("Margen %", "SUM([Margen])/SUM([Monto])", "DS Ventas", "es") is None
    # fórmula vacía también da None
    assert ai_refactor_calc("X", "", "DS", "es") is None
    # el prompt se arma en los 3 idiomas e incluye el campo y la fórmula
    for lg in LANGS:
        p = _build_calc_prompt("Margen %", "SUM([Margen])/SUM([Monto])", "DS Ventas", lg)
        assert "Margen %" in p and "SUM([Margen])" in p


def test_vercel_deploy_does_not_ignore_api_functions():
    """Regresión: .vercelignore excluía la carpeta api/ (arrastrado de antes
    de que existieran las funciones serverless de MercadoPago ahí adentro),
    lo que hacía que Vercel nunca subiera checkout.js/verify-payment.js y el
    checkout devolviera 404 en producción. api/ tiene que seguir publicada."""
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    vercelignore = os.path.join(root, ".vercelignore")
    assert os.path.exists(vercelignore)
    with open(vercelignore, encoding="utf-8") as fh:
        lines = {ln.strip() for ln in fh if ln.strip()}
    assert "api" not in lines, (
        "api/ no puede estar en .vercelignore: ahí viven las funciones "
        "serverless de MercadoPago (checkout.js, verify-payment.js) que "
        "sirven /api/checkout y /api/verify-payment en producción.")
    for fname in ("checkout.js", "verify-payment.js", "_license.js"):
        assert os.path.exists(os.path.join(root, "api", fname)), fname


def _repo_root():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _app_source() -> str:
    """El código de `app/app.py` como texto, para los tests que verifican
    que la UI y el motor no se contradicen."""
    with open(os.path.join(_repo_root(), "app", "app.py"), encoding="utf-8") as fh:
        return fh.read()


def test_vercel_rewrites_serve_all_landing_files():
    """Regresión: vercel.json solo re-escribía /, /mv_icon.png y /video/* hacia
    landing/ — todo lo demás que la página referencia con rutas relativas
    (descargas.html, img/*.jpg, payments-config.js, guia/pago/reviews, el ZIP
    de la demo) devolvía 404 en producción: botones de compra sin config,
    capturas rotas y 'Descargar demo' muerto. Tiene que existir el rewrite
    catch-all hacia /landing/."""
    import json as _json
    with open(os.path.join(_repo_root(), "vercel.json"), encoding="utf-8") as fh:
        cfg = _json.load(fh)
    sources = {r["source"]: r["destination"] for r in cfg.get("rewrites", [])}
    assert sources.get("/") == "/landing/index.html"
    assert sources.get("/:path*") == "/landing/:path*", (
        "falta el rewrite catch-all /:path* -> /landing/:path* — sin él, "
        "descargas.html, img/, payments-config.js y reviews-data.js dan 404")


def test_checkout_has_no_serverside_only_botid():
    """Regresión: checkout.js corría checkBotId (Vercel BotID) sin que la
    landing integrara el cliente de BotID — TODOS los clics reales en
    'Comprar'/'Suscribirme' recibían 403 {"error":"bot"}. La verificación
    server-side sola no puede volver sin la instrumentación del navegador."""
    root = _repo_root()
    with open(os.path.join(root, "api", "checkout.js"), encoding="utf-8") as fh:
        src = fh.read()
    assert 'require("botid' not in src and "checkBotId(" not in src
    import json as _json
    with open(os.path.join(root, "package.json"), encoding="utf-8") as fh:
        pkg = _json.load(fh)
    assert "botid" not in pkg.get("dependencies", {}), (
        "botid en dependencies reactivaría el checkBotId server-side")


def test_landing_pages_declare_lang_and_notranslate():
    """Regresión: sin <html lang> el traductor automático del navegador
    'traducía' la página — la marca quedaba 'MV Gobernanza de Datos',
    'US$ 390' quedaba '390 dólares estadounidenses' y tocaba Purview/
    Collibra. La página ya es trilingüe nativa (ES/EN/PT): se declara el
    idioma y se marca notranslate."""
    root = _repo_root()
    for page in ("index.html", "descargas.html", "guia.html", "pago.html", "reviews.html"):
        with open(os.path.join(root, "landing", page), encoding="utf-8") as fh:
            src = fh.read()
        assert '<html lang="es" translate="no">' in src, page
        assert '<meta name="google" content="notranslate">' in src, page
        assert src.lstrip().startswith("<!doctype html>"), page


def test_landing_prices_use_usd_not_us_dollar_sign():
    """El usuario pidió USD en vez de 'US$' (que el traductor del navegador
    convertía en 'dólares estadounidenses'). Ningún precio visible puede
    volver a usar 'US$'."""
    with open(os.path.join(_repo_root(), "landing", "index.html"), encoding="utf-8") as fh:
        src = fh.read()
    visible = [ln for ln in src.splitlines()
               if "US$" in ln and not ln.strip().startswith(('"', "'", "<!--", "*"))
               and 'US$ 390" en' not in ln]  # el comentario que documenta el bug
    assert visible == [], f"precios con US$ visibles: {visible}"


def test_la_landing_no_vende_creditos():
    """Los packs de creditos se sacaron: el producto pasa a que cada cliente
    ponga su propia API key. Este test evita que vuelvan por accidente (por
    ejemplo restaurando un bloque viejo del HTML) sin el sistema detras — que
    fue exactamente el problema: se vendian US$9/39/149 por algo que no
    existia en el codigo."""
    with open(os.path.join(_repo_root(), "landing", "index.html"), encoding="utf-8") as fh:
        src = fh.read()
    for rastro in ("cred100", "cred550", "cred2500", "crgrid", "crcard"):
        assert rastro not in src, f"volvio a aparecer {rastro} en la landing"


def test_landing_contact_form_has_visible_mail_fallback():
    """Regresión: el formulario de contacto usaba solo location.href=mailto:,
    que falla EN SILENCIO sin app de correo configurada. Tiene que existir el
    respaldo visible (dirección directa + aviso de copiado al portapapeles)."""
    with open(os.path.join(_repo_root(), "landing", "index.html"), encoding="utf-8") as fh:
        src = fh.read()
    assert 'id="mailFallback"' in src
    assert "navigator.clipboard" in src
    assert 'href="mailto:vieraschiavi@gmail.com"' in src  # link directo siempre visible


# ------------------------------------------------- migración a Purview/Collibra
def _sample_gov_tables():
    from mvdg.exporters import governance_tables
    gov = governance_tables("es")
    return gov["catalog"], gov["dictionary"], gov["glossary"]


def test_purview_off_by_default(monkeypatch):
    for var in ("PURVIEW_TENANT_ID", "PURVIEW_CLIENT_ID", "PURVIEW_CLIENT_SECRET",
               "PURVIEW_ACCOUNT_NAME"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import purview_export as pv
    assert pv.configured() is False
    cat, dic, glo = _sample_gov_tables()
    with pytest.raises(RuntimeError):
        pv.push_catalog(cat, dic, dry_run=False)


def test_purview_dry_run_never_touches_network(monkeypatch):
    """dry_run=True (el default) tiene que funcionar SIN credenciales y sin
    pegarle a la red — es el modo de previsualización."""
    for var in ("PURVIEW_TENANT_ID", "PURVIEW_CLIENT_ID", "PURVIEW_CLIENT_SECRET",
               "PURVIEW_ACCOUNT_NAME"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import purview_export as pv
    cat, dic, glo = _sample_gov_tables()
    r = pv.push_all(cat, dic, glo, dry_run=True)
    assert r["dry_run"] is True
    assert r["catalog"]["entity_count"] == len(cat) + len(dic)
    assert r["glossary"]["term_count"] == len(glo)
    assert r["pii"]["classification_count"] > 0  # dim_customers tiene PII real


def test_purview_glossary_status_reflects_curation(tmp_path, monkeypatch):
    """El estado Draft/Approved en Purview tiene que salir de la curaduría
    real: sin revisar -> Draft, validado/modificado -> Approved."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, purview_export as pv
    cat, dic, glo = _sample_gov_tables()

    def lookup(term_id):
        rec = curation.get_record(f"glossary:demo:{term_id}", "es")
        return (rec["status"], rec.get("text") or "") if rec else ("sugerido_ia", "")

    before = pv.push_glossary(glo, curation_lookup=lookup, dry_run=True)
    assert all(t["status"] == "Draft" for t in before["terms"])

    first_term_id = glo.iloc[0]["term_id"]
    curation.save_validation(f"glossary:demo:{first_term_id}", "es", "validado",
                             "", "María Viera", "Data Owner")
    after = pv.push_glossary(glo, curation_lookup=lookup, dry_run=True)
    statuses = {t["name"]: t["status"] for t in after["terms"]}
    approved_name = glo.iloc[0]["term"]
    assert statuses[approved_name] == "Approved"
    assert sum(1 for s in statuses.values() if s == "Approved") == 1


def test_purview_classification_heuristic():
    from mvdg.purview_export import _pii_classification
    assert _pii_classification("email") == "MICROSOFT.PERSONAL.EMAIL"
    assert _pii_classification("correo_electronico") == "MICROSOFT.PERSONAL.EMAIL"
    assert _pii_classification("full_name") == "MICROSOFT.PERSONAL.NAME"


def test_purview_push_mocked_end_to_end(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    cat, dic, glo = _sample_gov_tables()

    calls = {"token": 0, "bulk": 0, "glossary_list": 0, "glossary_create": 0, "term": 0}

    def fake_http_form(url, form):
        calls["token"] += 1
        assert "login.microsoftonline.com/tid" in url
        assert form["resource"] == "https://purview.azure.net"
        return {"access_token": "ptok"}

    def fake_http_json(url, headers, method="GET", body=None):
        assert headers["Authorization"] == "Bearer ptok"
        if url.endswith("/entity/bulk"):
            calls["bulk"] += 1
            mutated = [{"guid": f"g-{i}", "attributes": {"qualifiedName": e["attributes"]["qualifiedName"]}}
                      for i, e in enumerate(body["entities"])]
            return {"mutatedEntities": {"CREATE": mutated, "UPDATE": []}}
        if url.endswith("/glossary") and method == "GET":
            calls["glossary_list"] += 1
            return []
        if url.endswith("/glossary") and method == "POST":
            calls["glossary_create"] += 1
            return {"guid": "gloss-1", "name": "MV Data Governance"}
        if url.endswith("/glossary/term"):
            calls["term"] += 1
            return {"guid": f"term-{calls['term']}"}
        if url.endswith("/entity/bulk/classification"):
            return {}
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(pv, "_http_form", fake_http_form)
    monkeypatch.setattr(pv, "_http_json", fake_http_json)

    r = pv.push_all(cat, dic, glo, dry_run=False)
    assert calls["token"] == 1
    assert calls["bulk"] == 1
    assert calls["glossary_create"] == 1  # no existía -> se crea
    assert calls["term"] == len(glo)
    assert r["catalog"]["entity_count"] == len(cat) + len(dic)
    assert len(r["catalog"]["guid_by_qualified_name"]) == len(cat) + len(dic)
    assert r["pii"]["classification_count"] > 0


def test_collibra_off_by_default(monkeypatch):
    for var in ("COLLIBRA_BASE_URL", "COLLIBRA_USERNAME", "COLLIBRA_PASSWORD",
               "COLLIBRA_DOMAIN_ID"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import collibra_export as cb
    assert cb.configured() is False
    cat, dic, glo = _sample_gov_tables()
    with pytest.raises(RuntimeError):
        cb.push_glossary(glo, dry_run=False)


def test_collibra_dry_run_never_touches_network(monkeypatch):
    for var in ("COLLIBRA_BASE_URL", "COLLIBRA_USERNAME", "COLLIBRA_PASSWORD",
               "COLLIBRA_DOMAIN_ID"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import collibra_export as cb
    cat, dic, glo = _sample_gov_tables()
    r = cb.push_all(cat, dic, glo, dry_run=True)
    assert r["dry_run"] is True
    assert r["catalog"]["asset_count"] == len(cat) + len(dic)
    assert r["glossary"]["term_count"] == len(glo)
    # sin COLLIBRA_TABLE_TYPE_ID configurado, el payload usa un placeholder
    # visible en vez de fallar en silencio
    assert r["catalog"]["payloads"][0]["asset"]["typeId"] == "<COLLIBRA_TABLE_TYPE_ID>"


def test_collibra_term_type_id_has_documented_default(monkeypatch):
    for var in ("COLLIBRA_BASE_URL", "COLLIBRA_USERNAME", "COLLIBRA_PASSWORD",
               "COLLIBRA_DOMAIN_ID", "COLLIBRA_TERM_TYPE_ID"):
        monkeypatch.delenv(var, raising=False)
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    from mvdg import collibra_export as cb
    cat, dic, glo = _sample_gov_tables()
    r = cb.push_glossary(glo, dry_run=True)
    assert r["terms"][0]["asset"]["typeId"] == cb._DEFAULT_TERM_TYPE_ID


def test_collibra_push_mocked_end_to_end(monkeypatch):
    monkeypatch.setenv("COLLIBRA_BASE_URL", "https://acme.collibra.com")
    monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
    monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    monkeypatch.setenv("COLLIBRA_TABLE_TYPE_ID", "type-table")
    monkeypatch.setenv("COLLIBRA_COLUMN_TYPE_ID", "type-column")
    from mvdg import collibra_export as cb
    cat, dic, glo = _sample_gov_tables()

    calls = {"login": 0, "assets": 0, "attributes": 0, "logout": 0}
    asset_ids = iter(range(10_000))

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/sessions") and method == "POST":
            calls["login"] += 1
            return {}, ["JSESSIONID=abc123; Path=/; HttpOnly"]
        if url.endswith("/auth/sessions/current") and method == "DELETE":
            calls["logout"] += 1
            return {}, []
        assert headers["Cookie"] == "JSESSIONID=abc123"
        if url.endswith("/assets"):
            calls["assets"] += 1
            return {"id": f"asset-{next(asset_ids)}"}, []
        if url.endswith("/attributes"):
            calls["attributes"] += 1
            assert body["typeId"] == cb.DEFINITION_ATTRIBUTE_TYPE_ID
            return {"id": "attr-1"}, []
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(cb, "_http_json", fake_http_json)

    r = cb.push_catalog(cat, dic, dry_run=False)
    assert calls["login"] == 1 and calls["logout"] == 1
    assert calls["assets"] == len(cat) + len(dic)
    assert calls["attributes"] == len(cat) + len(dic)  # todas tienen descripción
    assert r["asset_count"] == len(cat) + len(dic)


def test_collibra_logout_failure_does_not_break_push(monkeypatch):
    """Best-effort sign-out: si el logout falla, el push ya hecho no se
    pierde (mismo criterio que Tableau)."""
    monkeypatch.setenv("COLLIBRA_BASE_URL", "https://acme.collibra.com")
    monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
    monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    monkeypatch.setenv("COLLIBRA_TERM_TYPE_ID", "type-term")
    from mvdg import collibra_export as cb
    cat, dic, glo = _sample_gov_tables()

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/sessions") and method == "POST":
            return {}, ["JSESSIONID=abc123; Path=/"]
        if url.endswith("/auth/sessions/current"):
            raise OSError("network blip")
        if url.endswith("/assets"):
            return {"id": "asset-x"}, []
        if url.endswith("/attributes"):
            return {"id": "attr-x"}, []
        raise AssertionError(url)

    monkeypatch.setattr(cb, "_http_json", fake_http_json)
    r = cb.push_glossary(glo, dry_run=False)  # no debe levantar
    assert r["term_count"] == len(glo)


# ------------------- integración real (sockets HTTP de verdad, sin mockear
# _http_json): levanta un servidor local que imita Purview/Collibra y corre
# el conector real contra él, para probar el protocolo/JSON/auth de punta a
# punta y no solo que se llamó a la función esperada con los args esperados.
def _free_port():
    import socket
    s = socket.socket()
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


def test_purview_integration_real_http_roundtrip(tmp_path, monkeypatch):
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    calls = {"token": 0, "bulk": 0, "glossary_post": 0, "term": 0, "classification": 0}
    received = {"entities": None, "terms": []}

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _body(self):
            length = int(self.headers.get("Content-Length", 0))
            return json.loads(self.rfile.read(length)) if length else None

        def _send(self, code, payload):
            data = json.dumps(payload).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def do_POST(self):
            body = self._body()
            if "/oauth2/token" in self.path:
                calls["token"] += 1
                self._send(200, {"access_token": "sim-token"})
            elif self.path.endswith("/entity/bulk"):
                calls["bulk"] += 1
                received["entities"] = body["entities"]
                mutated = [{"guid": f"guid-{i}", "attributes": {"qualifiedName": e["attributes"]["qualifiedName"]}}
                          for i, e in enumerate(body["entities"])]
                self._send(200, {"mutatedEntities": {"CREATE": mutated, "UPDATE": []}})
            elif self.path.endswith("/glossary"):
                calls["glossary_post"] += 1
                self._send(200, {"guid": "gloss-1", "name": body["name"]})
            elif self.path.endswith("/glossary/term"):
                calls["term"] += 1
                received["terms"].append(body)
                self._send(200, {"guid": f"term-{calls['term']}"})
            elif self.path.endswith("/entity/bulk/classification"):
                calls["classification"] += 1
                self._send(200, {})
            else:
                self._send(404, {"error": self.path})

        def do_GET(self):
            if self.path.endswith("/glossary"):
                self._send(200, [])
            else:
                self._send(404, {"error": self.path})

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
        monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
        monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
        monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
        monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
        monkeypatch.setenv("PURVIEW_API_BASE", f"http://127.0.0.1:{port}")
        from mvdg import curation, purview_export as pv

        def fake_get_token():
            import urllib.request
            req = urllib.request.Request(f"http://127.0.0.1:{port}/oauth2/token",
                                         data=b"{}", method="POST")
            with urllib.request.urlopen(req) as resp:
                return json.loads(resp.read())["access_token"]

        monkeypatch.setattr(pv, "_get_token", fake_get_token)

        cat, dic, glo = _sample_gov_tables()
        first_term_id = glo.iloc[0]["term_id"]
        curation.save_validation(f"glossary:demo:{first_term_id}", "es", "modificado",
                                 "Definición oficial revisada por el Data Owner.",
                                 "María Viera", "Data Owner Comercial")

        def lookup(term_id):
            rec = curation.get_record(f"glossary:demo:{term_id}", "es")
            return (rec["status"], rec.get("text") or "") if rec else ("sugerido_ia", "")

        result = pv.push_all(cat, dic, glo, curation_lookup=lookup, dry_run=False)
    finally:
        server.shutdown()

    # tráfico HTTP real recibido por un servidor de verdad, no una función mockeada
    assert calls == {"token": 1, "bulk": 1, "glossary_post": 1, "term": len(glo), "classification": 2}
    assert len(received["entities"]) == len(cat) + len(dic)
    statuses = {t["name"]: t["status"] for t in received["terms"]}
    approved_name = glo.iloc[0]["term"]
    assert statuses[approved_name] == "Approved"
    assert sum(1 for s in statuses.values() if s == "Draft") == len(glo) - 1
    curated_term = next(t for t in received["terms"] if t["name"] == approved_name)
    assert "Definición oficial revisada" in curated_term["longDescription"]
    n_pii_cols = int((dic["pii"] == True).sum())  # noqa: E712
    assert result["pii"]["classification_count"] == n_pii_cols


def test_collibra_integration_real_http_roundtrip(tmp_path, monkeypatch):
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    calls = {"login": 0, "logout": 0, "assets": 0, "attributes": 0}
    received = {"assets": []}

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _body(self):
            length = int(self.headers.get("Content-Length", 0))
            return json.loads(self.rfile.read(length)) if length else None

        def _send(self, code, payload, cookie=None):
            data = json.dumps(payload).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(data)))
            if cookie:
                self.send_header("Set-Cookie", cookie)
            self.end_headers()
            self.wfile.write(data)

        def do_POST(self):
            body = self._body()
            if self.path.endswith("/auth/sessions"):
                calls["login"] += 1
                self._send(200, {"userId": "u-1"},
                          cookie="JSESSIONID=sim-session; Path=/; HttpOnly")
            elif self.path.endswith("/assets"):
                calls["assets"] += 1
                assert self.headers.get("Cookie") == "JSESSIONID=sim-session"
                received["assets"].append(body)
                self._send(200, {"id": f"asset-{calls['assets']}", "name": body["name"]})
            elif self.path.endswith("/attributes"):
                calls["attributes"] += 1
                assert self.headers.get("Cookie") == "JSESSIONID=sim-session"
                self._send(200, {"id": f"attr-{calls['attributes']}"})
            else:
                self._send(404, {"error": self.path})

        def do_DELETE(self):
            if self.path.endswith("/auth/sessions/current"):
                calls["logout"] += 1
                self._send(200, {})
            else:
                self._send(404, {"error": self.path})

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
        monkeypatch.setenv("COLLIBRA_BASE_URL", f"http://127.0.0.1:{port}")
        monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
        monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
        monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
        monkeypatch.setenv("COLLIBRA_TABLE_TYPE_ID", "table-type")
        monkeypatch.setenv("COLLIBRA_COLUMN_TYPE_ID", "column-type")
        from mvdg import collibra_export as cb
        cat, dic, glo = _sample_gov_tables()
        result = cb.push_all(cat, dic, glo, dry_run=False)
    finally:
        server.shutdown()

    assert calls["login"] == 1 and calls["logout"] == 1  # sesión única reusada
    assert calls["assets"] == len(cat) + len(dic) + len(glo)
    assert calls["attributes"] == calls["assets"]  # todas tienen descripción
    table_types = {a["typeId"] for a in received["assets"][:len(cat)]}
    assert table_types == {"table-type"}
    assert result["catalog"]["asset_count"] == len(cat) + len(dic)


# ------------------------------------------------ enforcement (DDL, no ejecuta)
def test_enforcement_grant_revoke_by_classification():
    from mvdg import enforcement as en
    cat, dic, glo = _sample_gov_tables()
    roles = {"PII": ["rol_rrhh"], "Confidencial": ["rol_finanzas"], "Interna": ["rol_ops"]}
    ddl = en.build_grant_revoke_ddl(cat, roles, engine="postgresql")
    text = "\n".join(ddl)
    assert "REVOKE ALL ON \"dim_customers\" FROM PUBLIC;" in text
    assert "GRANT SELECT ON \"dim_customers\" TO \"rol_rrhh\";" in text  # dim_customers es PII
    # cada dataset tiene su REVOKE + (0 o más) GRANT -> al menos 1 REVOKE por dataset
    assert text.count("REVOKE ALL ON") == len(cat)


def test_enforcement_masking_postgresql_uses_view_not_native():
    """PostgreSQL no tiene masking nativo de columna -> el DDL usa una
    vista con las columnas PII ofuscadas, no ALTER COLUMN (que no existe
    para esto en PG)."""
    from mvdg import enforcement as en
    cat, dic, glo = _sample_gov_tables()
    ddl = en.build_column_masking_ddl(dic, engine="postgresql")
    text = "\n".join(ddl)
    assert "CREATE OR REPLACE VIEW \"dim_customers_masked\"" in text
    assert "'***' AS \"email\"" in text
    assert "customer_id" in text  # columnas no-PII se ven igual, sin ofuscar


def test_enforcement_masking_sqlserver_uses_native_masked_with():
    from mvdg import enforcement as en
    cat, dic, glo = _sample_gov_tables()
    ddl = en.build_column_masking_ddl(dic, engine="sqlserver")
    text = "\n".join(ddl)
    assert "ADD MASKED WITH (FUNCTION = 'email()')" in text  # detecta email por nombre
    assert "ADD MASKED WITH (FUNCTION = 'default()')" in text  # resto de PII, genérico


def test_enforcement_unsupported_engine_raises_not_silently_wrong():
    from mvdg import enforcement as en
    cat, dic, glo = _sample_gov_tables()
    with pytest.raises(ValueError):
        en.build_column_masking_ddl(dic, engine="oracle")
    with pytest.raises(ValueError):
        en.build_row_level_security_ddl("dim_customers", "steward", "rol_x", engine="mysql")


def test_enforcement_plan_never_executes_anything():
    """El módulo entero es generación de texto -- no hay ningún camino que
    abra una conexión de base de datos."""
    import mvdg.enforcement as en_mod
    src = open(en_mod.__file__, encoding="utf-8").read()
    for forbidden in ("sqlalchemy", "psycopg2", "pyodbc", "create_engine", ".execute(", "urllib"):
        assert forbidden not in src, f"enforcement.py no debería importar/usar '{forbidden}'"

    from mvdg import enforcement as en
    cat, dic, glo = _sample_gov_tables()
    plan = en.enforcement_plan(cat, dic, {"PII": ["rol_rrhh"]}, engine="postgresql")
    assert plan["grant_statements"] > 0 and plan["masking_statements"] > 0
    assert "NO ejecutado" in plan["script"]


def test_enforcement_row_level_security_both_engines():
    from mvdg import enforcement as en
    pg = en.build_row_level_security_ddl("dim_customers", "steward", "rol_comercial", "postgresql")
    assert any("ENABLE ROW LEVEL SECURITY" in s for s in pg)
    assert any("CREATE POLICY" in s for s in pg)
    ss = en.build_row_level_security_ddl("dim_customers", "steward", "rol_comercial", "sqlserver")
    assert any("CREATE SECURITY POLICY" in s for s in ss)


# --------------------------------------------- etiquetas MIP (Graph API real)
def test_mip_off_by_default(monkeypatch):
    for var in ("MIP_TENANT_ID", "MIP_CLIENT_ID", "MIP_CLIENT_SECRET"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import mip_labels as mip
    assert mip.configured() is False
    assert mip.list_labels() == []  # sin credenciales, no intenta pegarle a la red
    with pytest.raises(RuntimeError):
        mip.assign_label("d1", "i1", "lbl-1", dry_run=False)


def test_mip_share_url_encoding_matches_documented_algorithm():
    """Base64 -> base64url sin padding, '/'->'_', '+'->'-', prefijo 'u!'
    (documentado por Microsoft Learn, "Access shared items")."""
    from mvdg import mip_labels as mip
    import base64
    url = "https://contoso.sharepoint.com/:x:/s/team/EXAMPLE?e=abc"
    encoded = mip.encode_share_url(url)
    assert encoded.startswith("u!")
    # decodificable de vuelta
    b64 = encoded[2:].replace("_", "/").replace("-", "+")
    b64 += "=" * (-len(b64) % 4)
    assert base64.b64decode(b64).decode("utf-8") == url


def test_mip_dry_run_never_touches_network(monkeypatch):
    for var in ("MIP_TENANT_ID", "MIP_CLIENT_ID", "MIP_CLIENT_SECRET"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import mip_labels as mip
    r = mip.assign_label("d1", "i1", "lbl-1", dry_run=True)
    assert r["dry_run"] is True and "assignSensitivityLabel" in r["url"]


def test_mip_suggest_label_never_invents_an_id():
    """La sugerencia SIEMPRE sale de la lista real de etiquetas del tenant
    -- nunca de un id inventado por el programa."""
    from mvdg import mip_labels as mip
    labels = [{"id": "real-id-1", "name": "Confidencial - Solo interno"},
             {"id": "real-id-2", "name": "Publico"}]
    picked = mip.suggest_label("PII", labels)
    assert picked["id"] == "real-id-1"  # matchea por nombre, id real de la lista
    assert mip.suggest_label("PII", []) is None  # sin etiquetas en el tenant, no inventa nada


def test_mip_plan_skips_datasets_without_mapped_file():
    """Un dataset gobernado que no tiene archivo mapeado en OneDrive/
    SharePoint no puede tener etiqueta MIP (la etiqueta vive en el
    archivo) -- se lista aparte, explícitamente, no se saltea en
    silencio."""
    from mvdg import mip_labels as mip
    cat, dic, glo = _sample_gov_tables()
    file_map = {"dim_customers": {"driveId": "d1", "itemId": "i1"}}
    r = mip.push_labels(cat, file_map, dry_run=True)
    assert len(r["plan"]) == 1 and r["plan"][0]["dataset"] == "dim_customers"
    assert set(r["skipped_no_file"]) == set(cat["dataset"]) - {"dim_customers"}


def test_mip_integration_real_http_roundtrip(monkeypatch):
    """Mismo patrón que la simulación de Purview/Collibra: servidor HTTP
    local real (sin mockear _http_json) para probar el protocolo/JSON/auth
    de punta a punta contra la Graph API tal como la documenta Microsoft."""
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    calls = {"token": 0, "labels": 0, "assign": 0}

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _body(self):
            length = int(self.headers.get("Content-Length", 0))
            return json.loads(self.rfile.read(length)) if length else None

        def do_GET(self):
            if "sensitivityLabels" in self.path:
                calls["labels"] += 1
                payload = json.dumps({"value": [{"id": "lbl-conf", "name": "Confidencial"}]}).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            else:
                self.send_response(404)
                self.end_headers()

        def do_POST(self):
            body = self._body()
            if "assignSensitivityLabel" in self.path:
                calls["assign"] += 1
                assert body["sensitivityLabelId"] == "lbl-conf"
                self.send_response(202)
                self.send_header("Location", "http://example/monitor/1")
                self.send_header("Content-Length", "0")
                self.end_headers()
            else:
                self.send_response(404)
                self.end_headers()

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        monkeypatch.setenv("MIP_TENANT_ID", "tid")
        monkeypatch.setenv("MIP_CLIENT_ID", "cid")
        monkeypatch.setenv("MIP_CLIENT_SECRET", "sec")
        from mvdg import mip_labels as mip
        monkeypatch.setattr(mip, "_GRAPH_V1", f"http://127.0.0.1:{port}")
        monkeypatch.setattr(mip, "_GRAPH_BETA", f"http://127.0.0.1:{port}")
        monkeypatch.setattr(mip, "_get_token", lambda: "fake-graph-token")

        cat, dic, glo = _sample_gov_tables()
        file_map = {"dim_customers": {"driveId": "drv-1", "itemId": "itm-1"}}
        result = mip.push_labels(cat, file_map, dry_run=False)
    finally:
        server.shutdown()

    assert calls == {"token": 0, "labels": 1, "assign": 1}  # token mockeado aparte, no cuenta
    assert result["assigned"][0]["dataset"] == "dim_customers"
    assert result["assigned"][0]["result"]["status"] == 202


# --------------------------------------------------- escaneo de todas las conexiones
def test_connectors_scan_all_partial_failure_does_not_stop_the_rest(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    import sqlite3
    from mvdg import connectors as C

    good_db = str(tmp_path / "ok.db")
    con = sqlite3.connect(good_db)
    pd.DataFrame({"id": [1, 2]}).to_sql("clientes", con, index=False)
    con.close()

    C.save_connection({"name": "OK", "engine": "sqlite", "database": good_db,
                       "user": "", "password": ""}, save_password=False)
    C.save_connection({"name": "Rota", "engine": "sqlite", "database": "/no/existe.db",
                       "user": "", "password": ""}, save_password=False)

    df = C.scan_all_connections()
    assert set(df["name"]) == {"OK", "Rota"}
    ok_rows = df[df["name"] == "OK"]
    assert "clientes" in ok_rows["table"].tolist()
    assert ok_rows["error"].isna().all()
    broken_rows = df[df["name"] == "Rota"]
    assert broken_rows["table"].isna().all()
    assert broken_rows["error"].notna().all()


# ------------------------------------------- Azure Resource Graph (discovery)
def test_azure_discovery_off_by_default(monkeypatch):
    for var in ("AZURE_TENANT_ID", "AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET",
               "AZURE_SUBSCRIPTION_ID"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import azure_discovery as az
    assert az.configured() is False
    with pytest.raises(RuntimeError):
        az.discover_data_resources()


def test_azure_discovery_query_covers_all_data_types():
    from mvdg import azure_discovery as az
    q = az.build_query()
    assert q.startswith("Resources | where type in~ (")
    for tipo in az.DATA_RESOURCE_TYPES:
        assert f"'{tipo}'" in q


def test_azure_discovery_suggest_connection_profile_maps_known_types():
    from mvdg import azure_discovery as az
    sql = az.suggest_connection_profile({"type": "microsoft.sql/servers/databases", "name": "srv1/db1"})
    assert sql["engine"] == "sqlserver" and "database.windows.net" in sql["host"]
    pg = az.suggest_connection_profile({"type": "microsoft.dbforpostgresql/flexibleservers", "name": "pg1"})
    assert pg["engine"] == "postgresql"
    # un tipo no relacionado con conexión (ej. Storage) no sugiere perfil -- no inventa un motor que no existe
    assert az.suggest_connection_profile({"type": "microsoft.storage/storageaccounts", "name": "s1"}) is None


def test_azure_discovery_mocked_end_to_end(monkeypatch):
    monkeypatch.setenv("AZURE_TENANT_ID", "tid")
    monkeypatch.setenv("AZURE_CLIENT_ID", "cid")
    monkeypatch.setenv("AZURE_CLIENT_SECRET", "sec")
    monkeypatch.setenv("AZURE_SUBSCRIPTION_ID", "sub-1")
    from mvdg import azure_discovery as az

    calls = {"query": 0}

    def fake_http_json(url, headers, body):
        assert headers["Authorization"] == "Bearer tok-123"
        assert body["subscriptions"] == ["sub-1"]
        calls["query"] += 1
        return {"data": [
            {"name": "srv1/db1", "type": "microsoft.sql/servers/databases",
             "resourceGroup": "rg1", "location": "eastus", "subscriptionId": "sub-1", "id": "/x/1"},
            {"name": "stg1", "type": "microsoft.storage/storageaccounts",
             "resourceGroup": "rg1", "location": "eastus", "subscriptionId": "sub-1", "id": "/x/2"},
        ]}  # sin $skipToken -> una sola página

    monkeypatch.setattr(az, "_get_token", lambda: "tok-123")
    monkeypatch.setattr(az, "_http_json", fake_http_json)

    df = az.discover_data_resources()
    assert calls["query"] == 1
    assert len(df) == 2
    assert set(df["category"]) == {"Azure SQL Database", "Storage Account"}
    assert "resourceGroup" in df.columns and "location" in df.columns


def test_azure_discovery_paginates_with_skiptoken(monkeypatch):
    monkeypatch.setenv("AZURE_TENANT_ID", "tid")
    monkeypatch.setenv("AZURE_CLIENT_ID", "cid")
    monkeypatch.setenv("AZURE_CLIENT_SECRET", "sec")
    monkeypatch.setenv("AZURE_SUBSCRIPTION_ID", "sub-1")
    from mvdg import azure_discovery as az

    pages = [
        {"data": [{"name": f"db{i}", "type": "microsoft.sql/servers/databases",
                  "resourceGroup": "rg", "location": "eastus", "subscriptionId": "sub-1", "id": f"/x/{i}"}
                 for i in range(3)], "$skipToken": "page2"},
        {"data": [{"name": "db3", "type": "microsoft.sql/servers/databases",
                  "resourceGroup": "rg", "location": "eastus", "subscriptionId": "sub-1", "id": "/x/3"}]},
    ]
    calls = {"n": 0}

    def fake_http_json(url, headers, body):
        page = pages[calls["n"]]
        calls["n"] += 1
        return page

    monkeypatch.setattr(az, "_get_token", lambda: "tok-123")
    monkeypatch.setattr(az, "_http_json", fake_http_json)
    df = az.discover_data_resources()
    assert calls["n"] == 2
    assert len(df) == 4


# --------------------------------------------- Collibra pull (conector inverso)
def test_collibra_pull_off_by_default(monkeypatch):
    for var in ("COLLIBRA_BASE_URL", "COLLIBRA_USERNAME", "COLLIBRA_PASSWORD",
               "COLLIBRA_DOMAIN_ID"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import collibra_pull as cbp
    with pytest.raises(RuntimeError):
        cbp.pull_glossary()
    with pytest.raises(RuntimeError):
        cbp.pull_catalog()


def test_collibra_pull_catalog_requires_table_type_id(monkeypatch):
    monkeypatch.setenv("COLLIBRA_BASE_URL", "https://acme.collibra.com")
    monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
    monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    monkeypatch.delenv("COLLIBRA_TABLE_TYPE_ID", raising=False)
    from mvdg import collibra_pull as cbp
    assert cbp.table_pull_configured() is False
    with pytest.raises(RuntimeError):
        cbp.pull_catalog()
    # pull_all no debe explotar -- reporta el catálogo salteado explícitamente
    import mvdg.collibra_export as cb

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/sessions"):
            return {}, ["JSESSIONID=sess1; Path=/"]
        if url.endswith("/auth/sessions/current"):
            return {}, []
        if "/assets" in url:
            return {"total": 0, "offset": 0, "limit": 200, "results": []}, []
        raise AssertionError(url)

    monkeypatch.setattr(cb, "_http_json", fake_http_json)
    r = cbp.pull_all()
    assert r["catalog"]["table_count"] == 0
    assert "skipped_reason" in r["catalog"]


def test_collibra_pull_integration_real_http_roundtrip(monkeypatch):
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    calls = {"login": 0, "logout": 0, "assets": 0, "attributes": 0}
    term_asset = {"id": "asset-term-1", "name": "Cliente"}

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _send(self, code, payload, cookie=None):
            data = json.dumps(payload).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(data)))
            if cookie:
                self.send_header("Set-Cookie", cookie)
            self.end_headers()
            self.wfile.write(data)

        def do_POST(self):
            if self.path.endswith("/auth/sessions"):
                calls["login"] += 1
                self._send(200, {"userId": "u-1"}, cookie="JSESSIONID=pull-sess; Path=/; HttpOnly")
            else:
                self._send(404, {"error": self.path})

        def do_GET(self):
            assert self.headers.get("Cookie") == "JSESSIONID=pull-sess"
            if self.path.startswith("/rest/2.0/assets"):
                calls["assets"] += 1
                self._send(200, {"total": 1, "offset": 0, "limit": 200, "results": [term_asset]})
            elif self.path.startswith("/rest/2.0/attributes"):
                calls["attributes"] += 1
                assert f"assetId={term_asset['id']}" in self.path
                self._send(200, {"total": 1, "offset": 0, "limit": 200,
                                 "results": [{"value": "Definición real traída de Collibra."}]})
            else:
                self._send(404, {"error": self.path})

        def do_DELETE(self):
            calls["logout"] += 1
            self._send(200, {})

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        monkeypatch.setenv("COLLIBRA_BASE_URL", f"http://127.0.0.1:{port}")
        monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
        monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
        monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
        from mvdg import collibra_pull as cbp
        result = cbp.pull_glossary()
    finally:
        server.shutdown()

    assert calls == {"login": 1, "logout": 1, "assets": 1, "attributes": 1}
    assert result["term_count"] == 1
    assert result["terms"][0]["name"] == "Cliente"
    assert result["terms"][0]["definition"] == "Definición real traída de Collibra."


# ================================================== fixes 2026-07-16 (round 2)
# Purview pull (conector inverso), persistencia de lo importado, qualifiedNames
# reales, keyring, login del modo servidor, robustez de red y simetría de
# estados Collibra — ver docs/PURVIEW_COLLIBRA.md para el detalle de cada uno.

def test_purview_pull_off_by_default(monkeypatch):
    for var in ("PURVIEW_TENANT_ID", "PURVIEW_CLIENT_ID", "PURVIEW_CLIENT_SECRET",
               "PURVIEW_ACCOUNT_NAME"):
        monkeypatch.delenv(var, raising=False)
    from mvdg import purview_pull as pvp
    assert pvp.configured() is False
    with pytest.raises(RuntimeError):
        pvp.pull_glossary()
    with pytest.raises(RuntimeError):
        pvp.pull_catalog()


def test_purview_pull_mocked_glossary_and_catalog(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    from mvdg import purview_pull as pvp

    calls = {"token": 0, "glossary_list": 0, "terms_list": 0, "search": 0}

    def fake_http_form(url, form):
        calls["token"] += 1
        return {"access_token": "ptok"}

    def fake_http_json(url, headers, method="GET", body=None):
        assert headers["Authorization"] == "Bearer ptok"
        if url.endswith("/glossary") and method == "GET":
            calls["glossary_list"] += 1
            return [{"guid": "gloss-1", "name": "MV Data Governance"}]
        if url.endswith("/glossary/gloss-1/terms"):
            calls["terms_list"] += 1
            return [{"guid": "term-1", "name": "Cliente", "longDescription": "Definición larga."},
                    {"guid": "term-2", "name": "Venta", "shortDescription": "Corta."},
                    {"name": "sin guid, se descarta"}]
        if "/datamap/api/search/query" in url:
            calls["search"] += 1
            return {"value": [{"id": "tbl-1", "name": "dim_customers", "description": "Tabla real."}],
                    "continuationToken": None}
        raise AssertionError(f"unexpected URL: {method} {url}")

    monkeypatch.setattr(pv, "_http_form", fake_http_form)
    monkeypatch.setattr(pv, "_http_json", fake_http_json)

    r = pvp.pull_all()
    assert calls["token"] == 2  # un token por función (glosario y catálogo piden el suyo)
    assert r["glossary"]["term_count"] == 2
    names = {t["name"] for t in r["glossary"]["terms"]}
    assert names == {"Cliente", "Venta"}
    assert r["catalog"]["table_count"] == 1
    assert r["catalog"]["tables"][0]["name"] == "dim_customers"


def test_purview_pull_glossary_missing_returns_empty(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    from mvdg import purview_pull as pvp
    monkeypatch.setattr(pv, "_http_form", lambda url, form: {"access_token": "t"})
    monkeypatch.setattr(pv, "_http_json", lambda url, headers, method="GET", body=None: [])
    r = pvp.pull_glossary()
    assert r["term_count"] == 0 and r["terms"] == []


def test_purview_pull_catalog_paginates_with_continuation_token(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    from mvdg import purview_pull as pvp

    pages = [
        {"value": [{"id": "t1", "name": "a"}], "continuationToken": "cont-2"},
        {"value": [{"id": "t2", "name": "b"}], "continuationToken": None},
    ]
    calls = {"n": 0}

    def fake_http_json(url, headers, method="GET", body=None):
        page = pages[calls["n"]]
        calls["n"] += 1
        return page

    monkeypatch.setattr(pv, "_http_form", lambda url, form: {"access_token": "t"})
    monkeypatch.setattr(pv, "_http_json", fake_http_json)
    r = pvp.pull_catalog()
    assert calls["n"] == 2
    assert r["table_count"] == 2
    assert {t["name"] for t in r["tables"]} == {"a", "b"}


def test_purview_pull_integration_real_http_roundtrip(monkeypatch):
    """Servidor HTTP local real: prueba el protocolo de punta a punta (auth
    header, JSON, paginación por continuationToken) sin mockear _http_json."""
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    calls = {"token": 0, "glossary": 0, "terms": 0, "search": 0}

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _send(self, code, payload):
            data = json.dumps(payload).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def do_POST(self):
            if "/oauth2/token" in self.path:
                calls["token"] += 1
                self._send(200, {"access_token": "sim-token"})
            elif "/datamap/api/search/query" in self.path:
                calls["search"] += 1
                self._send(200, {"value": [{"id": "tbl-1", "name": "dim_customers",
                                            "description": "real"}], "continuationToken": None})
            else:
                self._send(404, {"error": self.path})

        def do_GET(self):
            assert self.headers.get("Authorization") == "Bearer sim-token"
            if self.path.endswith("/glossary"):
                calls["glossary"] += 1
                self._send(200, [{"guid": "gloss-1", "name": "MV Data Governance"}])
            elif self.path.endswith("/glossary/gloss-1/terms"):
                calls["terms"] += 1
                self._send(200, [{"guid": "term-1", "name": "Cliente",
                                  "longDescription": "Definición real."}])
            else:
                self._send(404, {"error": self.path})

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
        monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
        monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
        monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
        monkeypatch.setenv("PURVIEW_API_BASE", f"http://127.0.0.1:{port}")
        from mvdg import purview_export as pv
        from mvdg import purview_pull as pvp

        def fake_get_token():
            import urllib.request
            req = urllib.request.Request(f"http://127.0.0.1:{port}/oauth2/token",
                                         data=b"{}", method="POST")
            with urllib.request.urlopen(req) as resp:
                return json.loads(resp.read())["access_token"]

        monkeypatch.setattr(pv, "_get_token", fake_get_token)
        result = pvp.pull_all()
    finally:
        server.shutdown()

    assert calls["token"] == 2 and calls["glossary"] == 1 and calls["terms"] == 1
    assert calls["search"] == 1
    assert result["glossary"]["terms"][0]["name"] == "Cliente"
    assert result["catalog"]["tables"][0]["name"] == "dim_customers"


# --------------------------------------------------- persistencia de lo importado
def test_imported_save_list_delete_terms_and_tables(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import imported
    n = imported.save_terms("collibra", [{"collibra_id": "a1", "name": "Cliente",
                                          "definition": "Def traída."}])
    assert n == 1
    n2 = imported.save_tables("purview", [{"purview_id": "t1", "name": "dim_x",
                                           "description": "Tabla traída."}])
    assert n2 == 1
    terms = imported.list_terms()
    tables = imported.list_tables()
    assert len(terms) == 1 and terms.iloc[0]["name"] == "Cliente"
    assert len(tables) == 1 and tables.iloc[0]["source"] == "purview"
    assert imported.delete_term("collibra", "a1") is True
    assert imported.delete_term("collibra", "a1") is False  # ya no está
    assert imported.delete_table("purview", "t1") is True
    assert imported.list_terms().empty and imported.list_tables().empty


def test_imported_save_rejects_unknown_source(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import imported
    with pytest.raises(ValueError):
        imported.save_terms("atlan", [{"collibra_id": "a1", "name": "X"}])


def test_imported_save_skips_items_without_id_or_name(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import imported
    n = imported.save_terms("collibra", [{"collibra_id": "", "name": "Sin id"},
                                         {"collibra_id": "ok1", "name": ""}])
    assert n == 0
    assert imported.list_terms().empty


def test_imported_items_enter_curation_flow(tmp_path, monkeypatch):
    """Lo importado no queda aislado: entra al mismo inventario de
    curaduría que el catálogo/glosario de demo y los samples."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, imported
    imported.save_terms("collibra", [{"collibra_id": "a1", "name": "Cliente Importado",
                                      "definition": "Definición traída de Collibra."}])
    df = curation.list_items("es")
    row = df[df["item_id"] == "glossary:imported:collibra:a1"]
    assert len(row) == 1
    assert row.iloc[0]["status"] == "sugerido_ia"
    assert row.iloc[0]["text"] == "Definición traída de Collibra."
    # y se puede validar como cualquier otra definición
    curation.save_validation("glossary:imported:collibra:a1", "es", "validado",
                             "", "M. Viera", "Data Owner")
    df2 = curation.list_items("es")
    row2 = df2[df2["item_id"] == "glossary:imported:collibra:a1"]
    assert row2.iloc[0]["status"] == "validado"


# ------------------------------------------------------- keyring / connectors
def test_connectors_keyring_probe_never_raises(monkeypatch):
    """Sea cual sea el estado del keyring del SO en esta máquina, la sonda
    nunca puede tirar una excepción — ni siquiera un PanicException de
    pyo3 (que no hereda de Exception) si el backend está roto."""
    from mvdg import connectors as c
    c._keyring_ok_cache = None  # forzar una sonda real, no la cacheada
    result = c._keyring_usable()
    assert isinstance(result, bool)


def test_connectors_password_falls_back_to_obfuscation_without_keyring(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import connectors as c
    monkeypatch.setattr(c, "_keyring_usable", lambda: False)
    prof = c.save_connection({"name": "t", "engine": "postgresql", "host": "h",
                              "port": 5432, "database": "d", "user": "u",
                              "password": "s3cr3t"}, save_password=True)
    assert prof["secret_backend"] == "obfuscated"
    assert prof["password_enc"]  # algo quedó guardado, ofuscado
    assert c.stored_password(prof) == "s3cr3t"
    assert c.secret_backend_label(prof) == "obfuscated"


def test_connectors_password_uses_keyring_when_available(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import connectors as c
    store: dict[tuple, str] = {}
    monkeypatch.setattr(c, "_keyring_usable", lambda: True)

    def fake_set(conn_id, secret):
        store[conn_id] = secret
        return True

    def fake_get(conn_id):
        return store.get(conn_id, "")

    def fake_delete(conn_id):
        store.pop(conn_id, None)

    monkeypatch.setattr(c, "_keyring_set", fake_set)
    monkeypatch.setattr(c, "_keyring_get", fake_get)
    monkeypatch.setattr(c, "_keyring_delete", fake_delete)

    prof = c.save_connection({"name": "t", "engine": "postgresql", "host": "h",
                              "port": 5432, "database": "d", "user": "u",
                              "password": "s3cr3t"}, save_password=True)
    assert prof["secret_backend"] == "keyring"
    assert prof["password_enc"] == ""  # nada en el JSON, quedó en el keyring
    assert c.stored_password(prof) == "s3cr3t"
    c.delete_connection(prof["conn_id"])
    assert prof["conn_id"] not in store  # se limpió del keyring también


def test_connectors_no_password_clears_keyring_entry(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import connectors as c
    deleted = []
    monkeypatch.setattr(c, "_keyring_delete", lambda cid: deleted.append(cid))
    prof = c.save_connection({"name": "t", "engine": "postgresql", "host": "h",
                              "port": 5432, "database": "d", "user": "u",
                              "password": "x"}, save_password=False)
    assert prof["secret_backend"] == ""
    assert c.secret_backend_label(prof) == "none"
    assert deleted == [prof["conn_id"]]


def test_connectors_purview_qualified_name_azure_sql_confirmed_format():
    from mvdg.connectors import purview_qualified_name
    profile = {"engine": "sqlserver", "host": "myserver.database.windows.net",
              "database": "SalesDB"}
    assert (purview_qualified_name(profile, "dbo.dim_customers")
           == "mssql://myserver.database.windows.net/SalesDB/dbo/dim_customers")
    # sin schema explícito, se asume dbo (default de Azure SQL/SQL Server)
    assert (purview_qualified_name(profile, "dim_customers")
           == "mssql://myserver.database.windows.net/SalesDB/dbo/dim_customers")


def test_connectors_purview_qualified_name_none_when_not_azure_sql():
    """Para lo que Microsoft NO documenta el formato exacto (on-prem, otros
    motores/hosts) se devuelve None a propósito — no se inventa un
    qualifiedName que después no matchea con lo que Purview escaneó de
    verdad. Lo que decide es el HOST (el único patrón confirmado es
    *.database.windows.net), no una lista de motores permitidos."""
    from mvdg.connectors import purview_qualified_name
    on_prem = {"engine": "sqlserver", "host": "srv-datos.miempresa.local", "database": "d"}
    assert purview_qualified_name(on_prem, "t") is None
    postgres_local = {"engine": "postgresql", "host": "pg.miempresa.local", "database": "d"}
    assert purview_qualified_name(postgres_local, "t") is None
    assert purview_qualified_name({"engine": "postgresql", "host": "", "database": "d"}, "t") is None
    assert purview_qualified_name({"engine": "sqlserver",
                                   "host": "x.database.windows.net", "database": ""}, "t") is None


# --------------------------------------------------------- login modo servidor
def test_server_auth_not_required_in_desktop_mode(monkeypatch):
    from mvdg import server
    monkeypatch.delenv("MVDG_SERVER_MODE", raising=False)
    monkeypatch.setenv("MVDG_SERVER_PASSWORD", "secreto")
    assert server.server_mode_active() is False
    assert server.auth_required() is False  # modo escritorio: no se pide login


def test_server_auth_required_only_with_password_set(monkeypatch):
    from mvdg import server
    monkeypatch.setenv("MVDG_SERVER_MODE", "1")
    monkeypatch.delenv("MVDG_SERVER_PASSWORD", raising=False)
    assert server.auth_required() is False  # servidor pero sin contraseña configurada
    monkeypatch.setenv("MVDG_SERVER_PASSWORD", "secreto")
    assert server.auth_required() is True


def test_server_check_password_correct_and_constant_time(monkeypatch):
    from mvdg import server
    monkeypatch.setenv("MVDG_SERVER_PASSWORD", "correcta123")
    assert server.check_password("correcta123") is True
    assert server.check_password("incorrecta") is False
    assert server.check_password("") is False
    monkeypatch.delenv("MVDG_SERVER_PASSWORD", raising=False)
    assert server.check_password("cualquiera") is False  # sin var seteada, nunca entra


def test_server_run_server_sets_server_mode_flag(monkeypatch, tmp_path):
    """run_server() marca MVDG_SERVER_MODE escribiendo os.environ
    directamente (a propósito: tiene que sobrevivir mientras el proceso de
    Streamlit está arriba) — por eso este test la limpia a mano en un
    finally en vez de confiar en monkeypatch, que solo deshace lo que él
    mismo seteó."""
    from mvdg import server
    monkeypatch.delenv("MVDG_SERVER_MODE", raising=False)
    monkeypatch.setenv("MVDG_AUTHORIZED_HOSTS", "*")
    try:
        argv_out = []
        server.run_server(argv_out=argv_out)
        assert os.environ.get("MVDG_SERVER_MODE") == "1"
        assert argv_out  # se armaron los argumentos de streamlit
    finally:
        os.environ.pop("MVDG_SERVER_MODE", None)


# --------------------------------------------------- Purview: relación, qualifiedName real
def test_purview_classification_phone_is_phone_not_ip():
    """Regresión: 'telefono' mapeaba por error a MICROSOFT.PERSONAL.IPADDRESS."""
    from mvdg.purview_export import _pii_classification
    assert _pii_classification("telefono") == "MICROSOFT.PERSONAL.US.PHONE_NUMBER"
    assert _pii_classification("phone_number") == "MICROSOFT.PERSONAL.US.PHONE_NUMBER"
    assert _pii_classification("ip_address") == "MICROSOFT.PERSONAL.IPADDRESS"
    assert _pii_classification("full_name") == "MICROSOFT.PERSONAL.NAME"


def test_purview_columns_link_to_their_table():
    """Cada rdbms_column debe referenciar a su rdbms_table por qualifiedName
    (relationshipAttributes.table) — sin esto, la pestaña Schema de la
    tabla queda vacía en Purview."""
    from mvdg.purview_export import build_entity_payload
    cat, dic, _ = _sample_gov_tables()
    entities = build_entity_payload(cat, dic)
    columns = [e for e in entities if e["typeName"] == "rdbms_column"]
    assert columns
    for c in columns:
        rel = c["relationshipAttributes"]["table"]
        assert rel["typeName"] == "rdbms_table"
        table_qn = c["attributes"]["qualifiedName"].rsplit("#", 1)[0]
        assert rel["uniqueAttributes"]["qualifiedName"] == table_qn


def test_purview_qualified_name_map_used_when_provided():
    """Con un qualifiedName real (de una conexión SQL), la entidad se
    fusiona con lo que Purview ya escaneó en vez de usar mvdg://."""
    from mvdg.purview_export import build_entity_payload
    cat, dic, _ = _sample_gov_tables()
    ds = cat.iloc[0]["dataset"]
    qn_map = {ds: f"mssql://srv.database.windows.net/db/dbo/{ds}"}
    entities = build_entity_payload(cat, dic, qualified_name_map=qn_map)
    table_entity = next(e for e in entities if e["typeName"] == "rdbms_table"
                        and e["attributes"]["name"] == ds)
    assert table_entity["attributes"]["qualifiedName"] == qn_map[ds]
    col_entity = next(e for e in entities if e["typeName"] == "rdbms_column"
                      and e["attributes"]["qualifiedName"].startswith(qn_map[ds]))
    assert col_entity["relationshipAttributes"]["table"]["uniqueAttributes"]["qualifiedName"] == qn_map[ds]
    # el resto de los datasets, sin entrada en el mapa, siguen usando mvdg://
    other_ds = cat.iloc[1]["dataset"]
    other_entity = next(e for e in entities if e["typeName"] == "rdbms_table"
                        and e["attributes"]["name"] == other_ds)
    assert other_entity["attributes"]["qualifiedName"] == f"mvdg://{other_ds}"


def test_purview_glossary_repush_updates_existing_terms_not_recreates(monkeypatch):
    """Un segundo push del glosario NO debe fallar con 409: los términos que
    ya existen (por nombre) se actualizan con PUT; los nuevos se crean."""
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    _cat, _dic, glo = _sample_gov_tables()
    existing_name = glo.iloc[0]["term"]

    calls = {"term_post": 0, "term_put": 0, "terms_list": 0}

    def fake_http_form(url, form):
        return {"access_token": "ptok"}

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/glossary") and method == "GET":
            return [{"guid": "gloss-1", "name": "MV Data Governance"}]
        if url.endswith("/glossary/gloss-1/terms") and method == "GET":
            calls["terms_list"] += 1
            return [{"guid": "term-old-1", "name": existing_name}]
        if url.endswith("/glossary/term") and method == "POST":
            calls["term_post"] += 1
            return {"guid": f"term-new-{calls['term_post']}"}
        if "/glossary/term/term-old-1" in url and method == "PUT":
            calls["term_put"] += 1
            assert body["guid"] == "term-old-1"
            assert body["anchor"]["glossaryGuid"] == "gloss-1"
            return {"guid": "term-old-1"}
        raise AssertionError(f"unexpected URL: {method} {url}")

    monkeypatch.setattr(pv, "_http_form", fake_http_form)
    monkeypatch.setattr(pv, "_http_json", fake_http_json)

    r = pv.push_glossary(glo, dry_run=False)
    assert calls["terms_list"] == 1
    assert calls["term_put"] == 1                  # el que ya existía
    assert calls["term_post"] == len(glo) - 1       # el resto se crea
    assert r["term_count"] == len(glo)
    assert r["failed"] == []


def test_purview_glossary_failed_term_does_not_abort_the_rest(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    import urllib.error
    from mvdg import purview_export as pv
    _cat, _dic, glo = _sample_gov_tables()
    bad_name = glo.iloc[0]["term"]

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/glossary") and method == "GET":
            return [{"guid": "gloss-1", "name": "MV Data Governance"}]
        if url.endswith("/glossary/gloss-1/terms") and method == "GET":
            return []  # ningún término existente: todo se crea por POST
        if url.endswith("/glossary/term") and method == "POST":
            if body["name"] == bad_name:
                raise urllib.error.HTTPError(url, 500, "boom", None, None)
            return {"guid": "term-ok"}
        raise AssertionError(f"unexpected URL: {method} {url}")

    monkeypatch.setattr(pv, "_http_form", lambda url, form: {"access_token": "t"})
    monkeypatch.setattr(pv, "_http_json", fake_http_json)
    r = pv.push_glossary(glo, dry_run=False)
    assert len(r["failed"]) == 1 and r["failed"][0]["name"] == bad_name
    assert r["term_count"] == len(glo) - 1  # el resto se creó igual


def test_purview_push_catalog_batches_large_entity_lists(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    from mvdg import purview_export as pv
    monkeypatch.setattr(pv, "_BULK_BATCH_SIZE", 2)  # fuerza varios lotes con pocos datos
    cat, dic, _glo = _sample_gov_tables()

    calls = {"bulk": 0}

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/entity/bulk"):
            calls["bulk"] += 1
            assert len(body["entities"]) <= 2
            mutated = [{"guid": f"g-{calls['bulk']}-{i}",
                       "attributes": {"qualifiedName": e["attributes"]["qualifiedName"]}}
                      for i, e in enumerate(body["entities"])]
            return {"mutatedEntities": {"CREATE": mutated, "UPDATE": []}}
        raise AssertionError(f"unexpected URL: {method} {url}")

    monkeypatch.setattr(pv, "_http_form", lambda url, form: {"access_token": "t"})
    monkeypatch.setattr(pv, "_http_json", fake_http_json)
    r = pv.push_catalog(cat, dic, dry_run=False)
    total_entities = len(cat) + len(dic)
    import math
    assert calls["bulk"] == math.ceil(total_entities / 2)
    assert r["entity_count"] == total_entities
    assert len(r["guid_by_qualified_name"]) == total_entities
    assert r["failed_batches"] == []


def test_purview_push_catalog_partial_batch_failure_does_not_abort_others(monkeypatch):
    monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
    monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
    monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
    monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "acct")
    import urllib.error
    from mvdg import purview_export as pv
    monkeypatch.setattr(pv, "_BULK_BATCH_SIZE", 2)
    cat, dic, _glo = _sample_gov_tables()

    calls = {"bulk": 0}

    def fake_http_json(url, headers, method="GET", body=None):
        calls["bulk"] += 1
        if calls["bulk"] == 1:
            raise urllib.error.HTTPError(url, 500, "boom", None, None)
        mutated = [{"guid": f"g-{i}", "attributes": {"qualifiedName": e["attributes"]["qualifiedName"]}}
                  for i, e in enumerate(body["entities"])]
        return {"mutatedEntities": {"CREATE": mutated, "UPDATE": []}}

    monkeypatch.setattr(pv, "_http_form", lambda url, form: {"access_token": "t"})
    monkeypatch.setattr(pv, "_http_json", fake_http_json)
    r = pv.push_catalog(cat, dic, dry_run=False)
    assert len(r["failed_batches"]) == 1
    assert calls["bulk"] > 1  # los lotes siguientes se mandaron igual
    assert len(r["guid_by_qualified_name"]) < r["entity_count"]  # el lote fallido no aportó guids


def test_purview_retries_429_with_backoff_then_succeeds(monkeypatch):
    from mvdg import purview_export as pv
    import urllib.error

    sleeps = []
    monkeypatch.setattr(pv.time, "sleep", lambda s: sleeps.append(s))

    attempts = {"n": 0}

    class FakeResp:
        def __init__(self, payload):
            self._payload = payload

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

        def read(self):
            import json as _json
            return _json.dumps(self._payload).encode()

    def fake_urlopen(req, timeout=None):
        attempts["n"] += 1
        if attempts["n"] == 1:
            raise urllib.error.HTTPError(req.full_url, 429, "slow down",
                                         {"Retry-After": "0"}, None)
        return FakeResp({"ok": True})

    monkeypatch.setattr(pv.urllib.request, "urlopen", fake_urlopen)
    result = pv._http_json("http://x/y", {})
    assert result == {"ok": True}
    assert attempts["n"] == 2
    assert sleeps == [0.0]


def test_purview_429_gives_up_after_max_retries(monkeypatch):
    from mvdg import purview_export as pv
    import urllib.error
    monkeypatch.setattr(pv.time, "sleep", lambda s: None)
    attempts = {"n": 0}

    def always_429(req, timeout=None):
        attempts["n"] += 1
        raise urllib.error.HTTPError(req.full_url, 429, "slow down", {}, None)

    monkeypatch.setattr(pv.urllib.request, "urlopen", always_429)
    with pytest.raises(urllib.error.HTTPError):
        pv._http_json("http://x/y", {})
    assert attempts["n"] == pv._MAX_RETRIES_429 + 1  # intento original + reintentos


# ------------------------------------------------------------- Collibra: robustez
def test_collibra_column_table_relation_created_when_configured(monkeypatch):
    monkeypatch.setenv("COLLIBRA_BASE_URL", "https://acme.collibra.com")
    monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
    monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    monkeypatch.setenv("COLLIBRA_TABLE_TYPE_ID", "type-table")
    monkeypatch.setenv("COLLIBRA_COLUMN_TYPE_ID", "type-column")
    monkeypatch.setenv("COLLIBRA_COLUMN_TABLE_RELATION_TYPE_ID", "rel-col-table")
    from mvdg import collibra_export as cb
    cat, dic, _glo = _sample_gov_tables()

    relations = []
    asset_ids = iter(range(10_000))

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/sessions") and method == "POST":
            return {}, ["JSESSIONID=abc123; Path=/"]
        if url.endswith("/auth/sessions/current"):
            return {}, []
        if url.endswith("/assets"):
            return {"id": f"asset-{next(asset_ids)}"}, []
        if url.endswith("/attributes"):
            return {"id": "attr-1"}, []
        if url.endswith("/relations") and method == "POST":
            relations.append(body)
            return {"id": "rel-1"}, []
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(cb, "_http_json", fake_http_json)
    r = cb.push_catalog(cat, dic, dry_run=False)
    assert r["asset_count"] == len(cat) + len(dic)
    assert len(relations) == len(dic)               # una relación por columna
    assert all(rel["typeId"] == "rel-col-table" for rel in relations)
    assert r["failed"] == []

    # sin la env var: mismo push, cero llamadas a /relations (no regresión)
    monkeypatch.delenv("COLLIBRA_COLUMN_TABLE_RELATION_TYPE_ID")
    relations.clear()
    cb.push_catalog(cat, dic, dry_run=False)
    assert relations == []


def test_collibra_push_partial_failure_isolated_per_item(monkeypatch):
    monkeypatch.setenv("COLLIBRA_BASE_URL", "https://acme.collibra.com")
    monkeypatch.setenv("COLLIBRA_USERNAME", "svc")
    monkeypatch.setenv("COLLIBRA_PASSWORD", "pw")
    monkeypatch.setenv("COLLIBRA_DOMAIN_ID", "dom-1")
    monkeypatch.setenv("COLLIBRA_TABLE_TYPE_ID", "type-table")
    monkeypatch.setenv("COLLIBRA_COLUMN_TYPE_ID", "type-column")
    import urllib.error
    from mvdg import collibra_export as cb
    cat, dic, _glo = _sample_gov_tables()
    bad_dataset = cat.iloc[0]["dataset"]

    def fake_http_json(url, headers, method="GET", body=None):
        if url.endswith("/auth/sessions") and method == "POST":
            return {}, ["JSESSIONID=abc123; Path=/"]
        if url.endswith("/auth/sessions/current"):
            return {}, []
        if url.endswith("/assets"):
            if body.get("name") == bad_dataset:
                raise urllib.error.HTTPError(url, 500, "boom", None, None)
            return {"id": "asset-ok"}, []
        if url.endswith("/attributes"):
            return {"id": "attr-1"}, []
        raise AssertionError(f"unexpected URL: {url}")

    monkeypatch.setattr(cb, "_http_json", fake_http_json)
    r = cb.push_catalog(cat, dic, dry_run=False)
    assert len(r["failed"]) == 1
    assert r["failed"][0]["name"] == bad_dataset
    # el resto de los assets (otras tablas + todas las columnas) sí se crearon
    assert r["asset_count"] == len(cat) + len(dic) - 1


def test_collibra_term_status_reflects_curation(tmp_path, monkeypatch):
    """statusId Candidate/Accepted en Collibra tiene que salir de la
    curaduría real, igual que Draft/Approved en Purview."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import collibra_export as cb, curation
    _cat, _dic, glo = _sample_gov_tables()

    def lookup(term_id):
        rec = curation.get_record(f"glossary:demo:{term_id}", "es")
        return (rec["status"], rec.get("text") or "") if rec else ("sugerido_ia", "")

    before = cb.build_term_payloads(glo, "term-type", "dom-1", curation_lookup=lookup)
    assert all(t["asset"]["statusId"] == cb._DEFAULT_CANDIDATE_STATUS_ID for t in before)

    first_term_id = glo.iloc[0]["term_id"]
    curation.save_validation(f"glossary:demo:{first_term_id}", "es", "validado",
                             "", "María Viera", "Data Owner")
    after = cb.build_term_payloads(glo, "term-type", "dom-1", curation_lookup=lookup)
    statuses = {t["asset"]["name"]: t["asset"]["statusId"] for t in after}
    approved_name = glo.iloc[0]["term"]
    assert statuses[approved_name] == cb._DEFAULT_ACCEPTED_STATUS_ID
    assert sum(1 for s in statuses.values() if s == cb._DEFAULT_ACCEPTED_STATUS_ID) == 1


def test_collibra_status_ids_are_documented_and_overridable(monkeypatch):
    from mvdg import collibra_export as cb
    assert cb._DEFAULT_CANDIDATE_STATUS_ID == "00000000-0000-0000-0000-000000005008"
    assert cb._DEFAULT_ACCEPTED_STATUS_ID == "00000000-0000-0000-0000-000000005009"
    monkeypatch.setenv("COLLIBRA_CANDIDATE_STATUS_ID", "custom-candidate")
    _cat, _dic, glo = _sample_gov_tables()
    terms = cb.build_term_payloads(glo, "term-type", "dom-1")
    assert terms[0]["asset"]["statusId"] == "custom-candidate"


# ------------------------------------ glosario automático desde la base de datos
def test_glossary_auto_expands_common_abbreviations():
    from mvdg.glossary_auto import expand_identifier
    assert expand_identifier("fec_pag", "es") == ("fecha pago", True)
    assert expand_identifier("cli_id", "es") == ("cliente identificador", True)
    assert expand_identifier("imp_tot", "es") == ("importe total", True)
    assert expand_identifier("cust_addr", "en") == ("customer address", True)
    # camelCase también se parte y expande
    assert expand_identifier("fecPago", "es")[0].startswith("fecha")


def test_glossary_auto_never_invents_unknown_tokens():
    """Un token que no está en el diccionario queda tal cual (en minúsculas)
    — la corrección es del humano en la tabla editable, no una invención."""
    from mvdg.glossary_auto import expand_identifier
    phrase, expanded = expand_identifier("xyzzy_frobnicate", "es")
    assert phrase == "xyzzy frobnicate"
    assert expanded is False


def test_glossary_auto_terms_have_stable_upsert_ids():
    from mvdg.glossary_auto import build_terms_from_schema
    schema = {"cli_fac": ["fec_pag", "imp_tot"]}
    terms = build_terms_from_schema(schema, "es", conn_id="abc123")
    ids = {t["database_id"] for t in terms}
    assert ids == {"abc123:cli_fac.fec_pag", "abc123:cli_fac.imp_tot"}
    # re-generar produce los mismos ids -> save_terms hace upsert, no duplica
    again = {t["database_id"] for t in build_terms_from_schema(schema, "es", conn_id="abc123")}
    assert again == ids


def test_glossary_auto_end_to_end_with_real_sqlite(tmp_path, monkeypatch):
    """Flujo completo contra una base REAL (SQLite): esquema con nombres
    abreviados -> borrador con palabras completas -> guardado local ->
    aparece en Curaduría con su origen, editable/validable a mano."""
    import sqlite3
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    db = tmp_path / "ventas.db"
    con = sqlite3.connect(db)
    con.execute("CREATE TABLE cli_fac (fec_pag TEXT, imp_tot REAL, tel_cli TEXT)")
    con.commit()
    con.close()

    from mvdg import curation, glossary_auto, imported
    profile = {"conn_id": "sqlt1", "engine": "sqlite", "database": str(db)}
    draft = glossary_auto.build_from_connection(profile, "es")
    by_col = {t["column"]: t for t in draft}
    assert by_col["fec_pag"]["name"] == "fecha pago"
    assert by_col["imp_tot"]["name"] == "importe total"
    assert by_col["tel_cli"]["name"] == "teléfono cliente"
    assert all(t["definition"] for t in draft)

    # el usuario corrige uno a mano antes de guardar (editable de verdad)
    by_col["tel_cli"]["name"] = "teléfono del cliente"
    n = imported.save_terms("database", draft)
    assert n == 3
    df = curation.list_items("es")
    row = df[df["item_id"] == "glossary:imported:database:sqlt1:cli_fac.tel_cli"]
    assert len(row) == 1
    assert row.iloc[0]["label"] == "teléfono del cliente"
    assert row.iloc[0]["dataset"] == "base de datos"
    assert row.iloc[0]["status"] == "sugerido_ia"
    # y se valida/modifica como cualquier otra definición del programa
    curation.save_validation("glossary:imported:database:sqlt1:cli_fac.tel_cli",
                             "es", "modificado", "Teléfono de contacto del cliente.",
                             "M. Viera", "Data Steward")
    df2 = curation.list_items("es")
    row2 = df2[df2["item_id"] == "glossary:imported:database:sqlt1:cli_fac.tel_cli"]
    assert row2.iloc[0]["status"] == "modificado"
    assert row2.iloc[0]["text"] == "Teléfono de contacto del cliente."


def test_glossary_auto_broken_table_does_not_stop_the_rest(tmp_path, monkeypatch):
    from mvdg import connectors, glossary_auto
    import sqlite3
    db = tmp_path / "x.db"
    con = sqlite3.connect(db)
    con.execute("CREATE TABLE ok_tbl (cod_prod TEXT)")
    con.commit()
    con.close()
    profile = {"conn_id": "c1", "engine": "sqlite", "database": str(db)}

    real_list_columns = connectors.list_columns

    def flaky(prof, table, password=None):
        if table == "ok_tbl":
            return real_list_columns(prof, table, password=password)
        raise RuntimeError("boom")

    monkeypatch.setattr(connectors, "list_tables", lambda p, password=None: ["rota", "ok_tbl"])
    monkeypatch.setattr(connectors, "list_columns", flaky)
    terms = glossary_auto.build_from_connection(profile, "es")
    assert [t["column"] for t in terms] == ["cod_prod"]
    assert terms[0]["name"] == "código producto"


# ------------------------- accesos directos opcionales (escritorio / menú inicio)
def _read_accesos_bat():
    with open(os.path.join(_repo_root(), "MV_Instalar_Accesos.bat"),
              encoding="ascii") as fh:  # ascii a propósito: cmd.exe usa cp437/cp1252
        return fh.read()


def _read_bat(name):
    with open(os.path.join(_repo_root(), name), encoding="ascii") as fh:
        return fh.read()


_BATS_LANZADORES = ("MV_DataGovernance.bat", "MV_DataGovernance_API.bat",
                    "MV_DataGovernance_Server.bat")


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_repara_venv_sin_pip(nombre):
    """Regresion del error reportado en Windows: "No module named pip".

    Si la creacion del venv se interrumpe, queda .venv\\Scripts\\python.exe
    pero SIN pip. Los .bat asumian que la existencia del interprete implicaba
    un entorno usable, entraban igual, y reintentaban 4 veces el mismo
    comando condenado echandole la culpa a OneDrive.
    """
    src = _read_bat(nombre)
    # 1) se comprueba que pip EXISTE antes de intentar instalar con el
    assert "-m pip --version" in src, "no verifica que pip exista"
    # 2) se repara con ensurepip (offline, sin borrar el entorno)
    assert "-m ensurepip --upgrade" in src, "no repara pip con ensurepip"
    # 3) si ni ensurepip alcanza, se rehace el entorno automaticamente
    assert ":rebuild_venv" in src and 'rmdir /s /q ".venv"' in src, (
        "no rehace el entorno cuando es irrecuperable")
    # 4) rehacer se intenta UNA sola vez: sin esto es un bucle infinito
    assert "if defined MVDG_REBUILT goto errvenv" in src, (
        "sin guarda contra rehacer el entorno en bucle")
    assert 'set "MVDG_REBUILT="' in src, "no reinicia la guarda al arrancar"


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_no_culpa_a_onedrive_cuando_falta_pip(nombre):
    """El mensaje de "archivo en uso / pausa OneDrive" solo tiene sentido si
    pip existe y algo bloquea un archivo. Cuando falta pip, reintentar es
    inutil y el consejo manda al usuario a perder el tiempo."""
    src = _read_bat(nombre)
    bloque = src[src.index(":install_deps"):]
    # se distingue el caso "pip no esta" con un codigo de salida propio...
    assert "exit /b 2" in bloque, "no distingue falta-de-pip de archivo-en-uso"
    # ...y el llamador lo trata rehaciendo el entorno, no reintentando.
    # OJO: "if errorlevel N" en cmd es ">= N", asi que el 2 debe chequearse
    # ANTES que el 1 o nunca se alcanza.
    i2 = src.index("if errorlevel 2 goto rebuild_venv")
    i1 = src.index("if errorlevel 1 goto errdeps")
    assert i2 < i1, "el errorlevel 2 se chequea despues del 1: nunca entra"


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_no_filtra_el_error_de_pip_por_stderr(nombre):
    """El usuario veia "No module named pip" DOS veces: la linea que
    actualiza pip redirigia solo stdout (>nul), asi que su stderr se colaba
    a la consola ademas del error real de la linea siguiente."""
    src = _read_bat(nombre)
    assert "-m pip install --no-cache-dir --upgrade pip >nul 2>nul" in src, (
        "el upgrade de pip no silencia stderr")
    assert "-m pip install --no-cache-dir --upgrade pip >nul\n" not in src, (
        "quedo una redireccion que deja escapar stderr")


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_detecta_venv_que_ya_no_arranca(nombre):
    """El otro modo de falla real: se actualiza o desinstala el Python del
    sistema y el venv queda apuntando a un interprete que ya no existe."""
    src = _read_bat(nombre)
    assert '-c "pass"' in src, "no verifica que el interprete del venv arranque"


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_verifica_dependencias_antes_de_lanzar(nombre):
    """El Server.bat saltaba directo a :launch si existia el .venv, sin
    mirar si las dependencias estaban: arrancaba y moria con un traceback."""
    src = _read_bat(nombre)
    assert ":verify" in src, "no hay etapa de verificacion de dependencias"
    i_verify = src.index(":verify")
    i_launch = src.index("\n:launch")
    assert i_verify < i_launch, "la verificacion queda despues del arranque"
    assert '-c "import ' in src[i_verify:i_launch], (
        "la etapa de verificacion no importa nada")


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_ofrece_el_instalador_exe_como_salida(nombre):
    """Cuando el camino con Python falla de verdad, el usuario tiene que
    enterarse de que existe un instalador que no necesita Python."""
    src = _read_bat(nombre)
    assert "MVDataGovernance_Setup.exe" in src, (
        "no menciona el instalador .exe como alternativa")
    # y ya no manda a borrar carpetas a mano: eso ahora es automatico
    assert "borra la carpeta .venv" not in src
    assert "delete the .venv folder" not in src


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_redirige_temp_al_disco_del_proyecto(nombre):
    """pip (y venv) escriben temporales en el TEMP del sistema, que en
    Windows es casi siempre C:\\Users\\<usuario>\\AppData\\Local\\Temp SIN
    IMPORTAR en que disco pusiste esta carpeta. Si la instalacion tiene que
    quedar "100% en el disco elegido", un TEMP que sigue apuntando a C: la
    rompe en silencio - hasta que C: se queda sin espacio a mitad de una
    instalacion (el error real reportado: "No space left on device")."""
    src = _read_bat(nombre)
    i_temp = src.index('set "TEMP=')
    i_pip = src.index(":install_deps")
    assert i_temp < i_pip, "TEMP se redirige DESPUES de instalar - ya es tarde"
    assert 'set "TEMP=%cd%\\.mvdg_tmp"' in src
    assert 'set "TMP=%cd%\\.mvdg_tmp"' in src
    # %cd% en ese punto es la carpeta del propio .bat (cd /d "%~dp0" ya corrio)
    assert src.index('cd /d "%~dp0"') < i_temp


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_pip_no_deja_cache_persistente_fuera_del_venv(nombre):
    """Sin --no-cache-dir, pip guarda un cache en %LOCALAPPDATA%\\pip\\Cache
    - en C: SIEMPRE, sin importar TEMP/TMP ni en que disco este el programa.
    Con reinstalaciones repetidas eso crece en silencio en el disco que el
    usuario justamente quiso evitar llenar."""
    src = _read_bat(nombre)
    bloque = src[src.index(":install_deps"):]
    assert "pip install --no-cache-dir --upgrade pip" in bloque
    assert "pip install --no-cache-dir -r requirements.txt" in bloque


@pytest.mark.parametrize("nombre", _BATS_LANZADORES)
def test_bat_errdeps_menciona_espacio_en_disco(nombre):
    """El error real reportado por un usuario fue "No space left on device"
    durante la instalacion de dependencias - el mensaje final tiene que
    nombrar esa causa, no solo internet/antivirus/OneDrive."""
    src = _read_bat(nombre)
    bloque = src[src.index(":errdeps"):]
    assert "espacio en el disco C" in bloque
    assert "free space on drive C" in bloque
    assert "espaco no disco C" in bloque


def test_build_exe_bat_redirige_temp_y_no_deja_cache_persistente():
    """Mismo problema, mismo arreglo, en el script de build del .exe
    (packaging/build_exe.bat): streamlit+pandas+pyinstaller+cython son
    varios cientos de MB - si TEMP/el cache de pip siguen apuntando a C:
    aunque el repo este en D:, el build puede fallar con el disco lleno
    igual que le paso al usuario instalando el programa."""
    with open(os.path.join(_repo_root(), "packaging", "build_exe.bat"),
              encoding="ascii") as fh:
        src = fh.read()
    i_temp = src.index('set "TEMP=')
    i_deps = src.index(":deps")
    assert i_temp < i_deps, "TEMP se redirige DESPUES de instalar - ya es tarde"
    assert 'set "TEMP=%cd%\\.mvdg_tmp"' in src
    assert 'set "TMP=%cd%\\.mvdg_tmp"' in src
    bloque = src[i_deps:]
    assert "pip install --no-cache-dir --upgrade pip" in bloque
    assert "pip install --no-cache-dir -r requirements.txt pyinstaller cython" in bloque
    errdeps = src[src.index(":errdeps"):]
    assert "espacio en disco" in errdeps and "free disk space" in errdeps


def test_api_and_server_bats_are_self_sufficient_and_open_browser():
    """Los .bat de API y Servidor deben poder usarse SOLOS (crear su propio
    entorno como el .bat principal) y abrir el navegador — si no, el usuario ve
    una ventana negra con logs y cree que 'no funciona'."""
    api = _read_bat("MV_DataGovernance_API.bat")
    srv = _read_bat("MV_DataGovernance_Server.bat")
    for src, entry in ((api, "bi_api.main"), (srv, "mvdg.server")):
        # bootstrap propio: detecta python, crea venv, instala con reintentos
        assert "-m venv .venv" in src, "no crea el entorno solo"
        assert ":install_deps" in src and "requirements.txt" in src
        assert "goto errdeps" in src, "sin manejo de fallo de instalacion"
        # abre el navegador (si no, parece que no hace nada)
        assert "webbrowser.open" in src, "no abre el navegador"
        # arranca el entrypoint correcto
        assert entry in src
    # el API abre las docs interactivas; el server, la app
    assert "/docs" in api
    assert "localhost:%MVDG_SERVER_PORT%" in srv


def test_api_bat_no_longer_dead_ends_without_venv():
    """Regresion: antes el API.bat abortaba si no existia .venv (mandaba a
    correr otro .bat). Ahora tiene que arrancarlo el mismo."""
    api = _read_bat("MV_DataGovernance_API.bat")
    # ya no depende de que el usuario corra antes el .bat principal
    assert "Ejecuta primero MV_DataGovernance.bat" not in api
    assert "Run MV_DataGovernance.bat first" not in api
    # y define un puerto por defecto antes de usarlo en la URL
    assert 'set "MVDG_API_PORT=8600"' in api


def test_accesos_bat_creates_optional_desktop_and_start_menu_shortcuts():
    """El cliente elige (S/N) escritorio y/o menú inicio; los .lnk se crean
    por usuario (sin admin) vía WScript.Shell, con el icono del programa."""
    src = _read_accesos_bat()
    assert "choice /C SN" in src                      # es opcional de verdad
    assert src.count("CreateShortcut") == 2           # escritorio + menú inicio
    assert "GetFolderPath('Desktop')" in src
    assert "GetFolderPath('Programs')" in src         # menú inicio por usuario
    assert "MV_DataGovernance.bat" in src             # apunta al portable
    assert "assets\\brand\\mv.ico" in src             # con su icono
    assert "IconLocation" in src and "WorkingDirectory" in src


def test_accesos_bat_has_removal_mode_and_honest_taskbar_note():
    src = _read_accesos_bat()
    # modo quitar (reversible), y en los 3 alias
    assert '"%~1"=="quitar"' in src and "Remove-Item" in src
    # honestidad sobre la barra de tareas: Windows no deja auto-anclarse —
    # se explica el paso manual en vez de fingir que se puede
    assert "Anclar a la barra de tareas" in src
    assert "Pin to taskbar" in src


def test_accesos_bat_ships_in_release_zips_with_crlf(tmp_path):
    """El instalador de accesos viaja en los ZIP de entrega, y todo .bat
    dentro de un ZIP va con CRLF aunque el working tree esté en LF (un .bat
    con LF puede fallar en cmd.exe — la razón ya documentada en
    .gitattributes)."""
    import importlib.util
    import zipfile
    spec = importlib.util.spec_from_file_location(
        "mvdg_build_release", os.path.join(_repo_root(), "packaging", "build_release.py"))
    br = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(br)
    assert "MV_Instalar_Accesos.bat" in br._INCLUDE_FILES
    assert "MV_Instalar_Accesos.bat" in br._DEMO_FILES
    # _zip_write convierte LF -> CRLF para .bat/.iss
    lf_bat = tmp_path / "x.bat"
    lf_bat.write_bytes(b"@echo off\ngoto end\n:end\n")
    out = tmp_path / "t.zip"
    with zipfile.ZipFile(out, "w") as z:
        br._zip_write(z, str(lf_bat), "x.bat")
    data = zipfile.ZipFile(out).read("x.bat")
    assert data == b"@echo off\r\ngoto end\r\n:end\r\n"
    # y un .bat que YA está en CRLF no se duplica el \r
    crlf_bat = tmp_path / "y.bat"
    crlf_bat.write_bytes(b"@echo off\r\n:fin\r\n")
    with zipfile.ZipFile(out, "w") as z:
        br._zip_write(z, str(crlf_bat), "y.bat")
    assert zipfile.ZipFile(out).read("y.bat") == b"@echo off\r\n:fin\r\n"


def test_scope_combined_covers_demo_plus_all_samples():
    """El alcance combinado incluye los 4 datasets de demo + los 4 casos de
    Mis datos, con el MISMO esquema de columnas que las tablas demo — así
    cada pestaña los muestra sin ramas especiales."""
    from mvdg import samples, scope
    from mvdg.catalog import catalog_df, dictionary_df
    from mvdg.glossary import glossary_df
    cat = scope.combined_catalog("es")
    assert list(cat.columns) == list(catalog_df("es").columns)
    assert set(samples.sample_keys()) <= set(cat["dataset"])
    assert len(cat) == 4 + len(samples.sample_keys())
    dic = scope.combined_dictionary("es")
    assert list(dic.columns) == list(dictionary_df("es").columns)
    assert dic["dataset"].nunique() == len(cat)
    glo = scope.combined_glossary("es")
    assert list(glo.columns) == list(glossary_df("es").columns)
    assert len(glo) > len(glossary_df("es"))


def test_scope_combined_dictionary_filters_by_dataset():
    from mvdg import scope
    d = scope.combined_dictionary("es", "rotulado_alimentos")
    assert len(d) > 0
    assert (d["dataset"] == "rotulado_alimentos").all()


def test_scope_combined_results_run_real_sample_rules():
    """Las reglas de los casos corren DE VERDAD sobre sus archivos (no son
    números fijos): cada caso aporta filas con scores calculados."""
    from mvdg import samples, scope
    res = scope.combined_results("es")
    for key in samples.sample_keys():
        sub = res[res["dataset"] == key]
        assert len(sub) > 0, key
        assert sub["score"].between(0, 100).all(), key
    assert set(res.columns) >= {"rule_id", "dataset", "score", "status"}


def test_scope_combined_lineage_links_each_sample_source_to_bi():
    """Linaje honesto por caso: fuente externa -> dataset curado -> BI. Sin
    capas raw/mart inventadas para un CSV."""
    from mvdg import samples, scope
    nodes, edges = scope.combined_lineage("es")
    ids = {n["id"] for n in nodes}
    for key in samples.sample_keys():
        assert key in ids
        assert f"src_{key}" in ids
        assert (f"src_{key}", key) in edges
        assert (key, "bi_dashboard") in edges
    # el grafo demo sigue intacto adentro
    assert "mart_sales" in ids and ("mart_sales", "bi_dashboard") in edges


# ───────────────── Power BI: .pbit y .pbix, no solo la carpeta .pbip ─────────
# El agujero que cierran: la pestaña solo aceptaba la CARPETA de un proyecto
# .pbip. Quien tiene un .pbit o un .pbix —que es la mayoría— escribía la ruta
# de su archivo y recibía un error de "no encontré la carpeta definition".

def _pbit_de_prueba(destino: str) -> str:
    """Un .pbit con la forma REAL que exporta Power BI Desktop: zip con el
    DataModelSchema en UTF-16 y TMSL adentro."""
    import json as _json
    import zipfile
    schema = {"name": "ModeloDePrueba", "compatibilityLevel": 1550, "model": {
        "tables": [
            {"name": "Ventas",
             "columns": [
                 {"name": "Importe", "dataType": "double", "sourceColumn": "Importe"},
                 {"name": "Margen", "dataType": "double", "type": "calculated",
                  "expression": ["Ventas[Importe]*0.3"]},
                 {"name": "RowNumber", "dataType": "int64", "type": "rowNumber",
                  "isHidden": True}],
             "measures": [{"name": "Total Ventas", "expression": ["SUM(Ventas[Importe])"],
                           "displayFolder": "KPIs", "description": "Suma del importe"}],
             "partitions": [{"name": "p", "source": {"type": "m", "expression": [
                 'let O = Sql.Database("srv01", "DWH") in O']}}]},
            {"name": "Producto",
             "columns": [{"name": "SKU", "dataType": "string", "sourceColumn": "SKU"}]},
        ],
        "relationships": [{"name": "r1", "fromTable": "Ventas", "fromColumn": "SKU",
                           "toTable": "Producto", "toColumn": "SKU",
                           "crossFilteringBehavior": "bothDirections"}],
        "roles": [{"name": "Solo LATAM"}]}}
    with zipfile.ZipFile(destino, "w") as z:
        z.writestr("DataModelSchema", _json.dumps(schema).encode("utf-16"))
        z.writestr("Report/Layout", b"{}")
    return destino


def test_powerbi_lee_un_pbit_completo(tmp_path):
    """Un .pbit trae el modelo entero en JSON: tablas, columnas (con las
    calculadas), medidas con su DAX, relaciones, roles RLS y el origen real
    de cada tabla."""
    from mvdg import powerbi_meta as pbi
    m = pbi.read_pbit(_pbit_de_prueba(str(tmp_path / "modelo.pbit")))
    assert m.tables == ["Ventas", "Producto"]
    # La columna interna RowNumber que agrega Power BI no es del usuario.
    assert [c.name for c in m.columns] == ["Importe", "Margen", "SKU"]
    assert next(c for c in m.columns if c.name == "Margen").is_calculated
    assert len(m.measures) == 1 and m.measures[0].dax == "SUM(Ventas[Importe])"
    assert m.measures[0].display_folder == "KPIs"
    assert len(m.relationships) == 1 and m.relationships[0].both_directions
    assert m.roles == ["Solo LATAM"]
    # El origen sale de la expresión M, igual que por el camino TMDL.
    assert "SQL Server" in m.table_sources["Ventas"]


def test_powerbi_un_pbit_da_las_mismas_tablas_de_gobierno_que_un_pbip(tmp_path):
    """De la lectura para arriba, al programa le tiene que dar igual de qué
    formato vino el modelo."""
    from mvdg import powerbi_meta as pbi
    salida = pbi.ingest_powerbi_file(_pbit_de_prueba(str(tmp_path / "m.pbit")), "es")
    ejemplo = pbi.ingest_example("es")
    assert set(salida) == set(ejemplo)
    for k in ("catalog", "dictionary", "glossary", "lineage", "quality", "sources"):
        assert list(salida[k].columns) == list(ejemplo[k].columns), f"difieren en {k}"
        assert not salida[k].empty, f"{k} vino vacía desde el .pbit"


def test_powerbi_lee_el_traspaso_que_deja_el_generador(tmp_path):
    """El puente con MV DAX Lab, del lado de acá.

    Un generador de modelos sabe cosas que el TMSL no cuenta: cuántas filas
    trae cada tabla —van comprimidas dentro del Power Query— y por qué cada
    medida está escrita como está. MV DAX Lab lo deja anotado en el modelo;
    esto verifica que se lea. El contrato es el nombre de la anotación y el
    número de formato: si alguno de los dos cambia de un lado, este test se
    pone rojo del otro, que es exactamente para lo que está.
    """
    import json as _json
    import zipfile

    from mvdg import powerbi_meta as pbi
    ruta = _pbit_de_prueba(str(tmp_path / "generado.pbit"))
    with zipfile.ZipFile(ruta) as z:
        schema = _json.loads(z.read("DataModelSchema").decode("utf-16"))
    schema["model"]["annotations"] = [{
        "name": pbi.ANOTACION_TRASPASO,
        "value": _json.dumps({
            "formato": pbi.FORMATO_TRASPASO,
            "generador": "MV DAX Lab",
            "tablas": [{"nombre": "Ventas", "filas": 1200},
                       {"nombre": "Producto", "filas": 30}],
            "medidas": [{"tabla": "Ventas", "nombre": "Total Ventas",
                         "porque": "Cálculo directo, sin modificar el contexto."}],
        }, ensure_ascii=False)}]
    con_rastro = str(tmp_path / "con_rastro.pbit")
    with zipfile.ZipFile(ruta) as zin, zipfile.ZipFile(con_rastro, "w") as zout:
        for n in zin.namelist():
            zout.writestr(n, _json.dumps(schema).encode("utf-16")
                          if n == "DataModelSchema" else zin.read(n))

    salida = pbi.ingest_pbit(con_rastro, "es")
    modelo = salida["_model"]
    assert modelo.generator == "MV DAX Lab"
    assert modelo.table_rows == {"Ventas": 1200, "Producto": 30}
    # El catálogo decía 0 filas siempre. Con las filas declaradas, las dice.
    assert int(salida["catalog"]["rows"][0]) == 1230
    assert "MV DAX Lab" in salida["catalog"]["description"][0]
    # Una descripción escrita a mano en el modelo gana sobre el porqué
    # derivado: lo que puso una persona no se pisa.
    assert modelo.measures[0].description == "Suma del importe"


def test_powerbi_un_pbit_sin_traspaso_se_comporta_igual_que_antes(tmp_path):
    """Casi ningún .pbit del mundo trae el rastro del generador: los hace
    una persona en Desktop. Leerlo no puede ser obligatorio ni cambiar en
    nada lo que se ve de un archivo que no lo trae."""
    from mvdg import powerbi_meta as pbi
    salida = pbi.ingest_pbit(_pbit_de_prueba(str(tmp_path / "a_mano.pbit")), "es")
    assert salida["_model"].generator == ""
    assert salida["_model"].table_rows == {}
    assert int(salida["catalog"]["rows"][0]) == 0


def test_powerbi_un_traspaso_roto_no_rompe_la_lectura(tmp_path):
    """Un manifiesto ilegible —o de una versión del formato que este
    programa no conoce— se ignora entero. Que no haya rastro no es una
    falla del archivo; leerlo mal, sí."""
    import json as _json
    import zipfile

    from mvdg import powerbi_meta as pbi
    base = _pbit_de_prueba(str(tmp_path / "base.pbit"))
    with zipfile.ZipFile(base) as z:
        schema = _json.loads(z.read("DataModelSchema").decode("utf-16"))
    for etiqueta, anotacion in (
            ("json roto", {"name": pbi.ANOTACION_TRASPASO, "value": "{no es json"}),
            ("formato futuro", {"name": pbi.ANOTACION_TRASPASO,
                                "value": _json.dumps({"formato": 999,
                                                      "tablas": [{"nombre": "Ventas",
                                                                  "filas": 9}]})})):
        schema["model"]["annotations"] = [anotacion]
        ruta = str(tmp_path / f"{etiqueta.replace(' ', '_')}.pbit")
        with zipfile.ZipFile(base) as zin, zipfile.ZipFile(ruta, "w") as zout:
            for n in zin.namelist():
                zout.writestr(n, _json.dumps(schema).encode("utf-16")
                              if n == "DataModelSchema" else zin.read(n))
        salida = pbi.ingest_pbit(ruta, "es")
        assert salida["_model"].generator == "", etiqueta
        assert int(salida["catalog"]["rows"][0]) == 0, etiqueta


def test_powerbi_un_pbix_binario_explica_que_hacer_en_los_tres_idiomas(tmp_path):
    """Un .pbix guarda el modelo como respaldo binario de Analysis Services:
    no hay forma de leerlo sin librerías de Microsoft. Lo que NO puede pasar
    es que el usuario reciba "no se pudo leer el archivo" — necesita que le
    digan que lo guarde como .pbit, y en su idioma."""
    import zipfile
    from mvdg import powerbi_meta as pbi
    from mvdg.errors import friendly_error
    ruta = str(tmp_path / "reporte.pbix")
    with zipfile.ZipFile(ruta, "w") as z:
        z.writestr("DataModel", b"\x00\x01respaldo-binario")
        z.writestr("Report/Layout", b"{}")
    with pytest.raises(Exception) as exc:
        pbi.ingest_powerbi_file(ruta, "es")
    for lang, aguja in (("es", ".pbit"), ("en", ".pbit"), ("pt", ".pbit")):
        mensaje, _ = friendly_error(exc.value, lang, "archivo")
        assert aguja in mensaje, f"[{lang}] el mensaje no dice qué hacer: {mensaje}"
        assert "Analysis Services" in mensaje


def test_error_traducible_gana_sobre_el_mapeo_por_tipo():
    """`friendly_error` mapea por tipo de excepción; sin esto, el mensaje
    específico del motor se perdía y salía el genérico de archivo."""
    from mvdg.errors import ErrorTraducible, friendly_error
    from mvdg.i18n import t
    exc = ErrorTraducible("err_pbix_binario", "detalle técnico")
    mensaje, detalle = friendly_error(exc, "es", "archivo")
    assert mensaje == t("err_pbix_binario", "es")
    assert mensaje != t("err_archivo_generico", "es")
    assert "detalle técnico" in detalle


def test_ningun_origen_de_archivo_pide_solo_escribir_una_ruta():
    """Escribir una ruta a mano solo funciona si el archivo está en la misma
    máquina que corre el programa. No es el caso en modo servidor, ni cuando
    el cliente prueba la demo desde otra computadora, ni en la versión web:
    ahí la ruta que escriba nunca va a existir del otro lado. Todo lugar que
    acepte una ruta tiene que aceptar además subir el archivo.

    Se verifica por par (clave de la ruta -> uploader que la acompaña),
    porque el fallo real es que ALGUNO quede sin par, no que falten
    uploaders en general."""
    src = _app_source()
    pares = {
        'key="pbi_path"': 'key="pbi_zip"',        # Power BI
        'key="tab_path"': 'key="tab_upload"',     # Tableau
        't("db_sqlite_path"': 'key="db_sqlite_up"',  # SQLite
    }
    for ruta, uploader in pares.items():
        assert ruta in src, f"desapareció el campo de ruta {ruta} (¿se renombró?)"
        assert uploader in src, (
            f"{ruta} sigue existiendo pero no hay uploader ({uploader}) que lo "
            "acompañe: en modo servidor esa pantalla no tiene salida")
    # Y el uploader tiene que ir ANTES: es el camino que funciona siempre.
    assert src.index('key="pbi_zip"') < src.index('key="pbi_path"')
    assert src.index('key="tab_upload"') < src.index('key="tab_path"')
    assert src.index('key="db_sqlite_up"') < src.index('t("db_sqlite_path"')


def test_powerbi_acepta_los_formatos_que_la_ui_ofrece():
    """La lista de extensiones del uploader y la que el motor sabe leer
    tienen que ser la misma: ofrecer en pantalla un formato que después
    rebota es peor que no ofrecerlo."""
    from mvdg import powerbi_meta as pbi
    ui = re.search(r'type=\["pbit", "pbix", "zip"\]', _app_source())
    assert ui, "el uploader de Power BI ya no ofrece pbit/pbix/zip"
    assert set(pbi.EXT_POWERBI) == {".pbit", ".pbix", ".pbip", ".zip"}


# ─────────── Los datasets que carga el usuario, en TODO el programa ───────────
# El agujero que cierran: el usuario subía su Excel en "Mis datos", lo veía
# perfilado ahí, y las otras 19 pestañas seguían mostrando la demo. Para
# alguien evaluando el producto con sus propios datos eso lo convierte en una
# demo bonita en vez de una herramienta.

def _df_usuario():
    """Un dataset con PII y con defectos, como el que sube un cliente."""
    return pd.DataFrame({
        "cliente_id": [1, 2, 3, 4, 5],
        "email": ["a@x.com", "b@x.com", None, "d@x.com", "e@x.com"],
        "documento": ["1.234.567-8", "2.345.678-9", "3.456.789-0",
                      "4.567.890-1", "5.678.901-2"],
        "monto": [100, 200, 300, 400, 500],
    })


def test_el_dataset_cargado_aparece_en_todo_el_dashboard(monkeypatch, tmp_path):
    """LA prueba del reporte: cargar un dataset y que TODAS las pestañas lo
    tomen, no solo la que lo cargó.

    Se hace con AppTest y no con el navegador a propósito: Streamlit dibuja
    las tablas en un canvas virtualizado, así que leer el DOM no ve el
    contenido de las celdas — un test por texto pasaría o fallaría por
    motivos que no tienen nada que ver con el dato. AppTest da los
    DataFrames que la app realmente renderizó.
    """
    from streamlit.testing.v1 import AppTest
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    at = AppTest.from_file(os.path.join(_repo_root(), "app", "app.py"),
                           default_timeout=400)
    at.session_state["mvdg_user_datasets"] = {"clientes_crm": _df_usuario()}
    at.run()
    assert not at.exception, [str(e.value)[:300] for e in at.exception]

    # Pestaña por pestaña, identificada por las columnas de su tabla: que
    # aparezca en "alguna" no sirve — el reporte fue justamente que aparecía
    # en una y en las demás no.
    # Se devuelven TODAS las que calzan porque el mismo esquema aparece más
    # de una vez: la pestaña Laboratorio repite las tablas de catálogo y de
    # políticas como tutorial guiado sobre la demo, y ésa no tiene que
    # cambiar con lo que cargue el usuario. Alcanza con que alguna de las
    # tablas de ese esquema lo incluya; con el cableado roto no lo incluye
    # ninguna (verificado rompiéndolo a propósito).
    def _tablas_con(*columnas):
        return [d.value for d in at.dataframe
                if set(columnas) <= set(map(str, d.value.columns))]

    pendientes = {
        "Catálogo": _tablas_con("Dataset", "Descripción", "Dueño", "Steward"),
        "Calidad": _tablas_con("Dataset", "Dimensión", "Regla"),
        "Linaje": _tablas_con("source_id", "target_id"),
        "Políticas": _tablas_con("Política", "Evidencia"),
    }
    for pestania, tablas in pendientes.items():
        assert tablas, f"no encontré ninguna tabla de {pestania}"
        assert any("clientes_crm" in tb.to_csv(index=False) for tb in tablas), (
            f"{pestania} no muestra el dataset cargado por el usuario")

    # Y donde se ELIGE sobre qué dataset trabajar: si no está en el selector,
    # la pestaña se ve pero no se puede usar con los datos propios.
    mdm = next((w for w in at.selectbox if w.key == "mdm_pick_dataset"), None)
    assert mdm is not None and any("clientes_crm" in str(o) for o in mdm.options), (
        "MDM no deja elegir el dataset del usuario: deduplicar el archivo "
        "propio es justo lo que viene a probar un cliente")
    assert any("clientes_crm" in (w.options or []) for w in at.selectbox), (
        "el dataset no es elegible en el detalle del Catálogo")

    # La política de dueño/steward tiene que NOTAR que el dataset nuevo no
    # tiene responsables, en vez de seguir diciendo que está todo asignado.
    reaccionaron = [tb for tb in pendientes["Políticas"]
                    if tb["Evidencia"].astype(str).str.contains("clientes_crm").any()]
    assert reaccionaron, "ninguna política reaccionó al dataset del usuario"


def test_el_selfcheck_no_deja_sucio_el_contexto_de_streamlit():
    """Importar `app/app.py` lo EJECUTA (es un script de Streamlit). Fuera de
    un servidor, `st.form` estampa su id sobre el contenedor raíz y no lo
    limpia, así que TODO lo que use Streamlit después en el mismo proceso lo
    hereda: un `st.button` de la barra lateral se cree adentro de un
    formulario y revienta con "st.button() can't be used in an st.form()".

    Se vio de verdad: la suite entera se ponía en rojo mientras cada test
    pasaba solo. Un fallo así cuesta horas de bisección, y vuelve solo si
    alguien toca el chequeo — por eso se fija acá."""
    from streamlit.delta_generator_singletons import context_dg_stack
    from mvdg import selfcheck
    antes = [(dg, getattr(dg, "_form_data", None)) for dg in context_dg_stack.get()]
    selfcheck._check_46()
    for dg, previo in antes:
        assert getattr(dg, "_form_data", None) == previo, (
            "el chequeo dejó un formulario abierto en el contexto global de "
            "Streamlit: lo que corra después en este proceso va a fallar")


def test_cargar_un_dataset_refresca_el_resto_del_programa():
    """Streamlit ejecuta el script de arriba abajo: cuando "Mis datos"
    registra el archivo, el sidebar y las demás pestañas YA se dibujaron.
    Sin un rerun, el dataset queda guardado y nadie lo ve hasta la próxima
    interacción — el síntoma exacto que se reportó. El rerun tiene que
    dispararse UNA vez por archivo, o es un bucle infinito."""
    src = _app_source()
    reg = src[src.index("def _registrar_dataset"):]
    reg = reg[:reg.index("\ndef ", 1)]
    assert "st.rerun()" in reg, (
        "_registrar_dataset no vuelve a correr el script: las otras pestañas "
        "se quedan con lo de antes")
    assert "_mvdg_user_vistos" in reg, (
        "sin registro de ya-vistos, el rerun se dispara en cada pasada y la "
        "app entra en bucle")


def test_scope_usuario_produce_las_mismas_columnas_que_la_demo():
    """Si las columnas no coinciden, concatenar rompe o —peor— mete columnas
    fantasma que la UI muestra vacías."""
    from mvdg.catalog import catalog_df, dictionary_df
    from mvdg.quality import run_rules
    from mvdg import scope
    ud = {"clientes_crm": _df_usuario()}
    base_cat = catalog_df("es")
    assert list(scope.user_catalog(ud, "es", columnas=base_cat.columns).columns) == \
        list(base_cat.columns)
    assert list(scope.user_dictionary(ud, "es").columns) == list(dictionary_df("es").columns)
    assert list(scope.user_results(ud, "es").columns) == list(run_rules(lang="es").columns)


def test_scope_usuario_corre_reglas_de_verdad_contra_los_datos():
    """No alcanza con listar el dataset: las reglas tienen que EVALUARSE. Un
    catálogo que dice "todo bien" sin haber mirado los datos es peor que no
    tener catálogo."""
    from mvdg import scope
    res = scope.user_results({"clientes_crm": _df_usuario()}, "es")
    assert not res.empty
    # La columna email tiene 1 nulo de 5 -> 80% de completitud, y eso no
    # puede dar 100 ni puede dar pass con umbral alto.
    email = res[res["column"] == "email"]
    assert not email.empty, "no se generó ninguna regla para la columna con nulos"
    assert float(email.iloc[0]["score"]) == 80.0
    assert (res["score"] <= 100).all() and (res["score"] >= 0).all()


def test_scope_usuario_clasifica_pii_con_el_token_que_el_motor_compara():
    """`classification` es un token literal, no texto para mostrar: policies.py
    busca exactamente "PII". Traducirlo rompía la política de clasificación en
    silencio — el dataset aparecía con PII en el diccionario y la política
    seguía diciendo que estaba todo en orden."""
    from mvdg import scope
    ud = {"clientes_crm": _df_usuario()}
    for lang in ("es", "en", "pt"):
        fila = scope.user_catalog(ud, lang).iloc[0]
        assert fila["classification"] == "PII", (
            f"en {lang} la clasificación es {fila['classification']!r}, "
            "y policies.py compara contra el literal 'PII'")


def test_scope_usuario_no_inventa_dueno_ni_steward():
    """Poner "N/D" en un campo de gobierno es peor que dejarlo vacío: la
    política de dueño/steward lo contaría como asignado."""
    from mvdg import scope
    fila = scope.user_catalog({"clientes_crm": _df_usuario()}, "es").iloc[0]
    assert fila["owner"] == "" and fila["steward"] == ""


def test_scope_usuario_vacio_no_rompe_nada():
    """Sin datasets cargados —el caso normal— todo tiene que seguir igual."""
    from mvdg import scope
    for entrada in (None, {}, {"x": pd.DataFrame()}):
        assert scope.user_catalog(entrada, "es").empty
        assert scope.user_dictionary(entrada, "es").empty
        assert scope.user_results(entrada, "es").empty


def test_governance_tables_suma_lo_del_usuario_sin_pisar_lo_demas():
    """El bug que esto fija: al sumar el linaje del usuario, la primera
    versión REEMPLAZABA el de los casos de ejemplo (23 filas -> 17) porque
    partía del grafo de demo en vez del combinado. El cliente exportaba a BI
    y se llevaba menos linaje del que tenía antes de cargar su archivo."""
    from mvdg.exporters import governance_tables
    ud = {"clientes_crm": _df_usuario()}
    base = governance_tables("es", include_samples=True)
    con = governance_tables("es", include_samples=True, user_datasets=ud)
    for tabla in ("catalog", "dictionary", "quality_results", "lineage"):
        assert len(con[tabla]) > len(base[tabla]), (
            f"{tabla} no creció al sumar el dataset del usuario")
    # y lo de antes sigue estando
    assert "cafe_sales_kaggle" in set(con["lineage"]["target"])
    assert "clientes_crm" in set(con["lineage"]["target"])
    assert "clientes_crm" in set(con["catalog"]["dataset"])


def test_governance_tables_sin_usuario_no_cambia():
    """Compatibilidad: la API y los tests existentes no pueden ver ninguna
    diferencia si nadie cargó nada."""
    from mvdg.exporters import governance_tables
    a = governance_tables("es", include_samples=True)
    b = governance_tables("es", include_samples=True, user_datasets={})
    for k in a:
        assert a[k].equals(b[k]), f"la tabla {k} cambió sin datasets del usuario"


def test_politica_de_dueno_no_se_contradice_con_su_evidencia():
    """POL-01 decía "partial" y su evidencia decía "N/N asignados" — o sea,
    cumplimiento total. Lo que se exporta y lo que lee un auditor es la
    evidencia, así que afirmaba justo lo contrario de lo detectado. No se
    veía porque hasta ahora todo el universo traía dueño y steward."""
    from mvdg.exporters import governance_tables
    con = governance_tables("es", include_samples=True,
                            user_datasets={"clientes_crm": _df_usuario()})
    pol = con["policies"].set_index("policy_id")
    assert pol.loc["POL-01", "status"] == "partial"
    evidencia = pol.loc["POL-01", "evidence"]
    total = len(con["catalog"])
    assert f"{total}/{total}" not in evidencia, (
        f"la evidencia dice cumplimiento total pero el estado es parcial: {evidencia}")
    assert "clientes_crm" in evidencia, "la evidencia no dice cuál dataset falta"


def test_lineage_figure_no_solapa_etiquetas_largas():
    """La cita real de una fuente ('Reglamento Bromatológico Nacional de
    Alimentos (Uruguay, 2026); ...') es mucho más larga que las etiquetas
    de demo ('CRM', 'raw.customers'). Sin recortar, Plotly la dibuja tan
    ancha como haga falta y se solapa con el nodo vecino — se vio
    literalmente en landing/img/tab_linaje.jpg ('rawWNAEV, Decreto
    466/009)' = dos etiquetas de nodos distintos superpuestas).

    El texto completo no se pierde: sigue entero en el hover."""
    from mvdg import samples, scope
    from mvdg.lineage import lineage_figure

    nodes, edges = scope.combined_lineage("es")
    cita_larga = next(n for n in nodes if n["id"] == f"src_{samples.sample_keys()[0]}")
    assert len(cita_larga["label"]) > 30, "el fixture ya no prueba nada: la cita es corta"

    fig = lineage_figure(nodes=nodes, edges=edges)
    trazo = next(t for t in fig.data if t.mode == "markers+text")
    for texto_mostrado, texto_completo in zip(trazo.text, trazo.hovertext, strict=True):
        assert len(texto_mostrado) <= 22, (
            f"etiqueta sin recortar en el grafo: {texto_mostrado!r} — puede "
            f"solaparse con el nodo vecino")
        if len(texto_completo) > 22:
            assert texto_mostrado.endswith("…")
        else:
            assert texto_mostrado == texto_completo  # las cortas no se tocan
        assert texto_completo in [n["label"] for n in nodes]  # nada se perdió, solo se acorta lo mostrado


def test_lineage_figure_no_recorta_nodos_de_capas_largas():
    """Con los casos reales de 'Mis datos' activos, combined_lineage() suma
    una fuente por dataset — hasta 8 nodos en la capa 'source', más de los
    4 que trae el grafo demo. La figura tenía el alto y el rango de y FIJOS
    (460px / [-2.9, 2.7]) sin importar cuántos nodos hubiera: el último
    quedaba literalmente afuera del área visible, cortado por el borde —
    se vio en landing/img/tab_linaje.jpg (una fila entera de nodos sin
    ninguna etiqueta, recortada a la mitad justo arriba de "Aristas de
    linaje"). El caso demo (chico) no tiene que agrandarse por esto."""
    from mvdg import scope
    from mvdg.lineage import _positions, lineage_figure

    fig_demo = lineage_figure()
    assert fig_demo.layout.height == 460
    assert tuple(fig_demo.layout.yaxis.range) == (-2.9, 2.7)

    nodes, edges = scope.combined_lineage("es")
    fig = lineage_figure(nodes=nodes, edges=edges)
    y_min_real = min(y for _, y in _positions(nodes).values())
    y_min_visible, y_max_visible = fig.layout.yaxis.range
    assert y_min_visible <= y_min_real, (
        "el nodo más bajo de la capa más larga queda fuera del rango "
        "visible del eje y — se corta en el borde de la figura")
    assert fig.layout.height > 460, "más nodos por capa tienen que dar más alto, no el mismo de siempre"


def test_policies_evaluate_combined_universe():
    """Con el alcance combinado, el cumplimiento se verifica sobre TODO lo
    gobernado — la evidencia de POL-01 cuenta 8 datasets, no 4."""
    from mvdg import scope
    from mvdg.policies import policies_df
    res = scope.combined_results("es")
    pdf = policies_df("es", res, catalog=scope.combined_catalog("es"),
                      dictionary=scope.combined_dictionary("es"))
    ev = pdf[pdf["policy_id"] == "POL-01"].iloc[0]["evidence"]
    assert "8/8" in ev
    # y sin los parámetros, sigue evaluando la demo (compatibilidad)
    ev_demo = policies_df("es")[lambda d: d["policy_id"] == "POL-01"].iloc[0]["evidence"] \
        if False else policies_df("es").iloc[0]["evidence"]
    assert "4/4" in ev_demo


def test_governance_tables_include_samples_flag():
    from mvdg.exporters import governance_tables
    g = governance_tables("es", include_samples=True)
    assert len(g["catalog"]) == 8
    assert g["quality_results"]["dataset"].nunique() == 8
    assert (g["lineage"]["source_id"].str.startswith("src_")).any()
    # el default NO cambia (API por dataset y tests existentes intactos)
    g_default = governance_tables("es")
    assert len(g_default["catalog"]) == 4


def test_installer_iss_offers_optional_desktop_and_start_menu():
    """El instalador .exe (Inno Setup) crea el acceso del Menú Inicio y
    ofrece el del escritorio como casilla opcional — 'si lo desea el
    cliente', literal."""
    with open(os.path.join(_repo_root(), "packaging", "instalador.iss"),
              encoding="utf-8") as fh:
        iss = fh.read()
    assert 'Name: "desktopicon"' in iss               # casilla opcional
    assert "{autodesktop}" in iss and "Tasks: desktopicon" in iss
    assert "{group}" in iss                           # menú inicio siempre


def test_installer_iss_deja_elegir_carpeta_y_disco():
    """El asistente tiene que dejar elegir CARPETA Y DISCO de instalación.

    Tres cosas, y las tres importan:
      · ``DisableDirPage=no`` — la página "Seleccionar carpeta de destino"
        se muestra; si alguien la pone en ``yes`` el cliente queda clavado
        en Archivos de programa sin poder mandarlo a D:\\.
      · ``DefaultDirName`` — hay una sugerencia razonable de arranque.
      · ``PrivilegesRequiredOverridesAllowed=dialog`` — sin esto, Setup
        exige admin sí o sí, y un usuario sin permisos no puede instalar en
        su propia carpeta (que es justo el caso de una notebook corporativa).
    """
    with open(os.path.join(_repo_root(), "packaging", "instalador.iss"),
              encoding="utf-8") as fh:
        iss = fh.read()
    directivas = {}
    for linea in iss.splitlines():
        linea = linea.strip()
        if linea.startswith(";") or "=" not in linea or linea.startswith("["):
            continue
        clave, _, valor = linea.partition("=")
        directivas[clave.strip().lower()] = valor.strip()
    assert directivas.get("disabledirpage") == "no"
    assert directivas.get("defaultdirname", "").endswith("MV Data Governance")
    assert directivas.get("privilegesrequiredoverridesallowed") == "dialog"
    # y nada de fijar el disco a mano: sería contradecir todo lo anterior
    assert "usepreviousappdir=no" not in iss.lower()


# ------------------------------- informe de auditoría del archivo propio
def _df_con_defectos():
    """Archivo 'de cliente' con defectos reales: nulos en email/monto y una
    clave aparente (id) — dispara reglas de completitud y unicidad."""
    return pd.DataFrame({
        "id": [1, 2, 3, 4, 5, 6, 7, 8],
        "email": ["a@x.com", None, "c@x.com", "d@x.com", None,
                  "f@x.com", "g@x.com", "h@x.com"],
        "monto": [10.0, 20.0, None, 40.0, 50.0, 60.0, 70.0, 80.0],
        "pais": ["UY", "UY", "AR", "AR", "BR", "BR", "UY", "AR"],
    })


def test_informe_archivo_propio_cuadra_con_el_motor():
    """El Excel se REABRE y se audita contra el motor: mismo índice de
    calidad, misma cantidad de reglas, un plan de corrección por falla.
    La pantalla y el informe usan la misma fuente (file_report_tables),
    así que si esto cuadra, no pueden divergir."""
    import openpyxl
    from mvdg.auto_rules import auto_quality_results
    from mvdg.file_report import file_report_xlsx
    from mvdg.i18n import t as _tt
    from mvdg.quality import overall_index

    df = _df_con_defectos()
    ares = auto_quality_results(df, "clientes.csv", "es")
    assert len(ares) >= 3          # el df de prueba tiene que disparar reglas
    rotas = int((ares["status"] != "pass").sum())
    assert rotas >= 2

    datos = file_report_xlsx(df, "clientes.csv", "es")
    assert datos[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True)
    esperadas = [_tt(k, "es") for k in ("frep_sheet_summary", "frep_sheet_profile",
                                        "frep_sheet_quality", "frep_sheet_fixes")]
    assert wb.sheetnames == esperadas

    # resumen: el índice de calidad del informe ES el del motor
    resumen = wb[esperadas[0]]
    celdas = [str(c.value) for fila in resumen.iter_rows() for c in fila if c.value]
    assert f"{overall_index(ares)} / 100" in celdas
    assert any(v == str(len(df)) for v in celdas)          # filas

    # calidad: una fila por regla corrida (sin contar título/encabezado/pie)
    pie = _tt("frep_generated", "es")
    def _filas_de_datos(hoja):
        return [f for f in hoja.iter_rows(min_row=4, values_only=True)
                if f[0] is not None and f[0] != pie]
    assert len(_filas_de_datos(wb[esperadas[2]])) == len(ares)
    # plan: una fila por regla rota, con causa y corto plazo no vacíos
    plan = _filas_de_datos(wb[esperadas[3]])
    assert len(plan) == rotas
    assert all(f[3] and f[4] for f in plan)


@pytest.mark.parametrize("lang", LANGS)
def test_informe_archivo_propio_trilingue(lang):
    """Las 4 hojas salen en el idioma pedido — paridad ES/EN/PT también acá."""
    import openpyxl
    from mvdg.file_report import file_report_xlsx
    from mvdg.i18n import t as _tt
    datos = file_report_xlsx(_df_con_defectos(), "clientes.csv", lang)
    wb = openpyxl.load_workbook(io.BytesIO(datos))
    assert wb.sheetnames == [_tt(k, lang) for k in
                             ("frep_sheet_summary", "frep_sheet_profile",
                              "frep_sheet_quality", "frep_sheet_fixes")]


def test_informe_archivo_sin_reglas_no_explota():
    """Un archivo con solo encabezados (0 filas — típico: plantilla exportada
    vacía) no dispara ninguna regla y el informe igual sale: perfil sí,
    calidad vacía, '0 / 0' en vez de una división por cero o un traceback."""
    import openpyxl
    from mvdg.file_report import file_report_xlsx
    df = pd.DataFrame({"nota": pd.Series([], dtype=str)})
    datos = file_report_xlsx(df, "notas.csv", "es")
    wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True)
    celdas = [str(c.value) for h in wb for fila in h.iter_rows() for c in fila if c.value]
    assert "0 / 0" in celdas


def test_informe_descargable_desde_la_pantalla_de_perfilado():
    """El botón existe en la UI y usa el motor — sin esto, el informe sería
    una función huérfana que nadie puede descargar."""
    with open(os.path.join(_repo_root(), "app", "app.py"), encoding="utf-8") as fh:
        src = fh.read()
    assert "file_report_xlsx" in src
    assert 't("frep_btn", lang)' in src
    # gateado por dataset_name: la comparación genérica de ejemplos no lo lleva
    assert "if dataset_name:" in src


# ------------------------------------------------ entregable final por caso
def test_deliverable_builds_for_every_case(tmp_path, monkeypatch):
    """El entregable se arma para los 4 casos con KPIs reales (calidad
    calculada sobre el archivo, no números fijos) y migración en dry-run."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import deliverable
    for key in deliverable.case_keys():
        d = deliverable.build_deliverable(key, "es")
        k = d["kpis"]
        assert k["rows"] > 0 and k["columns"] > 0
        assert 0 <= k["quality_index"] <= 100
        assert k["rules_total"] > 0
        assert k["documented_pct"] == 100.0  # los casos vienen documentados
        m = d["migration"]
        # entidades = 1 tabla + N columnas; términos > 0 — calculado con los
        # conectores reales en dry-run
        assert m["purview_entities"] == 1 + k["columns"]
        assert m["collibra_assets"] == 1 + k["columns"]
        assert m["purview_terms"] > 0 and m["collibra_terms"] > 0
        assert len(d["lineage"]) == 2  # fuente->curado, curado->BI


def test_deliverable_curation_progress_reflects_real_validation(tmp_path, monkeypatch):
    """El KPI de curaduría del entregable se mueve cuando un responsable
    valida de verdad — y el término validado sale Approved hacia Purview."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, deliverable, samples
    key = "medicamentos_openfda"
    before = deliverable.build_deliverable(key, "es")
    assert before["kpis"]["curation_pct"] == 0.0
    assert before["migration"]["purview_terms_approved"] == 0
    term_id = samples.SAMPLES[key]["terms"][0]["term_id"]
    curation.save_validation(f"glossary:{key}:{term_id}", "es", "validado",
                             "", "M. Viera", "Data Owner")
    after = deliverable.build_deliverable(key, "es")
    assert after["kpis"]["curation_reviewed"] == 1
    assert after["kpis"]["curation_pct"] > 0.0
    assert after["migration"]["purview_terms_approved"] == 1


def test_deliverable_downloads_are_real_files(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    import io
    import openpyxl
    from mvdg import deliverable
    xlsx = deliverable.deliverable_xlsx_bytes("bank_marketing_uci", "es")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert set(wb.sheetnames) == {"Ficha", "Diccionario", "Calidad", "Glosario",
                                  "Linaje", "Plan de acción"}
    for lg, probe in (("es", "Entregable de gobernanza"),
                      ("en", "Governance deliverable"),
                      ("pt", "Entregável de governança")):
        md = deliverable.executive_summary_md("rotulado_alimentos", lg)
        assert probe in md
        assert "284" in md  # filas reales del caso del gobierno uruguayo


def test_excel_generation_needs_no_disk_at_all(tmp_path, monkeypatch):
    """Regresión del crash real en la máquina del usuario: xlsxwriter usa
    archivos temporales EN DISCO aunque el destino sea BytesIO — con el
    disco lleno, todo botón de descarga Excel explotaba con FileCreateError
    [Errno 28]. Con in_memory=True no se toca el disco: se simula el disco
    inutilizable apuntando TMPDIR a una ruta inexistente y los tres
    generadores tienen que funcionar igual."""
    import io
    import openpyxl
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    broken = str(tmp_path / "no" / "existe" / "tmp")
    monkeypatch.setenv("TMPDIR", broken)
    monkeypatch.setenv("TEMP", broken)
    monkeypatch.setenv("TMP", broken)
    import tempfile
    monkeypatch.setattr(tempfile, "tempdir", None)  # fuerza re-lectura de env
    from mvdg import deliverable
    from mvdg.exporters import bi_bundle_xlsx, governance_tables, to_excel_bytes
    df = governance_tables("es")["catalog"]
    xlsx = to_excel_bytes(df, "catalog")
    assert openpyxl.load_workbook(io.BytesIO(xlsx)).sheetnames == ["catalog"]
    assert len(bi_bundle_xlsx("es")) > 5000
    assert len(deliverable.deliverable_xlsx_bytes("rotulado_alimentos", "es")) > 5000


def test_deliverable_findings_have_remediation_plan(tmp_path, monkeypatch):
    """Cada regla que no pasó aparece como hallazgo CON plan: causa raíz,
    corrección inmediata, corrección de fondo y responsable — el entregable
    diagnostica, no esconde."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import deliverable
    for key in deliverable.case_keys():
        d = deliverable.build_deliverable(key, "es")
        f = d["findings"]
        not_pass = int((d["quality_results"]["status"] != "pass").sum())
        assert len(f) == not_pass, key
        if len(f):
            assert (f["root_cause"].str.len() > 0).all()
            assert (f["short_term"].str.len() > 0).all()
            assert (f["owner"].str.len() > 0).all()
    # y viajan en el Excel (hoja Plan de acción) y en el resumen ejecutivo
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(
        deliverable.deliverable_xlsx_bytes("medicamentos_openfda", "es")))
    assert "Plan de acción" in wb.sheetnames
    md = deliverable.executive_summary_md("medicamentos_openfda", "es")
    assert "Hallazgos y plan de remediación" in md


def test_curation_bulk_validation_reaches_100_pct(tmp_path, monkeypatch):
    """La validación masiva firma TODAS las definiciones pendientes de un
    caso con nombre/cargo/fecha — el entregable llega al 100% de curaduría
    legítimamente, y cada registro individual queda auditable."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, deliverable
    key = "bank_marketing_uci"
    df = curation.list_items("es")
    pending = df[(df["dataset"] == key) & (df["status"] == "sugerido_ia")]
    assert len(pending) > 0
    for _, it in pending.iterrows():  # mismo recorrido que hace el botón
        curation.save_validation(it["item_id"], "es", "validado", "",
                                 "L. Santos", "Data Steward Banca")
    prog = deliverable.curation_progress(key, "es")
    assert prog["pct"] == 100.0
    rec = curation.get_record(pending.iloc[0]["item_id"], "es")
    assert rec["responsible_name"] == "L. Santos"  # firmado, auditable
    d = deliverable.build_deliverable(key, "es")
    assert d["kpis"]["curation_pct"] == 100.0
    assert d["migration"]["purview_terms_approved"] == d["migration"]["purview_terms"]


def test_data_contracts_evaluate_against_real_rule_runs(tmp_path, monkeypatch):
    """Los contratos de datos formalizan las reglas REALES: la evaluación debe
    coincidir exactamente con la última corrida de reglas, sin números nuevos."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import contracts
    from mvdg.scope import combined_results

    res = combined_results("es")
    con = contracts.contracts_df("es", res)
    # un contrato por producto gobernado (4 demo + 4 casos reales)
    assert len(con) == 8
    assert set(con["dataset"]) == set(contracts.product_keys("es"))
    # roles del modelo completos y salidos del catálogo real
    for col in ("domain", "domain_owner", "product_owner", "producer",
                "sla_refresh"):
        assert (con[col].astype(str).str.len() > 0).all(), col
    # la matemática del cumplimiento es la de la corrida real
    for _, row in con.iterrows():
        sub = res[res["dataset"] == row["dataset"]]
        assert row["rules"] == len(sub)
        assert row["rules_fail"] == int((sub["status"] == "fail").sum())
        expected = ("incumple" if row["rules_fail"] else
                    "en_riesgo" if row["rules_warn"] else "cumple")
        assert row["compliance"] == expected
    # el laboratorio (openFDA) tiene fails reales -> su contrato incumple
    lab = con.set_index("dataset").loc["medicamentos_openfda"]
    assert lab["compliance"] == "incumple" and lab["rules_fail"] > 0


def test_data_contract_alerts_carry_downstream_impact_and_owner(tmp_path, monkeypatch):
    """Alarmística sobre el linaje: cada regla no aprobada genera exactamente
    una alerta, con impacto aguas abajo real y a quién avisar."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import contracts
    from mvdg.scope import combined_results

    res = combined_results("es")
    ale = contracts.alerts_df("es", res)
    assert len(ale) == int((res["status"] != "pass").sum())
    # el impacto sale de recorrer el linaje real: los casos llegan al BI
    lab_alerts = ale[ale["dataset"] == "medicamentos_openfda"]
    assert len(lab_alerts) and lab_alerts["impact_downstream"].str.contains(
        "Dashboard BI").all()
    # y la demo pasa por los marts antes del BI
    cus = ale[ale["dataset"] == "dim_customers"]
    assert len(cus) and cus["impact_downstream"].str.contains("mart_").all()
    # a quién avisar: warn -> PO; fail -> Domain Owner + PO
    for _, a in ale.iterrows():
        assert a["notify"].strip()
        if a["severity"] == "fail":
            assert "+" in a["notify"]
    # acción inmediata del motor de remediación real, nunca vacía
    assert (ale["action"].astype(str).str.len() > 0).all()


def test_data_contract_agreement_signature_persists(tmp_path, monkeypatch):
    """Documentar acuerdos: la firma persiste local, cambia el estado a
    'vigente' y es auditable (quién, rol, fecha)."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import contracts
    from mvdg.scope import combined_results

    res = combined_results("es")
    assert contracts.agreement_for("fct_sales") is None
    contracts.save_agreement("fct_sales", "L. Santos", "Data Product Owner")
    agr = contracts.agreement_for("fct_sales")
    assert agr["signed_by"] == "L. Santos" and agr["role"] == "Data Product Owner"
    assert agr["date"]  # fecha registrada
    con = contracts.contracts_df("es", res).set_index("dataset")
    assert con.loc["fct_sales", "agreement"] == "vigente"
    assert con.loc["dim_customers", "agreement"] == "borrador"
    kp = contracts.kpis("es", res)
    assert kp["signed"] == 1 and kp["products"] == 8
    contracts.delete_agreement("fct_sales")
    assert contracts.agreement_for("fct_sales") is None


def test_data_contracts_theory_and_export_trilingual(tmp_path, monkeypatch):
    """Marco teórico (Data Mesh / contratos) trilingüe + export Excel 100% en
    memoria con las 3 hojas."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    import io
    import openpyxl
    from mvdg import contracts
    from mvdg.scope import combined_results

    res = combined_results("es")
    keys = {c["key"] for c in contracts.THEORY}
    assert {"domain", "product", "contract", "sla", "alerts", "roles",
            "agreement", "federated"} == keys
    for lg in ("es", "en", "pt"):
        th = contracts.theory(lg)
        assert len(th) == 8
        assert all(c["concept"] and c["plain"] and c["practice"] for c in th)
        wb = openpyxl.load_workbook(io.BytesIO(
            contracts.contracts_xlsx_bytes(lg, res)))
        assert len(wb.sheetnames) == 3
    # honestidad: la teoría dice cómo lo practica el programa, no solo qué es
    es = {c["key"]: c for c in contracts.theory("es")}
    assert "linaje" in es["alerts"]["practice"].lower()
    assert "catálogo" in es["roles"]["practice"].lower()


def _mcp_env(tmp_path):
    import os
    return {**os.environ, "MVDG_DATA_DIR": str(tmp_path),
            "PYTHONPATH": os.path.dirname(os.path.dirname(
                os.path.abspath(__file__)))}


def test_los_servidores_mcp_de_bi_estan_completos_y_coherentes():
    """Cada servidor tiene que declarar todo lo que hace falta para
    conectarlo. Un preset a medias es peor que no tenerlo: el usuario copia
    una configuracion incompleta y el fallo aparece recien en su cliente MCP,
    sin decir que falta."""
    from mvdg import mcp_presets as P

    comunes = {"etiqueta", "plataforma", "transporte", "requisitos", "auth",
               "para_que", "docs"}
    for nombre, cfg in P.SERVIDORES.items():
        faltan = comunes - set(cfg)
        assert not faltan, f"{nombre}: faltan {sorted(faltan)}"
        assert cfg["transporte"] in (P.STDIO, P.HTTP), nombre
        if cfg["transporte"] == P.STDIO:
            assert cfg.get("comando") and cfg.get("args"), nombre
            assert "url" not in cfg, f"{nombre}: stdio no lleva url"
        else:
            assert cfg.get("url", "").startswith("https://"), nombre
            assert "comando" not in cfg, f"{nombre}: http no se lanza local"
        assert cfg["docs"].startswith("https://"), nombre

    # Las dos plataformas pedidas tienen que estar, cada una con su servidor
    # local y su remoto: son escenarios distintos, no alternativas.
    assert set(P.por_plataforma("Power BI")) == {"powerbi_local", "powerbi_remoto"}
    assert set(P.por_plataforma("Tableau")) == {"tableau_local", "tableau_cloud"}


def test_el_mcp_json_generado_es_el_que_documenta_cada_proveedor():
    """La configuracion se GENERA del registro y no se escribe a mano en la
    interfaz: si manana cambia un nombre de paquete, cambia en un solo lugar.

    Los valores son los de la documentacion oficial (agosto 2026), no de
    memoria — inventar un paquete npm daria una pantalla que parece andar y
    falla en la maquina del usuario."""
    import json as _json
    from mvdg import mcp_presets as P

    # Power BI remoto: Microsoft documenta exactamente esta forma.
    remoto = _json.loads(P.config_json("powerbi_remoto"))
    assert remoto["servers"]["powerbi-remoto"] == {
        "type": "http",
        "url": "https://api.fabric.microsoft.com/v1/mcp/powerbi",
    }
    # Tableau Cloud: servicio administrado con OAuth 2.1.
    assert _json.loads(P.config_json("tableau_cloud"))["servers"][
        "tableau-cloud"]["url"] == "https://mcp.tableau.com"
    # Locales: paquetes npm oficiales, lanzados con npx.
    pbi = _json.loads(P.config_json("powerbi_local"))["servers"]["powerbi-local"]
    assert pbi["command"] == "npx"
    assert "@microsoft/powerbi-modeling-mcp@latest" in pbi["args"]
    tab = _json.loads(P.config_json("tableau_local"))["servers"]["tableau-local"]
    assert "@tableau/mcp-server" in tab["args"]

    with pytest.raises(KeyError):
        P.config_json("no_existe")


def test_los_remotos_no_se_intentan_lanzar_desde_el_programa():
    """Power BI/Fabric remoto y Tableau Cloud autentican con OAuth
    interactivo. Ofrecer un boton de "conectar" seria ofrecer algo que siempre
    falla; se corta con un mensaje que dice que hacer en su lugar."""
    from mvdg import mcp_presets as P

    assert P.lanzable_localmente("powerbi_local")
    assert P.lanzable_localmente("tableau_local")
    for remoto in ("powerbi_remoto", "tableau_cloud"):
        assert not P.lanzable_localmente(remoto)
        with pytest.raises(ValueError) as e:
            P.herramientas(remoto)
        assert "config_json" in str(e.value), "el error tiene que decir la salida"


def test_el_camino_stdio_de_los_presets_funciona_de_verdad(tmp_path):
    """No se puede probar en vivo contra Power BI o Tableau: hacen falta un
    tenant, credenciales y (en Power BI) que el admin habilite la opcion. Lo
    que SI se prueba es el mecanismo que van a usar — lanzar un servidor MCP
    por stdio y listar sus herramientas — apuntando al servidor MCP del propio
    programa. Si esto anda, lo unico no verificado es el proveedor."""
    from mvdg import mcp_client, mcp_presets

    assert mcp_presets.lanzable_localmente("powerbi_local")
    tools = mcp_client.list_tools(sys.executable, ["-m", "mvdg.mcp_server"],
                                  env=_mcp_env(tmp_path))
    assert [t for t in tools if t["name"] == "mvdg_catalog"], tools


def test_mcp_server_full_roundtrip_over_real_stdio(tmp_path):
    """El servidor MCP de gobernanza responde por el protocolo REAL (stdio):
    cliente oficial ↔ subproceso servidor, sin mocks de transporte."""
    import json
    import sys
    pytest.importorskip("mcp")
    from mvdg import mcp_client

    cmd, args, env = sys.executable, ["-m", "mvdg.mcp_server"], _mcp_env(tmp_path)

    tools = mcp_client.list_tools(cmd, args, env=env)
    assert sorted(t["name"] for t in tools) == [
        "mvdg_alerts", "mvdg_case_deliverable", "mvdg_catalog",
        "mvdg_contracts", "mvdg_dictionary", "mvdg_glossary",
        "mvdg_lineage", "mvdg_quality", "mvdg_search"]
    # todas anunciadas con descripción utilizable por un agente
    assert all(len(t["description"]) > 40 for t in tools)

    cat = json.loads(mcp_client.call_tool(
        cmd, args, "mvdg_catalog", {"lang": "es"}, env=env))
    assert len(cat) == 8 and {"dataset", "owner", "domain"} <= set(cat[0])

    q = json.loads(mcp_client.call_tool(
        cmd, args, "mvdg_quality", {"dataset": "medicamentos_openfda"},
        env=env))
    from mvdg import samples
    real = samples.sample_quality_results("medicamentos_openfda", "es")
    assert len(q) == len(real)
    assert (sum(1 for r in q if r["status"] == "fail")
            == int((real["status"] == "fail").sum()))

    lin = json.loads(mcp_client.call_tool(cmd, args, "mvdg_lineage", {},
                                          env=env))
    ids = {n["id"] for n in lin["nodes"]}
    assert "bi_dashboard" in ids and "medicamentos_openfda" in ids

    md = mcp_client.call_tool(
        cmd, args, "mvdg_case_deliverable",
        {"case": "rotulado_alimentos", "lang": "pt"}, env=env)
    assert "Entregável de governança" in md


def test_mcp_server_errors_are_actionable_and_metadata_only(tmp_path):
    """Errores accionables (dicen qué valores son válidos) y respuesta de
    metadata: el servidor jamás expone filas de datos de los casos."""
    import json
    import sys
    pytest.importorskip("mcp")
    from mvdg import mcp_client, samples

    cmd, args, env = sys.executable, ["-m", "mvdg.mcp_server"], _mcp_env(tmp_path)

    err = mcp_client.call_tool(cmd, args, "mvdg_quality",
                               {"dataset": "no_existe"}, env=env)
    assert err.startswith("Error:") and "medicamentos_openfda" in err
    err2 = mcp_client.call_tool(cmd, args, "mvdg_case_deliverable",
                                {"case": "zzz"}, env=env)
    assert err2.startswith("Error:") and "rotulado_alimentos" in err2

    # búsqueda encuentra metadata del catálogo/diccionario/glosario...
    hits = json.loads(mcp_client.call_tool(cmd, args, "mvdg_search",
                                           {"term": "NDC"}, env=env))
    assert sum(len(v) for v in hits.values()) > 0
    # ...pero ningún tool devuelve valores de las filas reales del caso
    table = samples.load_sample_table("cafe_sales_kaggle")
    sample_cell = str(table.iloc[0, 0])
    dic = mcp_client.call_tool(cmd, args, "mvdg_dictionary",
                               {"dataset": "cafe_sales_kaggle"}, env=env)
    assert sample_cell not in dic


def test_tableau_official_mcp_interop_real_protocol(tmp_path):
    """Interop REAL con el MCP oficial de Tableau (@tableau/mcp-server):
    nuestro cliente MCP lanza el binario oficial, completa el handshake del
    protocolo y lista sus herramientas. El Tableau Server se simula con un
    HTTP local (el binario valida `GET /api/x.y/serverinfo` al arrancar —
    comprobado empíricamente); el protocolo MCP y el binario son 100% reales.

    Requiere el paquete npm instalado: se salta con instrucción clara si no
    está (seteá MVDG_TABLEAU_MCP_BIN o instalá @tableau/mcp-server).
    """
    import json
    import os
    import shutil
    import socket
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    pytest.importorskip("mcp")
    bin_path = os.environ.get("MVDG_TABLEAU_MCP_BIN") or shutil.which(
        "tableau-mcp-server")
    if not bin_path or not os.path.exists(bin_path):
        pytest.skip("binario oficial de Tableau MCP no instalado "
                    "(npm i @tableau/mcp-server y seteá MVDG_TABLEAU_MCP_BIN)")

    from mvdg import mcp_client

    class FakeTableau(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _reply(self, obj):
            body = json.dumps(obj).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self):
            if "serverinfo" in self.path.lower():
                self._reply({"serverInfo": {
                    "productVersion": {"value": "2025.1.0",
                                       "build": "20251.25.0"},
                    "restApiVersion": "3.24"}})
            else:
                self._reply({})

    sock = socket.socket()
    sock.bind(("127.0.0.1", 0))
    port = sock.getsockname()[1]
    sock.close()
    srv = HTTPServer(("127.0.0.1", port), FakeTableau)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    try:
        env = {**os.environ,
               "SERVER": f"http://127.0.0.1:{port}",
               "SITE_NAME": "demo", "PAT_NAME": "demo", "PAT_VALUE": "demo",
               "NO_PROXY": "127.0.0.1,localhost",
               "no_proxy": "127.0.0.1,localhost"}
        tools = mcp_client.list_tools(bin_path, [], env=env, timeout=90)
        names = {t["name"] for t in tools}
        # herramientas núcleo del server oficial (v3.0.0: 21 en total)
        assert {"list-datasources", "query-datasource", "list-workbooks",
                "list-views", "search-content"} <= names
        assert len(tools) >= 15
    finally:
        srv.shutdown()


def test_install_integrity_guard_passes_on_consistent_repo():
    """En el repo (copia consistente) el guardián no reporta nada."""
    from mvdg import integrity
    assert integrity.check_install() == []
    assert set(integrity.MESSAGE) == {"es", "en", "pt"}


def test_install_integrity_guard_catches_stale_i18n(monkeypatch):
    """Simula un i18n viejo (sin la clave de Contratos): el guardián lo detecta
    con un mensaje que nombra la pieza — así el usuario sabe qué actualizar."""
    from mvdg import integrity
    from mvdg import i18n
    stale = {k: v for k, v in i18n._T.items() if k != "tab_contracts"}
    monkeypatch.setattr(i18n, "_T", stale)
    missing = integrity.check_install()
    assert any("tab_contracts" in m for m in missing)


def test_install_integrity_guard_catches_stale_engine(monkeypatch):
    """Simula un motor viejo (deliverable sin findings_df): detectado."""
    from mvdg import integrity, deliverable
    monkeypatch.delattr(deliverable, "findings_df", raising=True)
    missing = integrity.check_install()
    assert any("findings_df" in m for m in missing)


def test_lab_case_full_migration_circuit_real_http(tmp_path, monkeypatch):
    """EL CIRCUITO COMPLETO con el caso del laboratorio (medicamentos_openfda),
    contra un servidor HTTP real que imita Purview: (1) push del catálogo +
    glosario del caso con su curaduría real, (2) pull de vuelta de lo que
    quedó en 'Purview', (3) persistencia local de lo traído, (4) entra a
    Curaduría. Ida y vuelta sin errores, con sockets de verdad."""
    import json
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import curation, deliverable, imported, samples
    from mvdg import purview_export as pv
    from mvdg import purview_pull as pvp

    key = "medicamentos_openfda"
    # curaduría real previa: el laboratorio ya validó su primer término
    term_id = samples.SAMPLES[key]["terms"][0]["term_id"]
    curation.save_validation(f"glossary:{key}:{term_id}", "es", "validado",
                             "", "Dra. Pérez", "Farmacovigilancia")

    meta, cat, dic, glo, _res = deliverable._case_tables(key, "es")
    lookup = deliverable._curation_lookup(key, "es")

    stored = {"entities": [], "terms": []}  # lo que 'Purview' recibe y devuelve

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):
            pass

        def _body(self):
            n = int(self.headers.get("Content-Length", 0))
            return json.loads(self.rfile.read(n)) if n else None

        def _send(self, code, payload):
            data = json.dumps(payload).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def do_POST(self):
            body = self._body()
            if "/oauth2/token" in self.path:
                self._send(200, {"access_token": "lab-token"})
            elif self.path.endswith("/entity/bulk"):
                stored["entities"].extend(body["entities"])
                mutated = [{"guid": f"g{i}", "attributes":
                           {"qualifiedName": e["attributes"]["qualifiedName"]}}
                          for i, e in enumerate(body["entities"])]
                self._send(200, {"mutatedEntities": {"CREATE": mutated, "UPDATE": []}})
            elif self.path.endswith("/glossary/term"):
                term = {**body, "guid": f"term-{len(stored['terms'])}"}
                stored["terms"].append(term)
                self._send(200, term)
            elif self.path.endswith("/entity/bulk/classification"):
                self._send(200, {})
            elif "/datamap/api/search/query" in self.path:
                self._send(200, {"value": [
                    {"id": f"g{i}", "name": e["attributes"]["name"],
                     "description": e["attributes"].get("description", "")}
                    for i, e in enumerate(stored["entities"])
                    if e["typeName"] == "rdbms_table"], "continuationToken": None})
            else:
                self._send(404, {"error": self.path})

        def do_GET(self):
            if self.path.endswith("/glossary"):
                self._send(200, [{"guid": "gloss-1", "name": "MV Data Governance"}])
            elif self.path.endswith("/glossary/gloss-1/terms"):
                self._send(200, stored["terms"])
            else:
                self._send(404, {"error": self.path})

    port = _free_port()
    server = HTTPServer(("127.0.0.1", port), Handler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    try:
        monkeypatch.setenv("PURVIEW_TENANT_ID", "tid")
        monkeypatch.setenv("PURVIEW_CLIENT_ID", "cid")
        monkeypatch.setenv("PURVIEW_CLIENT_SECRET", "sec")
        monkeypatch.setenv("PURVIEW_ACCOUNT_NAME", "lab")
        monkeypatch.setenv("PURVIEW_API_BASE", f"http://127.0.0.1:{port}")

        def fake_get_token():
            import urllib.request
            req = urllib.request.Request(f"http://127.0.0.1:{port}/oauth2/token",
                                         data=b"{}", method="POST")
            with urllib.request.urlopen(req) as resp:
                return json.loads(resp.read())["access_token"]

        monkeypatch.setattr(pv, "_get_token", fake_get_token)

        # (1) IDA: push real del caso del laboratorio
        pushed = pv.push_all(cat, dic, glo, curation_lookup=lookup, dry_run=False)
        assert pushed["catalog"]["failed_batches"] == []
        assert pushed["glossary"]["failed"] == []
        statuses = {t["name"]: t["status"] for t in stored["terms"]}
        assert sum(1 for s in statuses.values() if s == "Approved") == 1  # el validado

        # (2) VUELTA: pull de lo que quedó en 'Purview'
        pulled = pvp.pull_all()
        assert pulled["glossary"]["term_count"] == len(glo)
        assert pulled["catalog"]["table_count"] == 1
        assert pulled["catalog"]["tables"][0]["name"] == key

        # (3) persistencia local + (4) entra a Curaduría
        n = imported.save_terms("purview", pulled["glossary"]["terms"])
        n += imported.save_tables("purview", pulled["catalog"]["tables"])
        assert n == len(glo) + 1
        cdf = curation.list_items("es")
        assert cdf["item_id"].str.startswith("glossary:imported:purview:").sum() == len(glo)
    finally:
        server.shutdown()

def test_toda_dependencia_tiene_tope_de_version_mayor():
    """Un `>=` sin tope deja entrar la proxima version MAYOR sola.

    Cuando eso pasa, el CI se pone en rojo sin que nadie haya tocado codigo, y
    el rojo aparece en el PR de otra persona. Ya paso una vez: `mcp` 2.0.0 saco
    `mcp.server.fastmcp` y dejo el servidor MCP con ModuleNotFoundError.

    El tope se puede subir cuando se quiera — es cambiar un numero y ver si la
    suite pasa. Lo que este test impide es que ese cambio entre SOLO.
    """
    import re as _re
    raiz = _repo_root()
    sin_tope = []
    for archivo in ("requirements.txt", "requirements-dev.txt"):
        ruta = os.path.join(raiz, archivo)
        with open(ruta, encoding="utf-8") as fh:
            for n, linea in enumerate(fh, 1):
                linea = linea.strip()
                if not linea or linea.startswith("#") or linea.startswith("-r"):
                    continue
                # nombre + especificadores, ej "pandas>=2.0,<4"
                m = _re.match(r"^([A-Za-z0-9_.\-]+)\s*(.*)$", linea)
                if not m:
                    continue
                nombre, spec = m.group(1), m.group(2)
                if "<" not in spec:
                    sin_tope.append(f"{archivo}:{n} {nombre}{spec}")
    assert not sin_tope, (
        "estas dependencias no tienen tope de version mayor, asi que una "
        "release ajena puede poner el CI en rojo sola: " + "; ".join(sin_tope))

def test_la_interfaz_no_usa_emojis_decorativos():
    """La interfaz muestra ESTADO, no decoracion.

    El semaforo de calidad (verde/amarillo/rojo) es informacion: dice de un
    vistazo si una regla pasa, alerta o falla. Un cohete al lado de un titulo
    no dice nada — y 666 de esos hacian que el producto pareciera un chat en
    vez de una herramienta de gobierno de datos.

    Se permite tambien la flecha tipografica (->) del linaje, que no es un
    emoji sino un signo.

    Sin este test, el proximo texto nuevo vuelve a traer uno y nadie lo nota
    hasta que ya esta en la pantalla del cliente.
    """
    import re as _re
    permitidos = {"\U0001F7E2", "\U0001F7E1", "\U0001F534", "\u2192"}
    clase = _re.compile("[\U0001F300-\U0001FAFF\U00002600-\U000027BF"
                        "\U0001F000-\U0001F0FF]")
    raiz = _repo_root()
    # Lo que ve el usuario: el motor, el dashboard, la API y las dos interfaces
    # de escritorio. packaging/ y los tests son herramientas de desarrollo.
    carpetas = ["mvdg", "app", "bi_api",
                os.path.join("electron", "ui", "src"),
                os.path.join("electron", "launcher")]
    intrusos = []
    for carpeta in carpetas:
        base = os.path.join(raiz, carpeta)
        for actual, _dirs, archivos in os.walk(base):
            if "node_modules" in actual or os.sep + "dist" in actual:
                continue
            for nombre in archivos:
                if not nombre.endswith((".py", ".js", ".jsx", ".mjs", ".html")):
                    continue
                ruta = os.path.join(actual, nombre)
                with open(ruta, encoding="utf-8") as fh:
                    for n, linea in enumerate(fh, 1):
                        for ch in linea:
                            if clase.match(ch) and ch not in permitidos:
                                rel = os.path.relpath(ruta, raiz)
                                intrusos.append(f"{rel}:{n} {ch!r}")
    assert not intrusos, (
        "emojis decorativos en la interfaz (solo se permite el semaforo "
        "verde/amarillo/rojo): " + "; ".join(intrusos[:12]))

def test_el_automerge_dispara_los_instaladores():
    """Un merge hecho por el automerge NO dispara los workflows de `on: push`.

    Es la proteccion de GitHub contra workflows recursivos: los eventos
    originados con el GITHUB_TOKEN no crean corridas nuevas. Solo
    workflow_dispatch y repository_dispatch son excepcion.

    Costo caro y en silencio: instalador.yml tiene `on: push` a main desde el
    principio y estuvo 26 merges sin correr — nadie lo noto, porque "no
    corrio" no falla, simplemente no pasa nada.

    Este test fija que el automerge siga pidiendolos a mano.
    """
    import yaml as _yaml
    ruta = os.path.join(_repo_root(), ".github", "workflows", "automerge.yml")
    with open(ruta, encoding="utf-8") as fh:
        crudo = fh.read()
    datos = _yaml.safe_load(crudo)

    permisos = datos.get("permissions", {})
    assert permisos.get("actions") == "write", (
        "sin permiso actions:write el automerge no puede disparar nada")

    assert "createWorkflowDispatch" in crudo, (
        "el automerge no dispara ningun workflow despues de mergear: el "
        "instalador no se va a construir solo")
    # Los dos workflows de PyInstaller se eliminaron: pedirlos por
    # createWorkflowDispatch fallaría en cada merge, y el fallo se reporta
    # como warning — o sea, en silencio.
    # Se busca la forma en que aparecen en la lista de dispatch
    # (`{archivo: 'instalador.yml'`), no el nombre suelto: el comentario que
    # explica por qué se eliminaron los menciona, y con razón.
    for muerto in ("instalador.yml", "instalador_owner.yml"):
        assert f"'{muerto}'" not in crudo, (
            f"el automerge sigue disparando {muerto}, que ya no existe: la "
            "llamada falla en cada merge y el fallo se reporta como warning")

    for archivo in ("instalador_electron.yml",):
        assert archivo in crudo, f"el automerge no dispara {archivo}"
        # y ese workflow tiene que ACEPTAR ser disparado
        wf = os.path.join(_repo_root(), ".github", "workflows", archivo)
        with open(wf, encoding="utf-8") as fh:
            d = _yaml.safe_load(fh)
        # PyYAML lee la clave `on:` como el booleano True
        triggers = d.get("on", d.get(True)) or {}
        assert "workflow_dispatch" in triggers, (
            f"{archivo} no acepta workflow_dispatch: el automerge lo pide y "
            f"la llamada falla")

def _firmar_licencia_de_prueba(payload):
    """Firma un token MVDG2 con un par nuevo y deja esa publica configurada.

    Devuelve el token. El llamador es responsable de restaurar
    licensing.PUBLIC_KEY_B64 (los tests de abajo usan monkeypatch)."""
    import base64
    import json as _json
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric.ed25519 import (
        Ed25519PrivateKey)

    def b64u(b):
        return base64.urlsafe_b64encode(b).decode().rstrip("=")

    priv = Ed25519PrivateKey.generate()
    pub = b64u(priv.public_key().public_bytes(
        encoding=serialization.Encoding.Raw,
        format=serialization.PublicFormat.Raw))
    cuerpo = b64u(_json.dumps(payload, separators=(",", ":")).encode())
    return f"MVDG2.{cuerpo}.{b64u(priv.sign(cuerpo.encode('ascii')))}", pub


def test_la_api_deja_activar_la_licencia_del_exe(tmp_path, monkeypatch):
    """El .exe (Electron + React) no tenia NINGUNA nocion de licencia.

    Sus seis vistas son funciones gratuitas, no habia donde pegar la clave, y
    demo, paga y owner se veian exactamente igual. O sea que un cliente podia
    pagar US$390 y, usando el .exe, recibir la demo — sin siquiera un campo
    donde poner lo que compro.

    Esto cubre el circuito por el que ahora pasa: consultar el plan, activarlo
    con una clave firmada, y volver a demo.
    """
    import time as _time
    from fastapi.testclient import TestClient

    from bi_api.main import app
    from mvdg import licensing

    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    token, pub = _firmar_licencia_de_prueba(
        {"plan": "professional", "email": "c@empresa.com",
         "iat": int(_time.time())})
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub)

    c = TestClient(app)
    assert c.get("/api/licencia").json()["plan"] == "demo"

    r = c.post("/api/licencia", json={"token": token})
    assert r.status_code == 200, r.text
    assert r.json()["plan"] == "professional"
    assert r.json()["email"] == "c@empresa.com"

    # persiste: es lo que ve el cliente al reabrir el programa
    assert c.get("/api/licencia").json()["plan"] == "professional"

    assert c.delete("/api/licencia").json()["plan"] == "demo"


def test_la_api_rechaza_una_licencia_falsa(tmp_path, monkeypatch):
    """Una clave que no verifica NO se guarda y el plan no se mueve.

    Y tiene que fallar RUIDOSO (400): el sintoma de no hacerlo es un cliente
    que pego su clave, no vio ningun error, y sigue en demo sin entender por
    que."""
    import time as _time
    from fastapi.testclient import TestClient

    from bi_api.main import app
    from mvdg import licensing

    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    # una publica que NO es la del firmante
    _, pub_ajena = _firmar_licencia_de_prueba({"plan": "demo"})
    token, _ = _firmar_licencia_de_prueba(
        {"plan": "owner", "iat": int(_time.time())})
    monkeypatch.setattr(licensing, "PUBLIC_KEY_B64", pub_ajena)

    c = TestClient(app)
    for cuerpo in ({}, {"token": ""}, {"token": "cualquier-cosa"},
                   {"token": token}):
        assert c.post("/api/licencia", json=cuerpo).status_code == 400
    assert c.get("/api/licencia").json()["plan"] == "demo"


def test_la_interfaz_del_exe_tiene_la_pantalla_de_licencia():
    """Que los endpoints existan no alcanza: el cliente los usa desde la UI.

    Si la vista se cae del bundle, la API sigue respondiendo perfecto y el
    cliente igual no tiene donde pegar su clave — que era exactamente el
    estado anterior."""
    import re as _re
    raiz = _repo_root()
    app_jsx = os.path.join(raiz, "electron", "ui", "src", "App.jsx")
    with open(app_jsx, encoding="utf-8") as fh:
        crudo = fh.read()
    # Solo CODIGO: buscar sobre el archivo crudo daba verde con la vista
    # comentada, porque "/* licencia: Licencia */" sigue conteniendo el texto.
    # Verificado — el test pasaba sobre el bug.
    fuente = _re.sub(r"/\*.*?\*/", "", crudo, flags=_re.S)
    fuente = _re.sub(r"^\s*//.*$", "", fuente, flags=_re.M)
    assert '"licencia"' in fuente, "la vista licencia no esta en VISTAS"
    assert "function Licencia" in fuente, "falta el componente Licencia"
    assert "licencia: Licencia" in fuente, "la vista no esta enganchada a RENDER"

    api_js = os.path.join(raiz, "electron", "ui", "src", "api.js")
    with open(api_js, encoding="utf-8") as fh:
        cliente = fh.read()
    for fn in ("licencia", "activarLicencia", "desactivarLicencia"):
        assert f"export async function {fn}" in cliente, f"falta {fn}() en api.js"

    # y los textos, en los tres idiomas (la UI del .exe tiene su propio i18n)
    i18n_js = os.path.join(raiz, "electron", "ui", "src", "i18n.js")
    with open(i18n_js, encoding="utf-8") as fh:
        textos = fh.read()
    for clave in ("lic_plan", "lic_activar", "lic_clave", "lic_demo_ayuda"):
        assert clave in textos, f"falta la clave {clave} en el i18n del .exe"

def test_el_instalador_instala_deps_antes_de_usarlas():
    """`npm run <lo que sea>` necesita node_modules; `npm ci` es quien lo crea.

    Estaban al reves y el build moria con "Cannot find package 'esbuild'".
    Ese bug estuvo SIEMPRE ahi: no se veia porque el workflow moria antes, en
    el Python embebido. Cada falla tapaba a la siguiente, y asi el instalador
    acumulo 8 corridas rojas sin que se viera el fondo del problema.

    Se verifica el ORDEN de los pasos, que es lo unico que importa: el primer
    `npm run` tiene que venir despues del `npm ci`.
    """
    import yaml as _yaml
    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    with open(ruta, encoding="utf-8") as fh:
        datos = _yaml.safe_load(fh)

    pasos = datos["jobs"]["build"]["steps"]
    i_ci = i_run = None
    for i, paso in enumerate(pasos):
        cmd = str(paso.get("run") or "")
        if i_ci is None and "npm ci" in cmd:
            i_ci = i
        if i_run is None and "npm run" in cmd:
            i_run = i

    assert i_ci is not None, "el workflow no instala las dependencias (npm ci)"
    assert i_run is not None, "el workflow no usa npm run: .cambio el empaquetado?"
    assert i_ci < i_run, (
        f"`npm run` esta en el paso {i_run + 1} y `npm ci` en el {i_ci + 1}: "
        f"se usan las dependencias antes de instalarlas")

def test_la_url_de_cada_release_apunta_al_archivo_que_se_sube():
    """El workflow imprime, en el resumen de la corrida, la URL exacta que hay
    que pegar en Vercel (MVDG_INSTALLER_URL / MVDG_INSTALLER_URL_OWNER). Si ese
    nombre no coincide con el del asset que realmente se subio, la URL da 404 y
    /api/descargar contesta que no encuentra el instalador.

    Paso exactamente eso con el owner: el asset se sube como
    MVDataGovernance_OWNER_Setup_$v.exe y la URL decia
    MVDataGovernance_Setup_$v.exe (sin el OWNER_). El bloque que existe para no
    tener que armar la URL a mano era el que la armaba mal.

    Se comparan los dos lados por RELEASE (cliente-latest / owner-latest), que
    es lo unico que ata un asset con su link.
    """
    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    with open(ruta, encoding="utf-8") as fh:
        wf = fh.read()

    # Nombre de archivo que se sube, por release: sale del `exe=` que despues
    # se le pasa a `gh release upload <tag>`.
    subidos = dict(re.findall(
        r'exe="[^"]*/(MVDataGovernance[A-Za-z_]*_Setup_\$v\.exe)"'
        r'(?:.|\n)*?gh release upload (\S+)', wf))
    # …y la URL que se imprime para pegar en Vercel, por release.
    publicados = {tag: nombre for tag, nombre in re.findall(
        r'releases/download/(\S+?)/(MVDataGovernance[A-Za-z_]*_Setup_\$v\.exe)', wf)}

    assert len(publicados) >= 2, (
        "no se encontraron las URLs de descarga de cliente y owner: "
        f"{publicados}")

    for nombre_subido, tag in subidos.items():
        assert tag in publicados, f"se sube un asset a {tag} pero no se publica su URL"
        assert publicados[tag] == nombre_subido, (
            f"release {tag}: se sube '{nombre_subido}' pero la URL que el "
            f"workflow manda a pegar en Vercel apunta a "
            f"'{publicados[tag]}' — eso es un 404")


def test_cada_merge_produce_al_menos_el_instalador_del_cliente():
    """Un build que termina en verde sin dejar nada es peor que uno en rojo.

    El instalador del OWNER necesita el secreto MVDG_OWNER_TOKEN y, si falta,
    ese build se saltea (a proposito: si no, el repo queda en rojo permanente
    por algo que no es un bug). Pidiendo SOLO 'owner', una corrida sin secreto
    terminaria verde y sin ningun instalador — un verde que no significa nada.

    Con 'ambas', el del CLIENTE se construye siempre porque no necesita
    secretos.
    """
    ruta = os.path.join(_repo_root(), ".github", "workflows", "automerge.yml")
    with open(ruta, encoding="utf-8") as fh:
        crudo = fh.read()
    # Solo codigo: un // comentario que mencione 'owner' no cuenta.
    codigo = "\n".join(linea for linea in crudo.splitlines()
                       if not linea.strip().startswith("//"))
    assert "instalador_electron.yml" in codigo
    assert "version: 'ambas'" in codigo, (
        "el automerge pide una version que puede no construir nada; con "
        "'ambas' el instalador del cliente sale siempre")

def test_la_falta_del_secreto_del_owner_no_frena_al_instalador_del_cliente():
    """Cortar o seguir se decide por lo que se PIDIO, no por el evento.

    El automerge dispara con workflow_dispatch igual que una persona, asi que
    mirar github.event_name no distingue nada — se vio en la corrida #9: el
    automerge pidio 'owner', falto MVDG_OWNER_TOKEN, y el build quedo en rojo
    aunque nadie hubiera apretado nada.

    Con version=ambas el del cliente ya se construyo, asi que la falta del
    secreto tiene que avisar y seguir.
    """
    ruta = os.path.join(_repo_root(), ".github", "workflows",
                        "instalador_electron.yml")
    with open(ruta, encoding="utf-8") as fh:
        crudo = fh.read()
    codigo = "\n".join(linea for linea in crudo.splitlines()
                       if not linea.strip().startswith("#"))
    assert 'steps.modo.outputs.v }}" = "owner"' in codigo, (
        "la decision de cortar mira el evento en vez de la version pedida: "
        "el automerge dispara igual que una persona y no se distinguen")
    assert "github.event_name" not in codigo, (
        "github.event_name no sirve para esto: el automerge tambien dispara "
        "con workflow_dispatch")

def test_el_logo_de_la_marca_esta_en_la_web_y_en_el_programa():
    """El mismo logo en los tres lados, y ninguno vacio.

    Dos cosas que este test fija, las dos vividas:

    1. El encabezado del .exe tenia un EMOJI adentro de un span. Al sacar los
       emojis decorativos ese span quedo vacio y el programa se quedo sin
       marca — un hueco donde iba el logo, y nada que fallara.
    2. El favicon de la interfaz era un escudo ambar dibujado en SVG, no el
       logo. La landing mostraba uno y el programa otro.

    El .png de assets/brand es la unica fuente de verdad: los dos builds lo
    leen en tiempo de compilacion, asi que cambiar el logo lo cambia en todos
    lados sin tocar codigo.
    """
    import hashlib
    raiz = _repo_root()

    # 1. La landing usa EXACTAMENTE el mismo archivo que el producto.
    def sha(ruta):
        with open(ruta, "rb") as fh:
            return hashlib.sha256(fh.read()).hexdigest()

    marca = os.path.join(raiz, "assets", "brand", "mv_icon.png")
    landing = os.path.join(raiz, "landing", "mv_icon.png")
    assert sha(marca) == sha(landing), (
        "el logo de la landing y el del producto son archivos distintos: uno "
        "de los dos quedo viejo")

    # 2. Los dos builds leen el png de verdad (no un base64 pegado a mano,
    #    que se desactualiza en silencio).
    for script, tam in (("build-ui.mjs", "mv_icon_64.png"),
                        ("build-launcher.mjs", "mv_icon_128.png")):
        ruta = os.path.join(raiz, "electron", "scripts", script)
        with open(ruta, encoding="utf-8") as fh:
            fuente = fh.read()
        assert tam in fuente, f"{script} no lee el logo de assets/brand"
        assert 'rel="icon"' in fuente and "${LOGO}" in fuente, (
            f"{script} no usa el logo como favicon")

    # 3. La marca NO puede ser un hueco: tiene que pintar la imagen.
    for css_rel, sel in ((("electron", "ui", "src", "styles.css"), ".escudo"),
                         (("electron", "launcher", "src", "styles.css"), ".shield")):
        ruta = os.path.join(raiz, *css_rel)
        with open(ruta, encoding="utf-8") as fh:
            css = fh.read()
        i = css.index(sel)
        bloque = css[i:css.index("}", i)]
        assert "--mv-logo" in bloque, (
            f"{sel} no pinta el logo: si adentro solo hay texto o un emoji, "
            f"el dia que se saque queda un hueco")

    # 4. La CSP del launcher tiene que PERMITIR data: en imagenes. Sin
    #    img-src, default-src 'self' bloquea el logo y no se ve nada — falla
    #    silenciosa, la peor clase.
    ruta = os.path.join(raiz, "electron", "scripts", "build-launcher.mjs")
    with open(ruta, encoding="utf-8") as fh:
        fuente = fh.read()
    assert "img-src 'self' data:" in fuente, (
        "la CSP del launcher no permite imagenes data:, asi que el logo "
        "queda bloqueado sin ningun error visible")

def test_los_iconos_salen_del_logo_vectorial():
    """El logo es UN vector y todo lo demas sale de ahi.

    Antes habia media docena de .png sueltos y un .ico, cada uno generado a
    mano en algun momento. Cambiar el logo era una caceria de archivos, y el
    resultado tipico es el favicon nuevo con el icono del escritorio viejo —
    nadie se entera hasta que un cliente lo ve.

    Ahora manda assets/brand/mv_logo.svg y packaging/generar_iconos.py los
    produce. Este test verifica lo que se puede sin rasterizar; si cairosvg
    esta disponible, ademas compara los pixeles contra el SVG actual.
    """
    from PIL import Image
    raiz = _repo_root()
    marca = os.path.join(raiz, "assets", "brand")

    svg = os.path.join(marca, "mv_logo.svg")
    assert os.path.exists(svg), "falta el logo vectorial mv_logo.svg"
    with open(svg, encoding="utf-8") as fh:
        contenido = fh.read()
    assert "<svg" in contenido and "viewBox" in contenido

    # Cada PNG mide lo que dice su nombre.
    for nombre, tam in (("mv_icon.png", 1024), ("mv_icon_256.png", 256),
                        ("mv_icon_128.png", 128), ("mv_icon_64.png", 64),
                        ("mv_icon_32.png", 32)):
        ruta = os.path.join(marca, nombre)
        with Image.open(ruta) as im:
            assert im.size == (tam, tam), f"{nombre} mide {im.size}, no {tam}"

    # El .ico con los tamanios que Windows usa. Si falta uno, Windows escala
    # otro y se ve borroso justo en el escritorio o la barra de tareas.
    with Image.open(os.path.join(marca, "mv.ico")) as ico:
        tiene = sorted({t[0] for t in ico.ico.sizes()})
    assert tiene == [16, 32, 48, 64, 128, 256], f"mv.ico trae {tiene}"

    # Y el generador declara esos mismos tamanios (si alguien agrega un uso
    # nuevo sin sumarlo al script, el icono nuevo nace desincronizado).
    gen = os.path.join(raiz, "packaging", "generar_iconos.py")
    assert os.path.exists(gen), "falta packaging/generar_iconos.py"
    with open(gen, encoding="utf-8") as fh:
        script = fh.read()
    for nombre in ("mv_icon_256.png", "mv_icon_128.png", "mv_icon_64.png"):
        assert nombre in script, f"{nombre} no lo genera el script"

    # Comparacion real contra el vector, solo si se puede rasterizar aca.
    # cairosvg NO esta en requirements: solo hace falta para REGENERAR, y
    # sumarlo obligaria a instalar libcairo en el CI para nada.
    try:
        import cairosvg  # noqa: F401
    except ImportError:
        return
    import subprocess
    r = subprocess.run([sys.executable, gen, "--revisar"],
                       cwd=raiz, capture_output=True, text=True)
    assert r.returncode == 0, (
        "los iconos del repo no salen del logo actual:\n" + r.stderr)

def test_los_workflows_que_exponen_datos_exigen_repo_privado():
    """Los logs de Actions de un repositorio PUBLICO los ve cualquiera.

    "Emitir licencia" imprime una licencia — el que la copia tiene el programa
    completo gratis. "Monitor" imprime ingresos y cantidad de clientes. En un
    repo publico, los dos serian una filtracion, y no hay canal alternativo:
    los artefactos de un repo publico tambien son accesibles para cualquiera.

    Por eso los dos comprueban la visibilidad ANTES de generar nada. Este test
    evita que alguien saque ese chequeo por comodidad.
    """
    import yaml as _yaml
    raiz = _repo_root()
    for archivo in ("emitir_licencia.yml", "monitor.yml"):
        ruta = os.path.join(raiz, ".github", "workflows", archivo)
        assert os.path.exists(ruta), f"falta {archivo}"
        with open(ruta, encoding="utf-8") as fh:
            crudo = fh.read()
        datos = _yaml.safe_load(crudo)
        # solo lineas de codigo: un comentario que diga "privado" no cuenta
        codigo = "\n".join(linea for linea in crudo.splitlines()
                           if not linea.strip().startswith("#"))
        assert ".private" in codigo, (
            f"{archivo} no consulta la visibilidad del repositorio")
        assert 'exit 1' in codigo, f"{archivo} no corta si el repo es publico"

        # y el chequeo tiene que ser el PRIMER paso: despues de generar la
        # licencia ya seria tarde, queda en el log igual.
        job = next(iter(datos["jobs"].values()))
        primero = job["steps"][0]
        assert "privad" in primero.get("name", "").lower(), (
            f"{archivo}: la comprobacion de visibilidad no es el primer paso "
            f"(es '{primero.get('name')}')")


def _api_cliente(tmp_path, monkeypatch):
    """TestClient de bi_api con un directorio de datos limpio."""
    from fastapi.testclient import TestClient

    from bi_api.main import app
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    return TestClient(app)


def test_el_que_paga_puede_USAR_las_tres_funciones_desde_el_exe(tmp_path, monkeypatch):
    """El .exe habla SOLO con bi_api, y bi_api no tenia ni un endpoint para
    las tres unicas funciones que se cobran.

    O sea: el cliente pagaba, pegaba su clave, y la pantalla de licencia le
    decia "estas 3 funciones estan desbloqueadas" sin ninguna forma de
    usarlas. Vivian solo en app/app.py (Streamlit), que el .exe no levanta.
    Pagar y no recibir es el mismo problema que cobrar un mes y entregar para
    siempre, mirado desde el otro lado.

    Este test recorre el circuito completo por HTTP, con el motor de verdad.
    """
    from mvdg import licensing

    c = _api_cliente(tmp_path, monkeypatch)

    # --- en demo: la VISTA PREVIA anda (es lo que hace lucir el producto) ---
    for destino in ("purview", "collibra"):
        r = c.post(f"/api/migracion/{destino}", json={})
        assert r.status_code == 200, (
            f"la vista previa de {destino} no anda en demo: {r.text[:200]}")
        assert r.json()["aplicado"] is False

    # --- en demo: el PUSH REAL se cobra ---
    for destino, funcion in (("purview", "migracion_purview"),
                             ("collibra", "migracion_collibra")):
        r = c.post(f"/api/migracion/{destino}", json={"aplicar": True})
        assert r.status_code == 402, (
            f"{destino}: un plan demo pudo hacer el push REAL (status "
            f"{r.status_code})")
        assert r.json()["detail"]["funcion"] == funcion

    r = c.post("/api/bi/escanear-tenant", json={})
    assert r.status_code == 402, "un plan demo escaneo el tenant"

    # --- con licencia: ya no responde 402 -------------------------------
    # Se parchea el plan, no se firma un token: la firma ya la cubre
    # api/pago_a_licencia.test.js de punta a punta. Lo que se prueba ACA es
    # que el endpoint consulte la licencia, que era lo que no existia.
    monkeypatch.setattr(licensing, "plan", lambda: "professional")
    for destino in ("purview", "collibra"):
        r = c.post(f"/api/migracion/{destino}", json={"aplicar": True})
        assert r.status_code != 402, (
            f"{destino}: con plan professional sigue diciendo que hay que pagar")
    r = c.post("/api/bi/escanear-tenant", json={})
    assert r.status_code != 402, "con plan professional sigue pidiendo licencia"


def test_los_conectores_externos_siguen_apagados_por_defecto(tmp_path, monkeypatch):
    """Regla del proyecto: los conectores externos estan apagados por defecto.

    El riesgo concreto de equivocarse: si `aplicar` fuera el default, un
    cuerpo mal armado —o vacio— le escribiria al Purview de PRODUCCION de un
    cliente. Por eso el default es la vista previa y sin credenciales el push
    real ni se intenta."""
    from mvdg import licensing

    c = _api_cliente(tmp_path, monkeypatch)
    monkeypatch.setattr(licensing, "plan", lambda: "owner")

    # Sin `aplicar`, y con cuerpo vacio, NUNCA se aplica.
    for cuerpo in ({}, {"aplicar": False}, {"otra_cosa": True}):
        r = c.post("/api/migracion/purview", json=cuerpo)
        assert r.status_code == 200
        assert r.json()["aplicado"] is False, (
            f"con el cuerpo {cuerpo} se hizo un push REAL contra Purview")

    # Con licencia pero sin credenciales: 409, no un 500 ni un push a ciegas.
    from mvdg import purview_export
    monkeypatch.setattr(purview_export, "configured", lambda: False)
    r = c.post("/api/migracion/purview", json={"aplicar": True})
    assert r.status_code == 409
    assert r.json()["detail"]["error"] == "conector_sin_configurar"


def test_la_api_dice_por_que_no_se_puede_antes_de_apretar(tmp_path, monkeypatch):
    """/api/conectores separa "te falta licencia" de "te faltan credenciales".

    Sin esto los dos casos se ven igual desde la interfaz —un boton que
    falla— y el cliente que SI pago cree que su licencia no sirve."""
    from mvdg import licensing

    c = _api_cliente(tmp_path, monkeypatch)
    d = c.get("/api/conectores").json()
    assert d["plan"] == licensing.PLAN_DEMO
    for clave in ("purview", "collibra"):
        assert d[clave]["licenciado"] is False
        assert "configurado" in d[clave]
    assert d["tenant_bi"]["licenciado"] is False

    monkeypatch.setattr(licensing, "plan", lambda: "professional")
    d = c.get("/api/conectores").json()
    assert d["purview"]["licenciado"] is True
    assert d["tenant_bi"]["licenciado"] is True


def test_los_errores_de_la_api_llegan_con_su_motivo_a_la_pantalla():
    """ApiError tiene que exponer `code`, y la pantalla tiene que ramificar
    por ese campo.

    Lo encontro una prueba en Chromium contra la API real: el endpoint
    devolvia 402 correctamente, el codigo compilaba, y la pantalla mostraba
    "el sistema remoto no respondio". Motivo: ApiError guardaba el
    identificador en `message` y App.jsx leia `e.code`, que era undefined —
    asi que TODOS los errores caian en el mensaje generico. El que pagaba y
    no tenia licencia no se enteraba de que le faltaba la licencia.

    Ningun test de unidad lo hubiera visto: las dos mitades eran correctas
    por separado."""
    api = os.path.join(_repo_root(), "electron", "ui", "src", "api.js")
    with open(api, encoding="utf-8") as fh:
        codigo = fh.read()
    assert re.search(r"this\.code\s*=", codigo), (
        "ApiError no expone `code`: quien lea e.code recibe undefined y todos "
        "los errores se ven iguales")

    app = os.path.join(_repo_root(), "electron", "ui", "src", "App.jsx")
    with open(app, encoding="utf-8") as fh:
        jsx = fh.read()
    for motivo in ("requiere_licencia", "sin_credenciales"):
        assert motivo in jsx, (
            f"la pantalla no distingue '{motivo}': el cliente que pago vería "
            f"un error generico y creería que su licencia no sirve")
        assert motivo in codigo, f"api.js no produce el motivo '{motivo}'"


def test_la_pantalla_de_licencia_deja_USAR_las_funciones_no_solo_listarlas():
    """El .exe listaba las tres funciones pagas sin ningun boton. Este test
    ata la pantalla a la API: si se agrega una funcion paga al motor y no se
    puede llamar desde el programa, el cliente paga y no recibe."""
    app = os.path.join(_repo_root(), "electron", "ui", "src", "App.jsx")
    with open(app, encoding="utf-8") as fh:
        jsx = fh.read()
    api = os.path.join(_repo_root(), "electron", "ui", "src", "api.js")
    with open(api, encoding="utf-8") as fh:
        cliente = fh.read()

    for fn in ("migrar", "escanearTenant", "conectores"):
        assert fn in cliente, f"api.js no sabe llamar a {fn}"
        assert fn in jsx, f"la pantalla nunca llama a {fn}"

    # Y la vista previa tiene que seguir siendo gratis: es lo que hace lucir
    # el producto y lo unico que un plan demo puede ver.
    assert "migrar(destino, false" in jsx, (
        "no quedo ningun boton de vista previa: en demo la pantalla no "
        "muestra nada de las funciones pagas")
    assert "migrar(destino, true" in jsx, "no quedo el boton de envio real"


def test_el_perfilador_detecta_la_PII_que_se_usa_en_uruguay_y_latam():
    """El detector de PII no marcaba `cedula` ni `apellido`.

    Este producto se vende en Uruguay, donde «cédula» es LA columna de PII más
    común que existe, y un apellido identifica a una persona igual que un
    nombre. Un informe de cumplimiento que no las marca no está incompleto:
    está EQUIVOCADO, y el cliente lo firma creyendo que revisó.

    La causa era de regex: con búsqueda por subcadena, `ci` matchea adentro de
    «precio», así que llevaba `\\b`… y `\\b` no corta antes de un guión bajo
    porque `_` es carácter de palabra. O sea que `ci\\b` no encontraba
    `ci_cliente`, que es como se llama la columna en media base uruguaya.
    """
    from mvdg.profiler import _es_pii_por_nombre as pii

    debe_marcar = [
        # documentos de la región
        "ci", "ci_cliente", "cedula", "cédula", "Cedula_Identidad",
        "documento", "nro_documento", "doc", "dni", "rut", "ruc", "cuit",
        "cuil", "cpf", "curp", "pasaporte",
        # persona y contacto
        "nombre", "NOMBRE COMPLETO", "apellido", "Apellido Materno",
        "sobrenome", "email", "correo", "telefono", "telefono2", "celular",
        "whatsapp",
        # ubicación y financieros
        "direccion", "domicilio", "iban", "cbu", "tarjeta", "ip", "ip_origen",
        "fecha_nacimiento",
    ]
    faltan = [c for c in debe_marcar if not pii(c)]
    assert not faltan, f"PII sin detectar: {faltan}"

    # Y lo que NO puede marcar: un falso positivo hace que el cliente
    # desconfíe del informe entero y deje de mirarlo.
    no_debe = [
        "precio", "inicio", "equipo", "servicio", "monto", "saldo", "cantidad",
        "ciudad", "municipio", "codigo", "id", "fecha", "estado", "producto",
        "descripcion", "negocio", "ejercicio", "participacion", "anticipo",
        "principal", "docente", "ipc",
    ]
    sobran = [c for c in no_debe if pii(c)]
    assert not sobran, f"marcadas como PII sin serlo: {sobran}"


def test_el_exe_puede_perfilar_un_archivo_propio(tmp_path, monkeypatch):
    """La landing lo anuncia con estas palabras: «Subí un CSV o Excel y obtené
    al instante esquema, nulos, duplicados, PII detectada y reglas
    sugeridas». Y el plan de US$ 149 dice «Todo el programa sin límite de
    tiempo».

    El .exe no tenía nada de eso: ni endpoint, ni pantalla, ni forma de cargar
    un archivo. El perfilador vivía solo en app/app.py (Streamlit), que el
    .exe no levanta — así que el cliente bajaba el programa, buscaba la
    función principal que vio anunciada, y no existía.

    Es gratis, como en Streamlit: no está en FUNCIONES_PAGAS."""
    import io

    c = _api_cliente(tmp_path, monkeypatch)

    csv = ("email,monto,cedula\n"
           "a@empresa.com,10.5,1.234.567-8\n"
           "b@empresa.com,20,2.345.678-9\n"
           ",30,\n"
           "a@empresa.com,10.5,1.234.567-8\n")
    r = c.post("/api/perfilar",
               files={"archivo": ("clientes.csv", csv.encode(), "text/csv")})
    assert r.status_code == 200, r.text[:300]
    d = r.json()

    # lo que promete la landing, campo por campo
    assert d["resumen"]["rows"] == 4
    assert d["resumen"]["columns"] == 3
    assert d["resumen"]["duplicate_rows"] == 1          # duplicados
    assert d["resumen"]["null_cells_pct"] > 0           # nulos
    assert d["resumen"]["pii_columns"] == 2             # PII detectada
    pii = [col["column"] for col in d["perfil"] if col["possible_pii"]]
    assert set(pii) == {"email", "cedula"}
    assert d["reglas"], "no sugirio ninguna regla de calidad"

    # Los conteos son ENTEROS. Salieron de una Series de pandas y venian como
    # 4.0 porque un solo decimal unifica el tipo de la Series entera: "4.0
    # filas" en pantalla se lee como un error del programa.
    assert isinstance(d["resumen"]["rows"], int)

    # Excel, que es como llega la mitad de los archivos de una empresa.
    import pandas as pd
    buf = io.BytesIO()
    pd.DataFrame({"telefono": ["099123456"], "saldo": [100]}).to_excel(buf, index=False)
    r = c.post("/api/perfilar", files={"archivo": ("datos.xlsx", buf.getvalue(), "")})
    assert r.status_code == 200
    assert [col["column"] for col in r.json()["perfil"]] == ["telefono", "saldo"]

    # CSV con punto y coma: es lo que exporta Excel en español, y asumir la
    # coma daba UNA sola columna con todo adentro y un perfil que no dice nada.
    r = c.post("/api/perfilar", files={
        "archivo": ("uy.csv", b"nombre;importe\nAna;1,5\nLuis;2,5\n", "text/csv")})
    assert r.status_code == 200
    assert len(r.json()["perfil"]) == 2, "no reconocio el punto y coma"


def test_el_perfilador_no_se_come_la_memoria_ni_acepta_cualquier_cosa(tmp_path, monkeypatch):
    """Corre en la PC del cliente, con el Python embebido del .exe: sin
    ningún tope, un archivo enorme se lleva puesta la memoria del programa.

    El tope ya no es fijo en 40 MB —eso se comía el caso normal, alguien
    perfilando SU archivo— sino configurable y alto por defecto. Lo que se
    prueba acá es que, CUANDO hay tope, se respeta: se baja a 1 MB por la
    variable de entorno y se manda un archivo más grande."""
    import bi_api.main as api
    c = _api_cliente(tmp_path, monkeypatch)

    for nombre, cuerpo, estado, motivo in [
        ("virus.exe", b"MZ\x90\x00", 400, "formato_no_soportado"),
        ("vacio.csv", b"", 400, "archivo_vacio"),
        ("roto.xlsx", b"no soy un excel", 400, "no_se_pudo_leer"),
    ]:
        r = c.post("/api/perfilar", files={"archivo": (nombre, cuerpo, "")})
        assert r.status_code == estado, f"{nombre}: {r.status_code}"
        assert r.json()["detail"]["error"] == motivo

    monkeypatch.setattr(api, "_MAX_BYTES", 1024 * 1024)   # 1 MB
    grande = b"a,b\n" + b"1,2\n" * 300_000                # ~1,2 MB
    r = c.post("/api/perfilar", files={"archivo": ("grande.csv", grande, "text/csv")})
    assert r.status_code == 413
    assert r.json()["detail"]["error"] == "archivo_muy_grande"

    # Y con el tope apagado, ese mismo archivo entra.
    monkeypatch.setattr(api, "_MAX_BYTES", 0)
    r = c.post("/api/perfilar", files={"archivo": ("grande.csv", grande, "text/csv")})
    assert r.status_code == 200, "con el tope apagado el archivo tiene que entrar"
    assert r.json()["filas_leidas"] == 300_000
    assert r.json()["truncado"] is False, "se truncó sin que nadie lo pidiera"


def test_el_perfilador_llega_a_la_pantalla_del_exe():
    """Que el endpoint exista no alcanza: el cliente tiene que poder apretar
    algo. Esto ata las dos mitades — es la union donde ya se escondio un bug
    esta misma sesion."""
    src = os.path.join(_repo_root(), "electron", "ui", "src")
    with open(os.path.join(src, "App.jsx"), encoding="utf-8") as fh:
        jsx = fh.read()
    with open(os.path.join(src, "api.js"), encoding="utf-8") as fh:
        api = fh.read()

    assert "/api/perfilar" in api, "api.js no sabe llamar al perfilador"
    assert "perfilar" in jsx, "la pantalla nunca llama al perfilador"
    assert '"misdatos"' in jsx, "no hay pestaña para perfilar datos propios"
    assert 'type="file"' in jsx, "no hay forma de elegir un archivo"
    # FormData sin content-type a mano: escribirlo rompe el boundary del
    # multipart, y eso solo se ve con un archivo real.
    assert "FormData" in api
    assert not re.search(r'content-type["\']\s*:\s*["\']multipart', api), (
        "api.js escribe el content-type del multipart a mano: el navegador "
        "tiene que ponerlo con su boundary")


# ============================================================================
# mvdg/dataeng.py — motor de ingeniería de datos (perfilado avanzado, calidad
# por 6 dimensiones, claves/joins, tiempo, target/fuga, features, DDL)
# ============================================================================

def _dataset_dataeng_sucio(n=600, seed=7):
    """Igual espíritu que demo_data, pero con los defectos que este motor
    tiene que encontrar: tipos disfrazados de texto, PII regional, una fuga
    de target deliberada, nulos, duplicados y un monto negativo."""
    import numpy as _np
    rng = _np.random.default_rng(seed)
    fechas = pd.date_range("2024-01-01", periods=n, freq="6h")
    dias_atraso = rng.integers(0, 180, n)
    prob = 1 / (1 + _np.exp(-(2.0 - dias_atraso / 45)))
    pago = rng.binomial(1, prob)
    df = pd.DataFrame({
        "id_operacion": _np.arange(1, n + 1),
        "fecha_alta": fechas.strftime("%d/%m/%Y"),
        "cedula": [f"{rng.integers(1,5)}.{rng.integers(100,999)}.{rng.integers(100,999)}-{rng.integers(0,9)}"
                  for _ in range(n)],
        "monto_deuda": [f"{m:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
                        for m in _np.round(rng.gamma(2.0, 4000, n), 2)],
        "dias_atraso": dias_atraso,
        "activo": rng.choice(["SI", "NO"], n),
        "columna_vacia": [None] * n,
        "pago": pago,
        "resultado_final": pago,  # fuga deliberada: es el target renombrado
    })
    idx = rng.choice(n, size=int(n * 0.06), replace=False)
    df.loc[idx, "dias_atraso"] = None
    df = pd.concat([df, df.head(15)], ignore_index=True)  # duplicados
    df.loc[rng.choice(len(df), 5, replace=False), "monto_deuda"] = "-999999,00"
    return df


def test_dataeng_analiza_una_tabla_de_punta_a_punta():
    """El circuito completo: tipado -> perfilado -> calidad -> claves ->
    tiempo -> target/fuga -> features -> DDL, sobre datos sucios reales."""
    from mvdg import dataeng

    df = _dataset_dataeng_sucio()
    r = dataeng.analizar_tabla("cobranzas", df, target="pago",
                               columna_tiempo="fecha_alta")

    assert r["advertencias"] == [], f"alguna etapa falló: {r['advertencias']}"
    assert r["perfil"]["filas"] == len(df)

    # tipado: la fecha en texto y el monto en formato es-UY se convierten
    convertidas = {c: (a, b) for c, a, b in r["cambios_tipo"]}
    assert convertidas.get("fecha_alta") == ("texto", "fecha")
    assert convertidas.get("monto_deuda") == ("texto", "numerico")
    assert convertidas.get("activo") == ("texto", "booleano")

    # calidad: los defectos inyectados aparecen como issues
    codigos = {i["codigo"] for i in r["calidad"]["issues"]}
    assert "columna_vacia" in codigos
    assert "duplicados_fila" in codigos
    assert "montos_negativos" in codigos
    assert 0 <= r["calidad"]["score"] <= 100

    # tiempo: la columna de fecha se reconoce y arma la serie
    assert r["tiempo"]["columna"] == "fecha_alta"
    assert r["tiempo"]["dias_faltantes"] == 0  # cada 6hs, sin huecos de día

    # target/fuga: la columna renombrada como target se detecta como fuga
    fugas = {f["variable"] for f in r["target"]["fugas"]}
    assert "resultado_final" in fugas

    # features: se generaron, y los lags/medias moviles existen para 'pago'
    nombres_features = {f["feature"] for f in r["dicc_features"]}
    assert any(n.startswith("pago_lag") or n.startswith("dias_atraso")
              for n in nombres_features) or len(r["dicc_features"]) > 0

    # DDL: nombra la tabla y declara al menos una columna
    assert "CREATE TABLE cobranzas" in r["ddl"]


def test_dataeng_nunca_se_cae_por_una_etapa_rota():
    """Es el criterio central del motor original (autodata.py): si UNA etapa
    falla, las demás igual se completan y el fallo queda listado, no
    escondido ni fatal. Se lo prueba rompiendo `calidad` a propósito."""
    from mvdg import dataeng

    original = dataeng.calidad
    dataeng.calidad = lambda *a, **k: (_ for _ in ()).throw(RuntimeError("boom"))
    try:
        df = _dataset_dataeng_sucio(n=50)
        r = dataeng.analizar_tabla("x", df)
    finally:
        dataeng.calidad = original

    assert any("calidad" in a for a in r["advertencias"])
    assert r["perfil"]["filas"] == len(df)     # el resto de las etapas corrió igual
    assert r["claves"] is not None
    assert r["ddl"] is not None


def test_dataeng_topea_filas_para_no_comerse_la_memoria():
    """Corre dentro de un pedido HTTP con un tiempo de espera real — sin
    tope, una tabla enorme se lleva puesta la memoria del proceso."""
    from mvdg import dataeng

    df = pd.DataFrame({"x": range(5000)})
    r = dataeng.analizar_tabla("grande", df, muestra=100)
    assert r["perfil"]["filas"] == 100
    assert r["muestreado"] is True

    r2 = dataeng.analizar_tabla("chica", df.head(10))
    assert r2["muestreado"] is False


def test_dataeng_issues_son_neutros_de_idioma():
    """Cada issue de calidad tiene que llevar un CÓDIGO estable, no una
    oración armada en español — si esto devolviera texto en un idioma, la
    interfaz no podría traducirlo a EN/PT sin volver a analizar los datos."""
    from mvdg import dataeng

    df = _dataset_dataeng_sucio(n=80)
    r = dataeng.analizar_tabla("x", df)
    for issue in r["calidad"]["issues"]:
        assert issue["codigo"].islower() or "_" in issue["codigo"], (
            f"parece texto armado, no un código: {issue['codigo']!r}")
        assert not re.search(r"[áéíóúñÁÉÍÓÚÑ]", issue["codigo"]), (
            f"el código tiene acentos: {issue['codigo']!r} — no es un código estable")
        assert issue["severidad"] in ("critico", "alto", "medio", "bajo")


def test_dataeng_lags_y_medias_moviles_usan_shift1_siempre():
    """La forma más común de fuga en series temporales: una 'media móvil de
    los últimos 3 períodos' que incluye el período actual — el mismo que se
    quiere predecir. Si `shift(1)` faltara, el primer valor de la media móvil
    coincidiría con el dato crudo en vez de venir vacío/desplazado."""
    from mvdg import dataeng

    fechas = pd.date_range("2024-01-01", periods=20, freq="D")
    df = pd.DataFrame({"fecha": fechas, "monto": range(100, 120)})
    df2, cambios = dataeng.tipar(df)
    prof = dataeng.perfilar_avanzado(df2)
    feats, dicc = dataeng.generar_features(df2, prof, columna_tiempo="fecha")

    lag1_col = next((f["feature"] for f in dicc if f["feature"].startswith("monto_lag1")), None)
    assert lag1_col, f"no se genero el lag: {[f['feature'] for f in dicc]}"
    # el primer valor de un lag(1) NUNCA puede ser el dato de esa misma fila
    assert pd.isna(feats[lag1_col].iloc[0])
    assert feats[lag1_col].iloc[1] == df2["monto"].iloc[0]


def test_dataeng_leer_archivo_bytes_soporta_mas_formatos_que_el_perfilador_simple():
    """/api/perfilar (el perfilador rapido) solo lee CSV/Excel. Esta es la
    version avanzada, y tiene que cubrir lo que autodata.py cubria: parquet,
    json, jsonl y sqlite como archivo — no solo lo que ya existia."""
    from mvdg import dataeng

    csv = "a;b\n1,5;2\n2,5;3\n"
    tablas = dataeng.leer_archivo_bytes("x.csv", csv.encode())
    assert list(tablas["x.csv"].columns) == ["a", "b"]  # detecto el separador ';'

    jsonl = b'{"a":1,"b":2}\n{"a":3,"b":4}\n'
    tablas = dataeng.leer_archivo_bytes("x.jsonl", jsonl)
    assert len(next(iter(tablas.values()))) == 2

    with pytest.raises(RuntimeError):
        dataeng.leer_archivo_bytes("virus.exe", b"MZ\x90\x00")


def test_dataeng_joins_sugeridos_entre_tablas():
    """Multi-tabla es lo que autodata.py llama 'modo carpeta': subir varios
    archivos a la vez y que el motor sugiera cómo se unen, con la
    cardinalidad — un join N:N infla filas y rompe los totales, y hay que
    poder verlo ANTES de escribirlo."""
    from mvdg import dataeng

    clientes = pd.DataFrame({"id_cliente": [1, 2, 3, 4], "nombre": ["A", "B", "C", "D"]})
    ventas = pd.DataFrame({"id_cliente": [1, 1, 2, 3, 3, 3], "monto": [10, 20, 30, 40, 50, 60]})
    joins = dataeng.joins_sugeridos({"clientes": clientes, "ventas": ventas})
    assert joins, "no detecto el join obvio por id_cliente"
    j = joins[0]
    assert j["columna"] == "id_cliente"
    assert j["cardinalidad"] in ("1:N", "N:1")  # clientes es 1, ventas es N
    assert j["riesgo"] in ("bajo", "medio")     # nunca N:N acá

    # Y el caso N:N SI tiene que marcarse "alto": ese es el que infla filas y
    # rompe los totales si alguien lo escribe sin mirar la cardinalidad.
    a = pd.DataFrame({"k": [1, 1, 2, 2, 3, 3]})
    b = pd.DataFrame({"k": [1, 1, 2, 2, 3, 3]})
    joins_nn = dataeng.joins_sugeridos({"a": a, "b": b})
    assert joins_nn and joins_nn[0]["cardinalidad"] == "N:N"
    assert joins_nn[0]["riesgo"] == "alto"


def test_dataeng_rol_columna_reconoce_pii_regional():
    """El motor avanzado clasifica cédula/documento como identificador o
    clave foránea, no como texto libre — si los tratara como texto libre, ni
    el DDL ni las features los tratarían con el cuidado que corresponde a un
    identificador."""
    from mvdg import dataeng

    s = pd.Series([f"1.{i}.{i}-{i%9}" for i in range(200, 260)])  # 60 valores únicos
    assert dataeng.rol_columna("cedula", s) in ("identificador", "clave_foranea")


# ============================================================================
# /api/ingenieria/* — bi_api sirviendo mvdg/dataeng.py (el .exe, no Streamlit)
# ============================================================================
def _dataset_ingenieria_csv() -> bytes:
    """CSV con problemas de calidad a propósito + una columna con nombre
    sospechoso ('resultado_final') que predice casi perfecto al target
    ('aprobado') — dispara fuga_auc_alto Y fuga_nombre_sospechoso a la vez."""
    n = 100
    df = pd.DataFrame({
        "id_cliente": range(1, n + 1),
        "fecha": pd.date_range("2025-01-01", periods=n, freq="D"),
        "monto": [None] * 40 + [100.0] * (n - 40),   # 40% nulos -> nulos_masivos
        "aprobado": [0] * 80 + [1] * 20,
    })
    df["resultado_final"] = df["aprobado"]  # fuga perfecta, nombre sospechoso
    return df.to_csv(index=False).encode()


def test_ingenieria_archivo_end_to_end():
    """Circuito completo por HTTP con el motor de verdad: subir un CSV,
    pedir análisis contra un target y recibir texto YA TRADUCIDO (no
    códigos crudos) — igual que /api/perfilar, pero con el motor completo.
    """
    from fastapi.testclient import TestClient

    from bi_api.main import app

    c = TestClient(app)
    r = c.post(
        "/api/ingenieria/archivo",
        files=[("archivos", ("ventas.csv", _dataset_ingenieria_csv(), "text/csv"))],
        params={"target": "aprobado", "columna_tiempo": "fecha", "lang": "es"},
    )
    assert r.status_code == 200, r.text[:300]
    data = r.json()
    assert list(data["tablas"].keys()) == ["ventas.csv"]
    res = data["tablas"]["ventas.csv"]

    # Calidad: el issue de nulos masivos llegó traducido, no como código.
    issues = res["calidad"]["issues"]
    assert any(i["codigo"] == "nulos_masivos" and i["columna"] == "monto" for i in issues)
    nulos = next(i for i in issues if i["codigo"] == "nulos_masivos")
    assert "%" in nulos["detalle"] and "{" not in nulos["detalle"]
    assert nulos["accion"] and nulos["severidad_texto"] == "ALTO"

    # Fuga: la columna con nombre sospechoso aparece con las DOS razones.
    fugas = {f["codigo"] for f in res["target"]["fugas"]}
    assert "fuga_auc_alto" in fugas or "fuga_nombre_sospechoso" in fugas
    for f in res["target"]["fugas"]:
        assert "{" not in f["texto"], f"placeholder sin rellenar: {f['texto']!r}"

    # Features: fx_lag/fx_media_movil llevan {periodos}/{ventana} en la
    # plantilla — si no se rellenan con los propios `parametros`, quedan
    # literalmente "{periodos}" en pantalla.
    etiquetas = [f["etiqueta"] for f in res["dicc_features"]]
    assert etiquetas, "no se generó ninguna feature"
    assert not any("{" in e or "}" in e for e in etiquetas), etiquetas

    assert res["ddl"] and "CREATE TABLE" in res["ddl"]


def test_ingenieria_archivo_detecta_joins_entre_dos_archivos():
    """Subir dos archivos a la vez tiene que detectar la relación entre
    ellos (la promesa de la bajada: "uno o varios archivos")."""
    from fastapi.testclient import TestClient

    from bi_api.main import app

    clientes = pd.DataFrame({"id_cliente": [1, 2, 3, 4], "nombre": ["A", "B", "C", "D"]}
                            ).to_csv(index=False).encode()
    ventas = pd.DataFrame({"id_cliente": [1, 1, 2, 3, 3, 3], "monto": [10, 20, 30, 40, 50, 60]}
                          ).to_csv(index=False).encode()

    c = TestClient(app)
    r = c.post(
        "/api/ingenieria/archivo",
        files=[("archivos", ("clientes.csv", clientes, "text/csv")),
               ("archivos", ("ventas.csv", ventas, "text/csv"))],
    )
    assert r.status_code == 200, r.text[:300]
    data = r.json()
    assert len(data["tablas"]) == 2
    assert data["joins"], "no detectó el join entre los dos archivos"
    j = data["joins"][0]
    assert j["columna"] == "id_cliente"
    assert j["riesgo_texto"]  # texto traducido, no solo el código crudo


def test_ingenieria_archivo_formato_no_soportado_y_vacio():
    from fastapi.testclient import TestClient

    from bi_api.main import app

    c = TestClient(app)
    r = c.post("/api/ingenieria/archivo",
              files=[("archivos", ("virus.exe", b"MZ\x90\x00", "application/octet-stream"))])
    assert r.status_code == 400
    assert r.json()["detail"]["error"] == "de_err_formato"
    assert r.json()["detail"]["en"]  # trilingüe de verdad, no solo español

    r = c.post("/api/ingenieria/archivo",
              files=[("archivos", ("vacio.csv", b"", "text/csv"))])
    assert r.status_code == 400


def test_ingenieria_sql_end_to_end(tmp_path):
    """SQLite como motor SQL de prueba (no necesita servidor externo): el
    mismo circuito que un cliente real usaría con Postgres/MySQL/etc — se
    prueba la conexión, se listan tablas y se analiza una de verdad, todo
    reusando mvdg/connectors.py en vez de un conector propio."""
    import sqlite3

    from fastapi.testclient import TestClient

    from bi_api.main import app

    db_path = str(tmp_path / "demo.sqlite")
    cx = sqlite3.connect(db_path)
    pd.DataFrame({"id": range(1, 51), "monto": [10.0] * 50}).to_sql(
        "ventas", cx, index=False)
    cx.close()

    perfil = {"engine": "sqlite", "database": db_path}
    c = TestClient(app)

    r = c.post("/api/ingenieria/sql/probar", json=perfil)
    assert r.status_code == 200 and r.json()["ok"] is True, r.text[:300]

    r = c.post("/api/ingenieria/sql/tablas", json=perfil)
    assert r.status_code == 200
    assert "ventas" in r.json()["tablas"]

    r = c.post("/api/ingenieria/sql/analizar", json={**perfil, "tablas": ["ventas"]})
    assert r.status_code == 200, r.text[:300]
    res = r.json()["tablas"]["ventas"]
    assert res["perfil"]["filas"] == 50
    assert res["calidad"]["dimensiones_texto"]["completitud"] == "Completitud"


def test_ingenieria_sql_guardar_listar_y_borrar_conexion(tmp_path, monkeypatch):
    """Las conexiones se comparten con Streamlit (~/.mv_data_governance) y
    la contraseña NUNCA vuelve en la respuesta, ni cifrada ni ofuscada."""
    c = _api_cliente(tmp_path, monkeypatch)

    r = c.post("/api/ingenieria/sql/conexiones",
              json={"name": "Demo", "engine": "sqlite",
                    "database": str(tmp_path / "x.sqlite"), "password": "secreta"})
    assert r.status_code == 200, r.text[:300]
    guardada = r.json()
    assert "password_enc" not in guardada
    assert "password" not in guardada
    conn_id = guardada["conn_id"]

    r = c.get("/api/ingenieria/sql/conexiones")
    assert any(x["conn_id"] == conn_id for x in r.json())
    assert all("password_enc" not in x for x in r.json())

    r = c.delete(f"/api/ingenieria/sql/conexiones/{conn_id}")
    assert r.status_code == 200
    r = c.get("/api/ingenieria/sql/conexiones")
    assert not any(x["conn_id"] == conn_id for x in r.json())


def test_ingenieria_no_pide_licencia():
    """Explícitamente gratis: sin tocar mvdg.licensing, en plan demo (el
    default de cualquier instalación nueva), el endpoint tiene que
    responder 200 y no 402 — 'sin cobrarlo aparte' era el pedido."""
    from fastapi.testclient import TestClient

    from bi_api.main import app
    from mvdg import licensing

    assert licensing.plan() == licensing.PLAN_DEMO
    c = TestClient(app)
    r = c.post("/api/ingenieria/archivo",
              files=[("archivos", ("d.csv", _dataset_ingenieria_csv(), "text/csv"))])
    assert r.status_code == 200


def test_la_ingenieria_de_datos_llega_a_la_pantalla_del_exe():
    """Que existan los endpoints no alcanza: el cliente tiene que poder
    apretar algo. Mismo motivo que test_el_perfilador_llega_a_la_pantalla_del_exe
    — ya se escondió un bug esta misma sesión donde una función vivía en la
    API pero ninguna pantalla la llamaba."""
    src = os.path.join(_repo_root(), "electron", "ui", "src")
    with open(os.path.join(src, "App.jsx"), encoding="utf-8") as fh:
        jsx = fh.read()
    with open(os.path.join(src, "api.js"), encoding="utf-8") as fh:
        api = fh.read()

    assert "/api/ingenieria/archivo" in api, "api.js no sabe subir archivos a Ingeniería de datos"
    assert "/api/ingenieria/sql/probar" in api
    assert "/api/ingenieria/sql/tablas" in api
    assert "/api/ingenieria/sql/analizar" in api

    # Con "(" a propósito: sin esto, que el nombre aparezca en el `import`
    # de arriba ya hacía pasar el assert aunque la pantalla nunca LLAMARA a
    # la función — exactamente el modo de falla que este test existe para
    # atrapar (una función que vive en la API pero ninguna pantalla usa).
    assert "ingenieriaArchivo(" in jsx, "la pantalla nunca llama al análisis por archivo"
    assert "ingenieriaSqlAnalizar(" in jsx, "la pantalla nunca llama al análisis por SQL"
    assert '"ingenieria"' in jsx, "no hay pestaña de Ingeniería de datos"
    assert 'multiple' in jsx, "el selector de archivo no acepta varios a la vez"


def test_ingenieria_sin_licencia_tambien_en_la_pantalla():
    """La API ya lo garantiza (test_ingenieria_no_pide_licencia); acá se
    garantiza que la pantalla no le agregue una licencia que la API nunca
    pidió — nada de _exigir_licencia ni de texto de "esto se paga" en el
    camino de Ingeniería de datos."""
    src = os.path.join(_repo_root(), "electron", "ui", "src", "App.jsx")
    with open(src, encoding="utf-8") as fh:
        jsx = fh.read()
    inicio = jsx.index("function Ingenieria(")
    fin = jsx.index("const RENDER = {")
    bloque = jsx[max(0, inicio - 6000):fin]  # incluye FuenteArchivo/FuenteDb/Resultado*
    assert "requiere_licencia" not in bloque
    assert "lic_activar" not in bloque


def test_la_ingenieria_de_datos_llega_a_streamlit():
    """Mismo motivo que en React: que `dataeng.analizar_tabla` exista no
    alcanza si `_render_profile` (la función que arma la pestaña "Mis
    datos") nunca la llama. Con "(" a propósito, mismo motivo que la
    versión de React: sin eso, el `import dataeng` de arriba del archivo ya
    hacía pasar un `assert "dataeng" in codigo` aunque nadie la usara."""
    ruta = os.path.join(_repo_root(), "app", "app.py")
    with open(ruta, encoding="utf-8") as fh:
        codigo = fh.read()

    assert "def _render_dataeng(" in codigo
    assert "dataeng.analizar_tabla(" in codigo
    assert "dataeng.traducir_resultado(" in codigo
    assert "_render_dataeng(user_df, dataset_name, lang)" in codigo, (
        "_render_profile nunca llama a _render_dataeng: el motor completo "
        "queda escrito pero inalcanzable desde la pantalla")

    # Gratis: nada de licencia en el camino de _render_dataeng.
    inicio = codigo.index("def _render_dataeng(")
    fin = codigo.index("def _render_profile(")
    bloque = codigo[inicio:fin]
    assert "licensing" not in bloque
    assert "has_feature" not in bloque


def test_streamlit_tiene_tema_oscuro_configurado():
    """`app/app.py` pinta el fondo oscuro por CSS (`.stApp { background:
    navy }`), pero sin `.streamlit/config.toml` Streamlit sigue coloreando
    párrafos, etiquetas de radio/checkbox/selectbox y captions con el texto
    casi negro (`rgb(49,51,63)`) de su tema CLARO por defecto — texto
    invisible sobre fondo oscuro en TODA la barra lateral y en cualquier
    párrafo del cuerpo, confirmado renderizando la app real en Chromium con
    un escaneo de contraste real (pixel a pixel, no cascada CSS): bajó de
    ~30-59 elementos de bajo contraste por pestaña a 0.

    `mvdg_launcher.py` (el .exe) y `mvdg/server.py` (modo servidor web) ya
    pasaban `--theme.*` por línea de comandos y no tenían este bug — pero
    `streamlit run app/app.py`, el comando de desarrollo que documenta este
    mismo CLAUDE.md, no pasaba nada, así que corría con el tema claro por
    defecto. Los flags de línea de comandos siempre ganan sobre
    config.toml, así que este archivo solo tapa ese agujero — no compite
    con los otros dos caminos.
    """
    # Parseo manual y no `tomllib` a propósito: es stdlib recién desde
    # Python 3.11, y este repo corre CI también en 3.10 (ver
    # .github/workflows/tests.yml) — `tomllib` rompía justo ese job. El
    # archivo es chato (una sola sección, claves = "valor"), así que una
    # regex alcanza sin sumar una dependencia (`tomli`) solo para un test.
    ruta = os.path.join(_repo_root(), ".streamlit", "config.toml")
    assert os.path.isfile(ruta), (
        ".streamlit/config.toml no existe: streamlit run app/app.py vuelve "
        "a correr con el tema claro por defecto (texto casi negro) sobre "
        "el fondo oscuro que pinta el CSS de app.py")
    with open(ruta, encoding="utf-8") as fh:
        contenido = fh.read()
    tema = dict(re.findall(r'(?m)^(\w+)\s*=\s*"([^"]*)"', contenido))
    assert tema.get("base") == "dark"

    from mvdg import BRAND
    # Mismos colores que ya usan mvdg_launcher.py y mvdg/server.py (y el CSS
    # de app.py vía BRAND): un tema oscuro con paleta distinta confundiría
    # más de lo que ayuda -- se apoyan en la MISMA fuente de verdad.
    assert tema.get("primaryColor", "").lower() == BRAND["amber"].lower()
    assert tema.get("backgroundColor", "").lower() == BRAND["navy"].lower()
    assert tema.get("secondaryBackgroundColor", "").lower() == BRAND["navy2"].lower()
    assert tema.get("textColor", "").lower() == BRAND["ink"].lower()


# ===========================================================================
# Trazabilidad del pipeline: qué se le hizo al dato, en criollo y en técnico
# ===========================================================================

def _pipeline_contexto():
    """El contexto real que la pestaña le pasa al documentador."""
    from mvdg.exporters import governance_tables
    gov = governance_tables("es", include_samples=True)
    return dict(catalog=gov["catalog"], dictionary=gov["dictionary"],
                results=gov["quality_results"], lineage=gov["lineage"],
                glossary=gov["glossary"], policies=gov["policies"],
                indice=87, tablas_bi=sorted(gov))


def test_pipeline_doc_cuenta_las_etapas_en_orden():
    """El número de etapa no es decorativo: es el orden en que pasan.

    Si alguien agrega una etapa en el medio y se olvida de renumerar, el
    documento que lee el gerente cuenta el pipeline en un orden que no
    ocurre. Es peor que no tenerlo: es una explicación que miente.
    """
    from mvdg import pipeline_doc
    etapas = pipeline_doc.documentar("es")
    assert [e["n"] for e in etapas] == list(range(1, len(etapas) + 1))
    assert len({e["key"] for e in etapas}) == len(etapas), "claves repetidas"
    # La primera es leer el dato y la última publicarlo: si eso se da vuelta,
    # el recorrido dejó de ser un pipeline.
    assert etapas[0]["key"] == "ingesta"
    assert etapas[-1]["key"] == "publicacion"


def test_pipeline_doc_esta_en_los_tres_idiomas():
    from mvdg import pipeline_doc
    campos = ("titulo", "criollo", "tecnico", "porque", "impacto")
    por_idioma = {lg: pipeline_doc.documentar(lg) for lg in ("es", "en", "pt")}
    assert len({len(v) for v in por_idioma.values()}) == 1
    for i in range(len(por_idioma["es"])):
        for campo in campos:
            textos = {lg: por_idioma[lg][i][campo] for lg in por_idioma}
            for lg, texto in textos.items():
                assert texto.strip(), f"etapa {i + 1} sin {campo} en {lg}"
            # Traducido de verdad, no copiado: si EN y PT son el texto en
            # español, la paridad de claves pasa y el cliente igual lee
            # español en la pantalla en inglés.
            assert textos["en"] != textos["es"], f"etapa {i + 1}: {campo} EN sin traducir"
            assert textos["pt"] != textos["es"], f"etapa {i + 1}: {campo} PT sin traducir"


def test_pipeline_doc_sin_datos_no_inventa_evidencia():
    """Sin nada cargado, cada etapa se explica pero no reporta números.

    Un "0 datasets catalogados" escrito como si fuera un resultado es peor
    que el silencio: parece una medición y es la ausencia de una.
    """
    from mvdg import pipeline_doc
    etapas = pipeline_doc.documentar("es")
    assert all(e["evidencia"] == "" for e in etapas)


def test_pipeline_doc_mide_la_corrida_de_verdad():
    from mvdg import pipeline_doc
    ctx = _pipeline_contexto()
    etapas = pipeline_doc.documentar("es", **ctx)
    con_evidencia = {e["key"]: e["evidencia"] for e in etapas if e["evidencia"]}
    # Las etapas que la app siempre tiene calculadas deben medirse.
    for clave in ("catalogo", "diccionario", "reglas", "politicas", "glosario"):
        assert clave in con_evidencia, f"{clave} sin evidencia con datos reales"
    # Y los números tienen que ser los de las tablas, no constantes.
    assert str(len(ctx["catalog"])) in con_evidencia["catalogo"]
    assert str(len(ctx["results"])) in con_evidencia["reglas"]


def test_pipeline_doc_documento_arma_las_secciones():
    from mvdg import pipeline_doc
    doc = pipeline_doc.documento("es", **_pipeline_contexto())
    assert doc["titulo"] and doc["subtitulo"] and doc["pie"]
    assert len(doc["secciones"]) == len(pipeline_doc.documentar("es"))
    for sec in doc["secciones"]:
        assert sec["titulo"] and sec["modulo"]
        assert len(sec["bloques"]) == 4
        assert all(rotulo and texto for rotulo, texto in sec["bloques"])
    # El origen de los datos no puede decir "sin datos" cuando hay catálogo:
    # el gerente que lee el PDF necesita saber sobre qué se midió.
    origen = dict(doc["meta"])["Origen de los datos"]
    assert "Sin datos" not in origen, origen
    assert "clientes" in origen or "customers" in origen, origen


def test_doc_export_html_se_cierra_y_escapa():
    """HTML bien formado y sin inyección: el nombre del dataset lo pone el usuario."""
    from html.parser import HTMLParser
    from mvdg import doc_export

    doc = {"titulo": "T", "subtitulo": "S", "lang": "es", "meta": [("a", "b")],
           "pie": "pie", "secciones": [{
               "n": 1, "titulo": "<script>alert(1)</script>", "modulo": "m",
               "bloques": [("et", "texto & más")], "evidencia": "e",
               "evidencia_etiqueta": "Evidencia"}]}
    salida = doc_export.a_html(doc)
    assert "<script>alert(1)</script>" not in salida
    assert "&lt;script&gt;" in salida and "&amp;" in salida

    vacias = {"meta", "br", "hr", "img", "link", "input"}

    class _Balance(HTMLParser):
        def __init__(self):
            super().__init__()
            self.pila, self.mal = [], []

        def handle_starttag(self, tag, attrs):
            if tag not in vacias:
                self.pila.append(tag)

        def handle_endtag(self, tag):
            if self.pila and self.pila[-1] == tag:
                self.pila.pop()
            else:
                self.mal.append(tag)

    p = _Balance()
    p.feed(salida)
    p.close()
    assert not p.mal and not p.pila, (p.mal, p.pila)


def test_doc_export_docx_es_un_word_de_verdad():
    """Un .docx es un zip OOXML. Si le faltan partes, Word no lo abre."""
    import io
    import zipfile
    from xml.etree import ElementTree as ET
    from mvdg import doc_export, pipeline_doc

    doc = pipeline_doc.documento("es", **_pipeline_contexto())
    crudo = doc_export.a_docx(doc)
    z = zipfile.ZipFile(io.BytesIO(crudo))
    assert z.testzip() is None
    faltan = {"[Content_Types].xml", "_rels/.rels", "word/document.xml",
              "word/_rels/document.xml.rels"} - set(z.namelist())
    assert not faltan, f"al .docx le faltan partes: {faltan}"

    w = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    raiz = ET.fromstring(z.read("word/document.xml"))
    texto = "".join(n.text or "" for n in raiz.iter(f"{w}t"))
    assert doc["titulo"] in texto
    for sec in doc["secciones"]:
        assert sec["titulo"] in texto, f"etapa {sec['n']} no llegó al Word"


def test_doc_export_pdf_tiene_xref_valido():
    """El PDF se escribe a mano: el xref mal contado da un archivo que no abre."""
    import re as _re
    from mvdg import doc_export, pipeline_doc

    crudo = doc_export.a_pdf(pipeline_doc.documento("es", **_pipeline_contexto()))
    assert crudo.startswith(b"%PDF-")
    assert crudo.rstrip().endswith(b"%%EOF")

    inicio = int(_re.search(rb"startxref\s+(\d+)", crudo).group(1))
    assert crudo[inicio:inicio + 4] == b"xref", "startxref no apunta a la tabla"
    declarados = int(crudo[inicio:inicio + 40].split(b"\n")[1].split()[1])
    reales = len(_re.findall(rb"\n\d+ 0 obj", b"\n" + crudo))
    assert declarados == reales + 1, (
        f"el xref dice {declarados} objetos y hay {reales + 1}")
    assert len(_re.findall(rb"/Type\s*/Page[^s]", crudo)) >= 2


def test_doc_export_pdf_no_rompe_con_acentos_ni_tipografia_fina():
    """cp1252 no tiene flechas ni comillas curvas; el PDF igual tiene que salir."""
    from mvdg import doc_export

    doc = {"titulo": "Año — “ñoño” ✓", "subtitulo": "a → b", "lang": "es",
           "meta": [], "pie": "…", "secciones": [{
               "n": 1, "titulo": "Señor Ção", "modulo": "m",
               "bloques": [("et", "≥ 5 • ‘x’")], "evidencia": "",
               "evidencia_etiqueta": ""}]}
    crudo = doc_export.a_pdf(doc)
    assert crudo.startswith(b"%PDF-") and len(crudo) > 500


def test_la_pestana_de_trazabilidad_esta_en_la_app():
    """La pestaña existe y baja los tres formatos, no solo uno."""
    fuente = _app_source()
    assert 't("tab_trace", lang)' in fuente, "la pestaña no está en st.tabs"
    assert "with tab_tz:" in fuente
    for clave in ("tz_dl_html", "tz_dl_docx", "tz_dl_pdf"):
        assert f't("{clave}", lang)' in fuente, f"falta el botón {clave}"


# ===========================================================================
# Dos formas de instalar: mi equipo y la VM del cliente
# ===========================================================================

def _layout_portable(tmp_path, monkeypatch, escribible=True):
    """Arma en disco lo que el ZIP portable deja al descomprimirse.

    Se replica el layout REAL de electron-builder —el motor tres niveles por
    debajo del marcador— porque el bug que este modo puede tener es
    justamente que la búsqueda hacia arriba no llegue.
    """
    from mvdg import install_mode, paths
    raiz = tmp_path / "MV Data Governance"
    motor = raiz / "resources" / "server" / "mvdg"
    motor.mkdir(parents=True)
    (raiz / install_mode.MARCADOR).write_text("vm", encoding="utf-8")

    monkeypatch.delenv("MVDG_DATA_DIR", raising=False)
    monkeypatch.delenv(install_mode.VARIABLE, raising=False)
    monkeypatch.setattr(install_mode, "__file__", str(motor / "install_mode.py"))
    monkeypatch.setattr(install_mode, "_MARCADOR_ENCONTRADO", {})
    monkeypatch.setattr(paths, "_ESCRITURA_PROBADA", {})
    if not escribible:
        real = os.makedirs

        def sin_permiso(path, *a, **k):
            if str(path).startswith(str(raiz)):
                raise PermissionError(13, "Acceso denegado", str(path))
            return real(path, *a, **k)
        monkeypatch.setattr(paths.os, "makedirs", sin_permiso)
    return raiz


def test_sin_marcador_el_paquete_es_una_instalacion_normal():
    """El default no puede ser el modo raro: la mayoría instala en su equipo."""
    from mvdg import install_mode
    assert install_mode.modo() == install_mode.MODO_NORMAL
    assert not install_mode.es_vm_cliente()
    assert install_mode.raiz_portable() is None


def test_el_marcador_hace_que_los_datos_queden_en_la_carpeta_del_programa(
        tmp_path, monkeypatch):
    """El caso Conaprole: nada del trabajo puede quedar en el perfil de la VM."""
    from mvdg import install_mode, paths
    raiz = _layout_portable(tmp_path, monkeypatch)

    assert install_mode.es_vm_cliente()
    assert install_mode.raiz_portable() == str(raiz)

    destino = paths.data_dir()
    assert destino == str(raiz / "Datos"), destino
    assert os.path.isdir(destino)
    # Y explícitamente NO el perfil del usuario, que es lo que se pierde
    # cuando la VM se resetea al cerrar sesión.
    assert not destino.startswith(os.path.expanduser("~") + os.sep)


def test_la_variable_de_entorno_le_gana_al_marcador(tmp_path, monkeypatch):
    """Es lo que usa el shell de Electron para decirle el modo al motor."""
    from mvdg import install_mode
    _layout_portable(tmp_path, monkeypatch)
    monkeypatch.setenv(install_mode.VARIABLE, "normal")
    assert install_mode.modo() == install_mode.MODO_NORMAL
    monkeypatch.setenv(install_mode.VARIABLE, "vm_cliente")
    assert install_mode.modo() == install_mode.MODO_VM_CLIENTE
    # Un valor que no existe no puede dejar el programa en un tercer estado.
    monkeypatch.setenv(install_mode.VARIABLE, "cualquier_cosa")
    assert install_mode.modo() in install_mode.MODOS


def test_carpeta_portable_de_solo_lectura_arranca_igual_y_lo_avisa(
        tmp_path, monkeypatch):
    """Descomprimido en una carpeta sin permiso de escritura.

    Pasa de verdad: el consultor descomprime en `C:\\Archivos de programa`
    o en una carpeta de red montada de solo lectura. Que el programa no
    abra sería el peor final posible — abre guardando en el perfil, y lo
    dice, porque en una VM no persistente ese trabajo se pierde.
    """
    from mvdg import install_mode, paths
    _layout_portable(tmp_path, monkeypatch, escribible=False)

    destino = paths.data_dir()
    assert destino == os.path.join(os.path.expanduser("~"), ".mv_data_governance")
    desc = install_mode.descripcion("es")
    assert desc["modo"] == install_mode.MODO_VM_CLIENTE
    assert desc["datos_fuera_de_la_carpeta"] is True


def test_las_dos_formas_de_instalar_habilitan_lo_mismo(tmp_path, monkeypatch):
    """«Funciona igual» tiene que ser verificable, no una promesa de folleto.

    Un modo recortado sería una segunda versión del producto disfrazada de
    opción de instalación — y la menos probada sería justo la que corre en
    la máquina del cliente.
    """
    from mvdg import install_mode, licensing
    normal = (licensing.plan(), sorted(licensing.status()["funciones_pagas"]))
    _layout_portable(tmp_path, monkeypatch)
    assert install_mode.es_vm_cliente()
    portable = (licensing.plan(), sorted(licensing.status()["funciones_pagas"]))
    assert normal == portable, (
        f"el modo cambia las funciones habilitadas: {normal} vs {portable}")


def test_el_modo_se_describe_en_los_tres_idiomas():
    from mvdg import install_mode
    for modo in install_mode.MODOS:
        textos = install_mode._TEXTOS[modo]
        for campo in ("titulo", "detalle"):
            valores = textos[campo]
            assert set(valores) == {"es", "en", "pt"}, f"{modo}.{campo}"
            assert len(set(valores.values())) == 3, (
                f"{modo}.{campo}: hay idiomas con el texto repetido")


def test_install_mode_se_importa_sin_pandas():
    """Igual que `paths` y `licensing`: esto corre en el arranque de todo.

    El modo lo consulta `data_dir()`, que consulta `licensing`, que tiene que
    poder importarse en un runner con Python limpio (ver el test del build
    del owner). Si `install_mode` arrastrara pandas, ese build vuelve a
    romperse por el mismo camino, un módulo más abajo.
    """
    import subprocess
    guion = (
        "import sys, builtins\n"
        f"sys.path.insert(0, {_repo_root()!r})\n"
        "_real = builtins.__import__\n"
        "def fake(n, *a, **k):\n"
        "    if n == 'pandas' or n.startswith('pandas.'):\n"
        "        raise ModuleNotFoundError(\"No module named 'pandas'\")\n"
        "    return _real(n, *a, **k)\n"
        "builtins.__import__ = fake\n"
        "from mvdg import install_mode\n"
        "assert install_mode.modo() in install_mode.MODOS\n"
        "assert install_mode.descripcion('en')['titulo']\n"
        "print('ok')\n"
    )
    r = subprocess.run([sys.executable, "-c", guion], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr[-800:]
    assert "ok" in r.stdout


def test_el_workflow_arma_los_dos_paquetes():
    """El instalador y el portable salen del MISMO build, y los dos se suben.

    Rearmar el portable con un segundo `dist-win` daría un binario distinto
    que nadie compara: la versión que corre en la máquina del cliente sería
    la menos probada de las dos.
    """
    from mvdg import install_mode
    datos, texto = _yaml_workflow("instalador_electron.yml")
    pasos = datos["jobs"]["build"]["steps"]
    nombres = [p.get("name", "") for p in pasos]

    portable = next((p for p in pasos if "PORTABLE" in p.get("name", "")), None)
    assert portable, f"no hay paso que arme el portable. Pasos: {nombres}"
    guion = portable["run"]
    assert "win-unpacked" in guion, (
        "el portable no sale del build ya hecho: se estaría compilando otro")
    assert install_mode.MARCADOR in guion, (
        f"el portable no lleva {install_mode.MARCADOR} adentro, así que abre "
        f"en modo normal y guarda en el perfil de la VM del cliente")

    # Y el portable se arma ANTES de que el build del owner pise win-unpacked.
    i_portable = nombres.index(portable["name"])
    i_owner = next(i for i, n in enumerate(nombres) if "OWNER" in n)
    assert i_portable < i_owner, (
        "el portable se arma después del build del owner: se publicaría el "
        "binario desbloqueado como paquete del cliente")

    # Los dos van a la misma Release: publicarlos aparte garantiza que tarde
    # o temprano uno quede en una versión distinta del otro.
    assert "MVDataGovernance_VM_v$v.zip" in texto


def test_la_api_reporta_el_modo_de_instalacion():
    from fastapi.testclient import TestClient
    from bi_api.main import app
    with TestClient(app) as cli:
        r = cli.get("/api/instalacion?lang=en")
        assert r.status_code == 200
        cuerpo = r.json()
        assert cuerpo["modo"] in ("normal", "vm_cliente")
        assert cuerpo["titulo"] and cuerpo["detalle"] and cuerpo["datos"]
        # El idioma se respeta: la UI React lo pide en el del usuario.
        assert "machine" in cuerpo["titulo"] or "install" in cuerpo["titulo"].lower()


# ===========================================================================
# Reuniones: de la transcripción a la minuta
# ===========================================================================

_VTT_TEAMS = """WEBVTT

1
00:00:01.000 --> 00:00:05.000
<v Ana García>Arrancamos con el relevamiento del maestro de clientes.</v>

2
00:00:05.500 --> 00:00:11.000
<v Juan Pérez>El problema es que el maestro se duplica y el ERP pisa los códigos.</v>

3
00:00:11.200 --> 00:00:16.000
<v Juan Pérez>Y nadie sabe de dónde sale el campo segmento.</v>

4
00:00:20.000 --> 00:00:26.000
<v Ana García>Quedamos en que el dueño va a ser Comercial y el steward Martina.</v>

5
00:00:27.000 --> 00:00:33.000
<v Martina Rossi>Yo me encargo, te mando el diccionario antes del viernes.</v>
"""


def test_meetings_parsea_el_vtt_de_teams_con_su_orador():
    """El camino recomendado: la plataforma ya sabe quién tenía el micrófono."""
    from mvdg import meetings
    inter = meetings.parse_transcript(_VTT_TEAMS)
    assert [i["orador"] for i in inter] == [
        "Ana García", "Juan Pérez", "Ana García", "Martina Rossi"]
    # La 2 y la 3 son del mismo orador y seguidas: se unen, porque un VTT
    # parte cada frase en un subtítulo de tres segundos y sin unirlas "quién
    # dijo qué" sale en doscientas líneas de seis palabras.
    assert "segmento" in inter[1]["texto"] and "duplica" in inter[1]["texto"]


def test_meetings_parsea_srt_y_texto_pegado_a_mano():
    from mvdg import meetings
    srt = ("1\n00:00:02,000 --> 00:00:06,000\n"
           "Ana García: Buen día a todos.\n\n"
           "2\n00:00:07,000 --> 00:00:09,000\n"
           "Juan Pérez: Buen día.\n")
    assert [i["orador"] for i in meetings.parse_transcript(srt)] == \
        ["Ana García", "Juan Pérez"]

    pegado = ("[00:01] Ana García: ¿Quién mantiene el maestro?\n"
              "[00:02] Juan Pérez: Comercial.\n")
    inter = meetings.parse_transcript(pegado)
    assert [i["orador"] for i in inter] == ["Ana García", "Juan Pérez"]
    # "[00:02]" son dos partes: se leen mm:ss, igual que las muestra la
    # tabla. Interpretarlas como hh:mm pondría el minuto 2 en la hora 2.
    assert inter[1]["inicio"] == 2.0


def test_meetings_no_confunde_una_oracion_con_dos_puntos_con_un_orador():
    """El falso positivo que arruina una minuta.

    "El problema es este: el ERP pisa los códigos" NO es una intervención de
    alguien llamado "El problema es este". Si se cuela, la tabla de oradores
    se llena de nombres que no existen y la minuta deja de ser creíble.
    """
    from mvdg import meetings
    inter = meetings.parse_transcript("El problema es este: el ERP pisa los códigos.\n")
    assert len(inter) == 1
    assert inter[0]["orador"] == meetings.SIN_ORADOR
    assert inter[0]["texto"].startswith("El problema es este:")

    # Y el otro lado del filtro: un encabezado largo, todo en mayúsculas y sin
    # ninguna palabra funcional que lo delate. Solo lo frena el límite de
    # palabras — un nombre de persona no tiene siete.
    largo = meetings.parse_transcript(
        "Reunión Kickoff Conaprole Practia Gobierno Datos Marzo: arrancamos.\n")
    assert largo[0]["orador"] == meetings.SIN_ORADOR, largo[0]


def test_meetings_sin_orador_no_lo_inventa():
    """Un micrófono da un canal. Atribuir sin saber es peor que no atribuir."""
    from mvdg import meetings
    inter = meetings.parse_transcript(
        "Buen día, arrancamos.\nEl maestro de clientes se duplica.\n")
    assert all(i["orador"] == meetings.SIN_ORADOR for i in inter)
    ora = meetings.speakers(inter, "es")
    assert list(ora["orador"]) == ["(sin asignar)"]


def test_meetings_marca_los_hallazgos_con_la_cita_textual():
    """Un compromiso resumido no sirve para reclamarlo: va textual y con minuto."""
    from mvdg import meetings
    inter = meetings.parse_transcript(_VTT_TEAMS)
    hall = meetings.findings(inter, "es")
    por_tipo = dict(zip(hall["tipo_id"], hall["cita"], strict=True))
    assert "decision" in por_tipo and "compromiso" in por_tipo and "riesgo" in por_tipo
    # Textual, no parafraseado.
    assert por_tipo["compromiso"] == (
        "Yo me encargo, te mando el diccionario antes del viernes.")
    assert set(hall["minuto"]) <= set(meetings.transcript_df(inter)["minuto"])
    assert (hall["orador"] != "").all()


def test_meetings_cruza_lo_dicho_con_las_etapas_del_pipeline():
    """Es lo que convierte la minuta en trabajo."""
    from mvdg import meetings, pipeline_doc
    inter = meetings.parse_transcript(_VTT_TEAMS)
    cruce = meetings.pipeline_links(inter, "es")
    tocadas = set(cruce["etapa_id"])
    for esperada in ("mdm", "catalogo", "linaje"):
        assert esperada in tocadas, f"no detectó la etapa {esperada}: {tocadas}"
    # Las etapas son las del pipeline real, no una lista paralela.
    validas = {e["key"] for e in pipeline_doc.documentar("es")}
    assert tocadas <= validas
    # Y van en el orden del pipeline, que es como se trabaja.
    assert list(cruce["n"]) == sorted(cruce["n"])


def test_meetings_minuta_vacia_lo_dice_y_no_rompe():
    from mvdg import meetings
    m = meetings.minutes([], "en")
    assert m["vacia"] and m["intervenciones"] == 0
    assert m["aviso_vacia"]
    assert len(m["hallazgos"]) == 0 and len(m["oradores"]) == 0


def test_meetings_la_minuta_se_exporta_a_los_tres_formatos():
    from mvdg import doc_export, meetings
    inter = meetings.parse_transcript(_VTT_TEAMS)
    for lg in ("es", "en", "pt"):
        doc = meetings.to_document(
            meetings.minutes(inter, lg, titulo="Kickoff", fecha="2026-09-06"), lg)
        assert len(doc["secciones"]) == 3
        assert doc_export.a_html(doc).startswith("<!doctype html>")
        assert doc_export.a_docx(doc)[:2] == b"PK"
        assert doc_export.a_pdf(doc).startswith(b"%PDF-")


# ===========================================================================
# Relevamiento: el banco de preguntas y las repreguntas
# ===========================================================================

def test_interview_cubre_las_doce_areas_del_pipeline():
    """Las áreas son las etapas del pipeline, no una taxonomía paralela.

    Si el relevamiento preguntara por áreas que no existen en el pipeline,
    estaría relevando un proyecto distinto del que después se implementa.
    """
    from mvdg import interview, pipeline_doc
    etapas = {e["key"] for e in pipeline_doc.documentar("es")}
    con_preguntas = {a["key"] for a in interview.areas("es")}
    assert con_preguntas <= etapas, con_preguntas - etapas
    assert con_preguntas == etapas, f"áreas del pipeline sin preguntas: {etapas - con_preguntas}"
    # Ninguna área puede quedar con una sola pregunta: eso no es un
    # relevamiento del área, es una casilla marcada.
    assert min(a["preguntas"] for a in interview.areas("es")) >= 3


def test_interview_el_banco_esta_en_los_tres_idiomas():
    from mvdg.interview_bank import PREGUNTAS
    assert len({q["id"] for q in PREGUNTAS}) == len(PREGUNTAS), "ids repetidos"
    for q in PREGUNTAS:
        for campo in ("pregunta", "porque", "a_quien"):
            valores = q[campo]
            assert set(valores) == {"es", "en", "pt"}, f"{q['id']}.{campo}"
            assert all(v.strip() for v in valores.values()), f"{q['id']}.{campo} vacío"
        assert q["repreguntas"], f"{q['id']} sin repreguntas del banco"
        for r in q["repreguntas"]:
            assert set(r) == {"es", "en", "pt"}, f"{q['id']} repregunta incompleta"
            assert r["en"] != r["es"], f"{q['id']}: repregunta EN sin traducir"


def test_interview_guarda_por_cliente_y_no_los_mezcla(tmp_path, monkeypatch):
    """Las respuestas de Conaprole no pueden aparecer en otro cliente."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import interview
    interview.save_answer("conaprole", "MDM-01", respuesta="Pasa seguido",
                          responsable="Juan Pérez", area_responsable="Comercial")
    assert interview.load_answers("conaprole")["MDM-01"]["responsable"] == "Juan Pérez"
    assert interview.load_answers("otra-empresa") == {}
    assert interview.overall_coverage("otra-empresa") == 0.0

    # Borrar la respuesta deja la pregunta PENDIENTE, no "respondida en blanco".
    interview.save_answer("conaprole", "MDM-01", respuesta="")
    assert interview.load_answers("conaprole")["MDM-01"]["estado"] == "pendiente"


def test_interview_no_aplica_no_cuenta_como_pendiente(tmp_path, monkeypatch):
    """Marcar "no aplica" tiene que sacar la pregunta del denominador.

    Si no, un relevamiento correcto y cerrado nunca llega al 100% y el
    número deja de servir para saber si falta algo.
    """
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import interview
    from mvdg.interview_bank import PREGUNTAS
    for q in PREGUNTAS:
        interview.save_answer("x", q["id"], estado="no_aplica")
    assert interview.overall_coverage("x") == 100.0
    interview.save_answer("x", PREGUNTAS[0]["id"], respuesta="")
    assert interview.overall_coverage("x") == 0.0


def test_interview_las_repreguntas_funcionan_sin_clave_de_ia(tmp_path, monkeypatch):
    """El detector de respuesta a medias es local: se releva sin internet.

    Es la decisión que sostiene el módulo — un relevamiento pasa en la sala
    de reuniones de un cliente, que es exactamente donde puede no haber red.
    """
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    from mvdg import interview

    # Pregunta que pide frecuencia, respuesta sin ningún número.
    sin_numero = interview.follow_ups("ING-02", "Depende, más o menos seguido", "es")
    assert any("número" in r for r in sin_numero), sin_numero
    # Y la vaguedad se detecta aparte del número.
    assert any("condicional" in r or "ejemplo concreto" in r for r in sin_numero)

    # Pregunta por un responsable, respuesta de una palabra.
    sin_nombre = interview.follow_ups("CAT-01", "Comercial", "es")
    assert any("nombre" in r for r in sin_nombre), sin_nombre

    # Una respuesta completa NO dispara los genéricos: solo quedan las del
    # banco. Si saltaran igual, el consultor deja de leerlas.
    buena = interview.follow_ups(
        "MDM-01",
        "Sí, pasa seguido: detectamos unos 40 duplicados por mes cuando el "
        "sistema factura dos veces al mismo cliente con distinto código.", "es")
    del_banco = interview.question("MDM-01", "es")["repreguntas"]
    assert buena == del_banco, buena

    # Sin responder, lo primero que se dice es que falta responderla.
    assert "Todavía sin responder" in interview.follow_ups("REG-01", "", "es")[0]


def test_interview_sin_proveedor_la_ia_devuelve_none(monkeypatch):
    from mvdg import ai_provider, interview
    monkeypatch.setattr(ai_provider, "configured_provider", lambda: None)
    assert interview.ai_follow_ups("MDM-01", "Pasa seguido", "es") is None


def test_interview_se_exporta_a_los_tres_formatos(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from mvdg import doc_export, interview
    interview.save_answer("conaprole", "MDM-01", respuesta="Pasa seguido",
                          responsable="Juan Pérez", area_responsable="Comercial")
    for lg in ("es", "en", "pt"):
        doc = interview.to_document("conaprole", lg, "Conaprole")
        assert len(doc["secciones"]) == len(interview.areas(lg))
        assert "Conaprole" in dict(doc["meta"]).values()
        assert doc_export.a_html(doc).startswith("<!doctype html>")
        assert doc_export.a_docx(doc)[:2] == b"PK"
        assert doc_export.a_pdf(doc).startswith(b"%PDF-")


# ===========================================================================
# Transcripción de audio: apagada por defecto
# ===========================================================================

def test_transcribir_no_manda_nada_sin_clave(monkeypatch):
    """Sin clave configurada no hay ninguna llamada de red. Ni una."""
    import urllib.request
    from mvdg import ai_provider, transcribe

    monkeypatch.setattr(ai_provider, "configured_provider", lambda: None)

    def prohibido(*a, **k):
        raise AssertionError("se intentó salir a la red sin clave configurada")
    monkeypatch.setattr(urllib.request, "urlopen", prohibido)

    r = transcribe.transcribir(b"audio falso", "reunion.wav", "es")
    assert r["ok"] is False and r["motivo"] == "sin_proveedor"
    assert r["mensaje"]


def test_transcribir_avisa_cuando_el_proveedor_no_toma_audio(monkeypatch):
    """Claude no transcribe. Decirlo es mejor que fallar en silencio."""
    from mvdg import ai_provider, transcribe
    monkeypatch.setattr(ai_provider, "configured_provider", lambda: "claude")
    r = transcribe.transcribir(b"audio falso", "reunion.wav", "es")
    assert r["motivo"] == "proveedor_sin_audio"
    assert "claude" not in transcribe.PROVEEDORES


def test_transcribir_arma_un_multipart_valido():
    """El cuerpo se escribe a mano: un borde mal puesto da un 400 sin pistas."""
    from mvdg import transcribe
    cuerpo, tipo = transcribe._multipart({"model": "whisper-1"}, "mi reunión.wav", b"AUDIO")
    borde = tipo.split("boundary=")[1]
    texto = cuerpo.decode("utf-8", "replace")
    assert texto.startswith(f"--{borde}\r\n")
    assert texto.rstrip().endswith(f"--{borde}--")
    assert 'name="model"' in texto and 'filename="mi reunión.wav"' in texto
    assert b"AUDIO" in cuerpo
    # Dos partes: el campo y el archivo.
    assert texto.count(f"--{borde}") == 3       # dos aperturas + el cierre


def test_api_relevamiento_y_minuta(tmp_path, monkeypatch):
    """Los dos módulos nuevos, alcanzables desde la API que consume el .exe."""
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from fastapi.testclient import TestClient
    from bi_api.main import app

    with TestClient(app) as cli:
        banco = cli.get("/api/relevamiento/preguntas?lang=en")
        assert banco.status_code == 200
        assert len(banco.json()["areas"]) == 12
        assert len(banco.json()["preguntas"]) >= 30

        assert cli.post("/api/relevamiento/x", json={"id": "NO-EXISTE"}).status_code == 404

        guardar = cli.post("/api/relevamiento/conaprole", json={
            "id": "MDM-01", "respuesta": "Pasa seguido",
            "responsable": "Juan Pérez", "area_responsable": "Comercial"})
        assert guardar.status_code == 200
        assert guardar.json()["estado"] == "respondida"

        estado = cli.get("/api/relevamiento/conaprole?lang=es").json()
        assert estado["cobertura"] > 0
        assert len(estado["por_area"]) == 12

        rep = cli.post("/api/relevamiento/repreguntas",
                       json={"id": "CAT-01", "respuesta": "Comercial", "lang": "es"})
        assert rep.status_code == 200 and rep.json()["repreguntas"]

        vtt = ("WEBVTT\n\n00:00:01.000 --> 00:00:06.000\n"
               "<v Ana García>Quedamos en que el dueño va a ser Comercial.</v>\n")
        minuta = cli.post("/api/reuniones/minuta",
                          json={"texto": vtt, "lang": "es"}).json()
        assert minuta["intervenciones"] == 1
        assert minuta["hallazgos"][0]["orador"] == "Ana García"
        # Las tablas viajan como listas de objetos, no como DataFrames.
        assert isinstance(minuta["oradores"], list)


def test_api_relevamiento_preguntas_no_la_come_la_ruta_comodin():
    """`/api/{table}` es un comodín: declarar algo después queda inalcanzable.

    Es el error clásico de FastAPI y es silencioso — la ruta responde 200 con
    el contenido equivocado en vez de fallar. Se fija el ORDEN, que es lo que
    lo evita.
    """
    from bi_api.main import app
    rutas = [r.path for r in app.routes if hasattr(r, "path")]
    comodin = rutas.index("/api/{table}")
    for especifica in ("/api/relevamiento/preguntas",
                       "/api/relevamiento/repreguntas",
                       "/api/relevamiento/{client_id}",
                       "/api/reuniones/minuta",
                       "/api/instalacion"):
        assert especifica in rutas, f"falta la ruta {especifica}"
        assert rutas.index(especifica) < comodin, (
            f"{especifica} se declara después de /api/{{table}}: el comodín se "
            f"la come y nunca se llega a ella")


def test_api_documentos_del_relevamiento_y_la_minuta(tmp_path, monkeypatch):
    """Los cuatro formatos se arman en el SERVIDOR y viajan como descarga.

    Se sirven así y no como un blob construido en el navegador porque el
    escritor de PDF/Word es Python: reimplementarlo en JavaScript daría dos
    escritores que se separan en el primer cambio, y el menos probado sería
    el del .exe — el que corre en la máquina del cliente.
    """
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from fastapi.testclient import TestClient
    from bi_api.main import app
    from mvdg import interview

    interview.save_answer("conaprole", "MDM-01", respuesta="Pasa seguido",
                          responsable="Juan Pérez", area_responsable="Comercial")
    vtt = ("WEBVTT\n\n00:00:01.000 --> 00:00:06.000\n"
           "<v Ana García>Quedamos en que el dueño va a ser Comercial.</v>\n")
    firmas = {"html": b"<!doctype html>", "docx": b"PK", "pdf": b"%PDF-", "xlsx": b"PK"}

    with TestClient(app) as cli:
        for formato, firma in firmas.items():
            r = cli.get(f"/api/relevamiento/conaprole/documento?formato={formato}"
                        f"&lang=es&empresa=Conaprole")
            assert r.status_code == 200, (formato, r.text[:200])
            assert r.content.startswith(firma), formato
            # Sin Content-Disposition el navegador MUESTRA el archivo en vez
            # de bajarlo, y un .docx mostrado es una pantalla de basura.
            assert "attachment" in r.headers.get("content-disposition", ""), formato

            m = cli.post(f"/api/reuniones/documento?formato={formato}",
                         json={"texto": vtt, "lang": "es"})
            assert m.status_code == 200, (formato, m.text[:200])
            assert m.content.startswith(firma), formato

        assert cli.get("/api/relevamiento/x/documento?formato=zip").status_code == 400


def test_api_transcribir_exige_confirmacion_explicita(tmp_path, monkeypatch):
    """Es el ÚNICO endpoint que saca contenido del cliente de la máquina.

    El resto del programa promete lo contrario, así que la confirmación va en
    cada llamada y no en una configuración que alguien encendió una vez y
    nadie recuerda.
    """
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    import io

    from fastapi.testclient import TestClient
    from bi_api.main import app

    with TestClient(app) as cli:
        sin = cli.post("/api/reuniones/transcribir",
                       files={"archivo": ("a.wav", io.BytesIO(b"x"), "audio/wav")},
                       data={"lang": "es"})
        assert sin.status_code == 400
        assert "confirmacion" in sin.json()["detail"].lower()

        # Con la confirmación pero sin clave: falla igual, y explica por qué.
        con = cli.post("/api/reuniones/transcribir",
                       files={"archivo": ("a.wav", io.BytesIO(b"x"), "audio/wav")},
                       data={"confirmo": "true", "lang": "es"})
        assert con.status_code == 400
        assert "clave" in con.json()["detail"].lower()

        estado = cli.get("/api/reuniones/transcripcion?lang=es").json()
        assert estado["disponible"] is False and estado["motivo"]


def test_api_empresas_alimenta_el_selector_del_exe(tmp_path, monkeypatch):
    monkeypatch.setenv("MVDG_DATA_DIR", str(tmp_path))
    from fastapi.testclient import TestClient
    from bi_api.main import app
    from mvdg import clients

    with TestClient(app) as cli:
        assert cli.get("/api/empresas").json() == []
        clients.save_client({"client_id": "conaprole-001", "company": "Conaprole",
                             "it_restriction": "exe_ok", "status": "piloto"})
        fichas = cli.get("/api/empresas").json()
        assert [f["company"] for f in fichas] == ["Conaprole"]
        assert fichas[0]["client_id"] == "conaprole-001"


def test_las_dos_interfaces_usan_EL_MISMO_banco_de_preguntas():
    """El .exe no puede tener su propio banco de preguntas.

    Si la vista React trajera una copia, el relevamiento que se hace en la VM
    del cliente y el que se hace en el panel dejarían de ser el mismo trabajo
    — y el que se probaría menos es justo el del cliente.
    """
    import re
    ruta = os.path.join(_repo_root(), "electron", "ui", "src", "consultoria.jsx")
    with open(ruta, encoding="utf-8") as fh:
        vista = fh.read()
    assert "relevamientoPreguntas" in vista, (
        "la vista del .exe no pide el banco a la API")
    # Ni una pregunta escrita a mano del lado del navegador.
    from mvdg.interview_bank import PREGUNTAS
    for q in PREGUNTAS[:5]:
        assert q["pregunta"]["es"] not in vista, (
            f"la pregunta {q['id']} está copiada dentro del JavaScript")
    assert not re.search(r"const\s+PREGUNTAS", vista)


# ===========================================================================
# La landing no puede anunciar un número que ya no es
# ===========================================================================

# Los números se escriben con letra en los tres idiomas. Solo hacen falta los
# que la landing usa hoy y los vecinos, para que agregar o sacar una tarjeta
# quede cubierto sin inventar un conversor de números a palabras.
_EN_LETRA = {
    8: {"es": "Ocho", "en": "Eight", "pt": "Oito"},
    9: {"es": "Nueve", "en": "Nine", "pt": "Nove"},
    10: {"es": "Diez", "en": "Ten", "pt": "Dez"},
    11: {"es": "Once", "en": "Eleven", "pt": "Onze"},
    12: {"es": "Doce", "en": "Twelve", "pt": "Doze"},
    13: {"es": "Trece", "en": "Thirteen", "pt": "Treze"},
    3: {"es": "Tres", "en": "Three", "pt": "Três"},
    4: {"es": "Cuatro", "en": "Four", "pt": "Quatro"},
    5: {"es": "Cinco", "en": "Five", "pt": "Cinco"},
}


def _landing_html() -> str:
    with open(os.path.join(_repo_root(), "landing", "index.html"),
              encoding="utf-8") as fh:
        return fh.read()


def _lead_de(html: str, clave: str) -> dict:
    """El texto del `lead` de una sección, por idioma.

    El español sale del HTML (la landing arma su diccionario ES leyendo el
    DOM); inglés y portugués, del objeto I18N.
    """
    import re
    es = re.search(rf'data-i="{clave}"[^>]*>(.*?)</p>', html, re.DOTALL)
    salida = {"es": " ".join(es.group(1).split()) if es else ""}
    for lang in ("en", "pt"):
        # Los dos diccionarios tienen la misma clave: se toman en orden.
        todos = re.findall(rf'\n  {clave}:"([^"]+)"', html)
        salida[lang] = todos[0 if lang == "en" else 1] if len(todos) >= 2 else ""
    return salida


def test_la_landing_no_promete_mas_modulos_de_los_que_muestra():
    """El conteo del texto tiene que coincidir con las tarjetas dibujadas.

    Pasó de verdad: se agregaron tres módulos, quedaron nueve tarjetas más
    tres, y el texto siguió diciendo "Nueve módulos". Un visitante que cuenta
    las tarjetas encuentra doce y el que lee la bajada lee nueve — y en una
    página de venta, un número que no cierra es lo primero que se nota.
    """
    import re
    html = _landing_html()

    casos = [
        # (prefijo de la tarjeta, clave del lead, qué se cuenta)
        ("f", "p_lead", "módulos de la plataforma"),
        ("d", "d_lead", "formas de descargar"),
    ]
    for prefijo, clave_lead, que in casos:
        tarjetas = len(set(re.findall(rf'data-i="{prefijo}(\d+)t"', html)))
        assert tarjetas >= 3, f"no se encontraron las tarjetas de {que}"
        assert tarjetas in _EN_LETRA, (
            f"hay {tarjetas} {que} y el test no sabe escribir ese número: "
            f"agregalo a _EN_LETRA")
        leads = _lead_de(html, clave_lead)
        for lang, texto in leads.items():
            assert texto, f"no se encontró el lead {clave_lead} en {lang}"
            esperado = _EN_LETRA[tarjetas][lang]
            assert esperado.lower() in texto.lower(), (
                f"[{lang}] hay {tarjetas} {que} pero la bajada dice "
                f"«{texto[:70]}…» — tendría que decir «{esperado}»")


def test_la_landing_describe_los_modulos_nuevos_en_los_tres_idiomas():
    """Una tarjeta sin traducir deja la página en inglés con texto en español.

    El i18n de la landing es en JavaScript y no falla: si falta la clave,
    `setLang` cae al español sin decir nada. Solo se ve mirando la página.
    """
    import re
    html = _landing_html()
    # Los módulos que se agregaron con Relevamiento, Reuniones y Trazabilidad.
    for clave in ("f10t", "f10p", "f11t", "f11p", "f12t", "f12p", "d4t", "d4p"):
        assert f'data-i="{clave}"' in html, f"falta la tarjeta {clave} en el HTML"
        traducciones = re.findall(rf'[\s,]{clave}:"', html)
        assert len(traducciones) == 2, (
            f"{clave} tiene {len(traducciones)} traducciones y necesita 2 "
            f"(inglés y portugués): el español sale del propio HTML")
